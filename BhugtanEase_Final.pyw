"""
BhugtanEase — Restaurant Billing & Inventory
Double click karke chalao | Python + SQLite + Tkinter
"""
import sys, os

# .pyw file hai — pythonw.exe se seedha chalta hai, koi flash nahi

import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import sqlite3, datetime, os, sys, json, hmac, hashlib, binascii, secrets

# ════════════════════════════════════════════════════════════
#  EXCEL EXPORT HELPER
# ════════════════════════════════════════════════════════════
def export_to_excel(filename_prefix, headers, rows, parent=None):
    """Generic Excel export — openpyxl use karta hai."""
    import os, datetime as _dt
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        try:
            import subprocess, sys
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except Exception:
            messagebox.showerror("Error", "openpyxl install nahi hua.\npip install openpyxl chalao CMD mein.", parent=parent)
            return

    from tkinter import filedialog
    import tkinter as _tk
    today = _dt.date.today().strftime("%Y-%m-%d")
    fname = f"{filename_prefix}_{today}.xlsx"
    # FIX: Blank chhota 'tk' window band karo — parent ensure karo
    if parent is None:
        parent = _tk._default_root
    path  = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel File","*.xlsx")],
        initialfile=fname,
        title="Excel File Save Karo",
        parent=parent
    )
    if not path: return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_prefix[:31]

    # Header style
    hdr_fill = PatternFill("solid", fgColor="8B0000")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = hdr_align; cell.border = border

    # Data rows
    for ri, row in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="FFF5F5") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.fill   = fill
            cell.alignment = Alignment(horizontal="center")
            if ci == 2:  # second column generally name — left align
                cell.alignment = Alignment(horizontal="left")

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Freeze header
    ws.freeze_panes = "A2"

    try:
        wb.save(path)
        messagebox.showinfo("✅ Excel Save Ho Gaya!", f"File save ho gayi:\n{path}", parent=parent)
        try: os.startfile(path)
        except: pass
    except Exception as e:
        messagebox.showerror("Error", f"Save nahi hua:\n{e}", parent=parent)



# ════════════════════════════════════════════════════════════
#  DATABASE
# ════════════════════════════════════════════════════════════
# EXE (PyInstaller) aur .pyw dono ke liye sahi path
if getattr(sys, "frozen", False):
    # PyInstaller EXE — Program Files me write access nahi milta,
    # isliye DB LOCALAPPDATA me store karo (per-user writable location)
    BASE_DIR = os.path.join(os.environ.get("LOCALAPPDATA", os.path.expanduser("~")), "BhugtanEase_Restaurant")
else:
    # Normal .pyw run
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

os.makedirs(BASE_DIR, exist_ok=True)
DB_PATH  = os.path.join(BASE_DIR, "data.db")

# SPEED: Settings cache — bar bar DB hit avoid karo
_settings_cache: dict = {}

def _load_settings_cache():
    """Startup pe ek baar saari settings load karo."""
    global _settings_cache
    try:
        c = db()
        rows = c.execute("SELECT key, value FROM settings").fetchall()
        _settings_cache = {r["key"]: r["value"] for r in rows}
    except Exception:
        _settings_cache = {}

# SPEED: Persistent connection — bar bar naya connection banana avoid karo
_db_conn = None

def db():
    global _db_conn
    if _db_conn is None:
        _db_conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        _db_conn.row_factory = sqlite3.Row
        _db_conn.execute("PRAGMA foreign_keys = ON")
        _db_conn.execute("PRAGMA journal_mode=WAL")      # SPEED: Write-Ahead Logging
        _db_conn.execute("PRAGMA synchronous=NORMAL")    # SPEED: Disk writes fast
        _db_conn.execute("PRAGMA cache_size=10000")      # SPEED: 10MB memory cache
        _db_conn.execute("PRAGMA temp_store=MEMORY")     # SPEED: Temp tables RAM mein
    return _db_conn

def init_db():
    c = db()
    # Tables banao
    c.executescript("""
    CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);
    CREATE TABLE IF NOT EXISTS categories (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE);
    CREATE TABLE IF NOT EXISTS menu_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
        category_id INTEGER, price REAL NOT NULL, is_available INTEGER DEFAULT 1,
        FOREIGN KEY (category_id) REFERENCES categories(id));
    CREATE TABLE IF NOT EXISTS raw_materials (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL UNIQUE,
        unit TEXT NOT NULL, current_stock REAL DEFAULT 0,
        min_stock REAL DEFAULT 0, cost_per_unit REAL DEFAULT 0);
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT, table_number TEXT,
        customer_name TEXT, subtotal REAL, tax_percent REAL DEFAULT 5,
        tax_amount REAL, discount_percent REAL DEFAULT 0,
        discount_amount REAL DEFAULT 0, total_amount REAL,
        payment_method TEXT DEFAULT 'Cash',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT, order_id INTEGER,
        item_name TEXT, quantity INTEGER, price_at_order REAL, item_total REAL);
    CREATE TABLE IF NOT EXISTS inventory_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, material_id INTEGER,
        type TEXT, qty REAL, unit_price REAL DEFAULT 0, total_cost REAL DEFAULT 0,
        note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS suppliers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        phone TEXT, address TEXT, gst TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS purchases (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier_id INTEGER,
        supplier_name TEXT,
        invoice_no TEXT,
        payment_method TEXT DEFAULT 'Cash',
        payment_status TEXT DEFAULT 'Paid',
        total_amount REAL DEFAULT 0,
        note TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    CREATE TABLE IF NOT EXISTS purchase_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        purchase_id INTEGER,
        material_id INTEGER,
        material_name TEXT,
        unit TEXT,
        quantity REAL,
        unit_price REAL,
        total_cost REAL);
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT DEFAULT 'staff',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    """)

    # SPEED FIX: agar already initialized hai toh sample data dobara mat daalo
    already = c.execute("SELECT value FROM settings WHERE key='_db_init'").fetchone()

    # Migration: purane DB mein naye columns add karo
    try:
        c.execute("ALTER TABLE inventory_log ADD COLUMN unit_price REAL DEFAULT 0")
        c.commit()
    except: pass
    try:
        c.execute("ALTER TABLE inventory_log ADD COLUMN total_cost REAL DEFAULT 0")
        c.commit()
    except: pass

    # FIX: Agar shop_name set hai lekin setup_done=0 hai (DB migrate/copy case) — auto-fix karo
    if already:
        shop_row = c.execute("SELECT value FROM settings WHERE key='shop_name'").fetchone()
        setup_row = c.execute("SELECT value FROM settings WHERE key='setup_done'").fetchone()
        if shop_row and shop_row["value"] and shop_row["value"] not in ("Mera Restaurant", "") \
                and (not setup_row or setup_row["value"] != "1"):
            c.execute("INSERT OR REPLACE INTO settings VALUES ('setup_done','1')")
            c.commit()
        return

    # Default settings — ek hi executemany call
    c.executemany("INSERT OR IGNORE INTO settings VALUES (?,?)", [
        ("shop_name","Mera Restaurant"), ("address","Lucknow"),
        ("phone","9876543210"), ("gst","GSTIN123456"),
        ("tax","5"), ("currency","₹"), ("setup_done","0"),
        ("_db_init","1"),  # flag: init ho chuka hai
    ])
    # Categories batch
    c.executemany("INSERT OR IGNORE INTO categories (name) VALUES (?)",
        [("Starters",),("Main Course",),("Breads",),("Drinks",),("Desserts",)])

    # Sample items — category IDs ek baar fetch karo
    cat_ids = {r["name"]: r["id"] for r in
               c.execute("SELECT id,name FROM categories").fetchall()}
    samples = [
        ("Paneer Tikka","Starters",180), ("Chicken Tikka","Starters",220),
        ("Dal Makhani","Main Course",160), ("Paneer Butter Masala","Main Course",200),
        ("Chicken Curry","Main Course",240), ("Butter Naan","Breads",40),
        ("Tandoori Roti","Breads",30), ("Lassi","Drinks",80),
        ("Cold Drink","Drinks",50), ("Gulab Jamun","Desserts",80),
    ]
    c.executemany("INSERT OR IGNORE INTO menu_items (name,category_id,price) VALUES (?,?,?)",
        [(n, cat_ids[cat], p) for n, cat, p in samples if cat in cat_ids])

    # Raw materials batch
    c.executemany(
        "INSERT OR IGNORE INTO raw_materials (name,unit,current_stock,min_stock,cost_per_unit) VALUES (?,?,?,?,?)",
        [("Atta","kg",10,2,35), ("Basmati Rice","kg",15,3,80),
         ("Paneer","kg",5,1,320), ("Chicken","kg",8,2,200),
         ("Tomato","kg",10,2,30), ("Onion","kg",10,2,25),
         ("Cooking Oil","litre",10,2,130), ("Milk","litre",5,1,55)])

    c.commit()

    # SPEED: Indexes banao — queries fast hongi
    c.executescript("""
    CREATE INDEX IF NOT EXISTS idx_menu_cat    ON menu_items(category_id, is_available);
    CREATE INDEX IF NOT EXISTS idx_menu_name   ON menu_items(name);
    CREATE INDEX IF NOT EXISTS idx_orders_date ON orders(created_at);
    CREATE INDEX IF NOT EXISTS idx_inv_mat     ON inventory_log(material_id);
    CREATE INDEX IF NOT EXISTS idx_pur_items   ON purchase_items(purchase_id);
    """)
    c.commit()

def gset(key, default=""):
    # SPEED: pehle cache check karo, DB hit avoid hoga
    if _settings_cache:
        return _settings_cache.get(key, default)
    try:
        c = db()
        r = c.execute("SELECT value FROM settings WHERE key=?",(key,)).fetchone()
        return r["value"] if r else default
    except Exception:
        return default

def sset(key, val):
    _settings_cache[key] = val  # cache update karo
    c = db(); c.execute("INSERT OR REPLACE INTO settings VALUES (?,?)",(key,val)); c.commit()

# ════════════════════════════════════════════════════════════
#  LICENSE ENGINE  (Admin generator se sync — DO NOT CHANGE)
# ════════════════════════════════════════════════════════════
_SEC = b"BhugtanEase@2026#RestaurantBilling$SecureKey!"

# --- Base36 helpers (Admin ke saath match) ---
def _b36(n):
    chars = "0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if n == 0: return "0"
    s = ""
    while n:
        s = chars[n % 36] + s
        n //= 36
    return s

def _b36dec(s): return int(s, 36)

# Sign: 6-char HMAC (Admin ke saath match — 6 hi hai)
def _sign(d): return hmac.new(_SEC, d.encode(), hashlib.sha256).hexdigest()[:6].upper()

# Shop naam ka 4-char hash — Admin ke saath EXACTLY match karna chahiye
def _shop_hash(shop_name):
    normalized = shop_name.strip().upper().replace(" ", "")
    h = hmac.new(_SEC, normalized.encode(), hashlib.sha256).hexdigest()
    return h[:4].upper()

# Encode/Decode: Shop hash + expiry + type code
def _enc(p):
    shop_h = _shop_hash(p["s"])
    expiry = datetime.date.fromisoformat(p["e"])
    epoch  = datetime.date(2024, 1, 1)
    days   = (expiry - epoch).days
    tcode  = {"TRIAL":"T","MONTHLY":"M","QUARTERLY":"Q","HALF_YEAR":"H",
               "YEARLY":"Y","2YEAR":"2","LIFETIME":"L","CUSTOM":"C"}.get(p["t"].upper(),"M")
    return shop_h + _b36(days).zfill(4) + tcode

def _dec(enc):
    if len(enc) < 9: raise ValueError("Too short")
    shop_h   = enc[:4]
    days_b36 = enc[4:-1]
    tcode    = enc[-1]
    days     = _b36dec(days_b36)
    epoch    = datetime.date(2024, 1, 1)
    expiry   = epoch + datetime.timedelta(days=days)
    tmap     = {"T":"Trial","M":"Monthly","Q":"Quarterly","H":"Half Year",
                "Y":"Yearly","2":"2 Year","L":"Lifetime","C":"Custom"}
    return {"h": shop_h, "e": str(expiry), "t": tmap.get(tcode, "Monthly")}

# Serial format: BE-{enc}-{sign}
def verify_serial(serial, check_shop=False):
    s = serial.strip().upper().replace(" ","")
    # Dashes hata ke parse karo
    s_nodash = s.replace("-","")
    if not s_nodash.startswith("BE") or len(s_nodash) < 16:
        return {"valid":False,"message":"Format galat hai."}
    sign = s_nodash[-6:]
    enc  = s_nodash[2:-6]
    if _sign(enc) != sign:
        return {"valid":False,"message":"Serial tampered hai!"}
    try: p = _dec(enc)
    except (ValueError, Exception): return {"valid":False,"message":"Decode nahi hua."}
    today  = datetime.date.today()
    expiry = datetime.date.fromisoformat(p["e"])
    days   = (expiry - today).days
    if days < 0:
        return {"valid":False,"status":"EXPIRED","shop_hash":p["h"],"expiry":p["e"],
                "message":f"License {abs(days)} din pehle expire ho gayi."}
    # SHOP NAME MATCH CHECK — full name ka hash compare karo
    if check_shop:
        saved_shop = gset("shop_name", "").strip()
        if saved_shop and p["h"]:
            expected_h = _shop_hash(saved_shop)
            if expected_h != p["h"]:
                return {"valid":False,
                        "message":"Yeh serial is shop ke liye nahi hai!\n"
                                  "Serial kisi aur shop ke liye generate hua tha."}
    return {"valid":True,"shop_hash":p["h"],"expiry":p["e"],"type":p["t"],
            "days":days,"message":f"Valid. {days} din bache."}

def save_license(serial):
    # Pehle basic verify karo (shop check ke bina)
    r = verify_serial(serial, check_shop=False)
    if not r["valid"]: return r
    # Shop check sirf tab karo jab setup ho chuka ho (setup_done=1)
    # First-time activation mein shop naam abhi set nahi hota
    if gset("setup_done", "0") == "1":
        saved_shop = gset("shop_name", "").strip()
        if saved_shop:
            r2 = verify_serial(serial, check_shop=True)
            if not r2["valid"]: return r2
    s = serial.strip().upper()
    blob = json.dumps({"s": s, "chk": _sign(s)})
    sset("_license_blob", blob)
    return r

def load_license():
    blob = gset("_license_blob", "")
    if not blob:
        return {"valid":False,"message":"License nahi mili."}
    try:
        d = json.loads(blob)
        if _sign(d["s"]) != d["chk"]:
            return {"valid":False,"message":"License corrupt hai!"}
        # Shop check sirf tab karo jab setup ho chuka ho
        check = gset("setup_done", "0") == "1"
        return verify_serial(d["s"], check_shop=check)
    except (json.JSONDecodeError, KeyError):
        return {"valid":False,"message":"License data error."}

# ════════════════════════════════════════════════════════════
#  THERMAL BILL PREVIEW WINDOW
# ════════════════════════════════════════════════════════════
def _show_thermal_preview(parent, oid, bill_text, printer_name, bills_dir):
    """Thermal bill ka preview dikhao — phir print karo."""
    prev = tk.Toplevel(parent)
    prev.title(f"🖨️ Bill #{oid:04d} — Preview")
    prev.configure(bg=BG); prev.grab_set(); prev.resizable(True, True)
    prev.update_idletasks()
    sw = prev.winfo_screenwidth(); sh = prev.winfo_screenheight()
    prev.geometry(f"480x600+{(sw-480)//2}+{(sh-600)//2}")

    hdr = tk.Frame(prev, bg=DKRED, pady=8); hdr.pack(fill="x")
    tk.Label(hdr, text=f"👁️  Bill #{oid:04d} Preview", font=FH, bg=DKRED, fg=WHITE).pack()
    tk.Label(hdr, text="Yahan bill dekho — phir neeche se print karo",
             font=FS, bg=DKRED, fg="#ffcccc").pack()
    tk.Frame(prev, bg=RED, height=2).pack(fill="x")

    # Scrollable text preview
    pf = tk.Frame(prev, bg="#1e1e1e", padx=16, pady=16); pf.pack(fill="both", expand=True)
    txt = tk.Text(pf, font=("Courier New", 10), bg="#1e1e1e", fg="#f0f0f0",
                  relief="flat", bd=0, wrap="none", state="normal")
    vsb = tk.Scrollbar(pf, orient="vertical",   command=txt.yview)
    hsb = tk.Scrollbar(pf, orient="horizontal", command=txt.xview)
    txt.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
    vsb.pack(side="right",  fill="y")
    hsb.pack(side="bottom", fill="x")
    txt.pack(fill="both", expand=True)
    txt.insert("1.0", bill_text)
    txt.config(state="disabled")

    # Fixed bottom buttons
    tk.Frame(prev, bg=BORD, height=1).pack(fill="x")
    bf = tk.Frame(prev, bg=BG, padx=20, pady=12); bf.pack(fill="x")

    def do_print_now():
        ok, msg = print_thermal(bill_text, printer_name)
        if ok:
            messagebox.showinfo("✅ Print Ho Gaya!", f"Bill #{oid:04d} print ho gaya!", parent=prev)
            prev.destroy()
        else:
            txt_path = os.path.join(bills_dir, f"BILL_{oid:04d}.txt")
            with open(txt_path, "w", encoding="utf-8") as f:
                f.write(bill_text)
            ans = messagebox.askokcancel("⚠️ Printer Error",
                f"Direct print nahi hua:\n{msg}\n\nNotepad mein kholen?\n(Wahan se Ctrl+P karein)", parent=prev)
            if ans:
                try: os.startfile(txt_path)
                except: pass

    def save_txt():
        txt_path = os.path.join(bills_dir, f"BILL_{oid:04d}.txt")
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(bill_text)
        messagebox.showinfo("💾 Saved!", f"File save hua:\n{txt_path}", parent=prev)
        try: os.startfile(txt_path)
        except: pass

    btn(bf, "🖨️  Print Karo",   do_print_now, bg=RED,   py=10).pack(side="left",  fill="x", expand=True, padx=(0,4))
    btn(bf, "💾  TXT Save",     save_txt,      bg=DARK,  py=10).pack(side="left",  fill="x", expand=True, padx=(0,4))
    btn(bf, "✖  Band Karo",    prev.destroy,  bg=MUTED, py=10).pack(side="right", fill="x", expand=True)


# ════════════════════════════════════════════════════════════
#  PDF BILL GENERATOR
# ════════════════════════════════════════════════════════════
def generate_pdf(order_id, order, items, settings, path):
    try:
        from reportlab.lib.pagesizes import A5
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    except ImportError:
        messagebox.showerror("Error","reportlab install karo:\npip install reportlab")
        return None

    doc = SimpleDocTemplate(path, pagesize=A5,
          rightMargin=10*mm, leftMargin=10*mm,
          topMargin=10*mm, bottomMargin=10*mm)
    styles = getSampleStyleSheet()
    story  = []
    cur    = settings.get("currency","₹")

    def sty(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    from reportlab.lib.enums import TA_CENTER
    t_c = sty("tc", fontSize=16, fontName="Helvetica-Bold",
              textColor=colors.HexColor("#CC1111"), alignment=TA_CENTER, spaceAfter=2*mm)
    s_c = sty("sc", fontSize=8, fontName="Helvetica",
              textColor=colors.HexColor("#555555"), alignment=TA_CENTER, spaceAfter=1*mm)
    b_c = sty("bc", fontSize=10, fontName="Helvetica-Bold",
              alignment=TA_CENTER, spaceAfter=2*mm)
    ft  = sty("ft", fontSize=8, textColor=colors.HexColor("#666666"), alignment=TA_CENTER)

    story.append(Paragraph(settings.get("shop_name","Restaurant"), t_c))
    story.append(Paragraph(settings.get("address",""), s_c))
    story.append(Paragraph(f"Ph: {settings.get('phone','')}  |  GST: {settings.get('gst','')}", s_c))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor("#CC1111")))
    story.append(Spacer(1,2*mm))
    story.append(Paragraph("BILL / INVOICE", b_c))

    info = [
        [f"Bill No: #{order_id:04d}", f"Date: {order['created_at'][:16]}"],
        [f"Table: {order['table_number'] or 'Takeaway'}", f"Payment: {order['payment_method']}"],
    ]
    if order.get("customer_name"):
        info.append([f"Customer: {order['customer_name']}", ""])
    it = Table(info, colWidths=[65*mm,65*mm])
    it.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8),
                             ("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1)]))
    story.append(it); story.append(Spacer(1,2*mm))
    story.append(HRFlowable(width="100%",thickness=0.5,color=colors.grey))
    story.append(Spacer(1,1*mm))

    hdr = [["#","Item","Qty","Rate","Amount"]]
    rows = [[str(i+1),x["item_name"],str(x["quantity"]),
             f"{cur}{x['price_at_order']:.2f}",f"{cur}{x['item_total']:.2f}"]
            for i,x in enumerate(items)]
    tbl = Table(hdr+rows, colWidths=[8*mm,65*mm,12*mm,22*mm,23*mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#CC1111")),
        ("TEXTCOLOR",(0,0),(-1,0),colors.white),
        ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
        ("FONTNAME",(0,1),(-1,-1),"Helvetica"),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#f8f8f8")]),
        ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#cccccc")),
        ("ALIGN",(2,0),(-1,-1),"CENTER"),("ALIGN",(3,1),(4,-1),"RIGHT"),
        ("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2),
    ]))
    story.append(tbl); story.append(Spacer(1,2*mm))
    story.append(HRFlowable(width="100%",thickness=0.5,color=colors.grey))

    sumdata = [["Subtotal:", f"{cur}{order['subtotal']:.2f}"]]
    if order.get("discount_amount",0) > 0:
        sumdata.append([f"Discount ({order['discount_percent']:.0f}%):",
                        f"-{cur}{order['discount_amount']:.2f}"])
    sumdata.append([f"GST ({order['tax_percent']:.0f}%):", f"{cur}{order['tax_amount']:.2f}"])
    st = Table(sumdata, colWidths=[100*mm,30*mm])
    st.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),"Helvetica"),("FONTSIZE",(0,0),(-1,-1),8),
                             ("ALIGN",(0,0),(0,-1),"RIGHT"),("ALIGN",(1,0),(1,-1),"RIGHT"),
                             ("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1)]))
    story.append(st)
    story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#CC1111")))
    total_t = Table([[f"TOTAL:", f"{cur}{order['total_amount']:.2f}"]], colWidths=[100*mm,30*mm])
    total_t.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
                                  ("FONTSIZE",(0,0),(-1,-1),12),
                                  ("TEXTCOLOR",(0,0),(-1,-1),colors.HexColor("#CC1111")),
                                  ("ALIGN",(0,0),(0,-1),"RIGHT"),("ALIGN",(1,0),(1,-1),"RIGHT"),
                                  ("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]))
    story.append(total_t)
    story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#CC1111")))
    story.append(Spacer(1,4*mm))
    story.append(Paragraph("Dhanyawad! Phir padharein 🙏", ft))
    doc.build(story)
    return path


# ════════════════════════════════════════════════════════════
#  THERMAL BILL GENERATOR (58mm / 80mm printer)
# ════════════════════════════════════════════════════════════
def generate_thermal_bill(order_id, order, items, settings, width_mm=80):
    """
    Pure text thermal bill generate karo.
    width_mm: 58 ya 80 — character width choose karta hai.
    Returns: text string
    """
    W = 32 if width_mm == 58 else 48   # char width
    cur = settings.get("currency", "₹")
    shop  = settings.get("shop_name","Restaurant")
    addr  = settings.get("address","")
    phone = settings.get("phone","")
    gst   = settings.get("gst","")

    def line(ch="-"):  return ch * W
    def center(txt):   return txt.center(W)
    def rjust2(l,r):
        gap = W - len(l) - len(r)
        return l + " "*max(1,gap) + r

    L = []
    L.append(line("="))
    L.append(center(shop))
    if addr:  L.append(center(addr))
    if phone: L.append(center(f"Ph: {phone}"))
    if gst:   L.append(center(f"GST: {gst}"))
    L.append(line("="))
    L.append(rjust2(f"Bill #: {order_id:04d}",
                    f"Date: {order.get('created_at','')[:10]}"))
    tbl = order.get("table_number","")
    cust = order.get("customer_name","")
    if tbl:  L.append(f"Table : {tbl}")
    if cust: L.append(f"Name  : {cust}")
    L.append(line("-"))
    L.append(rjust2("Item (Qty x Rate)", "Total"))
    L.append(line("-"))
    for it in items:
        name = it["item_name"][:W-14]
        detail = f"  {it['quantity']}x{cur}{it['price_at_order']:.0f}"
        tot    = f"{cur}{it['item_total']:.2f}"
        L.append(name)
        L.append(rjust2(detail, tot))
    L.append(line("-"))
    sub  = order.get("subtotal",0)
    disc = order.get("discount_amount",0)
    tax  = order.get("tax_amount",0)
    total= order.get("total_amount",0)
    pay  = order.get("payment_method","Cash")
    if disc > 0:
        L.append(rjust2("Subtotal:", f"{cur}{sub:.2f}"))
        L.append(rjust2(f"Discount({order.get('discount_percent',0):.0f}%):", f"-{cur}{disc:.2f}"))
    if tax > 0:
        L.append(rjust2(f"GST({order.get('tax_percent',0):.0f}%):", f"{cur}{tax:.2f}"))
    L.append(line("="))
    L.append(rjust2("TOTAL:", f"{cur}{total:.2f}"))
    L.append(rjust2("Payment:", pay))
    L.append(line("="))
    L.append(center("Aapka dhanyavaad!"))
    L.append(center("Please visit again :)"))
    L.append(line())
    L.append("")
    L.append("")   # paper feed
    return "\n".join(L)


def print_thermal(text, printer_name=None):
    """
    Thermal text bill print karo.
    Windows pe win32print use karo, fallback notepad silent print.
    """
    import tempfile, subprocess
    # Temp text file banao
    tmp = tempfile.NamedTemporaryFile(suffix=".txt", delete=False, mode="w", encoding="cp1252", errors="replace")
    tmp.write(text); tmp.flush(); tmp.close()
    path = tmp.name

    if sys.platform == "win32":
        try:
            import win32print, win32api
            pname = printer_name or win32print.GetDefaultPrinter()
            # Raw print (ESC/POS compatible)
            hPrinter = win32print.OpenPrinter(pname)
            try:
                hJob = win32print.StartDocPrinter(hPrinter, 1, ("BhugtanEase Bill", None, "RAW"))
                win32print.StartPagePrinter(hPrinter)
                win32print.WritePrinter(hPrinter, text.encode("cp1252", errors="replace"))
                win32print.EndPagePrinter(hPrinter)
                win32print.EndDocPrinter(hPrinter)
            finally:
                win32print.ClosePrinter(hPrinter)
            return True, "Printed!"
        except ImportError:
            # win32print nahi hai — notepad silent print fallback
            try:
                subprocess.run(["notepad", "/p", path], check=True, timeout=10)
                return True, "Notepad se print hua"
            except Exception as e:
                return False, str(e)
        except Exception as e:
            return False, str(e)
    else:
        try:
            if printer_name:
                subprocess.run(["lp", "-d", printer_name, path], check=True)
            else:
                subprocess.run(["lp", path], check=True)
            return True, "Printed!"
        except Exception as e:
            return False, str(e)

# ════════════════════════════════════════════════════════════
#  COLORS & FONTS
# ════════════════════════════════════════════════════════════
RED    = "#CC1111"; DKRED = "#8B0000"; GREEN = "#276749"
WHITE  = "#ffffff"; BG    = "#f4f6f9"; CARD  = "#ffffff"
DARK   = "#1a202c"; MED   = "#4a5568"; MUTED = "#a0aec0"
IBGC   = "#f7fafc"; BORD  = "#e2e8f0"; WARN  = "#c05621"

FT = ("Segoe UI",20,"bold"); FH = ("Segoe UI",12,"bold")
FB = ("Segoe UI",10); FBD = ("Segoe UI",10,"bold"); FS = ("Segoe UI",9)

# ════════════════════════════════════════════════════════════
#  LOGO HELPER  (GIF data script mein hi embedded hai — alag file ki zarurat nahi)
# ════════════════════════════════════════════════════════════
import base64

# Logo GIF data — base64 encoded (alag file ki zarurat nahi)
_LOGO_HEADER_B64  = "R0lGODdh3AA0AIcAAP///////v7///7//v3////+///+/v/+/f7+//7+/v7+/f/9/v/9/f79/v79/f79/P3+/v39/v39/f39/Pv+/f78/P38/Pz8/fz8/Pz8+/76+vz7+vv7+/v7+vv6+vr6+vb6+vz49/n5+fj49/z19P3y8/nz8/b29vX19PL29fT09PLz8vrw7/Hw7/rt7Pvp6PXr6u3u7evr6unp6fjl5fnj4ffg4Ofl5OPj4uLj4uHh4ODg39/f3/bd3N/d3Nzc29vc2tvb2vfV1OvU0/TNzPPJyNna2dfX1tTU09DQz87Ozc7NzcrKyPTCwMbGxcXFxMXEw/O3tvKwsOu5t+qxr8LDwb6+vbu7ubm6t7e4t7W0srGxr62urPCqqO+ko/CVk+6PkOienOiOjKiop6ampKSko6OjoaCgnZubmZWVkpCQjoyMiYmKh+iFhOl/fel3duxxbO9qaOVpZ+tcWuNfXO9TUuRUU+1JR+FKRu1CQe89O+VBPuQ9O5GIhIaGg4WFg4KDgH9+fHt6eHRzcXBvbWxsa8ZVUmdoZGJiYF5eXFpbWVdXVFVUUlFQTvw3Mv01NP00Mf4zNv0zMf4zLv8yNP0yNfwyNf4xL/wyL/U4M/c1MvgzLvcyLvUxLe04NO4yLec2MuA2NP4vNP0vLvkvLfYvLfstLvUtKvgrK/UnKfEuLvAuJ+wuLOcvKvIqK+srKesmJE5OTExMSkpLSEhHRkVFQ0NDQUBAPT4+Ozs8Ojo7Nzg3NDQ0MzM0MDMzMjIyMDIwMC8vLi4uLi4vKy0uKy0tLCsrKiorKCoqKikpKScqKSgpJikoKigoKigoKCcoKSgoJygoJScoJConKCcnKScnJycnJSYnIyYmJyYmJSYlJSYmIyUlJCUlIiYkJSQkJCQkISMkIyMjIyMjICIiIiEhISIhHyEhHyAiHyIfICAgICAgHR4gHR8fHh4fGyEdHR4eHh4eGx0eHB0dHB0dGh0cHBwdGxwbHBoaFxkZFxgYFhcWFBcTExQUEQAAAAAAACwAAAAA3AA0AEAI/wABCBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzatzIseNDCQCelItmrWS1as7SxdIBQEONJl7c0METipWrU6RGTdqEhwgABwkTAGiBxANDCSCh4Lo2Lp+iFgCEejyYAOTUggKuai0IUqQyeecI7YgqFQOAM/z+AGijyhOdGgSawPH0SdSjU3N6ADjgwIIbPDX2DgQJZFi2bb1o5QqHzlihGQCQAjAzLxk0PzdQsGlnrB2XgWvkhaMVQ+AYeM54/Yh6CB67QZEBrFH37ZYMAEe25OpGrpAVJgKRDNu3RuAZd8isATdbhR4yboOOnDDzDRm84hYAVIBzSpSeN2/izP+h08MAAwAm5qSCk5WIp0iqugiM8mnUnBIAioSy8wIAFVCnvEEAAATE0ckqX/ykEQdMAIJLP05lYRYAMQiyTyIrAMABAGPMAkIXoETiyCpiACDEHqRsgskpcGhQkFA3zHLOMMUIY8w343CjyyJ+GBHVD7dcM4wwwfSDhhPDGCPMMOMM4wMAGKRRDD3RNLIFAGiow40iWEiQhCK+kBMMNvPUI888+uhyg0BZ9ELPMYc8wUECMiBSzjwXGsFFNus0osUMsYAjjDDA7PNHELocM+g23TARRih7RLoHH3zoEakedsAAQAhi8PHJK3QUQcQdrNgRBgtT0OHJKrDA0soeXQj/8IAAHoRxByif4hGGCwJItdWLVuEwAgAoCARUE22wAMB5v2aElFTPNivttNRWa+212Gar7bbcdtsQBiqEK+64KqAQLgonnDDCCCqsEMMMM8RgwgPe1mtvtV2V00wy3NDjzjCzFPLHLPwQIkAYnzwyySWRVDLJw5JYYgklkZBihwsFGDAQUFS04YACkSUABDHL0MPGQEoQoww9iCSQAQBlzINNLz9YZZHNGAQg0AqA6AIFSFV0s04VBFVFEMgKGV00ADo3AEcnmswRAgBZiQFLHS4AYABQAwmQldYDMYuQxgZxjRFITpSjTDj5iJOIIL/s88tnCRhtxTtYkGAHKY64/+IxAwwYYADINNiRyhwmJED2YAAEQQw09ACSAAYcYHCFOsbgE4hAZcRDjS+r4eYLNuuQoSEAf4hGWlVnwFMN6EYfUQg1+7iDDziMWEEWADLY8g08bEjwgVk5xKLOOu7wYpgz3DiRwAcAYBFPMt4sIVAMtohDjx8AZLfdKaPUwUIFGmhAghhv9DCACW6A4gkoeOyxSd9emKexBVPMUYolnYBSBA3y60QUDLK4qUgGACs4Ay5k4QSBsAEQEhBKH6BQgz1g4hKisMQoViEHvSjOXkLpHTyyAYxYMIIWwtCHPgQhgrrd64UwjKEMZ0jDGtrwhjjMoQ4rUrce+vCHQPzhDv+HeMOk4KMc6kiiEtXBDnvsox/86MYuGCGIMzghBykgQAl6EIUpCEBwDNEYCoZFxDL+Kl/NoMZJUAKNY0gjGk94gR048QhIPCISlLjEKBhWCVGMAhOraIMEGKAzghggARoIwxC65wdcsGMe8MBHP9gRj33AIxd0EwgZ4IEN1cRGIy7cHUF4oAQyIsRXEFkAAFLliVSg4hSteEWrVuGKUOjlADXsyjmQsYY1DQQBG1ICMKaBBBrgYROQqMMdVLE/SoxCFJKQxCM2sQchaO08PcDDFBQUmyAMYxrzUItQfMALauRjc9CLWTV28YMx4OId9xiHH6B3Bnzc4x79WMQIZLD/C33Agx+2gAxIyPALfQBjFuLYhzTAAQxerOEIiIDGM6iBDGCcYxEAuEEgbDEOetjjHvBIhxomAIAkHEMf9dgHL3BghVmk4x7rIIQKXOg0VIhCfL0CwANcQAKBhCAMc/AEK0CRilSwwg0lCMB5imAHV3hCD6pYxR58IoAStAEUr4ilK+xQhKicLSTnWEY0qmEMboxjEKUBSd1OwAhdtKAIn4BEKeYQGABsoAtuoIGs6BWFVtiBBWBkXGEgdzIAfGAMxSDHhVQAEnX6wgh1O8IvqqGOMZwudaMpzWTm8bnQCeJ2jJhQGsihDlqsAAESmAFH3cEGC2woKi3IkECqoA5+/zUQelhwB7+UEEHsfSMe3PPe00RxhyYUoQldBKzGEkACFgxAIC/IgyRc4QWtuYAOq8iEKvLghrraAA+igE9WYaEKVvDBiwWcCEmrgI9uLCIJAgGJCtRwjH0IAnoYEN4iYOMGVThzqstigADEoApIeKII5jEISG4wDGfAIxCvRUItsiEOY8CXQ/JoBi8gAwByUsMzUJKNPY4BDjTIYAmKOAc0guEEGUygBYPQBy7SMIYqBEEGExKKarVBDkUwYQwNbAQ42nGIqJxhHcgYhxZkUCwruMMY2PARAE5Qi3CwozgDgUMpMFGHGriABjYAQyn00FUxrILLWSuCJSpxii+4IP8EcoRDEQbkBlOgwg0A6MIq4AMHIrCABUIQg17Sy8MDAmAJgSiGPgoRhODQIhEniIwMGAEEIoACE6IQhSY0IQmGicITyQrsR2wWEVRGkCD55YI5ClGQExDiHrPIUCjjKxBUhoxxHgnAc8OmSmMZIAIHQZrWHPA1X1NrcgJpQRr+wJLfCIUJSejBHh6hCjxEIStme6FQdpAGRCxCEYQ4Aw5qbcYh2nogoi63utfN7na7+93wjre8503vetv73vjOt773ze9++/vf9QaJEmoxi4Ib/OCySDgtamELXNxiFowYhBq0wAMKgA0i5w4KUiaQcYDbKynkQEYyksGMZpjcGcX/wEUi1pAFIIBAICQQghTA8AY77METfJACN4MylBtMwAGzji8GUCmCI2BhDJDpuLkdwPSmA+7p2aZhvpxhjTWeJBrmQAMAHsUKVqjCFDqhBCUk0TCdfIIOWSM02ExABRYkIA2EGEQhCnGIQyDiEIEYw21AgoRBxEIb1BBGo0ndEaVzRGPIAoPiF694MYBhVxefIRqpsa9kbGMd9YDHPUhTBD5IYieViMTnJ0EJiV3iEpIAxTYrwJWtUwEACgAJD4KBjG8k4gpJYEIjxOENc2ABACKQAGU6uRoMWCC/A6nKBfILLQkYH9gEicCEIkNS3g0rASOIRTeQIIE5Jd/5aiU1/1IwgPyogF8qBQBAAeBACkjcQQpS8MIX6pCJOwRGY+YB3DXTm3/B6Z8ggON/akcR+cIM6gAMh2AGR1AsAGAEuKAPZwAAclBglGAK0AQJpycJlaCBloAgF3cshtAfGgMS3rQM8gAIGCACHHACslAO5bAIFtBY8VANv/AkLTAGbNBAAnECSEAGfxAIq8EBO7AFayAISuBVQvED3dYIb1MFgqAP/aAGUUEI6nAOaeADR4ADSBMDTpAGg4AIhHAFA8EBPECEgqCDIqAFftAlBOE03SE+5FMBU4AHbrABAxECQtAF6PMGXRACOSUULiAFYuAGbvAFRaABAWAAAkAARPAFbv/QBl5AA5F3M2ClDIqgO4ZVFgBwBSoBAlMACqQwB1EQBXLgCTYVCaLAMKpgB3oBFCxAB2+gVIzjTc4QORiAAl+CDupQC2OxIZRBDcAgC72AC9CADNpwDsBhBGewCNnADbYQAxlwBYeADcjwC40GAFbgDfBgCz+AAVCQC9SgDogwAyhwCP2QDM9QDdwwEp9xBfGAD4yABEgQC9zADcTQaDLABYPADVDWCMNYDMzADPjwGdnhAHBgU3WAHxtAAiYwIBpjTKlQXD3wNI/ACk0gEMbkCnUgBC/QBqVQTZviBk71BULgBp5QCnhgAwKAS5T4BOSgDMoADfaQC1cSG0YTCPj/AAUlYAemcAqs2BJFAAeqACo20D3n8QWtUF30IljDgA3dcAuIgAiNYA21AQi38TKdUw284EtIcAzOUFmGhQGYRRqRcQbz4AygAyWLYA7uIAjYxwFaUA7VAAxIIBQokD30IIWxgQKL8AtawAEfkAXrYHm3hQFagBzfIGUygAvYADzdY1dPEwmY0AqtsAqrwAfWxADpJwZ7EAbkYwN5IAqrUF0AQB+QwAlxIAQdQAREoAAaIAed0Alu0ARU0AVxMAoBMiAXoUvLsAzI8JXv2AhHCBIgoQPCQAgAIAauEAmnEAcs8AC79hNAQQR5wAo6F3Uk+DiR031C4QfxMCWWBTOu/5OWuPEL2KAOprMhf8AO4VALmmUG47kaCfADuJAP3xAISGAFt5AP0aAFIdY7v3UyIKEzK1AFV7AFXBAI4tAMzWNY0TM933CEFFIL2hNcOsV+N9VTXSMQz4UBRIBcUfAFmUA/2gEAEwkKnSCasGIiesAJmfAFUQB/tekThUSJTgAOzqALglAFw8IEi5AP45AGOuN8srELLdAEoCBNnxArQCE4JMAdnVBdURcyRqCd3EN+ALBJwnAPxSF87kANvBAEQnEEo8MOZNB9qMOeuKBZXOA6vJADAFB9aXAMbIAFWCBlHHBqvdOYeikQT4gNt2AFMrAF55AMjaIhCZAF01M9Qv/hW/PgBxawAQagAXGwCqSQkAPRA6DQChdJBa0wCXoQiUSgB5fgCvJxAAnQA1MTAl2QB5ZwCm1AAXRACqRAB8oyAF8jAAiAESDzA4wQDfpwDohwBa8FAE4wC/0wCJFmFjywC1bwAnjACZBwCrZKUlTwCZ9HXctyEDngB9uADd/AC4uACIlQDPlgD8BwJQmwA4PQDtRADYIwFlmADNggDo3ABZF2BfSwDcvQDcYADn0pDtCwDbdwAmXQDyShD/jwDcagcmywBEKhArNwD/9KD/hQHIHQD6lxCLagC8LwDcsgDbtgBDIACO3QDN9QCEhQUrtADuJgC2fQAiYgBXdgCpf/oAfKlAd6sAmjgJml2QqZ8AhxMBN5kCKesAdF4AWpMGZwIAVzcAqg0AQCEAJu0AqdcAo4WwehYAiLNKUT4Sv5mAjmsA+KQDQCwQXbkAgicDqEUBxycAqOgAl6IAVvOwmYoAnU1IoFMQI8oAM6sAM7oAM4MLgzAD2xcQI/8LeBuwMr0AKJC7iCW6xHYAZOYLjYiAW+0iZoEARYgAaEYAvCQA2/sBoXAAA/YEUFIQJWgAZiKBBOsAVrywE+8Ld+uwMtgLh+67fCAgM20Lu+27s1UAM9QAPVBwFE4AVSYAIC0QNS0B8wRwRd8AVfIAVFKRBk8wJN0AVUIInqNxXIJhA47JAGuhANahBpLVB3HwASVyAIAPAGp/AIpWcJl+AIqnAHUaCbwYYU+qu/QRcZ+xtBVbG/whYZzDcYknEEyxAPjJAEI9ABLaAEf+AL+qB1VgF+wFLABCwVsbe/QvG/ICE4IHwABwDCIAyAAWi9gYNuT3fCBZF/KTwtCVC6CIQGjZAGa1sGZwASMZAGFNAGrlAJlsAJnyAGJiAArKdtm8gL9WAP+rAP+1APuYAGJGV4HocR3xsEXLAaQTADCfAAPnABbwALHfQTA1gvNpO+g1GjVdwsSsc1YoND/bvGcjzHdFzHdnzHeJzHerzHHhcQADs="
_LOGO_SIDEBAR_B64 = "R0lGODdhqgAoAIcAAP///////v7///3////+///+/v/+/f7+//7+/v7+/f/9/f/9/P79/v79/f3+/v39/f39/Pz9/P/8/P78+/z8/fz8/Pz8+/76+vz7+v349/v8+/v7+/v7+fr7+/r6+vr5+fj5+Pn49/v39vf39/X39vr09PX19fX19PT09Pzw8Pfw8PLz8vHx8PDw7/Dv7+7u7ffr6+3t7ezs6/nn5vPm5+nq6ejo5+fn5ubm5ebk4+Tk4/jh4fff3/bd3PrZ2vTX1uXi4ODg397f3t3d3Nvd29na2dnZ2NjY2NjY1tfY1tfX1dTU0/XOzvPExN/My9DQz/W/vPW4t+u7uvKzssvNzMbHxcPDwb6/vbu7ubi4tra3trW0srCzsfGtrOqsq/Gmo+qlpPOenfSbme6dnemcnKurqqenpaSko6Ghn52dnJucmpmal/GWk/CRkO6MifCEg+96euqKh+p9fZeXlZaVkpSUkpOTkZGUkZKSkZGQjo2PjYyMiomJh4eHhYSEg+9vbeRxbupoZetgXu5aVuhZV+lUUexLSuVMSu1ERe5APfA9PeRDQuY9Ov02MfM5OO02M+Y3NYGBf39/fn5+fHp6dnp3eHd3dXV1cnFxcHFvbW1tamxsaWpqaGdnZWZmY2RmYmVlY2JiX15fW1xcWlhYVVVVUlFRT09PTE1NS0tKSUZHREREQ0NFQ0FBQEA/Pz4+PD08Ojo6OTg4Njg2NDY2M/8zNv8yNf4zL/8xM/0yM/0xM/gyMv0vLvcvLPwsLfgrK/oqKfYsK/UqJ/knJvUnJvYlIvchH/AxK/EuLPEtJ/EqJ/EnJfIhH+kxLekoJTIyMDExMDAwLy8wMDAwLS4uLCwsKioqKiopKSgpKCgoKCgoJiYoJicnKCcnJSYmJSUlIyQkIiIiISIhISAgHyEfIh8fHh4eHRseHB4dHBwcHBwcGhsdGxsbGhoaGRkaFxoZGhkZFxgaGBgYGBgYFhkXGBcXGRcXFxcXFhcVFhYWFhUVExQTEwwMCwAAAAAAAAAAACwAAAAAqgAoAEAI/wABCBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzWnwA4Ei2bLRAnVHCAkEeflx8QLqV65evWzBz5aqFrAsABQoAzAA0A0CBBxU8tZuGZAMALuXQ5REYSd40HACu2FECAEENLHbyyACgpEwfLFUFVpkkipMZNPrAsQgwSp8lI09YALCwRI0lTGaMVujaZ4tAKnioVF3QQIywXlECAFBBqI0AAgJgRGnzps0OAQUKAOgxRo6bLjB8CvjRJk6YHT4lIgCwAk+qedX23BBohVYrqACekEIR6FcvQz3HKGtEjM1NgQisUgJn7124Z9VgY4pRAc2zfPRUBTljrp46US88dP/SB0+fKj3W7KlDM2Tct3389qmjVSoLgBGl9rkrt2lEmXnpVAPLJt7MA04ZXMiSjzyvJKHFN/m0U0oQYCxSjDKQNOOMM8YcogEMhSyzyyKAFFKMM3Iw8QgzjwSSiDNTAJACIMkwgwwkh4ARAQEaGZQcQjn1aNGPQhZp5JFIJqnkkglx5JE22HxkjTijUJEDDD5AIQYchShyzDLDLBOHAEEWtNoQJzSJQAuUoCKLJigQqRECFaymkQACMBkRR0h0s0010b1iRVUVALDENVREkYwvgqQAwAVfNPEoAE04Y1yZHHUylBAAoCBJOvr0wZQ80BRRwQur2BNKVUhEE08kVU3/Ag80RgCAhTrqnAFABabw48kDCOAQyzyaACVQCBwIlIk9sthQVR7uePNEBR6QUg8qIPA4RjC+sPFDE02EVlUEIgxAaS/CfAEABHEkk4wjcKD2wyO2KJIII4owAkcJmV1E50BcjLJHAkTsUQMAW1DxBTG15DKMID0AYICeA3HERTj7yGOOPvy8IpidFIcs8sgkl2zyySinrPLKLG8EwBCrxCwzKzS7EgsttLxiCid6cFEECOY2xJFCQ7dstENOcsNNNtdc4w045rkAhzCNNFKLLTBdfQsvjPwAQAMDge2EFJ2iIYs95tCnyj7kTOKBQH3I8wwRAGxQwdB0VlAnAEDd/00QUAgAjhwCLOyxxANGVQVUBILzrffQQEEuhSDGDJPIIYk8sowhIvhUgAIS4KTZQKDjJMHoN4V+OkUcGaHNNqhoIVCdD1hASioaBEJMID1AEcgxwdjCyyNdNDABAHIAchzfAHjCTjWR4CEKOLGUsSsAkrxjDSlrWJLNO53cZ8Uz60gCgAmWlBMNEramI86gOrzyDigAVDFKNOG0wgkaUfEDywgjUEXatsICP5TjGqewgx+skY5TQCAnYhiGLtrQhC6MIWIF6AAhmBEjNvQiGDaZwSGKEaMxOENdXthcCgYAh2IQogSpeUjrrGENU+BBCFVABT8uEQIAfAAAl6hECv8WcQthsEEAF5DAYgbBjDEIAGwDuZtQpjEEgYygFeCgRhAqMAlSFQEBMUhVKBAQACRAIx6TgBupcAOATbRCEn24AgXqBgAbvKIemBjIHPaRDktsgRTpoMUK+KYHeHDjCQj4QCnuga0GFCYZiBGIAALBiBTMYBHGGEQX2MALYdikACoQQAnYoIhhBAIGiBjGH4ImkANUpFACyQEeYvGMNQjEDqnglAYuUYQ2FMMWunDELnDBC2S44QOoW4icIFK03LxiDgMZwR72ISpYIsdHPRpdnhCCJ4FoBnXbhGKRYOmCK9TgBDoAgAtI4AZGQCFPZRqZnSpggYod7Z74zKc+98n/z376858ADahAB0pQeSbnoAhNKEILylChmeChDz2BCVBgAhKMAAQjIMEKYmCDHMjgBA1Ipr9A1tCScUQJ4iBHSslBjnCcgx3PKMUkzlAFIdCAB0z4ghzkAEOREsQB4QEWQv1GkBc0E584SapS42nSjnCjG0rjBja6QQsh/GARzVAGMYbhC1zgQmvN8MLywgYAJ9AABZTQRCY4IQpSfKIPVRCIC/rwinR0owjMuwjjigQBMrzBDW6IwxveIFgy1PNkSePGN9KhDm20ohT4qEQGCuGLRsykETHJhS2AEQgM9CsnTCDDBBKgGU2s4xlLAIAHQDEOePilApKYR6kIQlKF/9QWYXQwChf4oYaB3PYgwPItABYAgC8cYxdTuEAKoOAMQTgghgQZnU+9WZCJSYQjRQDHKlIr3EqUIwhfcEYj/jAIZACDF7b4KjAK0TkFZAAQZAMbR5wnjRwIhBL2KEdcARAJe9DCDKTQxzu0YR87hAMd/OBD89aBD2oAYVd8eIcsLBELe4yjFZTAQirCYQ1wtKMcRCACKhwrP3xQ4gErKIU58sENLniCHu8QhxkmFUFfQAEAeXLA8SoQh0QkghDHMEYhUCMAL0AiEYFwBDPIAAAGjKEZiVgEMxAhKevKsCPZ2EY5ymGJrTyAIy+YBR+GWAxBlEAAUBCEMhgRBQAc7/8LjOjJ6Oarjmuo4hSxKIcrzKCBQjWFigCIgSzasSoACMFVsOKv3I4AAGhRAwlv24Q8XPECALAgFvfIo0BsEItNrGYT85DFbABgB3h4QzANSMU7TgECzYghGbV4BCSS4QwvGA8AbjhEDwrQhF2kCwA8cAQvBsEEFXhhBiIgRDH+wIQmvAEYj/AaUxfSum1o+R20AIc6oKlaAJgBFiSAwy+E8YcMjO54PVCEIVTQL4HMdyhK4AgLYCGOZ0BFEqSqlQvEKBAjIBoosqIVAiDAB2tcYxOl0McpOFVHWMxjE9RCgAfWsAlRiEIW4qAFDlaTB1M/oW6LxJarfdNmgdCACRf/AAAT3vAHQAyiEcH4wucaAIZDAKMYjGDCvIQhhhnsYAY8wAAAFMNMACDBHeGow2FXwId5mMIFdTPBKaowBWTc4hg3nsAD4ACMXjThcwbhBDqeUQUSxGAP4sgHJwrVh3Q8Iwh1S0U6Ch0EaoRDFVXIwirGEY0y6AAFlzAHJ86ABSUcTCA3mAU5UFGGOYzgE+6IRQ1icAp00CILD55DOqzB3VG04xRED0MvdHGaH0TBEYYIQRMgGQYEiKHrb+hBCQIRLwHAIZMOiIMxdgEHJvBgCmSIgEVAYAVNfOMafdhKC0IhC/sCYA13UAlx3gCFR/AC5oTonGZWAwQ9rEENaVjD/xzUsIUqAgAEaahDGtJgByGcYQ7hp0MLOlWHTZSBBEfBxMcBcAZa7KEPohAL3rAPhaYDlEAJ7CMQW8AJmUAVZ3AJjHYGeIAGaYAHQ1AGdoAGakAHQeAFZAAGIDgGX+AFYyAFz8UDcfAHXSAAPSAHyZUBTeAGgSAIcNAEArAaERAFKUgGKoARAzc0IVAGqTB1tvIJ6TQEejADiMALurALtiAMiBAF20RbelOFVVg0Vlgn9FSFyHE3e/NlRiEK/AAKV/AEVqAH0QBowPJlRcOGHMGGu2KFgWM3eoMAxiMBeJiHoeNNCgB2n9OHNwGIODEQBCA6gCgk/8I3ZbAHOmADWURgAh5wBTxgCL7wCG7QOccjTwBgBZlgCqcACmgwSEdVUhkxNINEJDsgbdMVMl/Whb9FikkCdrA4i7RYi7Z4i7iYi0sSEAA7"
_LOGO_SPLASH_B64  = "R0lGODdh8AA5AIcAAP///////v7///7//v3//v/+///+/v7+//7+/v/+/f7+/f/9/v/9/f79/v79/f79/P3+/v39/f39/Pn9/P/8/P38+/z8/fz8/Pz8+/z7+vv8+/v7+/v7+vj7+/35+fz5+fr6+vj4+Pb5+fz29vz08/f29ffz8vvx8Prt7PLw8Prp6Pbo6O/v7uzs6+np6Ofn5vjj4/fd3e3h3+Pi4uHg4ODg3+Dg3uDf3t7f3t3d3Nra2fbW1ebW1dfX1dLS0fXKyuzJx/PCwe7CwdDQz87OzMzMy8jIxsXEw8LCwcLBwMDAvvK7uvKvreuysL29u7y8urq6uLe3tba0s/Clou+fnemkpOWdmu6UkeSXlOyKiOOLirOzsbGxr6+vraurqainpqWlo6KjoJ+fnZqamJeXlZaWlJKSj4+PjYyNiouLioqKiImJh4eGhOyCge16eON+fOlycepsauhlYulfXOpcW+NaWehUUulNTPBGRuNFRPc6Ne09POc+OuU5NoCAfnt7eXl5eHl3dnZ3dHJycHJxbnBxbW5xbthST2traWdnZWBgXVxcWVVUUlBQTf01Nf00Nfo2NPo0NPw1L/k1Lf4yNP0yNP4zMP4zLfoyNfozLv4xNP4xLvwxM/kxL/4vMfouLvksLvQ1L/MyLfcwL/ExKvYuKvEuLvQrKvMpKe00MO0xLOg0MeI1M+4vLOYvLOkpJ0tLSEdGQ0BBPTw8Ojk5Njg2NjY2MzU0MTExLjIwLy8vLi4uLC0tLCwsLCwsKSsrLCsrKSorKioqKyoqKiopLSoqKCgqKikpKykpKSkpJykoKygoKSgnKCgoJicoJicnKScnJycnJiYnJicnJCYmJyYmJSUmJCckJiUlJSUlIiUkIyQkIyMkICMjIiIjIiIjHyMiIiEiHyIgICAgICAgHR8fHx8fGx4fHR8eHx4eHh4eGx0eGx4dHh0dHB0dGR4cHBwdGRwcHBwcGhwbHBscGhsbGhobFxoZGRgZFhsXFxYYExUVERERDgAAAAAAAAAAACwAAAAA8AA5AEAI/wABCBxIsKDBgwgTKlzIsKHDhxAjSpxIsaLFixgzatzIsaNHjQgA3MjF7BcxYsqUHTsWTNo5crdgISIDRYeIgSp2MMkS55AWCg0MNCwAwIUJBEgjKF0aEuGGFB+jVgxgoKrVq1alagUQAUANXOAYcXlRcAOAHLLumQGAZZUkSXhSner0iNImT5wsWXI0Cc8OAAkAPACw5NAPAA4GdiXk7pc3LxFAAODAaJyyZEMAmE1TTxg2Iy3+3HpHTpYSgV2SiWOHr5kUAE6mfWO3j1YPrgBCqNmFrxsufPiK+cK2zJeRHomWLaOGa9YsXEcA+PADa9o+fvXQvUPEImSYbOLa4f/T5SPEmVnaa3kBEJIBACqkKJW64qBChQckYqx4MAAADCty7PHKK6ic0kocMAhgAAICCFGHKa+48korrGiBGAAkaJHHgK+YgkcVFSyoUVc05EINL4ZAARUAOiQyTj2L1MAVAhcAEMgtLeyQRyeOfHJHDAIoAIABBZAARymlxOFBAE0VdEEEhchTzDdccNWVEcwoAw0uOXClhjvCUPNEBBsgwEY6yzijBJkAdNFNMPSwMdAZ8fxCTRIAMAlAD4zQIoUOOqz4JAJd+ZDLM+kMQlAPu0SzTYws+tJMO4gIdAECYtjTyzxfRIYAGOf8Yo4YEWjAAANUqKIHKFcIMFAMfLz/cgUAKNRRSiZ2/FCBCnh4AoocBAxJAhAjCLTDHJ9UskcMsFoyCh0JCuAqAAIQQNRW7HWlmRSNyPIaAE8gMgMAB+TgxwRVpEIJXpJ80ocVGQDgHrYVdXVEMuVQk0454azTTj3TrFFmk/QWbPDBCCes8MIMN+zwwxBHLPHEFFeMEVIYZ6zxxhlb7PHHDHdVgy7ULPMMNCgvk4xw5vgBQBWrQKJXXZp4sskmlFSiiSabSNLKGyEKVZCQK2CxgkBH/DHLONho0800tNgiDzvsMKIDewKdUc8x1TiBW0dIKWZQEl+0gLVHQvGgBR2nfLKJKH2sosorpIzyChPygnyQyLpY/+OMM8hs800iN8yYLQaE5PNEAUc6ksobJAiEAhV55LGEAA4o4IAAV7wChwdVFdTVIFJ+Y4YPQ0ShSDfbxKOI2TWmIc8wXUcQAgJ+pINMM0ho1qY3xpxThu9r1PNLNU8I5MQ3vphDRm4KGJHMMe2stUEEQ+hSzTqDEBr2DGeQ4QJSPviCTDqV0giAGPD0Ug9kIUQABjq/kBMGABe4R0UqkHiCBxxxkMMc5LADAzBAKCeYAhVgEBIU3METpZjDB4gyhVdAYhKpgMNhBEOYVTwiEnegQx3s0KtS9AEIefNIC7iQCF/wYxdrGJfyaNGMIuCPTItQAwDeYIq3nCIOxyrFJf8yMQlJqMIKFxKdEZyQBCU4EQlIOMIRikADbXUlB1BAwhOdMIQSIIGJTlTCEUqAtR6owQ9KCIFAQjCGNchIICBYAzDIEAUzGKIRvBBHLxI1owA8QRDiG0hIfICGQZSBBQIxAiC8BoAXQKGJSGhiEUBghCckIQmRREIKZCCETnYyCKAMAhCAIEoUDEkAMaBCG64AA2NpoQnFAswOptAGOMChDU04wZB2WYEfUCELWWiCGh9GKIG8QI23E0hIrNAEV81Lbwah0ZMiMChoWvOa2MymNrfJzW5685vgDKc4x0lOhIQkNIIARCACoc51AkIQ7gTEH/7wzkEYohCI+AMbxJD/BBdM4GBhK6dAGcI3bDQjGihLmTZmcbUVVOENcrBD3FpRt5pZYl2WCEUeUJiYhDQAADzwQQNAkAIXzAAHOQCUDnBgNoEoICSA4Ic92pEObERHWwejJsE+kgES+PSnJDjBCYI61Hhtk2/UOAYyuuEOcuACFoZIAyPyMQscqMAOo5DEJThBCUpo9RJ6qYTOKDEKO6hAAIEZCAQSoAUsJPECFygEPIAhjm8VRRbfCMY38LSZeiDDGk9AgFnUCpEmIUBIApEAQbgwi0WEoEZe4AY5vlCQYh5kp4K0rDIHQgEHpEoPppgVtQBghVfs4TAMQEBHBzKtBEwrhQbpz5CuRZAE/whtRF7RnjfKgYirabZGUhBHLFigo1H0oQ0jUMEU6CAKUUBCE5H4BB+YYNvE1KoOJ0DAbXFDCCllIxFn+AMjrFGNWaThKMVUw9aqIQszFEEMuyhZMogwEC5wwxjleJ5A0tCOYEyDkYxaxjrOMJA10K8bUQBAC85QCF5MYxux8MMfbgMAP3DjG2MQCBjI4Yt1FOJsYbBHMKBRCz8YgbHbIEY41lMB/ZFCE6CAQxCYMIUraCEGVUlME1IxijYIJAipeMSvKkAULZCiE3EQSAw2NCsBUGEVn6CDCgAwAjg09w5T3m69cuu3ZCBjHOTghjl8YQazdKVQzahUE1bh1VZQAf8AnXVpABT7gz2Ywg0UwCxuojSlLpBJMkVIxjKqgQsc7DcewkCe7860jGjw9XfBG55Z1oBobCTPLEi4RTmwwQgy+MEX9wAGZYs5BFw4Qx0f3uwGwCCIPxQiEYxoBjLikb4ahdh9lJUMqIYxKvzpTxSPmM8uBUICFHhgACERghay4AY3yEEPkwDFHDyAtRNQThWeiIQq7hAEAFyBFJHYwxJQoAIVrODcAgkASBSsiGTYAx+0YMPVAFAB6SxCH7Ogr1IAQIRdKGEEdRCFJYz45njtWC8RLJaW1+gEWGBDGM+QxSIQkYhatIMc7VCEC/Jkg0SAw06JoG8JBjGOXzhDEU//UKMNcBEOX1ijGM3YBiN8UY1m+IJKIZAFPoCBjXvEoxvNsAUj/sCFFIQEB7foRjGmQQ59yKIFL6AFOY5RC0TEAhe52EYxgMENRRjABYEYRy+wsQgkbOACZii5NRjhhRT0Mg6foAQm8ECHO+BhD6GwRCuqAIATyKEVetBDAPOwh05MIhR9kAMK4nAKUewBDkxwQydQAYcRIAAGc2jFJEoRijvYgQ+seIMaF54RzbpADLDARywYGQI/2IK+HADAFwjhbVM8YmedwMQmLIEJUNghCK81CDXhSnziRyCgyiw+XM+sfMUqhinKXEpXzGCOyRbkBYLYBjD0LRDoE0Qp3us+//jZo/wncaX8BzCAA9bPgPWzv/3tv20CDnhAgRiw/gKZf1UKoH4tF+BUtoV/CmNZIRAG8nYWTnA7ISAFHZAukLAKb5AgEUB6FhMSKVBHbIAGYEAEioV8A8VNemZ/HziCJFiCJniCKJiCKriCLNiCLviCMBiDMjiDNFiDNniDOHiCCMABJZABtJWDOXh02YAP8VAP9VCER0gP8XAO64AP18EP/KANt8AIgSAGToADKdABHrACkhGCCVECB/AQhKItXgiEFNMVOKAL2nBQ0CAN0kAN1LBQfrAFOtABxBYDS2AFb1AHecAKqoAKpdAJq1AFmMMQEAAAMiADBNUkNMBqi/8AC2MAAmUoTq8VfKzFTQW1DAmFMtOQDO7QCCkABHlwCqUwCpFQFxdVM5SwM56QCalQB0dDgYkhA7AEAGNwC7IwC7QQNbeAC7YQC4oABlBxRVuwOr+ADVDwNVqBfOq2FYkhBIdAB3NAB9I4jdUoQkIAW9eEVCazDMfQDNlADuwwD+DADFuwQ6rgCJmQCV1VCZMAVmJVCTdjCauwBPRWWQiQISiEWAAgCOtwjGzQAy9AA4YADrxwD4yQAgdwPcWTaF6jFHBFMISyfII0fDilGBbwfCHBAjTAHgjgArSgD4FVJgQxkV2BABIgkRYpSBFpEFNACpAwCm7wAzM2BXOQCqn/YI/zYgALcCpVIYADcX8/CZSAcUAGRIEWwTfPAA7tMAuD8AU6QEaN1Aj6oAgTIAR8EAmXAGOVQAmWMI+UwAmcUAmREFrDFgBdASAkEDp7NlfgABlnBwCFYDzkUCWbIQ+ekTwsMgZgsHEhEQEu8ARjwAavIViORAZsQF8E4wNmgAiLkAh+cARbMA3qYAsywgGLAA7pwAZQ5wIh0IwpMARg4AeIUAhe0IX4c5iRORAvEAZjYGhN4WKOMB8MQAEewABCkAd10EpCMwI6cQVt0AZToEuHWG1LgAVZoAVWIAQnMABUIS07UAVaoAVT8BeWeBF84w2NgCcAAALH130AUAS5/3ANRjACclAKkxAHTHAFdsBjmqAzmmAJj4B42XgqAKAjbrVdi4GX3/AtIWAG24AN+hAIl9IVsnMMw1AIs2AP5sAMw1APTwcARlAIjfAMvdAOBKYDfxALy8AL35CMl8IG8zAOsWAEMxAGJoINt/AFqGML/AANJgMO97APsEAWg7AP7wALX9AFsaANwIAPwwMASVAIDscL6NMI+EAO1EAM4rALtxEBLgZjTWaJiVEFrlAKeDAFSzAHkwAJrZAFF6IjHsIEP5AFkEAKbyAAEDACbtAKqgB5cJAKmDCIaopbI2MNxgAN7WALYfCXYlMG+/AHL8M/bxo5KMAEdUAKnxAK0/8lAEFhAGzKCvY4GN8HAKSDDEy3DvSAD95ADbDABWpUTLIzDNuQYAKRCGjCO92JAG4Sad0ZAWpQJ4oGAFtADsUwWRcQP0ZQDMewDn4QARnZA7gwDXxkKQDgBfpgDlIAV0QAc+qQPmWSKbwgD2IwEGyADip2P/nzHqoQCZogCq7gCqiACqzQbfb5A33gClVQmyrwQL/yAQLhBqVgCXowBR5AAFagQf5xCKFAClQAAULFBKMACalgrhkhMrvgN86wDMFQDOogC4wUEhsJC7bwAkv2CY5gClngARXgKjvwFxOYGEzQB32AWsIHAN01DN9wjgJRAokQDr2gD4Kwb2kQD7T/c2kAcCa70ztm4SbBIDzEU2nJgwAtEAvm0A2N0FJoAA7a4As+4DulNg3mkGoJ0QO9sAyUQn4AEAbtUw+5JnvlIAy9tq374wigkAUC8AEjcAIogAEEESwCESy80gnSBq8GgAJw4AqhoAeY8Hh/QVqtUAlzhweEewd3kAdT8AD8iJ25hQ3OEA/jAAtiIAa0EA/lUAiI1BWQ9Q5SYABw0EOZwAf1GTqJAQN2UAqnpYyVmrLf0AVjGAEzUAvggAzcYEMAULNh4jVmwWiO5jtd4A0/Swa2EwGUdrMKcClnkQjiUA/wYA62gAhGgDVdMQS5sAzpUAgQiRRqQA/pgA++0Aix/zAN1IMIELk+msIp3QkAuza2vwYJrJIAFEABDDACb5AHfCcAI2sKp9AHdUAHRSRtoCM0qJQFeIAJZ1oHH8AEqiAJooA3BYGUE4GG08ANtEAGiOQ7GyAGtbAPsfC0F/BSLBALtGcFq8AzSbKWCsAACpAFpjCbPgbBgkAPwtANYLABBxAAOsAI3NAL+DAI5qc1xxgdk5YOxxAN0VsjULANycANsmAELyAFscANxUDDLwAVTnAL9tAIZzAGW1AEMyCVuFFqywBhUdAFY7ABSMA88iCoIcAF0+Cwi/ACLlAjYGAPvDAPXZC+XrAOwUAOYMAeQkEFznW2KLACMBADWdAJqf+AREsWbUqSAXWmCWW1AysgAEvQBxoEr0xgCpqwBzsgAG6ACo9gCnIQBCqAAjFQBVbwnRxxAUSwBrHQDvxwC34wbwCQAmjwDbiAJ+ZnBrEgAkHQB5UgCZPAB4fxA3lAROvIB3/xTB65AYOgD/BADuXwIvpwzdPQWGAwTFyQDU1VDuDAD43AAooAD+NQDuSwD9kwb12QC/LQDtvQCFFwAX5wD+kQby+AAIBQC8UAzuOQDuPQDeMgHgQ2feHQVLJQBi/QFV3AC/EQD/xgGk5wkNWgCERgBLpgzuMQzrMQGt88DviQD0fQBKzACquwCn3ACuGqCqvACqkgqdTyA3agCqX/8AqH0AQ/cAeosApysAQnoAV8oAqkgAoRogpy0Er1BgNvsAdCTa49fa9hyBFoiVM+wAazsA+4MAZSaQOLgAs2ZBY6EAs9AAN5MK+aMHehIAl38QijwAdNkEQSwcoTsVNybSUIMAbZQGADkQI6kAa/EA9ooLrd1xR0TTAXyRGugliuIgAMsACsxQAHUJyLLUhBSQDTstgM0IxbsW8AYAJb0AjcMAhksbWwcDXXIwhOQAJ20AmSkG26Nwl94AYJMokN0xWIUA/DYAt+4AVdIAaFQAvtcA9pcDZm6BGWFQFbsAhrgEhEgAYXHAVKoNqgYMCk8HsAQACr9TEhAcXYoA4YQzcOuqAIp5EnxY0t6gMAM0CHRVEDG8kCK5AHfUAF1Eap2HTYYlPeCGNYD+wq2oXf/v3fAB7gAj7gBF7gBn7gCK4wAQEAOw=="

_logo_cache = {}  # variant → PhotoImage

# BhugtanEase Logo ICO (taskbar + shortcut ke liye) — base64 encoded
_LOGO_ICO_B64 = "AAABAAYAEA8AAAAAIADxAwAAZgAAACAeAAAAACAAOggAAFcEAAAwLQAAAAAgABINAACRDAAAQDwAAAAAIAAcEwAAoxkAAIB5AAAAACAAYDQAAL8sAAAA8gAAAAAgABavAAAfYQAAiVBORw0KGgoAAAANSUhEUgAAABAAAAAPCAYAAADtc08vAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAACpElEQVR4nIWS20vUaRzGP+/7+83PGXVc7cRkDVsQbFoYeVfo7gbaRXQgluoqkGrbCyMw6HARVHRRdKFBFLF3u1DLSkaQjf0BYRqUHTCxA5s5HbfRGnV0Zt73/Xbhgcig5/b5Pge+PEpEhO/BWMTT36R8YwzWWpRSiAjiHLmJCbK5HE5rgmgxJeEI+uscEdAa//z5C7S0nKWkpAQQPD/EosoKqiqWsyoeZ3m4kLhxlG7bivI8UGpGLMbgp1IpBgb+o6goSmZinGuXL/Hz42dw/zFeWwe5gZeMDw0SGrpIdN8fSD4PoRCSyZB78gwdBMFUI2FRLMblRAfB4Cs+XbnEx54eJkbH0NEYoydPY169Bq1RQKb9JqHyheixsTFqamrp6rpN/5N+qquqaF6ykMVr61DLluGVleJG0tj3Lxg504LyPEb//gdvThnegvnocDhMb28vicQNCgsLOXSgiZQxdNX/wg8FBUT//YtIw068+E8U791FrucBua47ROrWIcbA8eMnxPdDUl6+WOrr10symZRPw8Ny+NQpSW34Td6uXiNORMz/H8QZI+9q10v2Ua+IiDhjRAPEYjESiXaamppoaNjFh9QQOzZtpqOulqL+p6R+b8SbN5fh3Y1gLMHKSsRa0BrS6bS0trbK9u07pLm5WTo7O+XgwUPyJpmU67duSd+eRnlPVIaOnpQkEXmzdIXYzLg458RZK+rLJba1XaWvr4/q6tWMpkeoWfcrz+/eo3L/ETLPH+LNjVP65zkiWzZOCrRGWWvFOYfneSilyOfzdHd34/s+88vKKF7yI/bIMfz265ReayNYUYE4NzkoQDnnZhpYa/F9H4BsNotWilAQYN6+AxR+bMHkkKZuZhlMY7rR9MDUVNrM477ANw1mQQQE0GoW9RmnKk9BKeCJMgAAAABJRU5ErkJggolQTkcNChoKAAAADUlIRFIAAAAgAAAAHggGAAAATQocKQAAAQhpQ0NQSUNDIFByb2ZpbGUAAHicY2BgPMEABCwGDAy5eSVFQe5OChGRUQrsDxgYgRAMEpOLCxhwA6Cqb9cgai/r4lGHC3CmpBYnA+kPQKxSBLQcaKQIkC2SDmFrgNhJELYNiF1eUlACZAeA2EUhQc5AdgqQrZGOxE5CYicXFIHU9wDZNrk5pckIdzPwpOaFBgNpDiCWYShmCGJwZ3AC+R+iJH8RA4PFVwYG5gkIsaSZDAzbWxkYJG4hxFQWMDDwtzAwbDuPEEOESUFiUSJYiAWImdLSGBg+LWdg4I1kYBC+wMDAFQ0LCBxuUwC7zZ0hHwjTGXIYUoEingx5DMkMekCWEYMBgyGDGQCm1j8/R2zgUAAABu1JREFUeJzFl2tsVMcZht+z55y9eXeNufgChRiMvbiAMRgU1ZBWGBESGhEnqlS5pLSJIE1S2iZVxI+EWk1FEqpAlaah0JCqSQGZppWoxK0iqKVQOYTEqDZOrNZVbPZOfNld23s9Z+btj2NvAYNkokp80vw5mpn3+d6Z7+gbRUpJ3MWw3U1xANC++FIFinLjF/LOzdQAQLl5pykESWQzOeRzWUghoOs6XG43FJsNk8huvwm0VCqFzZsfQzQag91un5SFqtoghEAum0U2m4UwTDidDswqLcP8mmrULqvDl+vrUev3Y7amQRnfeCriis0GJZFIcOnSZQgGr0LXHdcBKCAlhDCg6XZUzp+PuoYGrGxsxIqVDaieNw9lqgpn9BrklU+Qab+IfOwaSg69DZvbPb7FbZwYFzcGBqEpioKioiK43R44HBaAoigQQsDr9WLv3tdQX1eHBZWV0EdGgZ5/gX9vx9gHryPV1Y14MARppmCDBiCLkZ9WoWTPy6BpAqo6WVxKKKqK/L97IeIJ6w6QElJaY8IBVVUxODiAVD4P/z2VCC//ChiIwMwkYIOADQ7YNCdUhxOqy21lKyXSb+yD+5uPwLFq5WSI8czNa58jc/p9+J7eZpXhhOs2mw2apkFVVRiGgXzewLYt30VHMIDp33gUZmYAzmnl0LxlUD3FUBx2a7EQgGEAigLFNJB8fico5Y1HQAJSggASu34O18b7odj1//0HSCKdTmN4eBDJZBIejwfr169HWUU5tm5+DPquVrhWfRXm6LA1XwpLQFHAifMWAop3GnLn/4pU23tQVBUwhTVfCCiahkTrLuj+atirF4KGCSSTSdbU+On1+rh8eQO3b/8BT5w4yXA4TJLs6LhMAHxh18vkZ/3s1z2MFJUy7JrFEOyM2Kcz4q1gpKiMEXcpI54yhvViRhetoBmPU0pJkc+TJMf+9GfGmr5OkpSGQSklMTIywqqqhZw37x52dXXx+siPL3znnXcJgB/9p5e53b9gH8DoijUcfnYHI2WVDEJjxDGDEd9sht2ljBbPYRDg2LHjlpgQzPddZbBsAbNXPrG+maYFkEgk6ffXUtftLC+fzQMHflMAMAyjALFt6zZW1yyiSTK67F5e1R2Mv7qHxuAQ4ztfYrhsvuWIt4JBaBz64fMU2ayVqRCM1N3L4Wd3WOJ5K3spJcePYBFdLg89Hh8BcMuW7zAeT5Akc7kcTdNkPp9nbY2fO3fvJi93sg8OBqEwseeXlltXAxza/mMGNDcHt36/YDNJDv1oB0O+coqxMUohrHEzgNPp5syZpSwpmUEAXLx4CTs7O0mS2WyWJPnppz2cXlzCK8EAx556jldhZwBujh75Q8G1zMeXKXK5gnjq2An2Axw58Nvx7PMF8QKA319Lp9NNAPT5prG5+VG6XEV0Ol08deqUtXEmQ5J881dv8r6mdWQiydCMuQzp0xjyljHb1U0KUThzSslcXz9DJeWMLlhMM52elH0BYOHCGvp8xdy58ycMhazb397+AZcsWUoAPHTo8A0QmzY+xP1tbeTvjrAPKsNaMWOr7qPIZChzecq8QWEYvLZ2IwMAE6/uvWX2UkpidHSUdXX1BMAHH9zIM2fOFOxMJpN8/PEnCIBvvXWwUBmRcJiNjWt4bWCAQw1rGHKUMAid8RdeKqyNv/gzBuFgSHEx+drrpCTFrQCEEAwEAtyzZy+rqqoJgCtWNBSsJ8l9+35NVdV48ODbhW+H3/09n2ttJS+0sw8aI94Khuw+5vr6mbl4iUGbi1HfbIZsHn7+QDPldaV3A4CUsrCpaZo8fvwE165tIgBu2PAAe3p6SJIXLlxgeXkFjx49Wpj/zFNP86PubqYebmFAdTOkexlb9xBj9Y0Mqx5GS+YyBDfDc/00R8coJSffASklhRCFep+IS5cusbn5Ec6cOYutra0kyWg0ytWr1/Ds2bMkyc96e/niK69QfPgxA6qbEe9shmw+hvUSRou/xBCcjFTWMn3+H5bw7RyYGEIIGoZB0zQLIB0dl7lp08NsalrH3t5eZrNZPvnk99jd3U2SfK/tKNs7/8lMyxMMwM5I8RxGiucwCI2xr22gEQwXKuNm8UkA1w/TNGmM1zJJnj79F7a0fIsnT55kKpXi/v37OTAwwGw6w2Nn3mf6b+cZ0r2M+CoYhM7Brc9Q5CxXJ/77dwRwKxDTNHn48BGeO3eOIyMjvHjxQwohGAwEGBse4vDq+xmEjcm9b1jCUt7S9uuHMtV3gRACmmY10bFYDF6vF06nE0II2HUdUlGQbvsjYJrwfLvFakam0KBOGWAipJRQb2q1JGk1o4oCBQBNAahTe3LcMQCAQt84mY5W5zNF8S8M8P+Mu/40u+sA/wXDj4+b4P8QfQAAAABJRU5ErkJggolQTkcNChoKAAAADUlIRFIAAAAwAAAALQgGAAAA6ErowgAAAQhpQ0NQSUNDIFByb2ZpbGUAAHicY2BgPMEABCwGDAy5eSVFQe5OChGRUQrsDxgYgRAMEpOLCxhwA6Cqb9cgai/r4lGHC3CmpBYnA+kPQKxSBLQcaKQIkC2SDmFrgNhJELYNiF1eUlACZAeA2EUhQc5AdgqQrZGOxE5CYicXFIHU9wDZNrk5pckIdzPwpOaFBgNpDiCWYShmCGJwZ3AC+R+iJH8RA4PFVwYG5gkIsaSZDAzbWxkYJG4hxFQWMDDwtzAwbDuPEEOESUFiUSJYiAWImdLSGBg+LWdg4I1kYBC+wMDAFQ0LCBxuUwC7zZ0hHwjTGXIYUoEingx5DMkMekCWEYMBgyGDGQCm1j8/R2zgUAAAC8VJREFUeJztmXuQVNWdx7/n3tvvx8x0z5swqAw+eCQThRAEJG65A5qUQdakUhW2FFCIGx8xKpkYXWvDKmvQhXI2YhJNSEyqEgHBjYMx7kqK6LoKxigoA4wQuvv2DMz70T3dfe893/3jdt+ZcWaYcZ2UZVV+VadqZu65fX6f3/n+fuc3p4WUkvgEm/JxO/BR7W8AH7dpAEAS5F8nFYQQzijYVK6lAYCi/HU3giSy2SyklNAUBZqmAVO0piaEwJ49e9DY2IhQKAzLss7pCAAnmkIIB15KCcMwkM1mkcsZAAi/14toNIoZNdMxa9aFuHjuHMyuq0PE44G0LIgpgNAA4OTJU9i/fz/cbi8MwwRAAOIDUwkhBFRVBQCYpglyJGw4GEbtzAswb9481F16KeZ+tg4zZ89GRXk5/Pk5Zsv7yFkS7otmgVIC4oPrfAgTwgbweDxQVRVFRUUwTXPc+aZpor+/D4qioCgcxozp0zFv7hzM/9wCXLZwIS6p+wyi/sDQC9kc+M5hpJ/Zg87XDyJ3+D3IlhYgFET5oVfgqq6yIT7sTpCAomDw3feGktiyLGeMBhUwTRPRaBQbNqxHXV0dLl34OdTW1sI1bJ5sPo6Bg29i8I03Yf75bcgT74Nn2gFkIaAAwg01EIDVlkTPnQ0ofeYXtjMfxiwLQtPQ/59NcNVeYANMxhRFQTqdwvqbb8KsCy8EzrYj89sm9L/5Z+QO/glm83Eg2QpkBiBACLiguL1AIACqIUASIl/tlFAE2Z3PILV7FYL/sBI0TSAvzck4n/r9f0MmW+G99ouTAyAJTdPQ3d2D61Zeh1df/19oJ1rQce1KqAAUuKGoLsDthghFbF1LaUeXBIy8LIUASQgAqsuH/nvug++qK6GEQva8c+WDlBCahlzL+0j9+Gco/cWPQcsqHGTn3kaRXzgajeLdo0dxf8O9CCxehODmh0GoECVRwONxokTDGAIARpwB+YgAPj/kqWPo/7d/t6uRlOeKIEBAZrPoXHMLQnf8ExS/H6BzEo8mL1QcVVUhpUQ6ncbZs23wen3Y/qMfYV9TE6IN34ay5ArInk4IVQWEgMwHY/hhRfsDRy5gmVD8EaQaH0f2aLP9vjUOhGVBaCo61t4Cd908+JZeDhomoCqjW4mC07lcDt3dXeju7oKqqpg7dw7Wr9+ARYsWwbQM3HH7Hejo6kLpkz+EDATsqAsxNMaA4PDdIEBNhUz1o//+f82/N0bwTRPC5UJv43bkDryKkkceAi0LUAviIfnYY40EwEikNL+OwtraC7lmzVo+/fQvefz4cRqGQZJsb2/n7NlzCIBrb7iRJNm57YeMQWMyVM2kv5y6J8qkt4zJcDWTwQom/eXjDj1QybgaYPrAKyRJaRiUUtojv2b6tTd4Gh6m/mv/qDmjADZs+Ab37dvHzs5OftAymQxJ8uDBgywpiVBTNe559lmSpL60nrrwUw+UU4/WMFlUzQRUJrViJkNVTAYrh5z2l1P3ldm/ByuZUIM8s2wFrcEMpWXZzpkmaUkaHZ2Ml09n+9fXjQaUkiMkZFkmgsEA6uvrEYlEkM1mYRgGLMuClBJutxuGYWD+/PnYunUrTMvCxns2oqO7G9Ht22B5vIBpgQrg/8EmBB/cDEyfBqu/EzKVgtC0oVwQwqk8ClQY7x6F1dlpSyxfAKgIdK5eBw6mUbztYfvvowvC0A5Eo2UEwNraWdy3b59NLCVN0xxBncvlSJJ33vltAuA3bt5Akuz4l82MwcW4N8rEjIuYPXmKVibD3ieeZOunFzIGN3VXCVvDttSSoSrq7hImItVMv/a6vZ5pUuY/v+v7D/MUwN7tT40Z/TElFImU0uPxUdNc3LhxI7PZLEkyl8s5L1mWRdM0mclkuGzZFwiAL730EmlZjM1dwIS7mHE1xHj5DGZbTpIkzVSavdt/Qv2COUzAzWSokro3yni4gulX/mfI+bzuB154iaeFi62XLaGVy9nPCvI6F0A4XMzi4giLiyMEwMWLl7C5uXkURCGpW1paWFpaxrmXzOZAJsPU/gM8LXxMFk1jQgSoX1THXLLVySOjo4td99zLuLuYMXfxUGLmnSTJ7Ok4E5+qZUxxc+D5F8aN/jCAx0YAFEZZWQXdbi/Lyyu5e/du2wHDoJWPREFKO3fuJADed+/3SJJtN2xgAi62RmoYh4dtf3cNrWyWMpOhzIOk/nCAA00v5p037OgaBqVhsO2qLzEG8Ez9Siehx3J+3B0oKYmypCRKTXNTCJVer58ul5ubNm0iSUdCwyG++c1bKQC+c+QIrfYOxkprqHsiTBRPYwwqO275lv1uNufIxJGNlEO6b3iACbgZdxWNWTYnBAgEQgQEhVC4dOkVbGj4LquqplFVXRRC4U033ew4b5qmA5NKpXjxxbO5ZNHlJMmexicYh0Y9VEU9XMU4XOzb8bQNkTMcvUspaeWB+n/7AmMiQF0E2bZsxYTRdwAaGxsphGBxcYTTpk3nV7/6NT733HNOdI8cOcIrrlhGIVQC4HXXrWIqlbKT0zSdfDh06E0C4FNP/ZQkmfjM5UyoQRvCW0q9qJLZ5mMjo2qapJTMxRNMVF5gJ7biZ/+vdw3lxkQAW7duIwCGw8W89tovs7V1KOnS6bSTwBs3foder58AuHz5Cvb29o6CePDBhxgOhNjZ18fBP/yRMeFjMlBBPVzFhPCzbenf0zIMSsOkNC0HpG35SsbhtQ+2YBWzp/4yUmLnAvj5jp8TAIVQqaoulpdX8NZbb+OxY8cckMJu7N27lzNmnE8AvOqqevb19TnJbeSdWbBgIW9edxNJ8sz1q5mAxz6NQ9WMQ2XPlm22lAYHSZLdDz3KOLR861FJXfiZmqD6OACWZdEwDL7xxkHefvsdPO+88wt9F0tLy3nXXXcxFos7kSbJkydPsr5+BQHw6quvYTqdpmVZDmRzczPDwTD/9M47tE6d5ml/lLqvjHqggklfKRPhSmaOHbd3+NXXGHcXDfVH4SrGoLHrO/dPTkJSFgqbbWfOnOHjj2/n4sVLHM1XVU3jli1bHN2TZDab5d1330MAXLVqlZPUhYNv8+aHnYTuvPteO6HDVUyG7IRu//pamqkUk5dcxoQaHOqVghVMqGG2fv7KCeXjABQWN4aVN9M02dTUxOuv/wo9Hh8BcN68T/P555tGAO/YsYOA4Nq16xypFQ68xYsu5zPPPkv29fF0aQ2TnoidD95SJko+xdYrr6GuhG3nC81deBrj8DJx3mwa3T2klGOewKMAhrcJuZytu4IdOnSI69dvoNdrg6xe/Y9sa2tznh848EeGw0VsaGggOdS1vvXWW1zy+UXMkux56BHGodo695dT95ZRF6ERXWoyXM0E3Gz97CJmjrxnOz9REp/roWEYju4L2r7tttsJKCwqKuauXbuGPTvG88+fye3bnyBJDuYT9IF/foBP/ORJMpvj6epa6u4S6oEK6vlW2v6foIJ6sIIxaDz7ldU0e+3CcK7ITwpgPHmdOHGCa9aspapqXLNmreOsrie5ZMlS/u53LzpyGhgY4LobbmSfkWP/1v9gHJrTjeq+MiZDlUx4o4zBx677NzmtxkTV50MBjAfy9ttvc/nyFZw/fwEPHz5Mkjx79ixvvHENm5uHSvDvX3yRv961m8xmGaucyaTbzoVkuJq6EmY8WMb+X/3GdnycrnNKAMYD2blzJ+vrl3Pv3udIkh0dndyy5RH29PQ4ufT0z3aw3zLZe9/3mYDG1kgNdXipz5zDwdcP2s5PUDKnDGAskL6+Pj766KPcs2cPpZRsa2vjyy+/7DxPJhJ8X0+QsTjjoQomoPHMlVc7rfb/x3kpJYWUH/07Msuy7CtzAEePHkUgEEBNTQ36+/uhqip8Ph+EEMhlMnB7vWhf+TXArSL6qx1QXK7J38yNYVMCAAx9SVK4vZZSOlfvLFxwwT7izbPtUMvLoQjYVyQf4Zp9ygAKRtK5+xn+84hF8zd9E14nTsKmHGBSVrjw+ojOA5j87fSU2hQ4XrBP/LeUfwP4uO0TD/B/r/ybVXrvG3kAAAAASUVORK5CYIKJUE5HDQoaCgAAAA1JSERSAAAAQAAAADwIBgAAANaJvGQAAAEIaUNDUElDQyBQcm9maWxlAAB4nGNgYDzBAAQsBgwMuXklRUHuTgoRkVEK7A8YGIEQDBKTiwsYcAOgqm/XIGov6+JRhwtwpqQWJwPpD0CsUgS0HGikCJAtkg5ha4DYSRC2DYhdXlJQAmQHgNhFIUHOQHYKkK2RjsROQmInFxSB1PcA2Ta5OaXJCHcz8KTmhQYDaQ4glmEoZghicGdwAvkfoiR/EQODxVcGBuYJCLGkmQwM21sZGCRuIcRUFjAw8LcwMGw7jxBDhElBYlEiWIgFiJnS0hgYPi1nYOCNZGAQvsDAwBUNCwgcblMAu82dIR8I0xlyGFKBIp4MeQzJDHpAlhGDAYMhgxkAptY/P0ds4FAAABHPSURBVHic7Vt9lBPlvX5mMvlOJtlks9ldVuqCfFMLy4cIBfRUe6+26MXrF2LVfklVUMG26r09rbY95bTqtZWCWmnR21puFWiPFgHtEaif1ZaCH3wKuEkmu8DuApvd7CYz8z73j8lkE3ZXFxT3j/o7Z85JMpOZ9/f8nt/zPu9kIgkhiH/hkAd7AIMdnwIw2AMY7PgUgMEewGCHYr8wTXMwxwFJkso28pOZnCQhBCVJ+kQuNpAQQkAIAVmW8UmMS5EkCblcDt/5znfR0tICRVEAEIB9cfu1XRHr9UAKZI1fgixLkCSr20zThK7nkcvloes6AMDn8yIajaKurg7D6usxeuRIjBwxAl6fH3Cc3i6VSLKzM4uzzhqJ5mYNgAOAONXTQZIkyLI1aCEEyN6tVVkZw1nDh2Pc2LEYP348xowehfphwxCvrUVQVcuEiULYSH68QUKSZUsDJAmoqAgjk2mHx+OBEP0DIEkoq77D4ShS1TAM5HI55HI5mKYJj8eDqlgMI0aMwIQJZ6Nh4kSMHzcew4cNgz+k9j65YSC3aw/yb78L/aWXoXx2PNQbvwYa5sfLhELyRi5XLoKmacIwjAELkGma6O7uhmlaVA4GVQyrr8f48eMxadJETJrYgLFjRiM+ZEjv74LQG5PI79wFY9tbMLb9E+bOPTCTKaAzA0AHXAG4pk6GZ8LZoGkC8scAgp18axuyu/f2AHAyIUkShDChqioaGhowYcLncM45U/G5s8/GWcOHw+PzlR0vAOitrTD2vgf9rXdgbNsOc8c7EPsPQrS0gNAhwQHILshuF+RgGFQUiONHcfyWJXBt2QBJli3qfZR2IAESQtfR/tvfQ51/9akDoOsGYrFKbNq0AV6vt2y/ns1CP/g+jHd2wti2A8aOt2Ds3gekm0G9E4AECS5IbjdkfwCUZUggIAgKYVXbNCEFVORf3Yr2ZY8gvGQRaBiAw3HqAJgmJKcTR+9dCvf0c6DEKi0RzGazaGiYhMbGBNxu94BaQFEUtLW1YOHChVi2bBm6cznk1z2D3OrVMA9qEAcbwc7jIAQkKJAUNySXE5KiWOcv2YrzTKHCpa9BAbqdqPz7S3APGwYhhMWGkw3DgOR0on3Vb5F/5XVUrlwOGsapO0HDMBAOR7B8+Qr8ad06eNxuyHW1yD27Eca7uwASjmAF5GAUUkCF5HYCgFVF0wSEKKppMeGS15IkQSIhOV3AsTa0f/f7FjinYpAKle/++zZkfnI/wkvvtWYXWe4B4FSNl9frw8Lbbkc6lYJ/5gwEVjwEMgcqikVlO2Ha1ym/UOm7XsZHkkBdhxyoQH7dOnSu3wBJUazzDTQKjDGOHsWRS66E+sPvQYlVWgWQpN4MGIj7kiQJDocDsizD7/dDS6Vw6+2LIQkB/41fhfNLcyDaWwFFsexUwdrayRerXaC79aYcnNJjAEBSXMjcdQ/Mri5AkgdWMVv0ZBlH5l4N1+fORnDeFWVa0guAvvrfTlhRFNjO8ejRo2hra0FraxsqK2NYu3YNfv2b38DpcEBddj+kyjiY6+6ZumQJLAG3FJAiAnaykoQSaKwq+gLQ39mGzIpfQXLIVgU/LH/ThKQoOHr3D5B/9TVUrHq4WPniVW0RnDhxEhKJHhEsdXS5XA7ZbBcAE16vH0OHDsXYsWNw7rnnYv977+GxlSuhqiHIDhl/e+UVnDVqJI498SQyN3wNSiBiDUSSIAqOurSP7QoTAEGckHpPMrLVDlI4hNj216BUxawK9yeIBdHrWPMntFwxFxUrH0fo69dDGAakkpmkTwCs7xvo7OwEANTV1WHKlMk477zzMH36uRgzZjT8/kDxJHPmXIoNGzbA4XDg/PPOw3N/fgbCqaD18q9AX/s0HMEoYJrFJOXCugAnUL6nHfqAgAQUBUamBb4ldyD6wNJeyRSjUPnud3fhyJRz4Zw2HfEXnwP6cJS9APB4POju7oaqqrjggi9g7tz/wMyZMxGPx0+4huUanU4n2traMGPGTKTTaXR2duKhX/wcCxctRFciidYpsyAfb4ekOC3lJYHuLkhuD+B2W5T8ADrbbCQJyJLlFZwOxHa8DteZn4EwTKsl7BACkiTB6OjA4ZkXwtyzE7F/vAHPuDF9W2prMdTJkSNH0+cL0O328qabbmYmk2FpGIbBfD5PwzBomiaFEBRCUNd1kuQbb7zJYDBEVQ0zGo1xz86dJMmjjz3OJJxMB2uouSuZqqhj84VfZjpyBlNwUpOCTAeqqQWqqXljTPuqmPZVUfNXUfNVlX2W9lUxHaxhSvLwyPyvWWPQ9eJYhGla70kemv91vg+w9Xs/Isny40o2+US03W43nnnmGaxYsQL5fB4AkM/nIcsyFEXptU53OBzQdR1TpkzGL37xc2SzWXR2dmLxt78DGgb8X70Wzn/7EkTmKOB2gZ3tkM/8DCLbXkbwwQchT5oAs6Md7DhuGZwCpaXi6lsqmystNuigLPeYJXv8Beof+59fovvJX8M5eiLU795mMaw/rShlgMfjYzgcod8fJADOmjWL27dvLzLAMIw+URRCMJ/PkyQXLbqViuKi4nTz0YcfIUl27NzNhBpnyhNlMljNRsg8smARSdI0DGae+TObL/wyk/AzBT+1YA3Tgeryyvuq2KTWMgmFh+bdYLHQ3gpMIMmOzX9lwlPBhOxh+1PrPrD6Qgj2CUAoVMFIpJKyrDAUCvOhhx4qtoLez8lM06RhGOzu7uasWbOpKC4OGVLH9/fvpyDZsvR+NkKhptYypdawERIPL1hYPK9JsmPjC2w6/9+ZgIcpR4hptZaaP24BoNYyAYXNV1xL0zDI0uQNgxSC+aZmpurHMQEHm794KQWtff0l3y8AqhqmqoZZURFlMBgiAM6dexnT6TRJMp/P93kywzBIknv37uOQIWdQlh2cN28eKQRz2SzTkz7PpCNg9Xuwho2Q2frf91CQFLmcVS2S7aufpjZ2sqUd3hi1CksvDl1yBc18njRNmnZidt8LweaL5jIJL5PuKLOvv2EB+wHV/1AAVDVcZAMgsb5+GF944S9FJpSK4Ymt8PTTa+h2++h2e7nmD38gSbZv3sqkEqTmj1PzVTEdqGYjnDz68GPWYLu6KEzTarlMB9t+8GMm1Wo2Amy6YA6NbFex4sVrFq7Xdu9SJuBhEn4e+co3LFA/JPkBAWBvkUglPR4fXS4Pf/rTn5XNDv2BsGTJEgLgmDHj2HLoEE2Sh7+5iAk4C/S2lD7hCLBjw/MWA/L5opKTZNeOt3n45lupt7VZ+0uuZ9p9v/EFJuQAU/44k2o1u/fsI4WgqX8w/fsEIBSq6JV8KFTBioooo9EYw+EIAXD+/GvZ0dHRpy7YetDZ2cmZM2cRABffvpgkmU03Uas9i2l3BTV/nKlANZPuCiajdezevbcnyQK1S0OUMq7Q97lkilrtcKa8USbgYcstiz9U+AYMQDgcYSRSSZ8vQGtikunzBVhVVU0AnDZtGhsbG/vUBdsfvPXW24zFqhgIqnz1pZcsuj68kkk4qQVrmPTGmFJrmJT8bGqYTr2jowcAm+66Xp58ARzTNNh84Rxq8FILVDMVjLN733tW9T9E/PoAYFQRALvibreXAFhbO4R33/1ffOKJ/+XQoZ8hIDEer6HD4WR9/TBu27atTxDsVnj00V8RAGfPPp+5ri7q3Tk2TTufKdlPLVDNpC/GdGGKO/KNWwqt8AEVLJy39Z6fsBEOpivOYAouHrl+wUlVv18NCARUAjLHjRvPBx54gE1NTUUaNjY28pJLLiUARqMxer1+RqOVfPHFFz+QCddccy0BcPmyZQVB/CuTSpApf5wpX8xyfMEapuBi+5P/138idt+/8CITjoAlqP44k4rK7Gt/I4U4NQBGjRpDr9fPYDDEuroz+NRTT5f1Xz6fL1aUJO+77376fH76fAGqapiBQJDPPvtsLxBskWxubmZ9/TDW1tYy2dho2dXrbmQKLssme2PUAnFqnihTlXXsPnCgN5UNgzQF84cOMzV0NFPOELVgDVOSn+lzv0DTbo8BJl8GwJlnDqPfH2QkUslIJMqLL76Yq1evZjbbWUxa1/WiIyTJl19+mWPHjqMsOxgOR+j1+rh27dpeINgsWL9+PQFwwTctqmYPvk8tegbTnsoe76/WMAkPD1001zIydjVLff7ca5iEm1rBKGlSkM1TZtE0So4dKABCCHZ3d3H27PMJgLKsMBgM0el0E5DY0DCJq1Y9zpxtVAoJ2WxobW3lvHnXEJCpqmF6vX6uW7euXxAWLbqNAPj6K6+QJFt+fJ9leNRaaxFUbAUnjz+2iqRlZszC9Y499AgTUJgO1vYslLwxav44c/ves8Y4QAEUQhC2meno6OAf//hHXnbZfzIUqiAgMRAI0ev1U5IcbGiYxCeffJJCWDO0YRhFUEjywQd/zkBApdvtZTCocv369WUg2MzJZDIcPnwEZ0z/PE3dYL69ndqoCUwpao/t9ceZdkeoxc9kLt1UNEfZ7W8x6auk5q20jFTJCrERLmZ+Z2mH2Y9T7ZcBdlJ2vPvuu7zrrrs5fPgISpDp8fjo8fjocCicNWs2N23aVDw2l8sVq7tly1aOGDGKgMxIJMKtW7eWgWAft3nzFgLgbx9/giR59HermYTLmsp8MSu5AguO3GitF4xslunJM61+P3GhFKxhAi62FBZYJyWC9gvTNIs9bkdbWxtXrlzJGTM+T5fLba3yFBfdbi+vvnoed+3aVTy2q6uLJJlKabzooosL02ctd+zYUdSPUhBuvvkW1lTXsKO9nYauMz39C0xJfmrB6uJ9AM1XxYQSZHbnTrbeu5RJyOXUL9A/HYgz6axgekwDja4uywYPUAfQ14f2zQ87hBDctGkTL7/8cqpqmJLkIADGYlX80Y9+zGw2WwaCrutcvHgxAXDEiJFMJJLFtrFdYiaT4ZAhdbz3+98nSR7f8DyTck91NZ/V1ylniMlRDUxFLLE8cYmseWOFlaKLqdEN1FtaranwowBwIitKW2Tbtn9ywYJvFd0gAE6dOo2bN2+2+q/wHZJ85JFHKUkyp06dykwmQ9M0y/Zv3LiRkUiEzVqaphBs+uKlTMFDLVhCcX+cmlJhVdrWiJJ96WANk3Cy+cIvMZfSCiJ4krPAQLYT22Pv3r28445vc8iQOgKgojh555139mLDxo0b6fP5OWfOJUV2lIIwf/5XeNON1rTYvvUlJpVgWaJaQeHT/qoy2mv+ODVvjEm42LJwSXFhdNI+4GQOtttDL1mkHDx4kHfeeRfjcYsRDQ2Ti9bYBmH79u2srIxx0aJbi6Jot8ORI0c4btx4vrd3LwXJpkuuYgoepoM1xaQ1b6zY9ylvzDI/SphJV5jHVjzW0/MnMf2dMgD9AXHgwAEuWXIHXS6PZXmXrygma+0/yKFDz+Qvf7m8+Ln9/ZUrV/LmBd8iSWZeeY0pp8omfzXTvhOY4KuiptYyCR9TtcPZ+Rer7UReP+nKf2QASnWiVDB3797N6667ngB41VVXl3kFTUtz6tRzym6q2Fb5+utv4L49e0iSTRdf1sMCG4DCdJeEk01TZzG3/0Ah+YHP+acFgP4Y8eqrr3LatGn87GfP5r59+4qfNzU18corr+L+/fvLGPLmm2/yZ0t/amnBhueZkn09832gmulAnInCDVGjcB/iZAzPaQegPyCWL1/OKVPO4caNPeYpkUjwnnvuYSaToRA93mDVqlVMJZMUpsmmKbOpyQGmQ0OY9lQyCS/b7v1J8U6RGMDdnkEBwN5stSfJZDLJJUvu4Jo1a4rJHjhwgM8991zZ7faWlhb+4w3rZuaxR1YyKXmYdkeZClax/fdPFRI/9X7va5OEOL3/F7B/PgOALVu2IBqNYuzYsXA4HDh8+DDcbjdCoVDx4cjuri64PR6Yx47jUP14IOBFZO1qeM+ZDOo6oJzSUz39xmkHAOj5yV2WZeTz+cLDmNZ7nvADqSzLYOGhhuOPrYJn1gy4R408LckDnxAAdggh4HA4yh6U6PeBjMJP3xLw8T0i10d8ogCcdJim9dvfaUoewKk9JveJxUd5JG6A8S//f4FPARjsAQx2fArAYA9gsONTAAZ7AIMd/w93HQM5GMK5zwAAAABJRU5ErkJggolQTkcNChoKAAAADUlIRFIAAACAAAAAeQgGAAAA8aMpGgAAAQhpQ0NQSUNDIFByb2ZpbGUAAHicY2BgPMEABCwGDAy5eSVFQe5OChGRUQrsDxgYgRAMEpOLCxhwA6Cqb9cgai/r4lGHC3CmpBYnA+kPQKxSBLQcaKQIkC2SDmFrgNhJELYNiF1eUlACZAeA2EUhQc5AdgqQrZGOxE5CYicXFIHU9wDZNrk5pckIdzPwpOaFBgNpDiCWYShmCGJwZ3AC+R+iJH8RA4PFVwYG5gkIsaSZDAzbWxkYJG4hxFQWMDDwtzAwbDuPEEOESUFiUSJYiAWImdLSGBg+LWdg4I1kYBC+wMDAFQ0LCBxuUwC7zZ0hHwjTGXIYUoEingx5DMkMekCWEYMBgyGDGQCm1j8/R2zgUAAAMxNJREFUeJztnXmYFNXV/z+3qrp7ZnpmmH0YVERFJQqiqCiuQYkb4u/VxMRdcUOjaF73FRU1iprXaEwiKhiXxBhNjAvuoIBb4h4B0YCRXWD2vbuq7vn9UVU9PT3dMz0LiEnO89QzMz1V996659yzfM+5t5XWWhggEhGUUpv92f9S9pQ6z8ZANt4fBv67M19kwNZZvyh1ngdUAP4TKVvGBhO/pQhCQP8VgH5SbzXXlqbpMgrAliap/0m0Oec+owD0R1I3t/Ck9vddF95s5n6gNEkXARiIyetucL1pv6/2NZvJ+a4LyUCNXw1kGLi56b+hY//pO+0EZuNZf9dXen8om3f/TgtAQN1pgf80DZHM9HTvLiKd7umXAPR3dQ3E6szUxn/qyu9J4JVS2SGB2Uxgf1fXQKzOTG18myv/uyR8myQM3JJooJjRm3a+S3P3b+EDdEcDxYzvElO7o0CQg59W6g2u6yZ+V0qllfzkzzPdk466uzfd/7Jpuzf9Z/ts6ufJzE+1oVsSZRMWB/9P/EzFAbbUl9uSqD9mJROT+iPI/aEuArBo0Wc0N7dgmgb9G48A3Ych3jyk3hM81/G8UsGkJ69EECFjW94kd20rXV+GYXRiiuM42LZNPB7Htm1EhNzcXAYNGkRlZSWDBw/+t1koFniTaBgGjuNw0kmnsGjRZ+Tm5qG1TvtQf9V+wMx0cyhChs8zqzev+ezUX/Llui627eC6dqf7cnNzqaqqYtiwYQwfPpwRI0aw8847U1lZSV5eXrd9fNfIgs5qPxLJIRLJIycnN6MAdKwq/68UnwA6mNLlSaUSTM7E7I7nO+4LGNwhS5lXdbKdMwzPz9VaE48HqzoOCDk5HqOHDh3KiBE7s8su32PXXXdl+PDhDBkyJCOzxXW9QRl996EHGsbua3sJDRCoTK01IroLYpTSXZfO0/2eaaDeTzr9hEzaIv3zyeMIVrRhKMBAa41t28RiMRwnDkBOTh6VlRVsv/127LLLLuy2226MHDmS7bffnsGDB6cdq9Ya7bqI1ijwTIVpokzT613rzBLcA2XWZn1jZPIzWbch0jUKSLdq+0q9dWyyFZ4OhhuICI7jEo+1E4vHAE0oFKa8vIIddtiekSNHscceuzNq1CiGD9+esrLytO1qrdFao0RQfh8YBmYo1EnPaMBeuQq3rp6c0aOgH0KQjjYlONaJRFCGkSoAivRqtW80UF5t4KQppRKru7293bfdBoMGFbLziJ0YOXIke++9F7vvvjs777wzVVVVXdrSWidCXZV0GUphhkKd7wWcujriy74i/smnOB98gvOPxbD8K5ymBkrfeJncfceC44L5HYJUfOa3VVen0wDfPhmGkVjdruvS1tZGLOat7nA4h4qKCnbeeUfGjNmTvffei1GjRrHddtsRiUQ6tROs6kAlGuCrPQWmgUqy4QI4to3z9QrszxZhf/gJ8Y//gfP5F+i16yDegsJAqRAqLxfTdqi/4BLCb72GFQ57wv5diAy0RpkmzZ8tJt5Q39UJ3NyUrM611jiOQ2trq2+7FYWFhey0046MHDmKvfYaw5577smIESOorKzs0pbjOAlmJ1a2L+2YZifdJkB8wwbsz7/A/uQfxD/8BPsfi5Gvv4aGesDFIIQRDmOEw5CT47m+IojWSH4B+qP3qb/tLsqmXw+OA75vsMWSz/z46jU0//k5Kq66JNUJzI4GErQI1LnWDoZhUVFRzujRoxgzZgx77703e4zZgx2H75h2dQeq3PAZbgKC8hy1pPfRgG5tw/nXV8Q+XYT94cfojz/D+fKf6G++ATeOQoEZwQyHUQVFCCB4jrChBXEdSBIs7bpY0WJid95N86QjyN97b2RLNgUiIKDtOBsuupyS667EyIl01gCbEolKB69q7VJRUcaoUbux555j2Hvvsey22yi23nrrLs8nr24DjwkWCgyV8MrxP9eAvWoV9pIviH/0KfZHH+Ms/hx31WpUcxMGGgihIhHM3ChK5XsMF+1Nkuug6Biv0NXLNlxBLAvVZtN40RXkzn8ZZVqoLdQUiOuiQiE2nH0hoZ2GkzdmNOLYnTVA6kumo4FY/aZp0tjYyI033sjFF19IQUFh58H6tl9EEsw2AZQBVldVbtfXY/9zGc4ni4h/+BH2J5+hl/8LqmuAOIKJsjpWtycl4q0KrdHJYWziRTO7w4kF47qQX4h+723qf/kbSq74X8RxOgnkFkGOgxEKUXvPr7EXvMeQRX9Huy7KMD0oOBkJHDt2XxYvXkxeXmYksC+ULDiGYRCLxdhpp5144425FBYW4jgOpml2gWWVUp0CUx2PYX+9kvhni7A/+Ajnk39gL10Ga79B4i2AYKgIRiSCsizEMALM2AvZ6Gzykle5KDAkc2IqlTSAoUC7SCRM2XvzydlpB8TV/QKJBpLEcVChEK0L32bjQRMon/cS0fHfR3yfpVdRQH9Wf/JzWmvy8vL49NOPmTJlCn/6058QEczklaM8ex5fuRL3n8t9Vf4JzuIluF+vhqYGBAeFhRGKQCQMOSVeX/7qFq0TTE99j9R36k5pZyw6AU+TWCGoq6P+smuoeO7JtG19K8kerTFCIdrXrKXm//2I6GmnkDe+g/lAVw2wzz7jWLRo0YBrgHRkWSa1tTXMmjWbM8+cjGPbmD7a1vzW29RPuQhzYy1udTUiNhYGYni2W1kW+JOqRYMIhnSs7q7QdBYl40n2O9P9GTWDaeI01VLwxKMMOuHHfY4KeuOQd3uvr/Vcx2bjhInoz7+g/PNPCJeWeOiq4T33reop19Xk5xdyxRWXs3TpUqxQCO2v3JzdRxMuKUVtXItRMAizoBhdMAjycj3Gu64nya6LcjUqKamZjnndObqJz3zB6e6ezBOuMa1cWq6bjlNf75mAPqz43kZjGUm7YJrUXnQF7W/NI3r3nYTLyzytaGSsCRQGEgruiUSEUChEfX09U6ZMwXG8UEu0xszPZ9BjDxArrYRYO2jtMd3VGXMPPTEpUzFHtlFQavudnGYtqNxcZPmXNNx2l4c9ZKlBB9w0OA7KClF/72+JP/Arco/8IYUn/wTtOF3C1E2yMyiVupNU13UZNKiYBQsWcPvtt2OapudcOQ6RYcMovONmdKwJVHbqNJU53b2PVt6VaZy9Wo14UYGZW0TsN/fT/tkiz0xlIQQDCsT54V7jy6/TcPmV6KJKBt11c5Jz3bmvzWICehIqx3EYNKiYW2/9Oe+++y6WZeHiebCDzjwd6/gfo5trwbIS3npA3dnoVH8glZR4V6fPsmSGSmcuRBDThOZW6q6+ET/myKq9ASGtwbJoX7acxrPPx4y3kHv5ZeTu8j1v9aeJTLaMWIUAGNKcf/5PaWlp8f4GDC0U/3IGUrW1ZwrMzmFhJnudXPyYDoRKhotTn+2O0tUKdjINrouRX4Q750Wa//osyrQgqc5yk5H/nm5zCzUnnw1r/oWx14GUXHKhBwJlCEuN5BfJlA3cHIWQWmvy8wv49NNPmDbtxoQp0NolMmQIBb+8AzfW6idkOjt3PY0vtRI2+dlU6s5/SG0vuf/O9wmWFabp2ptxWlr67BAmU7eCKYK4Gtcw2HjhJbh/fw8dKaTwzumYOTndJqqy0gDdF4dkR9k87zgORUUl3Hvvvbz++lxCluUJgeMw6MfHETntdM8UGKb30qltKuVfZMxsB++S7p1ExNMuSUztnRfgt6NddF4UveRTGn89E2UYnvfdD+p2AbouRsii7p5fE3/kcRQGkcmnkf/9g9A9IJObDQkMXiK1Jk9rnUj9gocStrW1sd122/Lee+9SUFCI+PfEq2tYt9eBhNavR4UiXnYrSb2L0XmSlAfwI6qrrQ/Gk0zi34tSGEHFEtkJgSjl5QHwkkhaKQzHwS2MUvHRO4S32sqz0QONELouyrJomjefuqOOxRRwh1Qx+O8LCJUW91i6tkl9gCDNa1mmp9K1pqWlhbq6Wmprq4m1xzBNE8dxOtXuRaNRli5dylVXXeMJB175VaSinOL7foEbjyX66KSOxYNyE8xTnlCkY37wbOcqZYWBx/wgQpAsVUDAfK18iFgEiUSQjd/QeNsvPGHLYIq6o0zmy+tMoyyL2IqV1J12NoZh4sbjFN4+nXB5qR/zd8/ijBogNze3T2o/uRDTcRza2tpxnBhgUFpawvbb78Buu3llWmPGjCEajXLccT9k1arVRKNRXNfFq8YyaWio59ln/8rRRx/t5QoALIv1507FeXAmRkGZh7il9A8p4Z2vCbobczIl0BAR36JkbwgS/frmxABEaYrffZO8USO9BNJAaAEf6dOOwzc/OAZ5511ENOYxR1P1zB+zTkoNiAkIKnhc16W9vZ1YrA2AkpIyvve9Eeyzzz6MG7cvY8aMYdiwYQkBCeidd97hiCOORCkjodINwyAej1NRUc7f/vYeZWVlCVNg1zfwzdiDMFaswojkdIm1E945HkMM3RUgSgejdnES6RzEJfIGWdbcCaAsC91Uh3X8D6n40+MDVj4mjosKWVRPvZS2++7DKChBmwbl7y8gZ4ft05qbdGNPbAxRSvm5gH1ZtKhnAUhmemtrK7YdIxzOYYcddmD//cfx/e+PZ99992WHHbbvMohEqtd/PhwO88gjj3LGGadTXFyaKPQIhULU1FQzefJkZs+e5ZkKwLAsmua8Qt2k4zDyCjCSxpkcmnmhpHQyAd0JQboQL9PEpVJqhjHo23MoDZxYCyVzXyT/wAN6dMx6IgnSu7Meo/nscwgXVxCvW0/eHXdSdvlFvWq/V8kgpVTClre2thKPt5OXF2XXXXflsMN+wBFHHMGYMWM61dN7VbtOFwcwmRzHIRQKMXXqxdx3372UlpZ7sDBgWRZ1dTU8/fTTHHvssV7CSCnPFJz1U+zZszALSj34s4d6ht4wMVtKBz+nOqViGNDUgHX44VS8/AyqH2ZAXBfDsmj94AOqxx+FKQbS3o7ac3cGL3gVwzI7oqEsKI0A7MOiRUu6CIBSing8TltbC5YVZtddd2HSpKOZNGkSe+65Z6dUblC9E2iJHl9KBBGN62oOP/wIFi58i6KiooRzGIvFqKqq4t1336a01HNuTMMgXlPD+j0PQK3fgBHOQSWNt8tqTq4LGEDKJABd/jZM3LYmBr34VwoOn9ApJZt1X1pjKOWZwP0ORS1fjpVXQLy1keLXX6Tg4AMS7Qbz35Mf1yUZJJJ+9TiOTWVlBeeddz4vvfQif//737j55psZO3asZ5dtG8dx0FpjmiaWZWXFfAgmT/mm4BGqqgbT3t6e0DZ5eXksX76Ma6+9rlNUEC4vp3DGLeh4rJPAJwMfCoU4Lrq5wQ+ZTG/is9AGaT8ncBCzTTH7voPy0tUtt93lVSAZRu9QYhGUCNowqD53KuqLz1GFxcQbN5Iz+TTyDz7AT/aYiTFl48SnIIFdkS3DMGhvb2PSpEl8/vkSfvvb3zBhwqGEw+EE08FT1ZZl9RkxDMLBoUO3Ydash/ztWyQ0U1lZObNnP8zzz7+QyBVox6HgxOPJOfZYdFMd4ks++GGZoRDXQVeVoyYegeSGcBprkaYGTw1bZlpHqbvQS4LQUGVOOHVa/X5hKa7GiBbiLnyLljkv++BQLyBi1wv56mf8H/afn0IVlUJLEwwZyqCbrktgIr0lo/MLpPq9Xlyek5PLq6++xs9+9r9s3LgRgHg8jmmanVR/f9DCoCLItm0OO+wwpk+fTm1tNZZlJf6fk5PDFVdcTl1dXUITGCIU/t+t6NJyiMc7Cjq8h8CyMDZuIPL9AylZ8jH5D87EOOZonNwc3MZaaGn1hMA0uizIdBOaFmfoZuI78g4KbSpEuzT8/E5c7ePz2cyZ63n8ja/No+m6mzCjgxAtuLFmCm68hkjV4D6XoaVJB6e/0bZtZs68n4MPPpiXXnqJcDic8OIHkgJNcOWVV/DjH59ATU0NoVAI13XJy8tj6dIvueGGGxICoF2XnGHDyL/xOtxYoz+pwcv4P7RJ8+VX0PLHpyk6+3Qqnn2Ssg8XEL1zBjJ6JE5zI7qpwZMd00xAf9kIdEfY38O9hoJYDInkUnjxBSjlM7+nVas1yrSIrVlL/eTzfAc4BM31mPsfSOHkU71kj5Wdr9Vl/KlhYHc4gGVZtLS04LouU6dewPTp08nPz8e27cRKHQjy4GFFfX0jBx88nmXLviQazcd1XQzDoLm5mZdemsP48eMTZWQCrJswCZk/HyNaCI6LVuJBtAoMZeA015F3+88pvfLSBOPseJy2Oa/Q+uAs7NfewHDiGHmFnufuulmp1R6jB8MA7WLbcQp//zBFxx+bHSAkAlrQCjYc9UOcV1/FKBiE4Wpsu90PK/dHOy6qj9hCignonhzHITc3l/z8fO6++24OOOBAFi5cSCgUQkQSApNNe93ZWU+zaIqLi3jkkdmEwxEcp4MZpmlwySWX0tzcjPJ3FBmmSdHdt+Pm5qJdx8P/AwxPQItg5RfTetV11P9mJgK4bW1Y4TCFx06i4sW/Uvz6C5g/PBbHiaOb60EptGkgKoB4M9QVdJeJBHA1drydwkcfouj4Y9F2HOkBhALPx8EyqZ12K/FX52AUFiNK4bbWk3PaKUQP3N+P+fsOLPX6yWC/XWlpOUuWfM5hhx3GTTfdlLDhTko8nonSlVUlU+AP7LHHHtx99900NjYkto/l5xfwySef8vOf34ZpmohSaMcmuvtuRC+9GN1W53n7ySSCFsGMFtJ0wc+o/93jmLm5iG2jbRtcl+jBB1Lx9B8oefNlQscfj2vHkaZGTxsYRtaxNfjRgp8ajsdbKXh4JkUn/AixbZQV6rn6yHUxw2Ean3uR9tvuwMovQbSDsuNI+WAGTbvK28ncR6c7kXxLM+ysGojH4+Tn5xOJ5HLjjTcyYcIEli5dmrDXA1FaZlkWtm0zefLpnHvu2VRXVxMOh3Ech9LSYu6771d88MEHWJaFFgVaM+jK/0XtMhrd1oL4K8NEJWUGFWa0gMazz6Px6b9i+Nu/lWF4YJLrEh23D+V/epTieS9iTDwc3dQITc1gmmhDkYqPakN5kUHgc4gkNIbd2krBA/dRfMqJiG1DNqYySPL8awUNUy7EDEc8j9Mwoa2Z6M8uJDJ0G7TTd0ApEJy0YWC2DQQOYGlpGQsWLOTAAw/id797JIEBDISDGLRz5513Mnbs3jQ2NvobSEzicZvLL78c27ZRhkK7LqFoPoPuug3HdTFEYSjlJWj8lLByNWBghXKoPWUyTa/PQ4VCHcCMYXh7/FyH/APGMfiFv1A85y+w7964jbWoWNwrTVMdSyWoSNbKzy76nr9uaaTg13dTcubp6Hhn5qerRfB/AcCJx6iePAXjm/XoSAQQVFsrDB9B4YXneYdWDEBOoUsY2N3CTSccIoJtezV97e0xJk8+g8mTJ1Nf34BlWQmcoM8D9NGs/Px8HnroISKRiA84QVFREQsWLOD+++/3QCOl0I5L/pGHkXvCCUhLHRhJ2EBwaY2EQliiqD/hVFo++QQjFOoo3TINlGF63rXrkn/U4VQufI3Cmb/FHVIFjdUe8wPQBfzCAQ+CNQwT3VxH9M7bKPnpOZ7aD3Ve+RlNgOvt4K27bjp6/lzUoCJwbJRhgh0j/5pLsQoLBuxgiqyygdmiSl6uwKC2toZRo3bjoYceZOzYsQlItz9lZa7rYlkWDz/8O845ZwplZaUJ5DEnJ8zChQvYfvsdcB0Hy7KIrV7NhjH7YTS1eYmRLuMXMC1oa8EZujUV818lss3WHtNT7g8+U0oR21hN04z/o+03D2K1tSKFg9DaBe3ZY9MwsZuqyb3pJsqmXY3Yfil2T++ulF/ObdH47PM0HHsSRn4UcTWmYaJbW1B77UXFW69gGgYoo8dKlWxyG2lMQPqGsiEv8eNSUlLOF198wYQJE5g584EBMQmBUzh58hmceeYZbNy4EcuyCIVCVFfXcPXV13jv4WuByDbbkHftVehYU8Ih7FzcobwVn5eP8a8VVB9/MnZzE0oZXoiWfKe/0rVtEy4vo+yun1M6/2U49Ps4jTWouINhWYhp0d5UQ+TSyyiddjU6W+bjC5ll0f7VV9SfexFmOOzBBIm5dSm44QqsUMjb+pbFWsqksZMpxQQk/7PvTpzj2ESj+YDBeedN4ZxzzqWlpSWtSeiNsxjkBmbMuJ1ddtmFlpYWRITi4hKeffZ5nnzySc8hRBDXpeiCc7HGHYA0N3oJEr/cS/nYgCh/K3hhEfK396g5fYqnyTMkjYI6f7Ed8vbek4rXXqDgvnuRsiKksR6jcQPhyWdSetfP/a1hWUYOIiCe3a878wKMjRuRSAQlXpm509KIefhhRI863BeUvqeSU4UixYtIhoL7Z18C0KakpIyHHnqQQw45NBElJAtBT0UZnQbrh4HFxcXMmvUQpmkkjoDJycnhuuumUVNTi2GaaFdjhkIU3HULjml6mTRNEmNVooQM28YsLMX+y1PUXH2956xlqoVQCiwT7TgYIhRfMIXyvy1AfnQccswPqXrwV5h+MUbWJs/VKMuk7vpbcObPRRUUJvwRQwRCIQquvQKDvpWVdUddTEBAA9FBUAtQUlLOhx9+xEEHHcyzzz6XMVTsyWYl5wvGjt2bG26YRk1NDaZpkpuby7/+9TU33zzdy0IaCtd2iO43jsiZk9Gt9aiUZFVy5a84DmZBGe133EXD7x73tpZ3o62Un1HUtkN4662oeupxKv7yezCtRGVxVuTj/M1/fZ72O/8PM7/U61fwtFZzE5FJRxM9YJx/AklHtm8gqJMT6LqeE5hNRVBvyTRNYrE48Xg7t9xyM1deeWUCPcw2bRxQ8JxpmvzoRz/ihRdeori4yC9UaebVV19lv/32w7HjWKZFbONG1o85gHBNLWKFOkq0lRe/JxirFKYWbFwPZh23T3Z5+yATZyjfPmfJHP/MnrblX1G936GYDY1eiKkFMRSGMrHj7ZQueIX8ffYeuHrCJOqkAfpaL5FN9Y1X+hUiGs3nqquu4uyzz05kFHvrHCanrO+77z6GDKlK1A+IwNVXX+NjAybadYlUVlJwwzU4seZOE5hI1Xa8CK5lYtku9aeeRWz9eo/5PS0Ew0BUL5nvp51dO079WRdibNgI4Yh33oDCy0421xM5+giim4j50MtcQCZKrR8ImJoa+gX2uqSkjFmzZnHkkUeyfv2GPuEFQVRRVVXFvffeQywWQ2tNYeEg3n77HWbOnOljA56HXXjmKahxByAtTf6JYRlqIFwXycvDWL6M2rPO84o3JMuNMb1Ry35+v2bazbjz52IWDkK5bmJUhha0FaLgkqmJ08l6Q9ne32eRSu0gqA0I6v5zc3NpampCa+mi4j0ot5w335zPIYccypIlS7o4h9m8RJB7mDTpaM4//zyqq6tRyjta7rbbbmfFihWYpuWZCytEwa034CpvA4coSTiBSkvicCcDhbIdVGEpzpwXqL1lhucUuk5WY8pq7lwXQhbNL75M64y7UfnFiOMkgHjDMKGlCWv8eHL227cDm8hyXiB7H6HPAhCsmuBcn8bGRurq6th55534xS9+wd/+9h433DCN9vZW2tvbCKWcwmnbNsXFpXz55ZdMmPAD3nrr7R4jhLQv4GuC6dNvYo899qCpqYnc3Fw2bqzm+uunJeoBteNQMP4gwif+GN1S38mu+2kCkkpJPKcwWkrrTbfS/NpcjFDYY0QfnK9ODPMzl7E1a6k79yKsUKhDwyTdp4Ho1CmYaUC4gdynmQYJ3IfFi7sWhSa/TKDmbdumqamJcDjE/vsfwLnnns3//M//kJOTk7h/3ry5XHDBVJYuXUpJSWkX79+yLNraPAF5/PFHmTRpErZtY6ac9dcdBSjhBx98xA9+8AMsy/JPImvg2Wf/yg9+8AOceBwrFCK2YiXr9zwAq7Utsccw2MyhVEfSSCmFGAYq1oZssxVl775JuKzUG3sGW5xVVbHrIqbJhkk/Jj7nBYyCIgwnyQcyDKStFRk9msHvvY5phYLB9anPnv6fEgYK3cX/ylCJDF1tbTWmaXDKKSfx6quvMG/e65xwwgnk5OTgOI5/Fr/NIYccysKFCznllFOora1OeO9Bv47jkJOTg9YuP/rR8Tz66GOJMDFbCoRxr73GMG3a9X7JmIllhbjqqmu87eaW5TmEw7Yl79KL0e1NKNPyNo5Ix5Vc4aO0RuVG0V8tp+7iSxHD8By9PtQFAIl9fHUzfkl8znNYBcWdmC/gRxIxoqefiBWOeOaih5Kz7kxCj2PSWovruiIiYttx2WOPMRIO50hRUYkUFhYlrkGDiiUUyhFAtt12mFx66WWyZMnnEpDrumLbtriuK1rrxGXbduKeBx98SIqKSiQUCktpaXmn9oP+TNOS++77tT+eru1lupL7nzhxkuTk5MnWWw8VpUy54447REQkHo+L67gSb2qStbvsJWusQbI6v1LW5Janv/IqZE1ehawtrJI1hKT2wdneu8bjWY2p0+V489C44G35Olwoq/P89v2+VueVy6q8ClkTLpHVpUOlfc0aERHRWb5/X6+st4Y5jsOIETsz9cILOWriUVRUVCQ+T94PmE7lBHG7ZVl8+uk/OO+883nvvXcoKSlLFJhAR+avsbGBO+6YwWWXXdYlkdSdxAfa5auvvuKgg75PW1sbhmEQiYRYsGAB2223Ha5tY4XDND7zHA3HnUAov8hL5qTaWbw8P4CBV8zp5ISoePdNIjvv5CF12YZlftt2fSPr9zsEtWw5Rk5u1y1tpoXdXEfojDMY/PBvve1fA3j0bDreZPXFkd6kawoK8tl22LYJ5idnC5PvTaXAWbRtm9Gjd2Pu3Ne56KKLqKurxfGzd8l9FhUVc/nll3Pzzbd42L7uOBiqJ6jYcRy23357fv7zW2loaCAnJ4cNGzYybdoNCYdQHIfosccQmjgRt7khkenrVM5teNGCEq9PCVmYdQ3UnX+xV6rlDSbjWDqRX7Fbd8mVqC8+x4jm+182kTLnBig0oeHb+aBM9/hDbyKSdDuyII0AqCSHKLmjcDjCu+++yyGHHMK4ceN47LHHiMfjnTJ93Q1IRBLxfl5eLvfccw+PP/4Y0WgejY2NnaKEACuYNu16pk2blhCCbCgQgtNOO5UTTzyB9evXU1ZWzl/+8gzz58/3xqA1BlB46/VIXp5XJBK8tz8JQQWBwi//dlwoLMJ9Yy4N9/7GC8t6GJMgiU2c9Q//ntgjj3p1fa7jVQmnTJeIYGCi//aBV1TS09buLJ1k6QbHSKkKthk3bn+WLPmcSCTSyRELcHilFM3Nzdh2jD322J3zzjufk08+mWg0CpBV7l/E2yAaCoVYunQpZ511Nu+883bCJASqyjAMamurmTZtGjfddFPiONlsXlgp2LChmv33P4C6ujpf+4xm7tzXvPfQ4u2uvehyYr+6B7OwPLHVPOg/+ac/4xgiOGGDsnffJOd7IzIidOLVrGNYFu1Lv2DDuEOw2uOJWoNUDRu0TyyO2nZrSj9+G6tg4Ao/MlGnkSulaG1tpbW1GcdxO23vUkoljmiPRqOUlJSyePHnTJkyhX322Yff/vb+RMq3p9y/Uh3RxIgRI5g3by5Tp3aYhCDt67ouJSVlTJ8+nenTb05okJ5UX9B/ZWUFv/zl3bS0tFBQkM+7777DI4886rXvny466OrLoGootLeh/ZWfyd9Q4p0QbjQ00nDJ1V4bkD517C9v17apPfdizPp6CFndHxUjgoqE0atWY3++1K80GpisXyZKCIDWGqUMHnzwAc455xwKC/Opra2mubnFP+WjI5PmfbGDS15elJKSMpYt+4qf/vR89t13X2bOnElbW1vCrncnCJZlJbaH33vvPTz22GOEwyEfWwijlHdIVHFxGTfcMI0777yzU4iYaWuWp608YTn66IlMnjyZ6uoaSkqKueOOGWzYsAHDstCOQ7iqkvzL/xcn3uzl77szY+BV7RQUE395Do0PP+Zl7LTbJTAMdvHW33on7sI3vVPKnZ5DWzEMdLyZ+HsfeH9v4uN6leu6ks6JW7NmDc888wxPPPFHPvjgA78KuJBwuGsqN9gF3NLSQizWxujRo7nssss46aSTOmmDTFm/ZJPwj398xplnTubDDz+irKxjm7hSirq6Ou699x6mTr0wq80o2s/SNTY2csABB7B27Tc0NTVx6aWXcPvtt3mbSkwTt72d9WO/j/riCyQn16sdSGj9zr4Q4Hnmto0uK6X8o7c7A0RKefv3LYuWBW9TPeEorFBO4giZ1HlOblsphWsaSFMd4R8fT+WTj22yJFDyJHWKCx3H6RS7a+3K3Llz5bTTzpCysnIBJBLJleLiUikuLu0Sy5eUlEokEhFA9t9/f3nhhTmJtmzbFsdxMsak8Xjci5UbG+WMMyYLKL/NMikqKpHi4lIxDENmz56ViOt7inODd5kzZ45EIrlSWVklpaXl8tlnn4kWETsWExGR+if/JCvIkTUFVV6MnoIHrPWvNbnl3u8FVbKSkKz/2RWiRcS1ba9P1xXtutJeVyerRoyR1VaBrI5WypqcjvbWJrW3NgkLWJNbLquiFbIqVCSrh+0qdmOj6D5gAdliJwkcIJ1gJOfcA2ldvnw5v//9H3jyyT+xZMliTNOioKAg4R8EK8RzAg2amhoRESZOPJJrr72WsWPHAnQL9QawLsA999zLNddcC0BeXp6veTStrS088cQTHHfccVlpgqDNqVMv5qGHHkIpxcSJR/LUU095PodSiKH45tCjkTe9rWXJdYFa4X0tTVK6N8BMbWzK3p5H7u67e9iA71xuOHcqbQ/OJJR0jlEmxzhZm2rDK2V329spefMlovuP26RaIKMAJA8uCMGCiW5paeG5557j4Yd/x8KFb9He3kZBQWGXSh/T367d0NBAbm4OZ5xxOldddRXbbLNNQu2n8+qTgaN58+Zxxhlnsm7dOoqLi3FdD2Z2HJvnn3+e8ePH9ygEwXjq6+sZN25/NmzYgOPE+fOf/8xhhx3m5QnCYZoXvEX9hKMwwnmdiiPE8E8VlRR8xLKQpjqYeBSDX/gzyrZRoRD1f36WpuNPwkoRpGzCNq3wNqA0VhOdMcP7FpJsN5T0gXoUgGRKttUBvf3228yaNZvnnnuOmppqcnOj5ObmdkL4Aq++oaGOIUOGcNlll3HBBRckdvlkChuDo2O+/noFZ5wxmfnz36S8vAKtNbFYG/n5+bz88suMHj26xxAx0ALPP/8CJ554EoYBu+++O/PmzcNQfq2dZbHh+FNxnn4aI3Da/Koh5ecJkgVADD/T2NxI0XNPUXD0UcRWrmLjfodAdS1mcvVR8qRnsP/gC4BhIE0NWP8zicpnntykoWCvBCCg5NUbOHbLli3jscce4w9/+APLli0nFAqTn58P4B/9pvyysBgtLU2MHTuWG2+8kSOPPBIgIwMDxrW3tydUeHFxMeGwRVNTI1VVVcydO49tt902UYSSiYK2zj33XB577PcYhsH999/PqaeejB2LEYpEaP3HZ1TvO54QXSt6OzFfJCEAqq0VPXIXqt57g/Wnn4/+0x8JFZZ6ew5TJzxDm8kCIOCBU+WllH34FpHgnP9NIAR9EoBkCmx/oILr6up46qmnmT17Nu+//z743/0XIHSBIDQ1NaK15uSTT+bmm6czdOjQRFupTNR+3aBSirvu+gXXX389oVCIaDRKbW0te+yxBy+99BJFRYPS1hgGKyyICmpqajjooIP56quvGDlyJPPnv0leXp53xpBlsfGcC7EfeggKSjpV6aQCQ0FeUCm8s3rHH4j11nteFjGLuUvF5sVPTSvLgrY2HMOk/IMF5I3cZZP5Af0WgIBSBcFxHF588UVmznyQuXNfJxaLU1hYmABzPCYp6utrqaqqYtq06zjvvPMB0tr0ZL/gueee45xzptDc3ExxcTFr165l4sSjeOaZvyQw70z2NtACzzzzDKeeeiqO4zBjxgwuvvhi7FgMKxymfflXbNzzIKx4vNNBzxmrpz38GLe1DTMv1weBsodpg7YFkJAFDfU4VVUUz76fgsMP9fYBbiIncMBaDYpEtNYJL/+YY45hzpznee21VznxxJ+gtUttbU3ifhFNSUkpDQ2NnH/+TzniiKNYvHhxl/MGoHNC6ZhjjuH1119lu+2G8c0337DVVlvxwgtzuPDCCxNjyIQWBmVkxx57LCeccCKxWIwHHnjAKy/3N4nmDt+B3DNORtobejxvLzj+BRFC+VE6Cs27UiZMXuFvYA1ZSEM17Ls3lQteofCICZuU+bAJzgoOYF4gURhy4IEH8oc//IEFC+Zz1lln+fsHqxMAUTgcpqSkjNdee42DDjqIu+66K8HwVOg3gJBHjRrF66+/zvjx41m7di1bbbUVDzzwADNmzEggjN2NUWvN9OnT2WmnnVm6dCn333+/t/HErwgquOQCdHEF2k46dygN8wKmKvqG2ikA/5wht6Ga0CmnUPn6HHKG74C27V4zv9flYpuy2CAZXHIcJwEILVmyRC68cKqUlnrAUn5+oZSUlElJSZnk5xcIIBMmTJBFixZlLAwJAJ729nY59dTTBEwZPHgrCYcj8tRTT/UIFAXPP/nkk2IYluyww3BZv369N95YTLSIbLzsWllFSNYUVPUI4KxNAYzSXWmLTvIHy9pwiaxU+VJ9yx3iioirRXQ3gNlAXptMANKhUako47Jly+SSSy6R8vKKhCCUlpZLaWmZKGVKaWm5/OY3v+2EJKa2p7UWEZErr7xKQqGIlJaWSVlZuXz44Ydpn0knBD/5yYkCyI033ug9E4uJaC3ta9bKqvJhsjpSIqujSYzMK5fVUe9nonIoSUCC35N/Jl+rc8tlVV65rCmsktUqX1YWD5H6p//ioX6OI9rtH/M9BDc7RHDAnMBeap1ODuNXX33Fr371ax5//DGqqzdSUDCInJwc2tvbaWpq4Cc/+Ql33303VVVVXcLFQAWbpsmvfnUfV111NQDDhm3L3LmvM3jw4IzhofYjhq+/XsF+++2HUvDhhx8xeHAl2rYxw2Gqr59O6y23YBWWofxkjkAXbCA1fZwpURU8TyiENNbCrrtQ8vgj5O4+0jtHIItDLLOl1CgjLWW7erORuN625ThOAv8XEfnnP/8pF1xwoRQVlSTyAOXllQJKtttue3nllVdERBImJbmPoJ0nnvijnzOw5KijJkosFhPHcTKOI9AC9933awHk1ltvTWgBrbW0f/ONrBi8nayJlMjaaGWPKj755+pohbfSEzV/Hta/uqBKVhGWdUcdJ7ENG735i2fWVJtSM5Puw4FU+9lcqaZh8eLFMnnymRKNFohSppSVVUh+fqGEwxG54YYbM5qEQAhefPElqagYLIBcdtnlCX8g3fhc100I4oEHHiQVFZWyYcMGb1x+oqj62htlFSFZVzBE1uZVyLpopazLIAydTEFehaxOLvrMr5Q1eRWyklzZcNGl4jheMa7uxkwNJA/S/W+zOIG9EYRkjfD+++/LD3/4QwmFIhIO50hJSZkAcswxx8i6dd+kdfSC5xcufEuGDh0moOSRRx7t1ikMhO/9998XpQy58847vft9XyC2erWsLttW1uSUZqUF0l2rCwbL6nCxrAgXSe19MxNZPnczOXuZri1KAJIFIVkjzJ07Vw455BAxDEvy8wvFssKy4447yVtvvZ3QBMnSHQjBp59+Kttuu61EIjny0Ucfd+sUBv1ddNHFUl5eKS0tLd6qicdFi8j6n10hK7FkbUFVrwVgbUGVrCIqq6u2l6ZXX0us+oHSvN+KAGyOwaeGj3/8459k9OjdRSlTTDMkZWXl8uijj6b1CwKGLl26VEaMGCE77rij1NTUJFR+uvdxXVdqamqlsrJKZs58wNMC7e2itUjbl/+UlQWVsiZS1mO4lxCEvHJZXVAlqwnL6j3GSdvSL0RExI3Htpi53iI1QDpBCDavtLS0yIwZM2TbbbcTMEQpU6655pqEkKQTghUrVsjWW28jEydOSvIHuvYTaI7Zs2fLDjvsKPHAMbNtTwuccpasIUfWpsEFUkO+tdFKWVMwWFYQkrXHniDxunqvcGQzOXv/VgKgdeDpd5iFlStXypQpUyQajQogp5xyqrS0tHRR88EzK1aslCFDtpZp027o1h8IhO3gg8fLE0/80bu3rU1ERJreekdWhQZ16wd44E6lrMkrlxXkyMYrrxNXxK8a+nbtfZ8EoDdbs/ozEJHMoWRXQehwFBcuXCgTJhwmgOy33/6ydu26LgwONMGXX/5Thg4dKi+99FJGfyC495133pXDDjvMMw2OI67jiuNqWfv9I2WtEZW1+YO7CkCuv40sVCQrc8ukfvbvOkq6MszPt+0HfGc0QLqJC5jluq7MnPmAlJSUylZbbS2LFi3uIgSB0Hz88cey5557ysqVK7uYjFQhuOiii2TevHmesLS3i4hIw++flFVEZG1+VRrmD5G1RGXl1sOl6Y35nlDH7U26vy/bhZO1AHzbEpnNWJI/D1axiMi6devk1FNPk6KiYnn55ZfTmANPCBYsWCCTJ0+WWCyWcP7Stb9q1SpvY6nW4vjt2I2Nsnr4brLG8k1BricA6wqrZDURWbP3gdK2bLnv7PVhE+lmvr6zGiD1Sg4bn332WRk5cqQ8+OCDXVZIIASvvPKKzJo1K6MWCKKPN954Q1atWpVgqIhI9XU3yUpCHpYfrZR1vqe/7kcnSbyx0etzAMGdTXl9K7mATUWSVLNYV1fHrbfeyqhRozj55JM77WAOikLee8/7Qsrhw4cnqoWS2wrOL2hqaqKkpATtuhimSfs/l7NxzP7euYOGIM2N5Fx+GcV33Ozl113HO3yin5TtEb396iOdAEg2SYQ+0qZsO6Dk0vL58+dTXFzMqFGjOhdf+iXvNTU1lJSUdNteJ0ZoDabJxh+fhvvUH7Bzisj/5QxKppzlfW8P3gaRzfGeA0EDogF687LZ3tvfCZSkLGF3FcO9XmX+KR8Nz71I/bnnUv77R8k59JCOE703AdM3pTBtNhPwba2IYKVnYnJP4+ryf/E2hzjNLTjVteRuN7TfXwXb2zH1dH9vnv/O+ABbmkrVSvn2Prtq3c1hz/tCWQvAlsaAb52CnUObcuNmVsPoH1+yHn1Px8B8G5S6ojbrCvOPm/m2qb+86PENtkS1FVDqy/f7EMfvGA3E2L8zPsB3nbZUE/rt67Be0Ja4WrMd05bC/NTxbnIBGEim9TY02hy0pTA2W8r6nMBN1eF/acuizWoC+rIq+7qS/yt42dFmFYDeMCX5qLZNQQNtIgaivf600dd56pMAbA77usmzYP0IGTMdh/ttjCmgvs5XnwRgS1evm0p4utvu9V2lLTIM7K+vsKkF9D9eALKZgIGapC0xzt7SNWBvaJOZgIzHqfSy/f5O9pbErP4sik1m1jYlFLylwp//pQ7apD5ANidjZqJ/Jzu7JdO34gT21YRkov8KS99pi4wCekv/CWamr++Y6bng880iAP1doUGB50C1N1C0OSOdvobG3dVCAvx/tGRLBOBPGNoAAAAASUVORK5CYIKJUE5HDQoaCgAAAA1JSERSAAABAAAAAPIIBgAAAKJfG8AAAAEIaUNDUElDQyBQcm9maWxlAAB4nGNgYDzBAAQsBgwMuXklRUHuTgoRkVEK7A8YGIEQDBKTiwsYcAOgqm/XIGov6+JRhwtwpqQWJwPpD0CsUgS0HGikCJAtkg5ha4DYSRC2DYhdXlJQAmQHgNhFIUHOQHYKkK2RjsROQmInFxSB1PcA2Ta5OaXJCHcz8KTmhQYDaQ4glmEoZghicGdwAvkfoiR/EQODxVcGBuYJCLGkmQwM21sZGCRuIcRUFjAw8LcwMGw7jxBDhElBYlEiWIgFiJnS0hgYPi1nYOCNZGAQvsDAwBUNCwgcblMAu82dIR8I0xlyGFKBIp4MeQzJDHpAlhGDAYMhgxkAptY/P0ds4FAAAK3JSURBVHic7P13vCRF1cePv6u6Z24Oe+/ussCCkrNEyUiQLKgPgjygBCVnBBZY4pIkBwlKEgQkCQaUtGSQjIIBEEHJIBtujjPdXfX7o6f79vRUh7l3n+f5/n6/7+G13Jnp6qpT6ZzPOVV1SiilNAmktUYIkfhdCIHWta8HaeLPor8Hz6L51UPRvJLKM/Gcll/wvom3pLrWw+9U3g/elVKm8hlNn1XvtHYz5WFq6yg/pveznifxGK1btO2Cz1n1i6aL1jUpXfRzVhnxvpzMGEuiqcyLgI/4WI7Wq4anNAGQVMD/JE1VMEyGTA0W/5yXkt6fqgBIKmMqaZYED3kFTZ6Bn2eymdJOdkLG380rNPLynldhZuVjEtSTHU/xMmT0wf8E1ZtvXIItaYprtGiHx5/nGQBxbRx9N15Wnnziv6UhrDh/eXlP4rkeyjOY86SL5mfiKa3+SWnSUEVcUCiljDxnCRhTeab3kso28ZnUVia+k/hJGxPG79onIxOTkdz1SN+pavs8kndJaEIhBEop48DL89uSoDiaCH7Lw2ceyqNFk/o6b52nMlaS8jKZJXmRRDw/07tZZlIaxcucSv+k5TsVktEvpkEw2YKStFQ03/8NmL8kykiyZdPgmunzkhjswfclocXjZdRL/5tmWlLZeds4C15HkWeeOZCnP039lQeZ/W+REAKplKphNC610jROlkTPajwpZc1zU/qsPJPST1bQZMG+JEpKlzSQom2e9G6axM+qa1Ld0yZK0vN62zEvL3naNk3g5WnzpDaOj3UpZeLkjgqLNNMxnndeFJGWR9Zvpmd5+JMmuyVewbQGjn7OsnnSmEyTvPG802xrE095GywJ3iYJMdP3JN6SBl/Q3sHgS6rXVBBEnNLaejIaKg9fpn6Io5q0d5MQZdogN7VZ2vtZYz5pPMbfTapDvQIhzm8eZRafJ6b6RD/LNA0xlQFnEghplCSIsvJPS5N3YKbxmNWgSdAx/n49kDFv+nrSRTWgiSfTBM1j8uQRrlnacSpjrZ52ipcd/G6aHGnId7I01fyyhMhkUFvmMqBpIue1s+KNW28D5IXvWeniEzdt0NdT9pIeIHn5SJqseQTMZNuzHlMqK23WOElKH7wTeO6D97KE2JJUZFESYsKpl2eeRBVF9J16eZxMnUyCTghBpgGeNfjSpI6pM/KYBabvWTzmbZAlmS7LpqyHsiBkWt6T0YJLIl2U0rTqZPJPGjtpyG9JCuIgr8AsqzfvuDAwmVmmNPXmnURZ5kbwe6IAyGNrxZnOoxHT8o2ny+rktM9RKZkmdKYyYZP4nkxeSfU1/RZduzbxMpWJkKbt6unXvP2c9CwtP9NkStKKWWMgjkSSeEt6Fk6kFEFh6tvJmHt5KG5SJc3RKgEQT2SaGKYBkDQo0ho7SRKlmRVRnkwdZqpkkm2XxFOSIEubECa+TO2YxYeJh2g94u/lNU1MlDURou0df5Y3X9PnPPXIK7gCT32c/8koHtNGoGh+eXlK0+5pzydTZpYwjfdhXHhG/8ogQZyhpMmVRtHK1QvRsiZoNE9T59djc5rKiNtkpnKi+eWZkFM1TaaCUOop5/+K8g70OM9xBBRNvyQRXb2aOWuMRctZEmNmKvyEabRPIXP1OiPSDnvEmVnSgzkPZTV4PE1cm6cJgXqFTr3vJH3PasukckzaZrKQf7I0mXYz/Z71vV4+6hnLAcXbKOt7Xr6CdydD9SB1AJnGcBbzeSv4P6HJ8lDcHjJRHF3Ef0t6N2sDkwl2TRbymijNbEszHeJIKou3vP07Fcoy30z8xcnU3vWQafJmUbycPOZeHj6mggCSJn+iAKiHsXoKrufdpPzqhWWTmWBxsyI+AJPyqIe3es2ALIGVVnaU7ywEA9XOqyTeTeVNdtJnTdQo38HnpPpO5llQxv+UKRTth8lq/yCfrHKCd9KEZZaPoUYAJMHfJaHBTRLTRGnaVylFfPtyVn5JnRGH/vVSkhY2+ShM6U0CK67J43xG3zWVE4WxaYPIhGDytkG8jvX6ZOLvJeWVJJTj5UwG+ucRRGnfTZTHdEzLK47cssZzUj+YfCTBuIjzZTwOnNfOjDMf/WtKnzQx4pTk5InyZiovSWtnaea8Az/vQE8bPEnwPK9Wik/cLAGUhWDqGdh5+cyjubP4C9JGB3W9Wjurz5MEbT15p9UlL1KNvltvf5h4yTMngnQy+iX+Uj1MBO8laZ96Oy+PFjPxtKTg3ZKCvaZ2SOoMU/r4b1k8JAmbvO9nUb39EaclYTrkzTNtPJoEZdqYXhLtOBmzIK2svH2RJphknoRpBSVBtnokWrz8LC1pmkTRjqpn4JnMgABKR+s12YFrssFM35P4yKK4Rstq+3ifpqXJU2ZA0TKz2ipvHacirII+jPOaZabFyw3yyMNLliBKGtdBOfHf8/R/Ul/H51KSQquJB2Ca0NFC4gMur6Y0Sdx6OzioSNKgzXpu4jneQEKI0M+QxPuSpiRbztS5WfUPPk9V02fxG+Uz+tlkX8cnX3wcmfJeEjxOVnOahEFeYWxSoqaxmWUK5TVL0gRZmjkeUI0nKIupvJIlrmlM2m8yNlJamrTf489NkytNM022vCwkFB0g8XQmbZUXiublK8uUSysj61mcl2hdJ9ueeXlIEkKmdp4KL0njxPR7koKN8mNSYnnMyCQBY+KnygQwSco4XDExmgSlkhiMU5KkjJY/FcqbR6Dxk/gKKG33WVreSQPClE9eIZcnzZIUsGl85BG4aSjRZLLUI4DiPJgmTpbQiaOU+O+TQSlZdcrS3Gltljbvkt4xkRCiOiBIvDPStFdWo+Zhrh5K6uQ0mowgyVsn06CID5jJmDhp0C8qdLM6Nk1A/z+JTDwm9UE97ZnHxMwjWIQQqZu+6hHy0X6Z7NhcUhQihCAeQBwu5IV2WVIp7VnwbnwLZj32T7A8lJfnpHzylBeQCZYnvTvVSWiyFePC8P96kpts+jgENaU1/WZCBNG8oug0rybOy9tUxlBSv5jqltVnJj7q7ec0Ey9KicuAeey+aDrTCa1ovmkowpR/HtvW1LlZZLKTsuqbBteyyo4P1jQhmef3uLaJ1r+eNlgSlKdPTd/j78frYBr8gRaO27pZ/CXxMhXtG1dWpnFsKicayCQr/zz5mSjapqZ34u1mFAD1UNYEjEtsU7q0jT+m/EzlJ/FvmjBxntPsraTnSe9nCa4kyZ40KOoxsUyCaaqoyERxLZY2OerhIcgnOtmzJkteyqsRk5SCCaEGv+fp//i7aWOqXlOnXqoSEtqn8Ie0AZzH7qznHRNjebRlWvnxv1nlSSkTd5mZOqse+F8vmTRivW0Q1TJZUDJqQiXxU48AWlLtMFWTLC9Npn2zoHleeB9Hr2laOq/QilPSPIqi9tyHgSYDYesdEPV2YprkNUXNiWukLGFhksZTgd55KE+Y9DhFNZYJ4aS9l5RXEiUN1DxtUA9feU2iOE/10FTG21RpSSEbSF8VME3+aD8YR1vaIMpre6WlzYKXab6BrDzS+MuC22l5pbVDVn2yyGQnZpHJDk0qL41XkwM2y2xKUwZp7Wh6L492mwylQfN6hEsWRdsoaSJGBXS8DUx9Z/IrxPlPql8eE6QKBea5HDSAytFC0mzurEkYpMsqcyppJgsLJ5tXnnonQcWoYIpGmQ3yXVI8TobvyVCaRgqem9KaEEySaVmP6ZIGt008Jgn7qYynqbZzPW2alw8hBHbWC57nAYRHcOMSpN6BZ7Jzo8yZ0pnKiD9Pej8t33p4nqyNaaK0ukyWJsNjPeUm9UFSXYJ8o//yDtI8JkKaxo3/Vg/iyJNXEsqayjyYCkXNv7x8hu8ECCDphSWlRf9f+v9vWtLaL03LTybfepFBXAhNVilm5VvPu1nPTHkKz/O0yQYJaP78+Xz00Uc1dqGZNEGZebRFzdta4ydJ9oiCBkTFLBFVaeN5BXx4nmeEkdU2F8SLymvnRzVc2lpvYEaZoK2Jv6wBNVkzLKoBTHmZfo/C8+hYUErheR6O41AqlcJ6NjQ00N7eTldXFyussAIrr7wyhUIhk8/4pEqivJPMNDGXpOn4P2F65REmS4rPRB+AUgrLsth22+145pmncxf0/9L/b1FUuKVpm5kzZ7L88suz2mqrsdZaa7HCCiuwwgorsPTSSzNt2jSam5snZbMvSds5T75RQZfneb3aPp5XFu9J7+bh1ZQ2TnbSw+B7S0srtt3AtGmdoT8gytySMBGm2sn53/cRSh7omORcWVKaKa2M+slHRRN/66NgEAVLkEq5OI6D6/p/Pc/Dtm3a2tpYaqmlWH755VlllVVYffXVWWWVVVhxxRVZZpllaGpqSuYwYbJk2a6TbWsTLM/zflL66GSLPp+M3Z8nTbQsEx9JdTO1ZxrZpkTVGSq0VriuWyUATIVNluJwPC758k2QYAJMnYf/b3m3zpIIuioa5CIK35Xy+7epqZGuri5mz57Naqutxpprrsnqq6/OiiuuyPLLL09HR4exhCCOglLK32QS2dGXyFUd8DxIn2fyxVcJkgRLfNxNZTzXa7enCaAsSuI3bS6ZqGoVwCyhl4SGqi4jvvacJvHjAyPasdUVzA/HovAtapdPhvzykp2neaBZXgdSvcIimOhBHV3XpVwuUyqV0NpDCIu2tlZmz57NiiuuyGqrrcbaa6/NWmutyZe+tDyzZs3CsqyafOMBU6K+AcuyMs+FxPsxmkc9sDiN0mB/9Hs0X5NmNaWN85OVJot/kzbPY/enjZMsXgOyTQPVb4igg9OdSXkp2ulpEjCJcVN6EzzPQ1kSMq0MUz5TpaxOzXKKRSdgkM51XcbGxiiXS2jt+3Pa29tZaaUVWWWVlVlnnXVYd911WWWVVVluudlMmzbNWK7rusZyAsFg4itrMqc5tqKfTRd15DEj4rykfY//ntQXaabiZFFcVODETYw0HgN+4nnkKS/6DoBtgttLAtZHKcue+r+iJGmb1ZlZ7WSCYXkRQVJ+8c9RGB9o9nK5jNYelmXT2dnJSiutyKqrrsraa6/FOuusw+qrr87s2bNpbW2tKSO6UmKa6FMmH0pSccJAjotVTMgwDVlks5Dc3yZzIa9ZkofyKBtT2Ul8ZvGTxnO0HWo2Ak28VP9ANcEtE9P12ElxKRnn06R98sCsNAGUVGY0rzQyIYy8Ez9IF43jHsB48OF3qVSmVBpHKRchLDo6OlhttdVYddVVWWeddVhvva+wxhqrM3v2sjQ1tdSUYZrskw2CWjtYK/X1FFpVlnWlRMTDmSvluywjfZJ1w278c9wMCd41adZovZLGUkBpS7nxuif9HkcG9fZ/HkpSLnmVq9baRwDmxKLqb1waZ0niPBXJgoJZDWhCL5MxC+KawVTmVCFtPhJYlckiKnsdyk6ZUmkM13UAQUtLC8sttyyrr7466623Huussw5rrbUWX/rS8jQ3N8c5wPPciuLVVT6BeL3ifWesk4ioBa39z7qSRnkIBVoKpGWBBaICHhRQXrwI998fMf7Sq8hVv0znrrsiPBeBjSd8QZHXNMyD3EyTPvrdlKcpTRbSM6U1jSOTYjTxF+c/LU1SuVmk0aBBCOn7AMyVzO9MqykgJ1SZCqwymRVJ6ZIaN2/DTdbGSyoz+lswIQMoP1aagPKFQgNLLbUUq6yyIV/5yjqst976rLPOOqy88spGT7yv2RVCTDjhpEyG8HWbJhpEBcZrpfwRYkuEsBBIdKUoZ2SI0ocf4f7lbUp//gvu2/9AvPcuzhdfUBwdYLR7No0vP0XTyivieg4SO1Gh5NXGedNM5V2T4MkyodPSpqWpR5mahEUqKRCWRLmeLwDyvJSnEJMEzGOHJNFUJ11WPnn4jL+fJFBMsC+eR7A0FgzswCPveS5CCNraWll55ZX4ylfWYcMNN2S99dZjrbXWYtasWTX5eZ5XFQ4NgqPEE0tveU4WJml74atkQENlr5iQMrTdhWWhAeW6jH/6Ke4//sHw3/6K9+c3cN95F/vDz9FD/bh4NNGAZTcgGhrQ02fTvHghfSeeSvE3dyOVRlkaG0kFUyROjrQ+SKOksWsyBeLjMv5uXiQY1fRJiCJL8+epZzCu0pByFc9aIyzJ2OIFjL73SbUTsB5KRg7VaSbz7H+D6h1ESXlAkgSXSCERUuBDcY9SqVSx3T2KxQZmzZrFKquswnrrrceGG27I2muvxUorrVQD5YPttsEADP7Ztl1VbtrgMtVXCoESGoFAoZEKPK2RKLQQSMv2tXzEfC/39lL6179x/vI2pTdex/3rm+j330cv7kV4DgJosBugWIS2TgqAEgpPKYR2YUxjtXdS+v0D9P78dmYc+gO0Uwa7mNi+QRtnweY8FBfmSc+y3k8TJEnl5EGaYL4GLUnw1INcXK2xNXjjI3x2+z3M3mfP7H0AeZyBaT6BJaXF0/LOA2PToFkeKWtKF/1uWVYVH67rMloq4ThlQNDe3sZKK63I2muvxUYbbcQGG2zAWmutxcyZM2vKc123ClFIKasmu4niGifCPVB7tlwI4TvplIsChLTB8iPESHzt7o2P43z8CeW//53yn99g/K9/R//zfYqf/Yfy+BACiZAFrGIB3dQEstkHDH4hoNyYJ0kACkcJ7KZ2hs86h9bttqZxpRVQ2kMI8wWW0T4IKH7ZZVzDJv0NkNGSEv5RygvFA16S9sPUI+BMgsfcfgo8jSjYfHr82Uzbdiuall62WgDE4UteSrL306RimoTLU34S5EkTRqY88mrMaB2rbXdQymN8fJzx8XFA0dDQyIwZ01lxxRXZYIMN2GijDVl7bd92b2mp9spHbzsOTIT4dVTJkzu7jqEvR2vQCjRIIUEKsATKKmIBHhp3wReU332fsTfeYPT1vyL/9g/Uxx8ievuQ2kVgYRca8IoNFNqmAwpd2YCF1uBVVjCC8hChj7CKI62h0EjDggUsOHkuX7r/7kBO5Z6YQdsk+QiSvschc5IAiX+Ofo+Pz/gYTDItTD6nOD8m/k28ZdWz6jcEynMpFBpYfNW1uJ9/wfQ9voXnueZ4AHmhUL1SdDISLSu/qSCMJKFTA5WlqFl3L5VKlMv+6be2tg5WXnkl1lhjDTbccEM23virrLrqqiy77LLx2lVp9yDPYK09aRCk1VFo3xFfmW4QWNIK394DsHzbPbgL1gP00CCljz6i/MbfKP/lr3h/exvn3X8jFi2AsTIWvp1vNTQgWjvQgFQaT2uE1gjXmZjcYsJlrHXlc2VCV9d+gpQqUWibifjtA/T84nam//BAlOsiLFn7YgqZJkDWuMiavFnlJCm2rP5aEkgz/nvcx1CTJwLtlbAKDQw99jg9P/4Jy7/0WNhPmQFBop0Rd3KkvmVgKgsW5ck3mk8gPePvT9ZJFPyVUiDExKR03TLDw6N4nr/Hffr0btZeey3WXXddNtxwQzbaaCNWXnnlmg02ge3u5ylDuz0JtQghEAiEzGfvaf/livL1sBW+k862CJz/GnCVh/fJx5Tfexfnjb/h/flvuG/+k/LHn2ANLkYiERQoNNiIQhHaG31RUtHsUimU9tftZWViK0GVdhcBu1qDEP6/yu/RflfSdy5aSuBYZZoaWhk841yat9uahi+tgPRcdB2bj0xaOq3N8ghV03upQrgOgRP/zYQW6uUpeZ5ptKeQdgOj773Lp3t8n65TT6BxpRXQrouQMhkB1OtwyGqEegVGEtyJ85YHFqWUGmr34LvnuYyPl8L98k1NzcyaNYtVV12FjTbaiI033pi1116b5ZZbruZ8u++Z99e0oxM+SrkGUsrY1BUNTFBvy0IABSTh5u2efgY/+jfu395k/LXXcd56C977CGvBYpQ3TgGJsIs0NhSgvRMP6U9gpUBplK4+9OVVTAchfFUfDrxYWwZ1MPEckAygMAKhFKqhkeb//IfFp5zF7Htvw0NjVfLOg/Dq7fu8iDFPuaaJmzYesyZ6XHklmQdRHhNNWwSeUkhL4PQPsPC/9qLpS8vRffJxaM8LkVaNAKiVJsaya+yaepnMIpMmTyorS1iZngcT1HUdxsfHK2vvmvb2dlZeeWXWXnttNt74q2y88casttoqTJ8+o+p9pVS4Tz7qE8gK6ps2GILJr4W/ciC0b2PjVfrBkkQ38WhAlUu4H33K2D/ewXv9rzhv/JXyu+/AJ18gRkbQKIoUkMUiXlMTlmwGXek7FMIFiYuIqHNBtXafwiHL6rpX8hIaPAlSAdqF9mnwq/vo+/ZudO2zN8pzEdIi3m1xyDuZsZXXro+Wl5ZPVCGZ7Ps86CANtaahh5rf8Sc+aH9VB4GHZtGhh+K99RbTn34CWWxEex6y4nBNOQwUwIvqQkyVjj4L8gqeRyF6UkXTnBymyqeli9cnyUYcHh7Gdct0dXWF6+7rr78+66+/vtFZF0D5IF/LskLPfxpiivOb3Hn4cF75PgIpbLRt+Va7pGLdQ3nB5zjvvIv7l7cZf/1vjL39DvaHH6H7ehFeGQsL2VBEForQ3oEAhPKPdAvlIZRAoFH+OEELEDrmrNO6CrqbnFhaa98cCPmvtYOT+lkIgdCgZOVNrSg0tjB02rm0bLMlDbOWwVUeUgosbaEl/sajBM2a17RMmvxJ/ZLX1k8SKEnlxOdM/Htavkl18ROBFj46VB5YBYsFp59L+b5fUTj0OFq32cY3CQQorRCIiX0AVXZohKksGyutwbLSRysdr3BaRZMktqlMk1QfGxtl9913Z//9v88aa6zByiuvXHPoJardofrwjSl/U3tVDXiqFamv7RVC4TvRpFXZaDOxHu6MDuP8+2NG3nwb9ac/Ufr7m+h/fgD/WYB2hrCxaZQFdEMB0dyCli1oHUx4DV7lJF9soirwd4MJ36YP7f2EdswLm6OUJaAn+AHL1ahiI9aH77P4rAtY5qafIrUH2L5gqpxMTeLHdCbA9H3ypmJ1+rSxl4QkkihNcWSlqTxERJZENQrtaqyizeJbbsP78QXIFdZm+llzQGkkGiUUorL7sq5VgCQtnmbbxFHBZBs9+Jy2uy1Jkpp46enp4dvf/jbga/fAOx/dLx+fyKkCpzLJRXB2wt/NH04+hfYnJhohrYlde1YFyjsupU8+wf3HO5T+9hbun/6GeutdvE8/Qg/2IwEhCtjFAqqxiGiagaUEnvBAeX4wj0rTCJ3hRw/rIRInfxblsc9N70x8Jlwo8CyQjsJqb6X8izsZ3ONbdO6yE8otoy3f2zgx3rM3OKWVXS/aNGnkJGGZNdGTzmBUI+5aBJGab5C3Bi0BF0TRZvCJJxg+/jSKQtBw2hyal10WHBfPlghsZOW9zIhAQlRP3iSNnNQQcUmWZgOnUVyAxDvGBFGj+QfflVK0trbx3HPPcMwxx3LNNVeH8Q/TYFoSBJt4B7SorEkrAdrBl7cgbLvi3ZchEnAWL2b8vfdw/vYmzht/Zezvb6Lf/wRrUS/SG/eX7AoFrEIB0TYNLUEo3x+A0qAdvHAK+cImrRXT7NSatGICNAhdWdfPMGHiZWX1qVLa34uA3yBKekht0yjKDJx+Hq1bbo5sbkZrsAQow8QI6hP8jQvqtC2ySZM5aeKZ5khW2viYib8/0RaqJl3a+K7KQwBK4QqQLkjbYvT99xk46HiKQz2o7bZj2gH7opXrn9ugGo0aNwKlQg4DRRlL81qaJk5eijZCWhlpFJTteR5dXdO54YYb+PrXt+Pb3/42juOES3Qm3oxmCL5tqpXyJYDQKFFAWgJBMZyQ7ugw5Q8/wXnzn4z95XWcv/wF9Y93YMEi7LESGo1lFSnaRXRTI57VFIh2UBpd0e6BHyBccvPleEWPm+tr4v1/muru28r/tVKo5lbsN/5Ez+XXsNS801BuvmVB0ySqx/zMQr2m96JjMUsxRstIUlSmtGltGfIuNUr7uzjd0WEWHXw4DZ/8h9G2FmaedxZWwcbzPCwpQ/QX8BAGBY0XOsHY1OymJMbNZVVTkvMkyTljehZ9N57OsiyOPPJINtpoI5Zddlk8z8u8my+03TVoy0ZLUdlkU3nuuZQ//ITSP9+j/Ne/o/72d8pvvo368FMYHERqhaCAVWhAFoqojkaEBtvzzQSlFbipLES5Cc2iLKGVZpNnTZQAZ6Slq2tcaO3vdTA9cj1kSzujV17DyDd3omWDDf1lq5gzuTq7ya00xcmEcJPGXxZMN03gWtSY7k8zIdkEzil4ClWQfHHoqRSffgXP8mj64dG0br4pnucghWXMK0QAU2nEtMGXBt2D52n5Jvkc6uU1Ln09z6O5uZn//Oc/HHHEETzwwAM1fJu3aIKQtm+3ex7lL/6D9/4HqLf/ydjf/o7661t4732I7ukBb5wCYNkN2IVGRMs0tPT8E3BaozRIx9frKgq7iWr5SBsZ2iBpQsTbMfrZ9LxqcOjoO/nbOq1fap6pislS+UmEOEaBVaAwNEDv3HNoePjXFZwjsXxma/I2ad4kFJuFTk3fs+oVT1PP2DSZlfWNbwGuiygUWHjlNbi33obV2Io7u5NZZ8yprJ5YxnaDiABIKnAyt9VCto1fD0xb0hSU6bouXV1dPPjgg1x99dUcf/zxuK5b5Q+IkhACpzzO0F33o/78KmN/eRf1wYfYixdCacR3/IlGrIYmrOYmtGiunILToDValUCJcMNOtGWzWiE0OWL1mKwgzCorL19TLVtEEGawrVl5HlZbK85jjzNw6+10HXwQnltC2A1+W06inLS0WeZtFb852juPCZyESnOT9h3PeC6yUGDw0fmUT59HR3s3/SML6Dj3ahqmd/s7/lJMqKrZnQkFM57ndaAEeZkqnwTv42bDZMjEv+d5tLe3ce655/K3v/0N27Zr1vzD9z0Pyy4y/vKrDF17Lbz0Bg0LeyjIAqK1G9W2FG5rC1gKz1Mo1wXP8/0EOgDS2Twl8k91uyW1a5TytlWAtuKCOasfk1BaGk+J+QnCJpKuQDY2MHTOxZQ/+xhbFvy2nEQ5ac/j9TMpJhP6jOeVV1HmcQpmkQ4mv1II22b07bfp+eExNErB6MgAxW/vScc+e6FdFxImf+izM/1YKcb4Ul5paZq4eWz9NMdJUllp76WlAYHWYFkFRkZGOeywwxgdHQ3t6po8KvWfcfklWBttQwtlZGMBV/gn+gruOAXHRWldWW/X/kabhHaryl9XDtlU0ELwXBs+m+sy8XvSQM7STGmDPv48nsaUPotXIy9aYGlFGYVtNyM+/Zj+Cy5DS4mO7QdIonrKzhJOaf6AtLY1jds8CjRtjkykw9+uLSXl/l4+O+Aw7C8W4okiTmcLnT+ehxUIiQQKltNziC2zXZSnAqZBMxmK+xGSyorzko8Enqdoa2vj5Zdf5pxzzsGyrDAARyxTcDV2SxPTrvkx400FLOXiWjZ2JchlsI9WhXvjkutT/WPs3/8ATcXsyvvOkuhvBVjaD6LS3NTO2G13MfrCS1h2AaGyN6bV4y9aEs7DpHyThEe9VFMHLcI9JYuPOJHWP/0Z0d1FaXgxTSedQMuqK+F6ZcKgjNTOl3C/S7QQE4NV5WZIy7hDJG9HZHVQkgbNY3JkoZaAHMehs7OTK664gocffphisVh7ExLgWRLlOLRuuhnFuXMYGRlEUMQVCZKbWg0arVOSUDUJtHqdTGl1jwtPU9p6zQ3T86jwTqpDUj4SjWfbWKUx+k45BzVe8p2FShPEDsgyUdLqHP9cL6rMeh7NNxq3wDQXTO0c/01rXVl6dsEu0vPjS3Hu+SWicyZW72LEJl+j+9gj0J5CSBslavs4+l1rPSEAkgZXwt2hqRWfzERfkpRmLqQJFBBYls3RRx/DokWLkFJWbzOFyjl5wPXoPPE4nK9tizfUg7QbJnEr3/8t1QM4TEI1T3/mMn8S+dNo16XQ1IHzwtP0/+KXYEt/12MCukz6LWnCa62N9vtk0GuS/2Cy494k4DzPQ9oFBu7/HeNnX0SxtQvtlnCKDXRdei52cxMeCqktP8piTCnHhXqmCTBxQ9DUKFp4Pc48k/2ZZJNGn0c/x23iuG1d4RClNM3NrXzwwfucfPLJoQCoKkeAEP7iidXQwKxrLsGb1k7BKaFzOoLqsQsh3dauh+LaWIvac/3RfKdSVj0UHwcT3/39jUp7NDe0MHzJJZS/+NyPdJwxhtLGQBICqwdNpJWXB61OyqTyXKRtM/Kn11l4yHGIooVXsHGG+ygc9gPattoc5bhY0qrsSq120JqQXo0JIMTUO9s0OU0NFNWuSbZckrZJgqzR/NP4ifMSkOe5dHV184tf/II77riDQqFgPn9QMQWav/IVOuedwfD4aI0mSeJxKgMtTlltbYLh0d+D0rXB15NkGkyV4nyYeK0ipaGxkcIH7zNw0RUgBVpV3C0xBZVmJtVTj+ikyVun+PumNHnGocmk1kqDJXEXLmTh/kfSPNSLKBbxRkZQq67FjLNOQXtqImpzTpO4xgTwP09u7T8PxQfXZIRNEmRbUgNVKUVLSwtz5szh/fc/wLZtoxAQloV2XTqOPgy5++64Q/1g1UbqXZJUD2LImsAi9jdaxpKErkuClOtCSxcjN93G8MuvIG3L3zXJxPbWOEXHWV6qx7+S9bxeIWAuQ1e2f3t4Gj4//Hjsf/wFu6UD6WnGXJf2C85CdnWD9vyDEykUR781M8l/ls/uz6tV47/F4VhSJ5mcING/8c+TIdP7SmmKxUYWLlzM0UcfE54UNNp2AmxpM+uKCynNWhrtjIOUYQtq7e/yC/7l0RQmHtPMprhtZ+obE4KSCKTG33Woa9PmKctUF7OJZa6XKU21AKk80wJPSuyxUXrP/jGecv2NQ5VQZSaazMSvl/c8eSa9n6/vFdrTYBfoOfsC7N/+FtkxDYXGGR6g9b/3pmvPbyE8D2HZJPiiE3mSwY/RAeZ/nrCNA4ajzAtRG9Y4vSLJ6UxwMPo9a/DXU1b0vbR0nucxbdo0HnnkIa677roaFBDyJC2U41JceUU6L5iHVxqr3AVQy58vz7P5T6IoRE6ycaOf4/1qciql8ZGnrLx8myjuY0jlD7BchdXWAY89wcB994NlobwytRhm4v2k69/zKLA8iibRbEmhLP9DJHc8pbELFv133cfYBZdARxvCk1jlcbylZjP9grP8PHL6QuJ8SlPlfBMgeWAmQsqUwZxky+fVgEuaossySXx7nkdbWwfz5s3jzTffDIVAzQCxJNpx6Prh99Hf3RM1NICwbD/ufg3v6cd2k2gykDoLyle365I3VybjSEucaICWGq0kDQ02w+dejNPTh7QKUIluU09+8QloQqF5fDVJ4zrpPZMwjdczuIlJK0XBLjD85z/Rf8QJ2I2NCM9C2TA2PkLb6XMofvlL4HqhAKi3DjU+gAntn5xR9HtagSZKgqZpHVDvQMqTPq59TJ+VUti2zdDQMEcddQzlsuNDekMseiUlQsHMy86jPHt51HgJLcXEWqwQIETk0IvZLMo3Wc3P4xA8CVmZy0nea2EaxEn9laddTe9kTjDwHX/KQTa2Id5+m56rrwnvPkzedlU72ePP4jymmVz19IMpfdKEnECXwg/5oBRIi1JPDwsPOBx7dBBdKPorNgODyG23Y9phB4KrEPbEhp80n4eJ95StwOk0GQdXHjMgT9olWaapg6Kd77ouHR2dPPfcM1x22WVVZwWi7wgpcFyPxuW+TOdlP8Z1xykKSIu8nqSh4z6SvFSvzZk0sSfTt3koK88AsidR8MRzXYoNnYxc91PG33nPh/mVCbQkuM6r0bMESjRd3vaUWqE1aCVQKL44/Bist96BthYsV/nBP5qKTL/kXESxAKiJm5hS+EriTWZVyJRJlj2U9VseqpevPPlM9n3Pc2lv7+THP76QV199zbg0KBHYtkCXXabtvQcN++/H+MgwliUnTgPW2TbSsKSTRkk+lGhd8sBiUx5RTR1NtyQERe4+CtJpjVcs0NDTR8+Fl/i/64pAzcgir2Mvy8+RpGnTBHqWMhL49yYIpbALkkXnX457/wMU29tRrgd2EXekn8Lxx9Gy0UZop4wO4ipMcpxLE1PVEKZ6lSXLroy/n5U2i9IgTVbZedKZJkXtJPCjAJfLZY499hjGxsZqygo6QVi+Nppx8TmoFVfFK42hKhNZ6mynkdaVWAHCv64rWD3QUtT8Q0xYvsGGnnhLpTkOk+BqdMBqISZ4iGUu9MS/qt8NgiLermmmQlC3aNtq7R+uCnjxvDKF5mk499zFwNPP+TBYecSbtcbplYAw4uPMZOakjZGkNFk2eM0YKnvIgkXf7x9h6Nwf09TahqcUFjZqbAhn7XWYefLx/jl/y0IKiWVw1KeVE6X/kQX/qWrdPHlNFlmYOjaPc9J3CLbxyiuvcP7552NZVs02YRB+nDtPYS21FNMuPwe3XJ6scM4k06pCWovksXHT3q1BGGLiXyqfU0QLJqEGCiUKFF3FwLzz0U4ZISyfl5zKycRj1m9plCRk0ygsQwhQClm0GXvnX/QdfgRtVgHPzwiBR1mV6broQgod7WitCPbrxEOym/hKemaMB1A7OcyZ1kN5pGc9lIQspuIAjD4z5e+6Lp2dHVx55ZX88Y9/NPoDALAtcB2mfftb6B/ujzPcj0y43dc8KTOrYKYcXVKPH0b4L4TP4s8DZKJy8JvmIDTymNAvwXsSiVTj6LZpqOf+SN8v70ZYEulVnxOYCnpM8w/F+Qo+T9p0Vf6EdoYGWHjA4RQXLsJrbEAohbBs3KF+rAN/QNc3dvDvUJTVjr8knrL4znQCZtlBeX9fUjb9VElKC0va2FYBy7IhiKobSNMMZ4qUNlrDsccex8DAAEKYQ5XLSgDGZS8+F1ZdCz06CtLyD7ikDTStCW7PMf2rek+AFqJyuUf9bWESPgEJUTExhJgoW0+YAkEIs6xi8zrKoiQjKxJaVGIlxvz8GoGSGuFqLLvI4CVXUu7vr8QNyA6TVi/FzaOkfKNtmocHX8j6V8opKfj86Dnw6vPo9i6k40fz0eVxvOW/zFLnnIZXMcsqjRDyFP1r4imJF6MJMNnGygN5snwOS5ICr7Jt2+HFH6XSOAOD/fT2LaKvbzGgUcpDiGT7MNp4ruvR0tLKX/7yBmeddVaNKQAVaC4kyvNomD6DzqsvoqypXHaRPvjDlqg4tOL//EcV+71SmA7qOgk4nwlR42VHeIz+vkQpyh+AMAUQ9SPhKu0hWpopvvMuQzfd7Ide9yYCh5jqmmSLZznqzKyaFUaevAIhqzywbIvFV1yHuP027I4uKJdwpIfAwhkfo/3csyguuwzadRN3+05mHgkVOe8bZVpr/5jkrrt+g/nzH6OjowMvIyRTGk3GmVdvOcGED+rhOC5jY6N4ngMImptbWGaZWay88sqsv/76rL322my44YY8++wzHHbYEXR2dvlHTXOQlJKRkSEeeughdthhhzCWYBU/gKc8pGXz+XEn4V59LcW2Lj9Uk4H3PNrSBLc1voMxKiCSKM17ncaH1n4BWogJm1NF3qujrCyKvhe9oyDRfrcsLLeM0z2NmS89TcOys/1lwRynM02OviSvfXx+BN/jZoDJz5RInge2zdD8p1j87W/TaDX7qEcppGWjh/pwdvsmX37gHvAUWBqk5e8nyUDgJv7idQ4FgKkhTAJgsrREBoPReTdxg4/jlBkdHUUpF9suMH16NyuuuBIbbrghG2ywAWuvvTarrLIyHR0dNeUce+yxXHPNNXR1TcdxnMxtzpZlMT4+xpe//GWee+45urqmVfEVklIoNN7IEJ9uuTMNb/8DWWzFxUWiEFrWzB5TPcNnEc9f1fXcKXEb8gqXONVotGDqB9AzoawlTZXjMOb+8JUoti0pDQwijj+G2Vde7B8esqxEhJI2roAq066eLe/xtk5GAgKtXIRlM/7BhyzYaifshQuRxWZc4WJpifA8hhs0yz7/LM1rrwnKRUmJVBIhiBlF+etWxcX/tQCoLdfvUB0daRGKXsjpOA5jY2Mo5VIsNjBz5lKsssoqbL755myyyVdZZ521+fKXVzCWaQr8udtuu/HYY4/T3T0D13UyHWa2bdPTs5iDDjqYm2++sSaisK8JQHsOolBk8Kln6dv5mxSKRTxlUdQu2nDy0jSIahqICBrQOtEHkMfhGTzLY7NmDqgI70tKGKTB88BH4Vn++vl40Wbp55+kca3VwXVB2kZJlVaPuF8nfvnH1P1Zwl+yBFynzBe77AHPvABt7WhvnKKWqKKFO7CYpnMvZOaZJ0HZRRds8nh662n/KhMg+pJSKhQAjz46n87OzioBMJVOjr+b9j2A9eB74cfGxnDdMkLYLLXUDNZYY3W22morttxyS9Zeey2WXnqZmrIC08W/vrv2gongQpBFixax7bbb8eGHH9HU1BS+Z2y4yoC0LIv+/j7uu+8+vvOdPWouF9ECtPKPs9rFAv/50Vycqy6HtunYZZe0k9dZbatEBLIaEEAegZtFWUIwj6c8/nsaJE3iMQ65o2VYSBwpsCQw1IPe/wCWue0mlOeAqJ40eTVjGpnqGOcrrRytNSg/ss8XR5+Ed921yPZp6LLCkQpb2uixUbx112XZ5x5FNjbgAba0EIl6v5oSFUi8Llk+gG98Y3ceffTRKSOAOHNmpibsSVk5UlsulRgbH0VrRUdHJ6uvvjqbbLIJ2267LRtssD7LL798VQ5KeRXpXX3BZxa5rkuhUOD1199gu+22QyltdPCFnIqJq6HK5RJLL700L730It3d3eHvQV20EChPIdE4QyN8ttVONL39Nk5zE7LisJoMnNYi6GhhjJefJQDin03v1CsA4mHU8tYlXlYeAeA7PyvPhcDSAmVpxl2HGY8/RNsWm6E9L/QF5BE0SXylCbAAMSTlpdFI7Z8L0Wh/n4hdYOHP72Dk4ENpbu3G0+OgpX+1nFaMOCW6H/ktXV/fGq9iziTpimrEWYf2F5Grwer1gGYVlPbc9FugnZWnGBsfp1QaQ0qL2bNns+mmm7Dddtuy1VZbsuqqq2FH1tSVUpEJ7/sELKs2Mk8WnLVtG9d12WCD9bnmmqvZf/8DmTatM9UeD/Jtamrmgw8+4LTTTuOmm27CcZwJqCj8ySkkaE9T7Gyn66oLGNh5Dz/yLRohQOra82xZsFxU/ktzBsXfj9uzQRrT73EyCY64AEmaZPVo3KQ6V/EXOAcjaV2hQdgUx4bovfQqWjbftPrikUlM/uh7SVo9rY7+b4HzVKFdhV0oMPzyq/T/aC6NTS24ykVoX1mJgsTtX0zDoYfR9fWt0Z7Cqoz3LKffZFCNUErpJEjzP4kAwLfnpZR4nsfo6CiOU6JQKLLmmmuwzTbbsOOOO/LVr36VGTNmVL0bQPOo131J2Zue52HbNsceexzXXHM1XV3TcV03dRALIbAsSX9/L/fccy977rmncVUArcHTULD47Kgf4f70pxTaK7e3RLz4aXWqxxmV9W7ahEgT4PVqzjy/Z/1WMz4FVZM/fK4F2OCNj9P1yAO0fX0bf+NMyu04eZFBEi9pVJWX56GljdvzBZ9tuxuFf/4T3diCdEEJjZQWtltieOmZLPf8s9hLz0BofzUjrc3TnH7xZ1FB7/tPfKp5IUsA5LUx4uRrGx8yjY2NUS6XaG5uZc011+DrX9+OXXbZha9+9as0NzeH7wQReUw2fL3IxcRPvP5KKVzXZZddduW5555j2rSuUKsn5SGEYHx8jGWWWYaXX36J7u7uCVgoQArfpNEVaT/e38fnm3+dpg8+Qjc0oj2vBuJNxdmUNvFMk3syAmCydnTSxK5HAFBp14oHu+oAu5AF5PAA5R13YPbD94GQPhIzHHOPlxP9HkB7E29RnoxtIUBoUdk3ofwoEJ4C6fHFd3+A/vXvkW1taM9FIUBLpC0pD/XRfsvP6f7BPniOi6wc9c0rAEztluaPqDkMVFuIwuR5DCFuQqPE01qWhWVZlEol+vp6GR0dYa211mTu3Lk8/vhjvPjiC1x00UVsvfXWNDc34zgOruuilMKyLGzbrpn8UZtnSVFQr8bGRu6443aWW242o6OjSGnVlBVtA6UUzc0tvP/+h5x55llVgyKE6VqD8G+4aeruZvol5zGunBAiqopWE9R6mpPaO85L9J8pTZzvybbRZCd/Wtnx/kwVuBCGNJMILCZ2LErlQUsr+oknGXz0cV/ruZ7RgW4yf+MTJkkQR59F//p9jX9DlPbvhhSeQtoWCy+4nNKv78dqbwfPw0JQ0AJpWTjDvbD99kzbb2//Rl9r4tbpNIGZJsiT5mgovOKrANFC/FWAXXn00fl0dHSG0jCvBgiWxEqlEiMjw0gpWWmlldhxxx3Yc8892XzzzSkWi2GeruuGjru0fPMMvjjciX9PqnOQb+AUfPLJJ9l112/Q3NwCkeM3SdLfsiwGBgb47W/vZ7fddk+8clxXNoD8Z7+DUb+8E9negXaVvw02hb8kEj4zVb/4B0bMgqAev0G9VA8qM9mvWdo2Ky/A30I7OoS3xeYs++RDCCn9NfQMXvKYJlmk0QT7s6XWeJ5GFmwGH3mMvm/ui12s+LCEQmoBQmJpl1Ft0f3Hh2nfcANcr4wURYSo9X/Ehc1UUHBNWHBjherIXAgROukGBwfp7e2hvb2Nvffem/vuu5/XXnuNa6+9lm222YZCoYDjOKFNH2j5LO2ep0OijWJyIpnSR/O1bRvHcfj617/ORRddTH9/b5VzMU2LFQoFTjppDr29vSGMrBnAQiCVouuCs/GWmY0qOwiZvHEli1RlY64GtNCAefUijfIK1jz55C0vTnknWyb6UwrR0o7zx+cZ/N3v/KOzyiVuAixpBAkV52wFzSklEAWb8ff+Te+hx9FoeWjLBvzAHwoNlqQ0OkDjcYfStuEGKNdFiiJIc8jz+OepUEJY8AlN5++0k8bCTdBjbGyM3t7FgGKnnXbg5ptv4tVXX+Gee+5mjz2+TUdHO67rhnZ9fNJHy4hC/izK1JAxCBUMIJNwCD4HKwM/+tFxfP/736enZ1HVCoSpHYKQ4u+++x5nnnl2zbJYKAikxPUUTcsvR+sFZ6DGy0gsFF4FOiZ75EOYX4G8/ljTaOVhSwGiiBKWHz/fsEMoj2A1tVOaIE1EFGIiIrKOpI2nz1qyjfdf1P+TyI+GZmkxeMXPcMslhLCAWjMjSzFM1tQU2gdlanyMLw4+BvvTz/AampHKwdIWlrZQ0oaxEZxV16DrpONAKT8eAgqha+dcdE4EPCUtVechowCYVEZS4rouX/7ylzj33PN4+umneeihhzjooIOYPXs2nufheRM2fXTHXBYPeTV+HooLMtMkjn4OJvA111zDuuuuy+DgYOjdTzJVPM+jq6ubm266iYcffjTxchEp/cNF0w7Yj8I3dkMNL0TYBWROHBCcxff5sHCaCgwO9+MNL8Jyx7BtgbSKufbE/29QRK/8LxSGb3e3tMKrrzD824d9b7rp0tcIJaGSSc0PzwVLsvi0c7CeewrR3hGeA1FCo4XGAsYdl655p1Ps7vZjAgQKdxJjul4ymgC19kb+AqS02Hrrbdhggw3QWjM+Ph7uKpTSMmqZoJy4kypNC9ZDaQ6xNMde8F1rTWdnJ7fddlvFQTmxLJjmRLVtm5NOOpG+vr5aUwBQQgJ+h3ddNg+naymEo6rsvrR6h9FxpMRxxrG225L2G27A2+0beEvNwB0coTTYA2Nj/ik5ywYRPVmnMc3IPHZltvPY9I5fcZOjMk//miZiYj8i8KTG0xJbWPRfciXeeMlfDkzgfUnA6pAfz0PYBfruupfyVddit3WiXWfislgU2BKGB5Df/iad/70X2vXQlo9SKpWqyT/u1Y+Xa0qfxCMkhASLTnr/6qV0x1nwuVAo8NFHH7H99ttzxBFH0NfXR2NjY+KW2vgAiDoZa47YJmjoegVDnGeT08cET13XZd111+Waa65mZGSodo0/RkopWltbeeedd8IIQlWdBwjpe3+169K4+mo0nTGH8tgAnl1IzLfKQ60rpwCVolAsYj04H+mM8uU//IZpLz5F4/13Uzj8h5RWW53S6Dju0CLE2DhYIGQBjcAVsiZUWR4onpcCHqWm6riyZiK2QIBiJktxxSEiQk54LjS34L3xOn133OPb5V7ylXRJAiUvKY2v+W2bob//jcXHzcEuNOFpfydowJlAIlyX8c5OZpx/hq/tQ4SUbtqYHKJpCtw0xgPKgQ11IhKJawnP82hoaKCtrY3rr7+eLbfckieeeIJCoRBu+FkSTpc8ki+J8podcbIsC8dx+P73v8fRRx/D4sWLKBaLqe+7rn/P4M9+dj3PPPNMTQQh/wivQFTapvuoQ5CbbwlDfZAgYKoGQFAHfM+zZRcZOvpYFlxwOY3LLkf7Hrux9M+uY5lXHqfrD/egDz+SsRWWxxsawRvuQ2iLBiFRsj6Ul0R5nK01FIO6Jp9DnnJq+hW/TTxcBDaNusTwXXfjaBchZc4d9dXl5XkmlIeWFmqgl4X7H0374kFU0fb3AKAnmtmWOKP9tB5zBC1rrQmeh5CSCd9b/WZHFgowImHtk/GlYBlw/vz5tLd31mjlNG1h25KRkWEsy+KYY47lrLPOCtf3pZShRjSVa5qkWYMq+p7JYZSm3fJ4v4N0WvunEHfccWdefvllOjs7UzcJWZbFyMgQq6++Os8++yzNzc3GzvU8/wjz8MuvsmC7XWlGhnAxizeJwEPiSosGygyM9NNxySUsNed4vPExrMamcNyV+noYeeIZhu69Hx5/Djm4CMtuxWpowEUjlAKt/UlSByxPm6xBmnBjlNZhJKOwDrr6nSDvtMGbSFqHayCqILGGRynNXp6lfnMnjet/BaU9bCFR1GrRtEmUJ63yXIRd4PMfHoZ3623+IR/XrfbrSIkYH8VZZVWWeekJiq3tNTsbTU7xPKZZ9HnW2NFaV5sA5sb2tYMJRphgeZCH63o0N7fR2NjMxRdfwlZbfY0XXniRQsGHt0nbipMmqOl7kgZISp/mfzCVGS8nSNfY2MgvfnEryyyzDGNjY6mx7D3Po7W1jTfe+AuXXnoZlmUZ625ZFsp1ad10Y1qOOxZnrB9ZsdmBSkiu6nqpyj9XaMDD9kq4AlpbuymdcjoLr7kO0diEKpdw3DLK8WiY1k3XXt9h+fvvZvqLT1I842zKKy7L+MgAenQYLS3/2KmOaCvSNZKpveMCt2q8CH+vw0Sos2SfR5YmrJ0MAi09hPCwrAKF0XGcZZdlxgN30LThuqA9LCknwmrFykqiJD6C/YVKeGjHP+G3+KZbcW69E7u1A+G42HoirdCe7/jzHNrPOoVCRyeeVjUOv6QxnNUmVYFqcyAC4Xmejkvx4K+PAHZj/vz5mWcBTFo2+FwoFOjv76ehoYG5c+dyyilzwj0AdkKwzKlQ0sBLko5ZCCFIH2wkCs4LPPnkU+y22zdpaWk2Dt6qhhYCx3F48snH+epXv1q16clPRHiVmDc8zCdb70zxrb+jG9uRnlvDM/hhw6vqVZlYUvo2Zml4jMarLmDWcUejSh5eQWAL/PMIQqMsGwm4vf30/fa3lG+6C++1lykoBa2tSCFwlYovRyeiLRPyqqe/JqXt/QzCuntCIBUIq0C5NEh55iyW/v09NG+wvn+FVmQvR9CnUZ7jdYmScfwIQFWcfgWb0ZdfZfH238amjMJGKK+i/X3UY0mb8kgv7LIbs/9wjy/BLfP2L5Mj3DSWTW1peh5t5yCNjHdakHkW5dGywWfHcWhvb8e2bc4883S23357Xn/9dQqFAlpr4zp5WhlZ/ORxGEbrnXRPYPR79HfLsnBdh69/fTvOPXcevb29RpMm+r6UEsdxmDNnDuVyuar8UHhI/7S33dFB50XnMa4tLKVCTU/snTh/AYdKaRQ2xdYCo8fPZfENNyEbLKTn+VrItkD4TihVdpBdncw86AfMev5hOh76FWr3b1ByHJyhAQquAMuqutEwaXIEvNXrUZiMA7BqgAdlA5ZSuAWgNIjTNYOZv76Dxg3Wx3VdsKqda1HBn8ZPXMhVjZOKA1XYFuXeRSz64TGI8iBSNCAiu4+F9gWTg0I1T2PaeXMR0q7cGF0tlKKfsxRZnN88Jky03jJpwC8pChj1KuuvXV3TeeGFF9l22+24/PIrQn+AW1kfNUJGzALGVLk0HvLYtPEOiP8LyLJsHMdlzpyT2Gef/6anZ3Eo0Ey8eJ5He3s7zz33PFdeeWUYayAcUJWTLcKycDyPzp23p/jDAyiNDiDsgu/OEhUeMrpIINHKAW3T0NLE0NGn0HPH7ViFAsrx8JTnR9m1JKJgITyFch0sWaBz511Y9ve/ovvJh2CfPRi1FWq4D0srsAsILF8gCRUuIkb/+eda/ItLVI72ruG9zvGnte+38J16HlgF7DGH0e4upv/+Xto22RjtuNiVG3RMQj7IJ5pnvIwoqcDHoBRaK1TlsM/Cw09A/ONNRFMlfqZUvmwSwr/xx7ZhuJeGIw+mZYMN8VwPS1g1kzNedtJcyGqXNArLyz4LYA4JZvIB5GEMCCf80NAAu+yyC1dd9RNWXXWVMH/j3vkplJeWz2ThKhBO4IGBAb72ta3597/fp7W1NXXZU2s/aMlzzz3LV76yjvGsgFb+yktp4Rd8scmO2P/5GFVsRnoKQaT8LEEg/PgIWruUHJeu22+mc+/vUHYc7MpR7GgbUAlWoYVGVi6lGH39dQavvZnSr39Nw2AvurnTv5HX1XjCqzmsUsNDtryqrnuGMK9Ki/Y1MAqhGtFFcJ1BVEsXXb+/m/bNN0M7DsK2U80MkxDI4k1rhSvALmlEo82iK65j5MQTKbZ1oj3DzjwhEc445dnLsewrT1Hs7kYrneuobz0xCRP9LpHnQZ2FEBPXgydBCq3NDWPKOK0S0TyCQd/V1c2jjz7CVlttyS233BLuEIxPoCRbx1RGFqQzwbsk5JM0WKICUinFtGnT+PnPf06xWMyMmWDb/onIOXNOxnES4g5KgfJcGmctQ/v5p+M6JQoVl7lpa28Saa3xlIfQNs2WR88BRzLwwB8oFgp+1Nyg7BAh4cN9y8JzXbTj0LrBBix7y0+Z8dxj6IMOp2zZiKFeEH604yWJGusW6KFzDdwGD1kaRTW1M+3Xt9C++WZ45RLC9q+6N9nTwV/TmMniUQiB5YJstBl85kVGTj+LpqZOlFcd8TnwE9jSolx26DjlRIrTZ/jjJHb3Y5IJmaWkkiZ80hytMgGiBU2WsqS2KX+tdeWmnS6Gh0c46KCDOOCAA1i0aFG4Xh6XVqa84oIrzURI6uSkZ2kdE/weoJmNN/4ql19+GQMDAzURiaLkui7t7W089th8brzxxvC8QZyk5e8N6Pz+3ohv7oY72gu2DL3Otb7v6r8B+VGHFK7VTIsus/CAw+l7+lks2/Z3nmld4xEXWoAt/U1KnovnerSsuy6zb76Gpf/4BN4hB+EIiR7qQ2KBZfkmRcocyppeITyX5r7VWof1VlojVAUBoNF2I7I8zlhrE9N+czcdW2+H55Sw7SKIiZ11UWGf5UwzUfQ9XQkZN/6f/7D40KNocCt7+LVVs8dAWAXU6ABys81oP+C/UZ6HkLXO7/h4iwqmPEKqHt9d6JtKThQkTLa7owVnMWh+JvA8RaFQpKtrOrfffjtbbLEFjzzyCIVCASFqb91JQitZjr80uy4PwjD9HpQTbBI66KAfcsghB7Fo0cKQ/6pBozVCaDzPZdq0Ti666CI++ugjbNuuvWdQ2pVDPpIZF5zNWGcXwiv5N9+IyimyqpYUWJqaG4SU8EWG0Bq3sYmW0SF69tqfkZdfRRRsPNdHAjpiWggqIcqEQEgLaUnwPFzPoWndtVj6xuvofm4+7LM/o844DA5QVDZCSITyiE738BYhEXGImcZTxXcQ7e0a5FUJraEFOFLjSAGWjXDGKFuNTLv7Zjq22QocF8suVkKoJ/djll8pkbQGBR4eXxw1h8J7/0Q3N6G0hyUIPfpaCCwFlqcpSYuOs+dgNTSCVqk3OWWN5aR6pKVLQroySWpMVtrkgSnxSam1Pym6urr45JNP+Na3vsVpp52G4zihhs2a9Gl85jFh6q1HQHFz4OKLL2b99TeoOjQUvBdtg0KhwIIFCzjllFOMAlSAf4bddWhe+yt0nHgM7tior5UNUE/4HyZejpHWGlwP1dBEy0AvC76zHyN//St2wUIrL4CCaQ2CFIKycpCOS/MG67L0XTcw44nf4uy6I0PlQayRQYTdgBLmTV6BCKgZsJX8tZ5Ya0hzwAJIrSgIC+1qxhFMu+d6pu24C67rVC31mcoL26wOs69qIiqFKFj0XXgl+rf3UWifjvbc2qU8rcEu4I7007jbt2nfaSe056BtWUE6yYg0jd8s/0VavWrbMQaz/ZenbtOl2SIBI9UwzIfHjY2NtLS0cOGFF7L99tvz1ltvhafpTIIjyReQxFO9gyGPUyrqqGlvb+fnP/85TU1NYSxBU3mu69HZ2cn999/Pr3/966ptwqFAQCOFRCmP7uOPR234VdTIMELadW1lDcvVGpQHTS3YX/yH//zXPpT+9QHSLtQiEAPPQlgUKKBtCV4Z6bi0bbU1X3rot8z43X2Ut9qS0lA/1viIv4kpjPK8ZHwEgdbUQoDVANqhpBzab7uJzt2+hXJK2JZd1TJ5kFwWTYwpjfI8pG0z+MSTDJ57IU3NnbiqFEZ9igoxCSjPw2vroP2sE/E3+kqk8gOEirSY8P9LVHUaMFrRgLk0D2k90DmgqKCplUoS3zflLxe++OLLbLPNttx888/DuAHxCz3ik9c0mU1OPJMTKG4WxPlMqw9MHBpaf/31uPzySxkaGkLK+J5+ifb3faK1pqWlhbPPPpuenp6qCyh8JvAdRZ7Abm1m2o/PYkxYoD0sJbC0DLfQKuFvDFIVxSIN/wAEfhwC0dpKwwcf8vF3v4ezaAHCsvBcD6X9fQfE2iL87G+gRchi5RZkF5SmY7ddWObpB2m76TpKK6xKeagHoRws6Z++05ow7JmKmTACkEr7ZyP0xK3DOuIPEIBXqYSNH1ZNj5fovOU6pn/3v/zAqnbtMmwSYqsXwYLwr42zLcY+/ogFBx9Fo1YoKUFVxk/l1fCaNtvGHe2h8ft707LeV/zNQuHKSy0v8fkQ1/jB71lh101jOane4SpAFuWF2yZKKjwtveu6dHR0MD5e4pBDDmb//fdn8eLFRqdZmpY2IQQTnEpq/LT8TO0WmCw/+MEPOPDAA6uCiMQHplKapqYm3nnnHc4991zDPQQCoTXKFniOR/uOO9Ky73+jR3vRhQIwsYU0LyLQFXNBuA6ivYuWN15hwX6H4o2VQIrKxE8zBYKnOqgwCPwAlhToPvhAln75cRrPOodySxve0ACeDQUhqgNc1IkKNGAriyKgpEtpdIim665i2vf+G1WuHPP1zxrXspyBRo3VFNW8+kuNFnguC484noaPPoamJn81pepFv42lkEinBNNn037yCfinamVoqmX5riDZ+Vw3/wljHJhYBYgmrrrZJgVC1WtHmyqUlo/ruti2TVdXF3fccQdbbrklzzzzTLjhph44n+U3SHovjZLQUeAPuPzyS1lnna9UVgZqg4gI4ddx2rRp3HLLLTz//PMxUwCUlD70lRqhFV3nnY5aekWEM+rHEwjaIS+iDQScFmi3hGxfCjV/Pl8cchgS7cckVF6iPyBp4grLRkkP5ZQodnWy1DmnMvOFJ1D7fR932MUdHcGy5cRVaHFzOYXl0C9gSTxbUhoao/3S8+k67CCUM462izVp0yivtzxIpagE9SxIei6+CuvhRyi2T/PjOorqMjW+Q9OyLNTYGA0/OoKmLy+PdjVYteVrXX0MPpgPaYomDZUn1S9pfMu4VIgyVHm1prC0yWJy9iWRqez4cz9Et6Krazoffvgxu+66K5dffnl4p0C08Uz8ZzWWif+0AWKSrPEOC763t7dz88030tDQYPQHaC1QCvxQVYK5c+cyOjoa8iuEr5ElYEl/Sa64/Gzaz56DGh9BWAJPAFoitQihfjKIpSqNpcFzPArtXbh33sWnp5zua3TlVUJVVwZ/ilANBYJQ/oWWlr8FVjllmlZdmaVvv4Fpj9yP2mILSoN92GUHYdkhGtBMrBIE5kF1Gb6w8qSFpQWlwUFazz+T6Sce75+ys4rh9WqmcZTXcx5PK4RAVcw0/4SfzcCTTzF83sUUWiZukdZ6Ii6DhUBqiRYW7vg47qqr03n4QSit0JXVSGHgKTp+skzONP5NlKYMIVc8ADMDS4rywjPXdWlubqZQKHLSSSfx3e9+l4ULFxpP1yXZQ3n5nqxvowpaSUm5XGajjTbi0kt9f4Bt157x9/0afvCQl156iZ/85CdVy4K+SV+Jey8L4Lm0/3B/1Ne2QQ8PYgkLxGRcgn7uAg2uS6FtBu5l19Hzk6vxCkW0W0IJhdTJcRkDR2WQU3DET1jCX4bzPHA82nfclmWf/gPNV1/FWHcXenCxf8d9zOdh9MdoiWdBg9SMjyygeOpJzDz9FN/bLyvx/gP/RgaszvIV1cYl9JGVtAqMf/YpA4cdR6OncSRo7ZrHhHSQsoBTHqblhKModnWhPU3g9Df1U9wsNZFJaORFMkH66N+Q3eiDtAbMW9iSFg7R/ILzBN3dXdx3331stdVWIWyeuCKsNo96GzfJtMh6L06FQgHXdTn0UN+H0dvbh13Zllqd54TP47LLLuOtt96q2Rugta8+hAZZaKDj/LMYbmj0Y85XJt6kSWu8ij+i76QzGPzVb7ELDShXo0V6wEmTozQQDEJKbEv6V3VLm+5jDmPWy08iDjqIsdIYjI75++N9tViDXLT2J5sti5SHBhDHHcfSF56Lcj0sWRtTsrZa6f1q+h4VRlL7fGmvzBeHH4/97/fRzY1YrovWE8JcygnDSMgG1Fgf1lc3peOA71UCfQi0llX7EkzzLvp7nMc4SkijuAIM/prQRRUCqH4YFJbuAU/7rV6bOk8ZWlNlEuy88y5cffU1VdeMBe/FNUpez+9k+TbaWBUtd9lll7LyyisyNFQbTsznFT8gyPAop59+RijQqrSDwN9157h0bLUl7Yf8EGd0EGnL6nBbIvgnQoeTSRAG6ZQQSOXhWRZtdpH+Q49k4JWXsQs22iuHsQhMSCrKY5TCtEIgbdt3o5UdmpZfjqVv/intD/6a0fXWxh1ahPQEnlVAaBmEQUEjUEiwC4wMLcb93veYffmloBRIgW8dJU/oOJzPS1qIylU4/rFvYVssuvAKePBBaJ+GdpwK0ploS/9wlP+exA/u0nTq8RQaG/1VE1EREAb+suz4qAmbpcRMSMokDKKUuRPw/2kUTBbHcWhubsaybI477lj23/+ACsy2w41DebR/lLI6pR57MuB14rxAJ9ddd10iUgHwPEVXVxfz58/njjvuqAkhFu59r1yt1n3GKXirrARj476tHO1soGLGJ9eXyHMBQitUsUjDSImefQ5g7P33sexGXO1VDfh4HaN/TaVIrUEIZMFGeR64HtN22p4vPfsoxbPOYtwCOdQDRSjoArYSWEpTsG3cgcXoXXdnmZuuQ2giOxbTB/ZknLoIgVSVc4WexirYDDz5OM4Fl1No7UIpB4HwBSbVk1lXTld6IwPY236djm/uhlIeUtoVT3uFJ8NBtyQ+60HTJiSRp/4JPoCJBk7jIc1Wnow5kMdWj8KgiUNF07njjjvYbrvt+Mtf/kKhUDCeyMua2FkCox7vapCv1jr0B2yzzTaceeaZ4dFhk3T2PI+mpibOPvssPvnkE3OcASlBKYozp9N57hk4ZReNOQBEXARkta1yHXRLkeIHn7NwnwNw+wf9ew21Dh1+9ZpD0SdCSoTtmwVWcyuzzjmTmX98FHb7JmqgnzE1ilewKNo2crAHttic5X95A3aDjYptnokLn6R7JGrMkyTNqLV/2EppCpZk7OMP6TvoRGy8ijOSGg+rqKAshEBqzZjdQOvpJ2LbBdAqXBUI0mbNiyzIHq1D/HO0LUxpTZRwM1C0phODKJpZPd7yPM8DbZn0bpKUDy7y7O7u5m9/+zvbbrsdv/zlneHae54QSWkwydSAeZ2K0TYKbho65ZST2WOPPRgY6KtyCoZwUmmKxQY++eQz5s07J+Habo2shBDr2Hsv5G474Q0PVQKJKpQWkXBbyXA4mia8Vw+BKHvojnbkq3/m8yOO8bWAVpXda/m0alIf+v8AS6Lw8ByH1vXXZ5k//IqG226hPGsZSkNDjI30UF5nXZa7/07sadNAgbTTtWcehWPyWVTxqiqC0NMsPGYOxY/ex2toAteh1kvhzwypQVoW3sgQ9m670r7N1njKRQSbwEL5UrtBJ4nv+FwwmQJRlJuUl6m+VeVEE+Yd0FFJNll7Oc7wZB2IQfnlskNbWzuu63HAAQcyZ86c8BIS02m7+Ptpz+ODLK9jNC4w/bsRBNdddy3LLLM0IyMjhtgHuiLQpnPPPffyxBNPVq10hJ0vKrf+CEnnRfNw2lvB9ShbNpbOF5VHGP6h/byF42C1d8I997J43rlg2WhXJQrDPEK7ytFMZbOMbeN5LiiX6fvvy7IvP4U48HsMfXktOu+7C3vWLPA8pG0h9NS3ziYJjMDFglLYtsXCSy9H/P4RrI4O8FyUEL4QrBmjGi3BUuA2Fug+8Xi/DO1vfEob0/U42KPvVPmFclDaPK3aCJQEgeOTNe0qojQ4YnIQZsHnePq0Aei6/uGhjo52LrvsMnbbbTc+//zz0Bufl7IEUZopEYdn0WfBRJ41axY/+cnVlErlqvfj0tqyLObOncvQ0FBMQPs7BAUWXtmlda21aT72KMbHB2hSRUTKnXLRcjIFmOdRbO1g6MLL6LvnfmTB9pf2JlLkapMkmoDuNmiJ64zSMGsWX7r1BlZ49TFaVlsRXVnuC+zsPPlH2ypexyREqDVoTyEKNsPzn2LsnPOwW1vwXIVI8aQIBLa00MODNO21Fy2bb+Qf9xXJEYjS+IvXK0sxRudRXgUazbNmK7C5wHTvYzyPep4lURK0SYPxugJ3g6u55s9/jK233pqXXnqpLr9AmrDJ8hEkPQ/qEsQG/MY3vsGxxx7DokULsG3bOFBbWlp54403wo1PYVht/KnnSenv21EeXSceB2ushVsewhPJF4tU1QdC1R8YelU11hoXSWuhhaHDjmHoxZeRtj0R9EInw8s8WmriPQUSLKsJPA/teTROr5ywM5ybX9Lkm/0ewrIoffoJi486jiavsrVYuwQbeGJv+f+Ef36l1NpCywnH+k8UE/siyJ4HJg1dD8LOUqZJCgASjgP7v01oG2LOn+R36oMzeWB/XMKloZOJZ/5pu66ubj755DN23nnn0Kse1dJxjR0v1yS56yETr0KIEAmce+48tthic4aGBkOfxcTEoRI3YBrXXHMNb775Zs2qgI327UwFhc4O2s8/m3G3XPFJ1cYiqEInEC4VBjH6g39RkkqBZVMYHWXRAYcy/vlnSMvGqyzHmeqZtz2i/eWbBAIpK6HKlPIPUsUmURKZzI0sdBCQh/Z3L2qHBUecgPXvf0FzM0JpRMSfUlWeH2UBbdl4w/3Y3/k2zeuug+u5YPuTP17f6CW4Wcg3yeYP8omnnywZrwabLKVpAtPz+ADNoqjvIfpbEgW7B7XWHHDAAcydO7dqC3Gaxs6DFJIoCa3EhWRTUzM33ngTbW1thstFROjDGBkZ4fTTT6/Z9hymlL5XvWuPb1HcfXf/1p+IQDHyl7M+Qmv/Np3mVhr/9R6f//AoPNdFoPxLRJLeq7NNA4gfwvxYX5vMx2geWWZrEm8C/Ft7CjYLL7gM+eBDFNum4bhOGHY94KcqD/w7HKVbZryljc5jD8fCj8AEZvQa3TORx5cSHZ/1zpV42ya9939/IDlCU5FmSe96nodl2XR2+hF4vvvd7zIwMFBzqjCPMzAvmToqzp9/dNhhzTXX5MILL6S/f8AYDNV1PaZNm8ajj87nnnvuMd+oJCbmc9dFZ6M7ZiLKrh/cs07UUlO+qBwldh1E+3Ss+Y+y6Ix5WFYB5alw4k4GIVVVoU7kmPVeJoLTGq2VPz4K/vn+sfOuQLZ04GkVnpUwM+AjAmlZqJEhint8k6YN1vfPDFjVSC7KY5IQSyOT0ktLmzf/Kh+A6UGAvbSeEH557fYkpvK8a5KcprR5y/Z9AgrP03R3z+C+++5j++2/znvvvRf6BYKy0rRI2qCrt3MCktIPJfaDHxzI9763L4sWLQ5vToqSUj5amDfvXBYuXFijQTT+YNSOR+Oaa1A8/nDK4/1o27/+Ko/TMmnwCIJDOgLtlWho7Wb8kitZfPd9ULB9xxnVgzT+OQ+lmZVRfvPmH++veB8qIfyjvJZF6T+fMXTYCTQIQPqxCazY0eLosV9dCUumPQ+npYGuY46MLKdV82DSxPVQPe2YNV5NfVzjBKztiHxwO85EfgdQLU3Gn5BURpSPcrlMd3c3r7/+Bttuu23NhZ3xzkriI94ppobNgp7B98AcueyyS1hllZUZHR2tui1JCN8UaGpq4oMP3ueiiy6qOgUJhGfMLQR4iq7jj8BdYxXkiBv6A+LlJvknEtu8ojGlp2hsamT42BMY/+tfEAU7vG03a9ClUR5hH80zbkvHB3/c91GTrwaBRCqPL44+Efv999CNTUjXENU3yocQSAHCstDDwzTsvjvNX13fDwVuWT4SS0Eleesdf14P/DfybXguhEjaCDR1yoItk5ncSfZe/rz967k6Ojro6elht91246677goP3uRtgywbPw+PAQWTecaMGVx99dUoZb5TwI8b0MUtt9zCiy++GDsx6DudlO1rr4bOabSdfibj7jhK1p47mDRpGw8HXSjS2NPLgh8egRoY8H0QSlV2CtbnrJsKGTVa7ERfle1MxZkndLi3YMGlV+L+5nfQ3o32xtAi/dp3v4YC6WnKxQJtRx/th/oSkXBgOVBJlOd6/Fp5KK+i1rpyM9BUtHW9NBUYFCyXBcFC03g2OVr85P7x28bGZiyrwAEHHMhFF10SHtDJI2SSbPwkIRKF2vH3YCKq8I47bs8JJ/yI3t6eqlWB4K8QAtdVnHLK3HAPQWCDV1QP2BLleUzbey/4+tdgeBgdOXxU5YfQ1f+E0ghVXX+JwKrEGhAotLQQjobWDsTrf+aLE08H6U8IJfxbcqI42GRmpLVRvI3rFfYm7a+1Dmauf7uS66/3D/7xJYbO+zGNLZ24ysHSomavQZQn/7y/b7p5Y0NYW3+Nps2CdX8Z7ro0mR8m0lqHAisJHSQpGZMTMY/jLz6+l6gTcCpOvCQSYmLi9/f343ke06Z10dfXV9lZl1wF8+ARBB52KSWtrW3MnXsKRx11dAjJsy73gNrOiXdifBBkCdlgafDUU09l0003rYkqDL5Ds62tjZdeeomf//znNbEQwpZQYNmS7nln4RRtpFJoIfAigTNS+yl0xIvqH8CPxy/A8xwa2mbg/Pxmem68BV200a6L1iL0CUQpCtnNwtncfnkhsum3qjwrml9pQFiUexYzdPjRNJf9JdAgCEqw6SetjaTWuBrajjmici7Dqypvsg5NE/9JaeO+MpPTM/pekilRNXuSmE9zuuWdBFEtkNWpQvhr5bZtUyqV6evrATR77/1dHnzwD/zpT6+y33770dfXU/HyV0+UtM6LViXQ2F1d0/npT69j772/G9rgWWcIsn7Lg06iaQJB1tTUxDXXXB06KON7wj3Po6Ojgx//+EI+/PDDcG+DEBNOKyEFnuvRvuWmiP33oTTSj23ZYWDQLNLE9kdEtHGINrR/5VihqYWBE05n+LVXsQsN/tKgyBZ+8QGcyVOdtrJRo6IpeBoswYIfnYZ++20/nr/yQvNFazNUDx2JUqBGR9Cbbk7HTl9Hegpt2bnqkFcwpLWZqa558zNRTTyAOFyldn9YTdr4b3GG40ykMR5o9bGxUXp7F9Pd3cVxxx3Hs88+y91338XWW3+NWbOW4vbbf8FNN92EbUsGBvooFOzM/E3CR+tg7/0M7rvvfnbf/VssWrQ4PEOQ5Gyqsi3rHMgmaKf1xC1D6623Hueddx79/f1VpknAS7FY5IsvvuC8886bGJjRtpeCIMjGrDNPhWWWR5XHqzzUcQqj9UbS+DxWmy7h4AR/BcBqoGV0gL6jfoQzPOyfHFRuTV+Y4KlJECa1Y1p+0bzS+kK6GlGwWHDDLTh33InV1gWuE0L7MOy4EV77F4AiJK4q037IAVAs+Pf7ZWj9yfq7TG2TlP9kUccS3QhULyNRCjboDA0N0dfXw8orr8Rll13Ga6+9ylVXXcW6634F13VxXTc8AXjwwQfz7LPPssUWW9DTszi8WzCtTkn8ua7L9Okzeeqpp9hxx5348MMPKRQKOI4D1A7aPOZOVluYBGgQWvzoo49izz33pL+/vyaqsOM4dHd386tf/YqnnnoqMSya9hQNyy1H86nHUy6N+CHFllCcByFAC4lUDqJtGvZrf6LnjHP8gCUZRcQVQ9ZkTssjSeHUkPLt/pG/vsXoKWfQ3NiKo5x0RiOkwResY+M4K69B63d2QmmNlGJKdnQyUk32lywJJ6HRBIgmyLq0oB7BkVbJYHD39fUxPDzMFltswZ133sVLL73EiSeeyNJLL43ruiHUD5x/wZ76ddZZhyeffJJTT53L0NAgY2NjNWvpVRAuxSETLBO++ebf2WGHHfjrX/+W6yBRmu0VJ5OXujovKm2vueKKy5k9ezZjY2OhYKuS3tLizDPPZHR0BCFElRNSCIGwJNpTTDvkB6ivboYeGa4cGa6Hksy64HeBdsvYbdMYu+ZnDNz3W/8sfMLSYF6hmOQriH+P+4Cq2jb4oPyNPe7oKP2HHUvz0DBuwZ6I528QQLXN4AfzKLsjNH5vb4rt07Fcp+JXSb54NIuS2iPJZje9k7sOhucpszywgQUm7JhWQJI9FhYqZXg+vrd3MULA97+/L48++jDPPPM0++67D62trZTL5XDiRxsiyD9Yw7csiwsv/DEPPPA7Zs9eJvSiJ03GNN6DZcKPP/6EnXbaiT/+8fkqJBAIn6z2yOrYJNLajzGnlGKZZZbmiisuZ3y8hJS1nd7S0sKrr77GNdf8tOqwUDh5AK0UdmMT084+jTGpK0E+ZbjnP4jIa4oNEFztbZ6E/n2DBNazZ2EXGlh8wsmMf/gh2L4poA0rAqa2i46RvOZivD3j+auKCeMpjbIs/jP3LNQrL6Ha2rBdB4kI/1XxE5gDWk8E80Ciyi6icwbTvrc3Gj9P/5n56vqkiZok1EyCw2Tvp5napudp9D+yFTjR4SAllmUxNjYW2vdz5szhxRdf4I477uDrX/86Wiscx0EpFd4GFG3MqgFegczgT9xddtmFP/7xj3znO3vQ27s43EsfpyzbynFcWlvbGBwcYvfdv8kjjzxCsVgMzY+0xk0bkNH2SUIkAV+BP2D33XfjqKOOpKent8YUcF1Fe3sHl19+Oe+++25VwNGQLIlyFe277kBxj+/gjgwgChLbk+HNO/40Nv9Lo+C5RuJpl0KhgcZPP6P3+JPA06Ak/mm6ZASY15xK5CGiGGqEsNZ+ANGCzeCvfo139c0UmqfhumWE9sOtm+oYHozCv4kYKsitNIy189dpWGUF8FQoPEJvWYaSyYuAsvwJefNJe5ZqAsSS111API1t21iWVWXfX375Fbz66qtccsklrLnmmqF9H5gF9fgTgndc12XWrFncd999XH31Ndi2zcjISM3Oujz8u65LU1MTnuex1157ce+991ZtHc5qjzwdkFZHHwn4S5Lz5p3N+uuvz8DAYIiGAsdUoVCgt7eXefPmGYWKJwQK/8bernNOweuchl12UFbtmncSn/HPxsGNxtMODR0diAd+T8+1P0UULFA6ch1WbX6mvJIEgum3cDOUqR2VQhQKjP3zn/QfeyJ2o6QkXQqemLigpLYioTM00P4akGiEZdPyvX38SaMIV0fqpbg2jyu3pLRB+jTEYMrHlGcOEyBMnlkJ81vCv6hRQX9/P6XSONtsszW//OWdvPrqK5xwwo+YNWupKvs+yZ5L8/xGKdCanudxzDFH8/jjj7H22muF13NFHYR5hIvruhQKBaS02G+//bjttttCn4DJgWdqo3odNhP5THj929paueaaq7Ht2vsag3Bov//973n88cdCh2DAn62Ff2jFc2ldfXWajjuY0vgYwrIqQUVEpqZPqouO/BO+8wJXKXRrJ4Nnn8/In1/zLx71SkEmqZoxq2+S3q39kcpSJahyif8ceTJywRdIq0DR8QW45+/aMZdT+YeomAhS4IyPUl5zdVq33QqtFdoy81IvmsmryZPQYiBA6lHI0fdl8iAWsb/ZFJXqpXKZ3r7FSEuy//77MX/+ozz++GN873v70tzcHMLp6MSPO8ji8DBPRQMN6bouG264Ac888zSHH34kg4ODeJ5rRANp5Hketl2gubmVQw45jBtuuDG8rThLyiZR2gCPPwuQzcYbf5VTTz0lXBWIaw7btjn33PMYGxsLHYL+Q3+SW9JCK03XscfirLkaemwY//SArrGB85BmQksGSEIogXA1QhSwhsZYfMQJeCPDgIVSFZ+BwS+UZDOHZdWBrgQCrTTKUwjbYvHFV9Lw1GOI1i6Up9HS939IX2rV5hfxf0gtQEiUJXDdcRq+/U0KLS04pqvAg/InIfTT8pmMUAyrEkMIJjQh0wZxUobxzOOT1IfiMzjjjDN48cXnufXWW9hyyy1xHBfHcdBaJ8TGr9WqJjiTx04K0EBbWxs/+9l13HbbLygWi/T19VWtEiR1WNTWCi4kaWlp4cgjj+Ciiy4KHZBJ3too76aGj7dnEgV+Dtd1Oemkk9h666+F+wOCZ370oBZeffVVfvrTCYdgVRlCoJWH3dVFx+knMe66VY7FvLZlOHGJqIbIb1oIXM+h0NqMeO0Ves6/yI8nqNzKzcW1kzkJ7ic5wNKQgEL7wqZg0//4k4xdcDm0tYKng6v5qqZuPL+qz7piDpQ9xtuXon2v/0JpjTVJ11nSOIh/j84F09gPPtczd5PIWJO0mH/xAqOMBJ+VUnR0dLDllluy5pprAFSW0nSVBs5bhun3PJMnmAiO47DvvvvwzDPPsOWWwZ6B9G3EcQqcf+3tHcydO5dzzz2vKsKQiaLCcbJOLpho12KxyFVXXUVzczOOU67i33U9Wltbueyyy/jggw+MDkEpJK7n0v7d7yJ23Ak9PISQdjgx87BYBfsj/4LfACwUngcNrdMpXXE1Q08+hV0o+ifmonllFFgPpNbav6MP5WFJSXnhQnqPPImCVtiqiKyo9SifpjyqFJDwl/5l2cVaYWmKKy6PFgoprKrISXm1vkm5xesbT5eWV5LZGf2bRAHPxo1Ak5l4UeaLxSL/+Mc77Lzzzmy22WbcdNNNDA+PhLHwA80Zl+xZDGfxF6egjOCwzVprrckTTzzOnDknMTDQR6k0XhFIZoecuZEFXV3TOfvssznrrLNCTZzqjCJZYOYVDIFtv+aaa3D++ecwMNCPbU/4NJTSFAoNLFy4mHPOOdfoP1ESQGLbFl3nnkypuQWF50fbFRM+rSgaqMpHREKHoatDj1d+VwIUEonCk2ArWHzcHJzeXj84SaS+efoxCQHW5KF97Y8CLQU9J51G47/ew2pqwsMFofEi/Jv8EVXfhb+IgdJYxQLWx59R/vgzBBZK6xoUMVkhn6Y8TN9NfRN9nuUTiJsW0pSJebKZ4VgS801NTXR3d/PGG29w6KGHsvnmm3HNNdfQ29sbLu+ZgnQmMZ2WLqsDAhvZ8zwKhQKXXHIJd999D21tbQwMDFAsNqQimuhvwc0+3d3TOe+885gzZ04V5M6qR9rNQFkmVmAKHHroYfzXf/0Xvb19YdlRh+BvfvMbnn/++Sq+Av4tKVCOom2TTSj8cF+s4UF0IWlBLE0wp6fXWqM9D93SSuGtv7Po3IuQlr8qEAjctLLyaMOwzytwxFMesmDTe8NtuHfcjuho8+8ljOQ50ZfpW221JhRWsmAj+/op/+kvFchsXgkyIcoshRmtZ5rZGK9zfLxH50ia3yD6nta6+l6AeiBIEkXfdRx/KW3atC4+/PBDjj/+eDbddFPOP/98Pv/881AQBA7BesqIC6moZEtyGgZ7ChzHYa+99uL5559nm222YdGihZknC+Pl+9d4dXPZZZdxyimn5IopEO3gLFvOJJCiv19++eUstdRSlEqlKn9K4BO44IILDIFOBEqDEj7c7z75R5SWXhY5XjIui03Wrgzf1xrtapqauyhdfwP9jz+BsC20p0LHYXwQp5WZKOQFaKUo2AVG/vJ3+uaegdXUjecKBPnHlbEOgCsFFh7O3/7m82G4MDVNCZl8W/HPpvEb8hAbM4m8xuZvHtMkRABRTRFISJ+iOwLNBcZJV+xJUQmZ7LoeDQ1NdHZ28emnn3HmmWey2WabMW/ePD766CMKhUIIcU3e9aRGTWuEJArQgOM4rLzySjzyyEOccsrJ9Pf3USqVa87hJ+Xvx5NTdHfP4JJLLuXUU+fWnEOYrG2YVs8AOX3pS1/i0ksvZWRkpEoAKuVvDnrqqWe47bY7qlEAFZtWCpTn0rDc8rSccAxuaQhhWNaqEVhaT8QOiPOlQShdHV9Ag9QKz5I0e5K+E0/H6e+rBBAx3zcYn0hJiKy6Lzx/oo4M88Xhx9HY34+wQSgXUYllEI97YHIERPOXwl8jEfhBVmwKuH/9K1oppCj6PpPYu8Ff0xjIEugmiqPAtPGUpe2TSEYb2kz5pE9QYJKJ4HkerutSLBbp6upi8eLFnHPOOWyyySbMmTOH999/P1yrj+64S4NppnKyGjhIG5gEUkouuugi7rnnbjo62sOAoXnqOgG7p3PxxRdx6qmn1qzDZzlr6vF9BN+DNvrud/fke9/7Hr29vVUrG57n0djYyIUXXsiiRYurQogJ4a9r20KgPEXH4Qeh1tkANTrshwcnMjkyzKoo/2nmoPY8dHMLxb+/Qf+PLwEp/Q1COrmOSeNooryJ9EppsCx6551PwyuvYLd3IFwXqfNNgizSSiGKBfRb71H+z2f+DkqDHyDNNq/lP9v2j7bxZOZAHsVSdTOQiYJ9+3GpG80oqzLV5oWPCAqFAl1dXQwPD3PZZZexySab8KMf/Yh//evfoSCoJ0xXUkXTKID8ruuy11578dxzz1ZWCRbV7ElII8/zKkLgYs4444y6Q4ylUVxzRCeyUopLL72E1VZbjfHx8SqHZFNTE++//z6XX3551f2Cgc2spUQol0JrK61nn0zJ9Q/MoJiA5zl5zK6nRrsusqWTsWtuZPTFl7DsQuWsQHa+pkkhhM+f8lxsu8DAIw8zfuWNyNZWXM9BCVlxei4B0hoairBgIc5b7yChsruyll8TfK+H8phB9eSV9SwMClqrkfzvo6PDlEpjeK7CtgsIUR3CyATJkuBQpHi09sNbWVaBrq7pjI+Xueqqq9hss8054YSTQkQQ+AiyKp7V8Gk8BScLV155JebPf5QTTjiRwcFBXNfFssxLfdF6+8eTPbq6pnPBBRdwzjnnVsfsS9EIWQMkyCMeUz4QkNOnd3P55ZcyPj4e8qmUwvMcurunceutt/D2229X8RNyYNto16P729+i8I1dUCNDWFYBLTQeGiVF1eUfeQdzrbLwTUopLaSr6Dv5TLwxPz6BMrRtHJWa+kzj+XBcCsYXLKDv6NOwJEhlIyubjqIII8m+zl03IdHeKKMv/6kS9y9lBaEOMpUd94clodg0ShpnNWM/LVOlFOeeex777bcfzS2N9PYuYmxsFCknjuVOnSZi/HV1TadUKnHllVew6aabMmfOnHBNO0kQmGyjLNgXNxECv4BSikLB5vLLL+Puu++iubmZwcGBcPkySzIrpejqms68eWdzySWXhLv40mzC4N3o33jdksoLTIEdd9yRH/7whyxcuLCydVmEpsLw8DDnn39+UkP4fy1J19mnoBpbUaKMwJoQEpr8UMBYxITH3XM1dnMz7gvP0nPNtWhpI1wVatM8cH2irSRCK5A2PSfNpfD+u4jGNpQuh6cYTe/lIVNaTykEFuWXX6vEAVgyV5bVixKj42cy5k2NYlSqNnxDdOAFfz/++GN+85vfcM899/LnP7+O65ZpaWmjoaGhEsk2/wBO8zsE3vhyeZzh4WFmzpzJAQccwNFHH83yyy8P+JA9CtHj/E4Wevuv6fAk4ttv/4Mf/vAgXnnlJaZPn1k1mZO0ut850N/fx9VX/4Sjjz6mgiSqrwKPavPJ8apDM0AIweDgENtuux0ffPABjY1+nwRat7+/n1//+td84xvfqOJFiMqRYc9DWhafHXws6uc3Itu7wfPr6l8ZzoSwiPEQrVP8tygFZSFAeoqxpgIzX3iKptVXQ3keVuQasKw6A2hPIW2L3ptvYfCQ4yi0t2CV8M/nC1VzlVcaf2lCdoIkuOO4Sy3F0n96nsJSM/y7AmJ9mWS3Ryds0pzIY/MnjZnou6b3E9/zPE/HmY9SsJQUOMYcx+Gpp57ml7+8g8cee5yFCxdQLDbR2toSasF6IEp84k78BcuyKZdLDA8PsfTSS3PIIYdw1FFHMXPmTGBCEABVdq6J6nVk+n4Km9HRUU444QRuuukm2trasSw73BocrUf0fV84weDgINdffz0HH3wQjuPk3gWZx+kTTeefV7B57rnn+Na3vkWx2FgRAH7QkKGhIdZdd12efvrJKjQTtknlgozxf73Hws12omFkGCUlwlO40hcAGn9+Jk2sJP7i6TQg7CJ6cAFqj71Z9v47UK6LtJMDlfgYQlBxUPgOOMti5O13+M/XtqNxxEVIicbDUqLmfsN6+I3PheB6sGAFpVwq0zX/N7Rtt51/E5A0823Kq14ymj4xxZE37yR+aoKCRl8Awig8wZZa27bZaacdueOOO3j11Ve4/PIr+MpX1mZwcIC+vr5Qe2YxlgSpo17ewAbv6upmaGiIc889l0033ZQrr7ySoaGhcELlETr1alrb9uvc3NzM9ddfz/XXX4/WipGREePtPbVlCdra2jniiCO44447jJGFohAuyl+WMAso6g9wHIevfe1rHHLIISxatKjiu5gIIvrnP/+ZO+6403xOQFrglmlaeRWKhx+EMzaAti1cSSVUdnpd63JMAbglaJ2B+t1vGXp4PlbBDiMIGcsQINBoLfFw/bHhuPQdcwJNPQNgW2gU0hDWu15KdGij8aSPAsafecGvyySLqsefYno3bx5JSKFKMGufjA+TMg2WzwIoWS6XeOKJJ7nttjt4/PHH6OvrpamphcbGRgA8T6G1yq3VohWt/FpBBBbj4+OMjo6y5pprcsIJJ3DAAQdU3e5jsq8n09hRiRkItVdffZWDDjqEt956i+7u7tRQYT4SsFDKo1Qa45e//CXf+c53ciOBelFAMKmHh4fZaqut+fe//01LS0sY/tx1XWbMmM6zzz7DzJkzQ6QSUgXOOosX8fFm29L6yRd4DUUsT+PJSjkkD/osGF2jXKSFWxqFdddm2WcfQzY1hceKk8ahEiBKDrqhSM85lzA272yK7R14rh8bwSIdQufR/lGegzIrKUDaiJFB2Hprln7yDwhkaLYk1TNvmXGbPstMiNcrbd6a6hZ8D/cB5JXkgcNsAhWUKRaL7Lrrrtx77928+OKLnHnmWSy//HL09fUwONiPlCKEnmlMwoQPIFZqBRF4FIsNdHV18+9/v8/BBx/MNttsw0MPPRwilWAzUZTfOP9p34PfgkYNBF25XGbjjTfmmWeeZs89v8PixQv9gJCxq5qr6+Qfdy4Wixx44IE88cQTVeHFktogyUmYlDZoN601HR0dXHHFFeHED35vaGjg3/9+n8suu6zmajEAJS20UhRnLkXHqccz7pRACjwx4VFPg/+m68WDgRe3T/0CXaymNtSf/8yiG2+rbA5y0FqgtRkNCLeMaCgy9OzzDF10IXZzJ8pTWKr6SLPJyZtXmCb6MPADq1iFJspvv0Pp088Q0j96rA1nSeKTLsv3kAXp03xGScLLVE6iE3AqtgoQ2sUBKhgYGOCBBx7g1ltv5aWXXqZUKtHa2kaxWDTuh88qP175YHAPDQ2hlGKXXXbm9NNPY9NNNwVIvD0oLmmT4LbJieO6E/EEfvzjCzjnnPNoaGgMw4UlDQDLglKpTENDAw8//DAbb7xx6IwzOX+idU2y+ZLaKwhicvLJp3L11VfT3d1NuezfIhQsDz777LOsu+66VfcOVDL1LW2nxGfb7IJ47XVoakJ62Yoh0JRCiPD+gTiP0XbWQiJQSK/McHc3y7z0NE2zZ+PhIUUBWQHzYT/hLxmWBwf5/Gs7UvjHvxBNDQhPpfoO67GR0+olESjhe//d4WHaf3cXnd/cDc/1EFZ1WJWoczaJp7waO4vnehCE6b2ay0GjifLadoGmDDSL67p0dHSw//778/TTT/PEE09w8MGH0NzcTG/v4jDKbXSCZk3+OAXltLS00NbWxkMPPcR2223HIYccwr/+9a/KcljtLT/xOtVE0k3gJ7C1Pc/D8zxOO+107r//Pjo6OhgcHDRGIg7I8xQNDQ2Mjo6yxx578Oabb4Zmi0k6J/GTKslDYWPheYrTTz+NtdZai6GhoVDI2bbv1Lzggguq6hW2iRD+2npDEy2nnYinK9sAUnw1SXVOS+8LCY3UAqe5geLnHzJ4wWUgpX89mdZV5gD4/SQsi+Gzzsd+82/Q2lzxwpvLq9fOjiOFqveF71uwtEBJjaVLjD/7QsXbULlUJIZ8ooI1PklrfDBQ07d5+I8riui7aXO6SqGYlgGnQtFKxiP+fPTRR9x1193cffddvPXW2wghfc+6lLielwj98pBvkngMDPQzY8ZMjj/+OI4++mja29vDrbnxSzYCMkncNCkd+EEKhQLvvvseBx54IC+99CLTp8+sOeHo5+1/tm2boaFBll9+eZ544gmWX375MBxamsYwlZ/Er49UPGzb4oknnuSb3/wWra2tlUGnsW2L0dFRfv3rX7PDDjuEPomwvtq3p7WQfP7NvbAfehTZ2oHnOZXYQdrIgxJRu3KCzyQSWuNZElsJpFAMa5fpTz1Ey6ab+Zd1WAUQ/hFf7ThYhSJ9f3iQoT2/R6GhFU+VEZXInaZWmwqajdZNxbIRQqBHhlBbbsGyTz0CFv5x6mCpJIPiZsFU+AzyCeZX7bjL9iOEAiALlmRR1oQJHGkAIyMjPPTQQ9x66y947rk/Mjo6kmoepOUfr5Bt25TLZYaHB1ljjTWYO3cu++23H0AI0+sJBJJGwfLb8PAwRx11FLfffjvTpnXXwLGAf58/SX9/P+uttx6PPPII3d3doRCIp0/q0DzOoIC3Y445lmuvvZallppVCYkmGRkZZd111+Xxxx8PfTNV24Vdz79E46UX6d/um3gFG8sDS3sgDCaTqF56i15BloQUgmVFrTUUbLzBftydduFLD/3Kn9RWZVArX4CWPv+cj7bchc7PPsdrsrBc4YckS7CNJwu/46Riml0IiVAepZZmZr38FI0rreSvYMjkiWwyN6MCe0kKgywfQpSEiF0PnscuSbLp0gJiBL6BKGz/7ne/yyOPPMzjj8/n4IMPoqmpkd7eHkqlUrjzz8/L7NxIgsOBbd3V1cX777/P/vvvz84778QLL7xQddjIVD8TpZkogXe9tbWV2267jcsuu4Lx8fEq/0PcCeY4Dp2dnbz++uvst99+jI+PI4RINFWS2tP0W5VtVzlxN2/e2ay++hqVswI2nucvbb7wwgvceefEsmBVX9oS5Xo0b7Y5eq9vIYeGkbbEqxyCqaX6FIdflvaDh0jwXI9iazvysfn0P/gQ0rb8c/zBmJSCnjmn0/zR+5Ram3CVrFruS4O78XFtateksR89sjzRtr7Ashb3UHrjjcroTEdvwRiIKjYTdDfVJ8vBZyrHlD4JDSTuA4gnjEqqeiRWnImoIPA8j80335ybbrqJl19+ibPPPpvZs5elt7eHkZFhhBDYtjSZoDWSNP4sCOs9bVoXjz32BNtvvwNHHXU0n3/+eWivRyddUgNlNfrEaojLiSf+iN/85te0tLRUIvZMLPdF83Ech+nTp/PYY49x6KGHGhGDqZ6mvKK/xZ2FSmm6u7vDVQEhJtBYW1s7V199NQMDg6HvJsjDDxWq0Bq6Tj0Rr6Md4ZaR2q6Z6jW2aob2jyarcIpA4WDRYEmGz78MZ3QULSXKdbBsi95f3Im4536sjlaE42ErltxBn0mRQOAy+tKrgL/aYxIyNW9NQsPH+zSL8joCA8rVjCaHRNwBYXJgRf/G0wa+Ac/zcByHFVdckXnzzuaVV17m5z+/mU033YSRkSH6+noBQqdefvJjEXieorNzGo2NTfzsZ9ez+eabc91114Ue/bSoRHF7Lem5L6j8zTg777wTTz/9JOuttz6LFy8OhU00bXD4qLu7mzvuuIPjjz8+8wRhljSPpwlsQ9d12WWXnTjwwP1ZtGihf9hHC1pb2/jnP9/jhhtuDAVAtE8taSE8l5Y116Rw4D6URsfASkBhOnLe3shZLX/Bjj60RiCRnotqbkX/6RX6fvozH0HZFiPvvsviU86FpmZQgoJWgPI3BdZpspomaT15TCg/kKKI+8pruK6HtqwqZBRHp3nKyNO/JkpSXFllhO9P1gmYZKNmOdiSpFNg+wcTRmvN/Pnzufnmm3n88acYHOynpaW1cvYgOaxWUoMHQqdUGmNkZITNNtuMc889l+233x6grq26aRQsww0ODnHEEUdy11130tXVHWmL6MDTWJbN4sWLufjiizn55JOXGB8BBW2xePFittxySxYuXExjY2OIklpaWnjxxRdYdtllavtQ+TO69NFnfLbx12gaHELYtu99r1DaYDUphrBvIn4Dge8UVFKiXYfx7mks/8ozFGcvx4e77YV++FHstjYKrgtChqZB0jjLqwHTzMrgWdVtycH/pS+wxhubWfqVpyiushLC8/w4B5Ogyfrf0vxCefPMjAeQVnhepuMIwoQogrgDQeAQrTU777wz999/P8899wzHHnscHR3t9Pb2hGff63Xm+Vrfj0Pw2muvseuuu3LooYfy2WefUSgUQngc1G8yFNShvb2NO++8gzPPPIuBgX6U8qqiDfn/fLtwxowZnHXWWdx5512JW4ZNbRZPk2TrKqWYOXMml112GY7jhD4H27b57LPPuPba68J0VWVJgXY9Gr68HM2HH4xXGkYm7H3PQ2l18AORaGSxGevzj1n0s5vp/cUvsR96mIaOdizX33Cjc+qreid/HlJCoKVGaw+r2IDsW8jgcy/4CwApGjwPYkuiepx6efOM5p2IAJIkSlxLpEmhpHeTmImniW8u+vTTT7nnnnv4xS9+wVtvvYVtF2htbasKMFprB9c6aLTWlYi6/qm9L395Bc4443QOOuiHQO0mouikyNvAgTCxbZu7776bY489lrGxMZqbW6suGoXqWIX33XcfO+20g3GjUN52S+PlqKOO4sYbb6SrqysUdg0NDTz11FOsssoq1ZexalCeQlugFvTw6WZb0/z5ApxiAaFUaO4LIRI9/2ljI0QF+JrWQmBp36YeaGlAY9E5Mg5WpY7ULvnl9UXlTR8fQwEC0PgWkNTgFhvQQwOUZ3TRffdtdGy7NahaBJDUN0kIJW9f10OZeeYVAHkKSOrkrMMt0TSmPALNFGjQ4eFhfv3rX3PzzTfzyiuvVTRue9hA8ag5aQ1dKBQYHx9nZGSYnXfeiQsvvJD11lsPrXXV8ly9AiAoJ5h4r732Gvvssw8ffPARM2bMCDVxdDCUy2UaGxt5/PH5rLvuujXHiJPKMJlepnTgmwKbb745CxYsoLm5Gcuy6OnpYb/99uPmm2+u2u2IBqUVSnkIu0DfpT9h6OS5NLR1oF1/X4Egv7ZNEgDgTzSpRSVPgVvxEtpSkHbqZjLjNA+FwkkEHGn/UJIsUh7qwdlgA5a77Saa1l4z8VSgSQCkjaHJmgJZ9cgUAHFnVxbsSPqbRxulMZdlx8T9BEppHnvsMX72s58xf/5jlEol2tvbq+7vy0f+akNfXy9tbW2cdNJJzJkzh8bGxkQ0kLdOMOEX+OSTTzjggB/w3HPPMX369PDcQtBm/p6CEVZY4Us88cTjLLPMMlXbdfP6WYw1FBNbmX/3u9+x77770tHREbZTqVTi6aefZqONNprYoShACNCeAKFwB4b5ZJNtaX7/Y2iw/ANElZiRyrAUmHcMRElVJpytQUn/XECALiZjqtZTflX+WoMQlSvGFQgbW2lGR/sQe+3N7J9diezuAsdB2xKJrFvImHhJUjR5kHa0HlkKNyhHBoMvqjXzvBwvMK7N4s+T8onarmkaIkgfeMv9TT2w88478cADv+OJJx7jgAP2Q0pBT8+iqmPJ2Z2icRyXzs4uQHDWWWexzTbb8uyzz4WrD8H5+jwTLU7Bcd3llluOhx9+iIMO+gGLFi2oSh/Y5e3tbfz73/9mn332YWhoqKY/6pn80WfRU4Hf/va32W+//ejrG8CyCpW4Cw5XXHFFVbsLBGjh73VXmuK0DjqOPZRxb7QCd6udbvW2S/B7tI+kBruSs9T+3sN8fVhb91CLJ4wvE3862J8gKpuVtEZYEuGWGSmP0DJvHsv/6jas7i6054Bl+zsBM/KP8mX6XMWDoT3jijFPGXG/ULwMrbWPAKISYapbUieTZiplxf0Eb775Jtdffz2/+tV9LFq0kJaWNhobG2tOCablXygUGBgYwLIsjj32WM466wxaW1sTDxglCcVoRwRtG2xwuuCCCzj77HNobW2tumw08E8sWrSAvff+b+66667wWZogTapT3PwK0n/xxRdsttkW9Pb20tTUhGUJRkdHeeCBB9h6661rVyM0gMIdGePTzbem8Z1/oRuafds3gSYzaSfzXlZewWeTKRovU0P1tV9FCzEwxmhXG90/u5KOPb/jXzyKDq89D/KZTH3z+AaiVC/KziKZx4Y3fY4yE31ugvFLgqKaskpjVA4hBYd01l57ba699lpeeuklTjnlVDo6OujpWUS5XA7vH0jLHwgvFW1sbOTSSy9hq6224tFHH42gAfOux6jENR1JDgag53mcfvrp3HrrrQDh7scgrX92fya/+tW9zJ07N/XmobzaIJpeKcXSSy/Neeedy8jISMXE0JTLZS699NJQUFW/CMrzKLS10nbc0ThuCWTlKrApdHHQZpNxbqVRXJsmIdvoWNJao4I0QiALRegborzmiiz12B/o3PM7eG5FGEv/iHo0n/hcSuI5aQLHx3aSwE/KO+lZUj5A9XFg06EC40spcC7t3TxQNanMIP8slBJ3GH766afccsut3Hbbbbz//r9pamqmqak5dS/BRHlg2wWGhgZQSnHkkUdy3nnn0d7enogGTLzHBWJgwhSLRZ599jn2338/Fi/2/Q+BcxDAsiQ9PT1cd921HHbYYTVaOa/tbzKzgmXA733ve/zmN7+ls7MTz1P09/fxwAO/Y5dddgmdoEFZWuH7AkbH+c9mX8f+xzvopgYsT1XcgdXlpvGZpP2gYlgIf59A3P9Xj3+pXhIItNAgLSxhURpaCLt9g5k/v4HGmTNRjoO0Kku5U9Rr0Xqk2fxZJnlW3lnPUuMBpEEnE6VNzjxM5RkopgYxSd24IFi0aBG33347N910M//85zs0NDSFEXPSBYEOo+z29/ez7rrrccklF7PDDjuEvpOgffJ0WPR54Bx89913+e539+att96mu7u7skzoX+ChlKJUGuN3v/sd22+/vXFlIF5mnkET8P3pp5+y9dZbMzg4HJo+W2yxGY8++mit2aFBKQfLLrLottsZOfBImluaKemaTYK5BYBRY/kZVFYGat/JQ5MSAMLf4ad0ifHRERp+dDxLXTwPWSjiOQ7Stqe8DJlc9uR9PXHKKwCEMOwDyCooSVoFz/JWyJRvfDtqklCI55k2AYJJGgiC/v5+7rjjTm644QbeeuvvNDQ0hcdl4wdy/LxAKS+MajQ0NIzneRx77LGcd955lau6q9FAPVI7EAILFizke9/7Pk899RQzZswIeZFSUCqN0dbWzpNPPslqq61WG8ijTgraLUABN954I0ceeTQzZsxAaxgZGeTee+9l1113rapbcGuQQOOVy3y+1S5Yb7yOaGpBe25NUIw8fJgoeDcuACZTz3pI2AX0+DDjDQ20X34J0w89EK1cPDQCP0hJUvyBrLFYj0ZPQtqTRQPR9+NUFRMwjRGTQymNstBA3rzyoIKk8qNCAKgSBMPDw9x1191cf/3PeOONN2hoaKSlpTUxWpGU1W0wMNDPZpttzpVXXskmm2wcIo74sd40dBNQEDdhdHSMww47jDvvvIuZM2fgVQJlWpZkYGCAddddl0cffYSOjo7cGiF90GiU8gfnbrt9kyeffJIZM2YwNDTAJptswsMPP1x1qhE0aIGnHWyrSO+vfsXwfx9CY3MzJdzKOj5V4bmi7aAB4UcY8ZfZslCLEBNHhqlFFdHf8iCgKueeX5WKp18gCxZqcAB3ueWZdvsNtG/zNRzX8c9ECIlGVfY85J8DWX2U1zSKC4565kSSqRX6peIvTEVy59X4SwIypVHU5o1WOHoSsbW1lUMPPYQXXniBG2+8kdVXX43e3h7GxsbC48hRfv1sRGXCCLq7Z/L662+w/fY78OMfXxjmH9/GG3j9TdI84C9wYjY3N3HHHbczZ85J9PT0hulc16Ojo5M//enPHHbYYakOn/rayefBtm2uuOJyOjs7KZfLtLW189xzz/Lb3/425oD0nV+WLKCVR8e3v4Xa/Ks4Y0MIYafeISKE8O16Pxv8PQZmJ1j4vYbfZMdarvrG/oEvrIRt4Q4uxt16C2Y++zDN23wNz3GwwyPpuoJupqaB42TUyHXMjSztbvo9HnOzSgCYBlWSoyarsLikiQ74+AA2TZAgbR5JF9cCcVMk+jwqCBzHoampiUMOOYQXX3yJm266idVWWy0MWxY9gVhdpqZcLtHc3IwQgtNPP43ddts9vM4sugkpCv/i9Yi3T7CScfHFF3H55ZcxPDwc3n3gHyHu5r777uOss84Mw5NlUdKk8XmauBZtzTXXYM6cOQwPD1fCtdlcddVVlMvlqjYI66X8/fAdJx1NWfvbeMMIuQm8BE0QH7hRmBxXIvG2T/strc41bSAqV5UJgTPUiz7wEJZ9+Pc0rbACwnGxLInOORmTxq8QtcFn8iC3uKaPm8XxOVUPxds49TRgXlgRMJMHAdRr3/9PUsBLEN4LYGRklDvvvJOf/vSn/PWvf6GxsYnm5pbEY8O+QLEZHOxn1qxZXHLJxeyzzz5hvqbJk8VTYKr86lf3ceihh6GUorGxsWIqSHp7e7n11lv4/ve/H+7umwoaCAZFueyw44478cYbr9Pa2srixYv51a/uZY899qhxPmp/nzAozac77Y717B/RLe1oz8XWEWEQ6cdQ+wdtl+DgM2k242/4R3NrkIKhjqKS3vcrSLRdxC6PMaxd2s6bx8yTj/d3MyqFFLJiI8iglCr/xmQpqMNkzN8sKG9KkyfvzOPAeSXWVDyUSQ6SpEqnLVfGG62exo6HLbvtttu4+uqr+ec//0lTUwtNTU2JgsC2bUqlEiMjwxx22GFcdtmltLa2Zu7nT+rkwDn4zDPPsO++32dwcJC2trYQXSilePDB37P55psnOgXThE5cAAeC5IUXnmennXahubmFsbExNtlkYx599JEaQSYQKK+MLDTQ/9CD9Hzrv2ltaMPVHtYkBbfJrk+08anY8FrX3FeQtC9BoLGUhW6wEEODjHVNo+Nn19K15zfRnudDlIiwimYTd+Yl8Zc2IYPxGK9PNP96aCpKMng305Ucrbgpk7TveZgIysiTT2BPJ0HFeIPEGzUKnUwdF2zycRyHlpYWjjzySF55xb/9aNlll6W3d3F4pDietz9hbTo7p3HDDdez/fbb88Ybb4THg+PBPpIGe/B7sH14m2224Q9/+D0zZ84MdycKISiVShx44IF89tlniRuF0vwEcRMl8F9sscWW/OAHP2DRooV0dnby0ksvMX/+o7UboIRA2za4Hq077UzjFltRHh/wbyRK6cug7U0D13Qlex5TMQ/5QsNCNRRxB/px1lyDmU89TNee38Rzy/7W5mjeCWM76dBaHO6b5kaaqZM2d5KeZwkNU1nRd8FgApgknSljE4SPD+J4GtPnNEQQfZZW2bS8THXLej/4GyyTAfT09PKzn/2M66+/ns8++5TW1naKxWLl1qNq91exWGBoaJCmpiYuuOACjjzySKA26EiSAIzyFNxR+O6777HHHt/hvffeo7OzE4C+vl622WZr/vCHB7FtKxyEpkkXLS+af7z+AIsX9/C1r32Nnp6eEAXMnz8/9J/46RRo/4yEZdsMPPB7+vbYl8bmDpQqE+iWKLoz1TGtf0zCO0yHRmkq8LyC2AOhVhWg0w8/risRhlT/F7i77M7St9xIcdYMlOOEV6JntU0ef1Reqhc1B/XO+27euSDjEzhtUAa/myZlXALGHRXxQZCm9U1OjjRJlySM4pSlEePlRlcNuru7OOOM03n55Zc45ZRTaG5upLe3F0FwcerEu47j0tzciusqjjrqKA4++GCGhoYoFAo1cQhNvpToX9v2NfOqq67Co48+zPrrr09Pz2IApk3r4vHHn+T0008PnYKmOsY1QVpf+cFDZnD++ecxPDxEW1sbzz77PPPnz69yPOrKPXzC8m8Uatt1J6xNN8MZHUIKGd4kFBc+Jq1kElDR34310L4HW+qJyV8zxjQILLCKgKbc30fhmBOZ/bu7Kc6agXZdpG2Hkz9NU5ooOv7zIOQo1bOPI2vuRNvPJCxMbR2gOWl6mFaRtAkURw7RtFkXeJrKTqt0WtqkcuINaYKWpkEYDHzHcZg9ezYXXXQRL774IkceeThKe/T2LkIIqmz9wHvf1dXNz3/+c3bYYQf+/ve/16wSRClJQAXwfPbs2fzhD39giy22ZNGihQghWGqppfjpT3/G7bffHl47ZkI6eSFncGJwzz33ZNddd6Wnp5empkZuvPFGo1YRAMpDFhpo/dGhCOWAKKCErtkuW2//x5FknP+0seubKfgCyhvBK7s0/+RyZlx9MdJu8Ce/ZQ69ljUWozyYFF30nbR+SPo9rZ2yKG4iJ435MH1wPfhkC4tu9kkyCUzPTGmSKA/ENwmepI6MSu085kf0GVSvGvzlL3/hkksu5je/+R2u69HZ2Rk5ruw3fqHgX9Hd1tbGT37yE/bdd99QIEa3EadRYA7YtsXw8AgHHHAgDz74IDNmTKdUKuM4JR5++GE233zzXIFETHWOagchBO+++0+23npbQFIuj/Hggw9W5R/lH63wymX+s9WOyDf+hm5q8j3qTGzmydPPpnonPUsiJTRSCHShATXYh7vULKbfcj1tO2+P6/ibe7Sc8OvnNSGzKM3UykLWWXNlqrwFZcfbsQaHZEEz0+esypkKz7J/o+/FzYg0CZ3VUPFTeqay4mXGn0djF6633nrcddfdPP74Y+yyy8709/cyNDRYtZmoXHZobm5lfLzMfvvtz8knn4JSE1eNxcsytY3WGsvyNwy1trZw9913svfee/H5559h25JSqcwBB/yAL75YEE7OPP1j0qy+sHFZbbXVOfnkkxkcHGBsrMzVV19T1RZR4aWVh9XYRPPhh1D2yljgX/GlgYw+ifMX7YM0tBJNW8nJT4+FtBqw+ntQG23AUk/+gbadt0c7LrZlVy7xmMjXvNfDzGcaLyaUm2QCm+qRRnnN4bT5aUIjYUSgOKSpSlSn5EmCZnkQQb2UpqknizjykhAitLsD594DDzzABRf8mNde+zMtLc00NjaGkD+YmP39vey88y7ceuutzJq1VOIpP5PEDsoMnh1yyCHceuutLL30MvT09LLbbrvxq1/dE6bPsjXTkBn4l5puvfU2/PWvf6WlpZk//vE51lxzzXD7cpBOaYXQEnd0kM823YHiu2+jG1qxPB1q5Dw8mFBkbv61RksLYQncoQGsvfdgqet/gtXZjVd2kYXKKkMGmDC1ex5a0mM6iY8AecXjRGTF8jDxV3U5qElK5a1QkiatN680+z3v72nSdkl1UPA32FoZXHTyrW99i+ef/yM333wjyy+/PD09i0KTIdjpN21aN/PnP8aOO+7AG2+8bnQOQrVGiZcZCOxbbrmF448/noULF7HMMsvw8MMPc95551c5BdNQU9L+gWDwNTU1cs4554T7HG677bYq/oDKmrkE5VJs66DlkAMou2VsIfGk50/MHPZ/kgmQZsdGtT+2ja0V40ND2Ccfz9J3/xKrsxvXK2PZFUZ1bR5xmuzkr+d5HiRdj6bPi8JrfkvaCDRZKZgnj7w2b9xWN+U3FakbfzdqB2Y1vuk9oCqqbm9vLz/5ydXccMP1LFiwgPb2dizLj/5TLBYrUYKbuOqqn7DPPnvX3F2Yx5wJ0Me8efO4+OJL6O7uZuHChdx1193sueceVfZ6FI7mHTTBhqTDDz+c22+/g2WWWZYnn3yC5ZdfrnoDkvBv+BFYuIsW8NlXt8b+vAdZFCjln2o0tV9aHVOfQRiHUCgQBYkcKzPeUKD1qovoPvgHCM9DocOAnb6gymdjZ40Fk9ky1fkSpbjfoC4klJCPiWquBluS8DhNm0+mnKk0cHzAp3VqkDZu8yXZ6tHG951v/qUbXV1dnHPOPF588QUOPPBAxsfHGR4eruwf8GhsbGR0dIx9992XM888K/QbBGggy98hpQw1/bx58zjttLl8/vlnNDU1ccwxx/Duu++FMRRNlKcfArRx5plnsswyy/DBB+/z85/fUjVBZOUfSLTnYs1ciub//m+0OwBWMZz8eX1AUTL1Q5xUQwFvaJTRZbqY9od7mX7wD9BOGU8KhJhwhqaVlGQGJU2sepfx/rcoCWkktV/iYaD4ZEgrxFRYvZAo7Xl8wkYHbpqZEZ/0ca2TB4VE3zHZ4+b6RA8cuay44krceuutPPHEE2y33Xb09fVTKpVQSmPbBTo6Ojn//PPYb7/9GB0dDZ2Medom4MN1Xc4880zOPPNMhoeHGRsb48gjj2J8fLzqnTx9Ey0nsDWXXXZZjj32GJTyuPfee+np6QvPIOggfw2VyN40/3AfVOdMhONUzbxo26dp/iQtXcW7EEjbhoE+9CbrM+uJR2nfdhtcp4xnS6RSxlmfB2InpTObH1NTTvEy4+g3K3JV8F4aYkmqc66twHlt+aRJmUUmARNv9HhjJzVKki8iq7PSJnfae1l52LYV2v5+bMGHueWWm1l22aXp7V2E5zkUChYzZszknnvuZbfddufjjz9O3S8Q5S/UwlJSLpc555xzOP/881FK8eKLLzF37mlVW4Xr6Zug3YITg4ceeihf+9o2vPvuO/zhD79HiPitxhqkf7Nww2qrYu++C+74IFIU0ASDLXvSJI05LcA/ma/9ZTwpcAYHkfvuzTKPPUjTKiuj3HEs28bSNkrUXmYaLdckYLI+T2Vc5Bl3SYgnbVLHv0dXNUxpor+lCoC4TVQPZWnPNE2UR+DUw08SlDfll8RXFmJIq2/cUbjfft/n5Zdf4rTTTkNKQX9/H1rDjBkzeeGFF9luu6/zpz/9KfGasOjfaPk+cnCZO/dUTj31FEqlca699lruvffeGkdjEqUNvMbGRi688EKam1u49dZfUCqVqvcbBO0jFTaC1kO+jyq24Iiyf32WEIiUSz5MFG1XicaRoOwC2lOUR4ZoPudklrrzVuy2NvA8hN0A+OVIoasCiiwJik7QJWG+pDlho5/TTJQ8ZSeaAGkwK49ETGIiDr+zKplVfh6qB9qaKNB2aYc6ommT8ow78oJnQd6O4zJtWhcXXHABzz33HN/61rfo6+tlYGCA7u7pfP7553zjG9/gwQcfDCdummCa+CyR0t81eNppczn11FNxXYdTTjmFd999N/QHJKGhNIHvIxKPzTfflIMPPpTnnnuGp556qspnASAQCGHjaEXLZlvA5ptQGB5FWwJPSN9MSBhLqaTBEwJhN9A8VMYpWnT88mZmnHUW0vPQWqGt2rP30bpk+R2yfC7xz3nGt2mcm+qfF/LnaS/T/ZZJdTZGBMqC8kn2RJo0yqux8zZcGl9JeWalMwmewKbKi2DieZgGlW1PBCT5yle+wv3338/dd9/FCiuswBdffFY5dqzYe+//5rrrrpuwtXXWMk9gt1s4jssFF5zHmWeexUcffcQRRxzB+Ph4zQCLD+g0JBSERTvttFPp7Ozk2muvrWojqHjZAeH5h4QaD9qXMS2RWlBQKnc03bBkLfCEfw2ZZTVgDfQzsOIsuh/5NV377IPnuv6x/crOvijAyGu6TkogZfGfMl5MyiOu8ZMoLzo2naw0vSfTBrFpZ5MJeqYVFK9cUj5Jv0V/r0eIxJFEFhTL4iGvGZKHxyBN9KjwXnvtxfPPP8fcuXNxnDJDQ8M0NbVwzDHHctZZ88LbkNOO/U60kz9ZPc/j3HPP4YQTTuCpp55i3rx5YZm175g93fE0nuex1FIzOeecc3j44Yf5xz/+UX0CUQikAFmJndf1zW8i11oNPT6MZ4mqA0JJpLVGBH2HR9G1wG7AG1xE+f9T3NWF2FVd4e+ce2eSSWIktZJqoG1e/IG+NNJqbNPgUIWQ+CClFiUdpD4oPhT6oi9pMX1I9UUNtgVToWhKXkwlBkklFoTMQ5WC8U20BaHEaolYf6KTmTv3nD7c2ZN916zfc85NFwxz7zl7r7X23uvnW/uen107ce1fT2LzzltXHtPdQ4kRsghWF2vWI9JGQ4Bc39whvYFJ4+9B3vQ7iyCj9wKMMkwJ7R4A2l7KMlY7QN90ie5LaBTl6a0BtTnKM2d+6/HZs2/h4MFf4+WXX8bGjRuwtDTA3Nwcnn76MKanp9lr/RP/XM7o+YWjK/YefPBBHDlyBKdOncKePXvGeKQ15dALN6aUGHbs2IHZ2VkcPnx4lV9RZPeHDCsU/R4+OngIXz76KKav+CqqbE9jdW6UJ+7URYk+alz44r+Ynvsptv7uKUxt2oTl5UX0etPi3NN5t4I/N38W5XxpP06mxFvSz8ODg/mWrY0dHw6Hdd7ZImos2sCk/p4AoJ2fVACw9OJ0kPTIdfQufF3XYzcaHT36Jxw8eBDnzr2PxcVF7Nu3F88//xy2bNmy5lFgnAHn53q9Hvbv34+TJ0/i7bffxrZt29a808ALf9PFQfPz87jvvvvw5ptvYvPmzWO2UQMoqhpFWeDiP/6J93f+EBu+uIC6TJfkrczbqK4Zn5eRRWLY76G3vIzFixew7uCvsPWXj6DGENWwQq8cv5NPC7bcPK/KJxSxZa5v4huZT6ttNDiFgpj0XgAtunHKR6JOdJK1CZIWPiqbO87NhdSHa+OdQ8ozZdBer4fz58/j0KFDePbZP+DChS9w22234ejRo9i2bdva9/cpPJP8ffv24eOPP8brr7++esty0pXSWDYnlO4FuP/++3HTTTfhoYceGtOnrkdZva4rlL0ePpj7Gaqjx1Bs2jLarQfGsn7aGxj5fo2i10f95RcYbNyAK357GFv23w0MB0DRGz29p6qAzCmaJiG6btI6WtfZa35g6abZdySQaDql71SnNRcCcbvXXNajbehxrj3XhlOakyHJopHey59bZNrX21bST5p8jdLVfUVRYDAY4Oqrr8aTTz6Jl146iV27duG1117D7bffjnfeeWf1/v/E34KGRVHgxIkTKMsSBw4cWN0PkOZQ0zs5xYEDB/DGG29gaWlprKQoimJlR3C0Z7Fp/71Y6vdRVqm2H922m/5yvlP9PvD5Z1j4+jdw1anj+Mr+uzFcWgZQAkUJrCAMOjbJRuicaCWpFehzGRRVaAiRIkLunKZzE6KIhFvnkh6MGq0FpzjiEIZF+QAi/TR+FmmoI+cRGbt0nnO8/FXos7OzOH36NB5//HG8++672LlzJ86ePbt6WbFl/AmaT01N4ZVXXsH8/DxOnz69+mbiXC8a0LR52L59O/bu3Yu33nprzd4QABRlD1W9jJndP0C54ztYXriA0eW5NSoAdVGsPn67AFBMTeHip59g8P3RZt/GW7+H5cESyqkS9cpLTKkx53MmnePmmI7ZohwR0uOWDCuJepOrJkPSSVtH9lcA7wRqC+FRlGsnOUZET02W5dSWXpp+WiaV9NH0SO3Sb+3r16/Hww8/jLNnz+Lmm2/G7t27cebMmdWrBi3+KWtv3rwZx44dw/Hjx3H+/Pk1vy7QdZUuVkn99u7di4sXL7IwuUCBuqrRXzeNzXP3YqleBopi7Mm9BUa38aI/hcEnH6G45ye49tSfMbP9m6iGy+j1p0fXF6TWRsbk1pJzMG2+LFtP48+P56UITXBa1qc8OPma7pSkbM/ahPVYcI2itXyukLWIHLRLA7PqsSYyLWpSa3YlJz+WbxI+88wzeOyxx/DEE0/grrvucr2xGLj0i8N7772H+fl53HPPPa5+VMdERVFgYWEB09PT/E0yK+0X//MhPvz2bvQ//hDoXwEUK0GrLNGvlrDw5QDrH/kFth56FMOiQD0cYKrojb2go6t1SOWulhCk4J7GDNj34DeRzekQsV8JXXN6ll7mWgaOOJZHVqQe98q34GwErWg6aBDU0lXTLRlDfu3AAw88gFdffRUvvPACXnzxxeyCIR3SpjsIt2/fjltuuQXnzp1TDdKzLzAzM6PeIVcNK6z72jWY/tFe9JYGKHsVKlSjR4svLuKzuo8rjzyFrb85OAoY1RC9FeeXbM/K4hp51tsqb6UgZNmst+6XkLGGHjl0IV0UVBSF/npwTUGqnKVULpQqmEdRa9KkqBhBFZSvFsWtTCCdp5lKy1yW7tx5igZOnDiBG2+8Edddd91YP6oXvX4jlRfe21s1Z5SCZF2vXNjT6+HCG3/HJ7N7sK7oo+6XqD/9HJ9fuw1X//H3uOKOWRTLA6AsRw8YKYrVx35T+R5b9dh0xHa9PGk7SWeLl5ZsaJmhyVaRDFcCWIpFJiYSADwTknjmvLwIpivo7jWspGOkvwdVUYdOEX5hYQHr1693jSEPqGte+uEch3dMq20qABjiX3fcid6Zv2Hd8jKG3/0WrnzuOWy44XoMlxaAqRmUdQ2gBhzZVdKh60BhtZWydtSPIm09AYB+p6X0mpuB6CA48ipN4XASmvOgmUMjCm3TZ+7mB00PjaeHPE7KycwRTFQfDgEAlzboqqrCzMyM2Y8ez4OoCBOFoJB/15xilVc1ehT3prk5fL78KQY/vhNXnf4LZm64HsuDAeqpdSjrlXv4FfhrBU/pez5eytMqfzj0acmzyhTten1PAGtSslH05343YDSatcm43OA9WV7rb5GXf3RcFgRryrdLsiBlIqkUSee0dqO2FYqih8UP/o2Pjh3HNT9/CMV0HxguoiingLpEXawkCObSYI8tNp1Hj30n0gKLxttqI/H26hhtB5ASII+O1Gi5Gi8fkLUoqV9bR8v7U17WInVNk1iQLvtK/AAeqnLzKWUPSz/OlirUKIclhv0KfZTAsB5dDdDrAQJPyodDLnmftrA7kmg43gDYy6s5/bm+0bW2Sg9rPKs7P5pyaVBUQA6lckikGUhkcjnnlhw+10F7ezAnownMSvI8MqTxe+o3rUTj5lyDyNJ5T0mTzmuQluORn+/VBapejaoqMBhUo1dyleWY89M+1CZz/bl5kq5ZoJ+lMedj9CQUrjxhr4dQSgeNPycn/26tp7VmoesAPNlIa+OBSdxxqnRUJ4tPNOqn+onLTLkMagzeLNWUkrxcB0+fpFsibS2i689+r3HpBZ4N50Ka26b9tTFEeGp1eRRZWKgnyofTYc29ABrR6NqlEUsTZGU0L2+Nj/XeQo4XtyjU0CWE0hX8a0rU2ZtAXqmfNYaiKC49vCMgr83cSEhLkjOJcs1KcBwypKhBGoeU4blyntIYAtAgT6SeiZDEi6vvtP70WJSXxEc6FsmG0ejPlVcefdsQZ3RNrnKzdPOUG57xc/Y4Scp18iJODxKmUJ46tNfuvb5IeY/dS6rVRZxCXU86V9tRok7FwU2tJtJkRwwpMnYtA0nO3Sa4RgKONO4mqICSZKSefkk3qw3XLi+FuPPREqdJgIk4oiXLawseHakvFFVV1d5oRQ2rbeT1GGrEmJvqw/VrUrNF5Vn6NoGUVqbq0sglksYnZW3OrtJxib923qsf/WyRZRMa367ml5MhQX4Prd4LwAnK/9OomveLRiiu7uG+S8c8Mtr28ejmOReVGyVOz0gWiTiTZ5zUsRNR57GQgaYPh/Ki1MRZLCQnlZyafGv+rXmgvKJy2duBE/OoYVsO7BlsNLp7HXVS2RyY3OadZ/G98DY/T/nQutOS6W2fy5SCvibH4tmGIkEkGvgkedo5KQh6ZWsyaKDJ27GXAmtCOIfTok7T2imywJIz0IgcNdr8PxecJHkWz7ZZxyuPa6vtkUR40ezOUVuobumgydT6RIJeRA8p2Um+Y/mN15+4Y/kvdlofEQFwHbWaUzoP+OCzBmOkmofKl3hLkFQjLcPSBZX+SzypfhI1Od+0nNHW2tIjv/aAm2OpxqffIzIl8toghexNeGrtpXnI/9IxTr62iekhL9Jj3wykUZPsIQ0yz8z5xGj1mRS5o5N0Oev2pnsbFkyNQFMp++R8uKwVIanGl+RGj1EZebvLUeJpvsL5hWW/NAh4E5tFWgKiOrFvBtIYUQbaq6YsJalCXAYd27BgHlhK+0YMz6Mjx4Pqp5UenGwuE6T23npZqzml7GPxkZzJCkIW6uCMW2pnEaerFShTew5RUn1yXaje2rzQDXLO0em5nB/3XgaOl0SUl9Y310d8PbhHUJs2uTxvtKMRV9M3Un9KBiQZBjXmSHCx5tiChxo68spIbai8XAeOnyW3DUUzOS0dIxRBX54A1QSpcdk+EnglOeFyJfpMwFyIFf21DMUZNjfpmpwmOlvnuCzRVDbX1zN/2jmPPl07KjdHNEjRcq6NLHo3HScjybb0tbK8VYtTktpF7ZYLLpL9cxS1aVEPKQDQCfQa7qQoMqguZXofHHk5qI2TUcfhxuQJOpKhp3Pcdy9RVBJFHW3q5lwHCQ225R2lomh2w1JkzsQ9gCbQSiIK3TWZXpqk41FH4eYiCtdyB2pjTE2DjgfK0vacI2swOP/s6UPlUceX+NN+ETkRysevvbLbyyeRVb5K/azjTchVAuSLA1za+GtryMBa4/JAvHQuGjAudwQH5E1CKQOn85To3FsyJT5N9J00+vPw9ELstiVW2xJ0UqiwDfprVAJwnSVjBi49nJKWDFx2n6QzRoKI1t/bdtIyohSpj3OKrIlV22oJgtb1TfhHbCda53v003Tx6s7tU6Q2FnVpP/abJQ3SIKE0OZ5Ja7oZIi12V5Pmyb5SkPRAvrY6Nu1vGb10jpaMHCSnhv//QGJJtmUX6VzTF35I9Xr+X9vL4hBNG0Rn6VfUK9wlSNpVpPHAMU8g8DgTRQBcwOEeG2ZBSq982p/2ldo1ycwSQrPmktbcmh5eO5D6Jtke0tpHsznt5xmHtV5dlRXScc6+tMDS1j/LaERuGr2lyJYbohY9aR+vTA1mNslGUh8u8+V6eGpcSZ6XcjnaXGr6aPsPml6a8UulYHTuaQCL9PO2S3/ScyU5tGCNyTPveVuPXXaFotTXwXiNgZIURfPjZVmOPRddg5d5H46XJF8yFprRtaxjOSa3qBqky/tL+nFBj56TdLaCE+XhdUi6plqZ45l3qQ/lrZV0ls6ULPSW8/Ts+nOoQpKRO3bq4y0zkhxuPj1+Q2lsLesRjZ2MwAoJ9nlefijx4ozbA/korLXgrAVPpcXkoLQkw9vGakv1lHSXSh4tuGljsXhYFIHqXVBuA+k7pwu3PtGSQuqTn/f+atPlvEulBadraBOQU6ZtDZLz1DKzZ3E0R24bKdtQdM6ovpYxJ5K+e8sLT9kg9fFkVU0XyrPt3EfKtIjDeGRocyIlty6Jji099JZNAtrPgJHIRAceXeimGUYijZ/XwDw6daG3xzm4+bb6aMTNPf3v0cNqE0FcTfVvomeXukySrHFYwcbySfWBIJGaOFr3UqWpcUfgWIQivDmHo7woemmSNShktQwbWPuI9ibzIB2PIiMN4Uj1q1cf6bzVXkMz6bzVxqOH1Mazhhol/aRykPuu8ZXWlN0EnERUpLWZJDftH+SOlU8onVxPQEr9OH5tKOfVdsFznp7jTSCrpFs+11a5YK2hRLlB0+ybH++q7OJ4eRxKS2RWCWzNkWeeNf21tjl/Cd1wfqM+EYh21jZRKOMopQFTQ6aBwwPrKE96XjMEa/G44x6jkeZNQgQepOQxRouiCJCjfIzcGnB6ckGde1suDdz0uKY7h8pymcAlJMW1lR6Cy+nH6ci1TXo1ubfAKs00PfPjeVsWAUgwQxqcx/C6KjW0c1wEzGtaby2lLY5k5NznnK8nU3gd3YLg0jEPRZyM9uPmmHtXH6ef5DxSkEvBwlNGektSKkcLsBI/DcloaNFzTLIDK/FRyudNvQ4g7+CBhlE+0mePgUu6cKVCNDNqzscda+JsHmOX2loBgaOmCMajm5aZImSNg95rIsnNyw2J6Hltp1zTr2l518V85bwob68/lsBax+EaUsZ0ATyZn06y1odzZEpSjZeyA82+HkOl+mm1dmQRaUCynN1TA+ZrwOlMeXEZyiNHo5ThJdQU4e8NRokvbSNlWA4l0HYcGmuSLDQb09ad+oQH7nPokvue96NrXwyHw1pyEKum8baRlPE4Y5LB9bUmKZ+gqI4e3hLfSEakPGgA4vh4AlBXWdmiNnLqmr8nw0NSEuH0obZAL87Jg2IkWWjtPP2jcqUxe+2b4x1+HoBl8NYA2mQbTa7Vpq1eEeOQgo/m7Pn55BjUOLqeP89Y2tIkkkjE+aPk1YXToSvb19pa6DlKrl8BcmPUSgT6OUoSrJPaeuscSy8uokp8tM/csRwaU2juMRbPXHBklU2RoGKteVuKoqi65jdqm+gkBREORUiBXOLHlWNNdJL0iJBkR65NwMQgCm29balRWtGPfpZ4Ria/aRnjCR4RfpSXVQdy/bS2XWcQSpzjSPpYwV3r65HdlCwn0+yzCx0Sn0kF2bEEE30qMMc44uhNS4VIvcRB7XTcahuFo5w+TWpKK+NpdW2UH0fWuLsqRbrcM4qUgLn8dDyqz6RKMK7c60K+hTQB4zoAT7aIGln+OZLd8gmyBkYdmtu/yMcWyaYe4vTzbAhpvKQ2Usbl2moZ11Ona3yi2TpKnrFzMrn54coyz/p44bcn0XA6W+PxIGPO9mnb/FjJOTllSC/m6BJqRShahnCyaECh/Lw6pXb5lWST2EyjfC2I7XWUvL8ni0q86B6LtecyKQjfdj+oyyRgBcNcjzwwNNlrsRCDtZcxVgK0rQ+7gnccAvEaaRQSWsfbUgqg+Qs0Ld04uO0td6Tjkxgfp69XP06npuVXRIbWxgvBpdJO0iPXM333jk06l+Q1bZ/I9WqwaFak3z3QqWnQsWonC3alz1JQiIyd/lk6EA7Z31o9NeOyPkfJWypoOnWZCDz9LgdZ1+97bLzNGrWZa8k//ge18TFO944+EQAAAABJRU5ErkJggg=="
# BhugtanEase Logo PNG 64x64 (tkinter iconphoto ke liye)
_LOGO_PNG64_B64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAABCGlDQ1BJQ0MgUHJvZmlsZQAAeJxjYGA8wQAELAYMDLl5JUVB7k4KEZFRCuwPGBiBEAwSk4sLGHADoKpv1yBqL+viUYcLcKakFicD6Q9ArFIEtBxopAiQLZIOYWuA2EkQtg2IXV5SUAJkB4DYRSFBzkB2CpCtkY7ETkJiJxcUgdT3ANk2uTmlyQh3M/Ck5oUGA2kOIJZhKGYIYnBncAL5H6IkfxEDg8VXBgbmCQixpJkMDNtbGRgkbiHEVBYwMPC3MDBsO48QQ4RJQWJRIliIBYiZ0tIYGD4tZ2DgjWRgEL7AwMAVDQsIHG5TALvNnSEfCNMZchhSgSKeDHkMyQx6QJYRgwGDIYMZAKbWPz9HbOBQAAASdElEQVR4nO1ba5RU1ZX+zq1b71dXVz+qW9DFyKuBbgSVpzgqcbJEm0gcMy6NJmqMShhHo84kISIoRBQyScZJ4pNkNI6KYxKyyMQ4yxEjMzE8BRSl22B3Vd1qWrvpR/Wr6j6++XEfXUU30h3B/jGcte6qW1W36uz97W9/e59Tt4RhGMT/4yGNtQFjPU4DMNYGjPU4DcBYGzDWQ7ZPDMMYSzuKhhACQojPZC4HAJfL9ZlMOJpBnvoKLdsTNTU1QVXVESM/EuOEAOzL7O+VJMk5NwwDqqoil8tBVVWQhMfjQUlJCRKVlfB6veaXnMIhCyHQ39+Pyy67HMlkEl6vdxTIE4A47nMhBCTJpDNJqKqGXC6HfD4HgPD5/CgvL8eECRMwrWYqpk+bhpqpUxGLRCDJMiBJgwieikEWa4BhGCD5CXogLCeHeUfYuStBCPP7crkcBgYGQOoQwoV4PI6amhrU1dVi9qxzUDejFpMmTkRFddVQ23Qd1PVBRp5sJpAQkjQIgBktyREgm77F8wrYESYJIYSjHbpuIJ8fdNjj8aGqqgrTptXg/PPPw5zzz0ft9Ok4c/yZgFysNwSgdnVBPdSI/J63oe3aC+1//hfu+fMQ3/RTQNOAk6lRlvO5j9sGAbCdsulvM6+YgYNPhBBQVRVdXZ0w6RzA+PHjMHPmTMydOwdzzjsP02tqEK+sHDK/CkBXMtAOvgd1526oO/dCP/AujJQCI98DCQKQ/eh//z30LatHsH4JeLJAMAwIlwu5Dw4j33KkGICRDiEENE1DRUUFFi1aiDlz5mDu3LmYOnkywtFo8XwAtNwA9A+boe97B/mdu6Hu3gv9vQag9WMQOQAuCNkH4XHD5Y2bHBMCop/ovnclvBctgisQHI6SoxvW5/XubvS89EvE7lj+lwOg6zqCwSCefvppuN1u5z0dgNreDr3xA2h79kHdsRvavgPQD38IdHcB0CHghvD4gFAQQoRggIBhmATTtEEn/UFoh95BduO/ILZm5adnga5DuN3ouH8dgl/8AqRgoFADRv49hmEgEAjg3XcPYPny5XjyiSeQ1zQMvPAfGPi3Z6B/mIaRTANaHwgBIXkheTxAuASWQprRsB6FEKa8WuewzqFpkP0x9P3gUQSu+xK8kybCMAwI6S9oYDXT+e4nNgG9ffAvWgBq2tBWeKRAaJqGkpI4Nm36GX67dSs8bjekM8ch/9o2GMk0hNcLVyQOKRyD8PtBSQJ03YywDYA1oe28PRwgSMDtBrKdyN63trixGM3QDQi3jNz+d5B9+IeIPrgKNAxAkgYBGF70TjQIj8eLu+6+Bx3t7Qj89SIEHloLQ+sFXC5Q0wFNN4WnwLnBTxcX1WObMCEEoGmQgjHkXn4ZfW+8afYHuj4KEwkIwBgYwMdfvBahe++Au7LcDIIQwzFgZBSwrwuFQmhsbMB9q1ZDAhD6h+VwnTcfek8X4HIVOT5YYcxHyYo+Cq459pwkIAHCALKr1po9ymiYoOuAy4W2W1ZACocQue1rZmAsLRkCwPG6QLtPkGUZkiRB0zR0d3cjm82itLQMjz/xOF5/7TV4/H6Ef/SwOYFtrBCAJMACcAsBsSZwQC0MghDCpHAogvwftqH35V9D2N99gkEr77O/eBG9v9iE2KafWq3c4LyfqCZCmIskWZZhGAay2SyOHm1DNptFJBLG5z63GNOnT0NfXx/cbg/u/OY96M1mEVwwD/6//wb03nYTCHKwSz4WhIL5DAymxLFMNEDA5UXvug0wcvkTt8m6DsktI9fQiI6bbkJ4xV3wz5oJapoJoD0PSfb19WHWrHOdtYAdbVVV0dvbC8PQEIvFUVdXiwsvvBAXXLAQdXV1SCQSSKVSWLBgIbq7s+ju7sTK76zE2nVrkevsQtvci4DDhyG8fgjLWYIQHOrgscwbNhVdLujZdkR+/jQiX/kyDFWFkIep5CRAwtA0fLTgEmipJBLv74e7JGql1GDchwDg8/nQ09MLTcshGo1hwYL5WLp0KRYvXoxJkyYWzaOqKtxuN373u1dQX78U0WgUuVwOb77xOmadey66tmxF95VXQw6VANY6w25DYYvZCXLZJg5JU1gH+iBNnYrKXX+AsPuPY8HSNMDtxtE7v4WeHz2M2NPPInLTl4fvI0iyt7eXkydPpd8fZCgU4fz5C/i97z3EAwcOsHDous58Pk9VVanrOg3DYD6fJ0l+97v3UQgX/f4gL7nkc9TyeWokW6/6MtPwMhOuouIroxKuouItZRpuZjylzESqmQklmAlUUAlUMGMdStB69Jc7r2UCFcyEq5iCl93P/LtpUz5PwzAGD1UlSXb/8jdshsSWBYupqSoNTaNh2Vx4FAEQCITo9fp5/fU3OE7ncjnm83lqmjbkw4ZhUNd1qqpKVVV58cWL6fMF6HK5+cRjj5Ekew81Mh21nA9UMB2pYNvqdWy/85+oJP6KaXioSBETmGBlkbPKMaDYACgiwMzsC6gf45SuaabNTc1MVU1gs+xn7+t/IEkaqjqs/UUA+HwBxmJxAoJTpkzltm3bSJKaph0XAMMwqFkTf/DBn5lIVDMYDHP8+LOYSaVokGxf+wiTkKlEzmDKHaUyqZYDqTTVox3s+MGjzEybzRTcVFxRkxEFQBQBEKxkJlDOlBxh9+aXzfltu3SdhsXMls9fySTA1i/dQIOkoR7f9iEARKMxlpaW0ePx0ecLcP369Q4b1OOgWJgKmzdvpsfjoyx7ecvXbjEj0tNDZcb5TLlCTEermYSf6bOmMn+k1fzebJadP36cysRaJuFm2l/GTKSqOPLBSmaClUzCx86f/8KMamFQ8ib1j67bwGbITEUT7H/3IGkYowOgpKSUkUiJ8wiAV165jK2tprH5Y3NuGBBuv305XS43Q6EQt732Gkmyc8tWNsNPJZRgOlLFJPxUzl1A9ejRQYA7Onn0/rVMlVQzBR8z4SpmQgkzFUIJNsPNjn9+1HS+0A4r73u3/5FJX5wpeNh2+52fSP0TAhCJlDhsAAQnTpzE7du3O0zQhxEUXdepaRq7urp4zjmzKUkyFy26kLn+fqqGwdalX2IKPiqhBDPhaqbg5pFl11A3dBr9Aw4QA+83sPWq65iCn4o7RiU2jinIbP/2/YPO2/NrGqnp1Lq6qdTMZlqEmI6NZ+7wh2b0PyF1TwiAfZSWltHr9TMYDPKpp55ydGE4EFQrGtu3b2ckUkJZdvPHj5pR69l/gOlgGRV/uZnbkWo2w8W2b60yHcvlaFgsMkh2P/ci02dOYTPA1htvs/JZLVbzvEqD5Me3rGAKfqbgZ/s3vzWi6I8YgEgkylgsznA4SgC85557nGgNJ452KqxZs4YAOGHC2VSSSdPQu7/DJGRmItUOtZPwsvuFlwaN1jSH1jlFYduaddT6+wfFzmacXfJe/CWT8DEdrGQ6Pp65ZIq0qsKoAYhGY0MAiEZjjMXijMfLrZQAr7rqKmaz2WHF0U6FgYEBXnTRxQTAFd9YQZLsb/2IyrgpzHhiVIKVTIcqmfLFLdF6zwJBK8ptmxFFkdc00jCYa04yXXEWlUA5k/Cy7Y57Rxz9EwJga4AphhIBMByOsrw8QQBcuHAhjxw5MiwIdirs3fs2S0vjjERLuOOPfyRJtj/6mFn2wlVM+cuZjlQxJYJsOX+RGenC+m6VtyLn7ZJHg0cuW0bFEtdUsIIDhxrM6H+C8o8IgNLSMofygUCIN910Mzdu/D6j0Rhl2cPKyioKIXHGjDo2NTUNWyHsVHjkkQ0EwCVLLqeuqsz39jJTO5eKK2yyIFDOTKSaKchsu+c7J46gXfJ++BM2Q2amZBzT8PIju+6PgPrDADDF0YBQKEIAjMXivPXW27h3716Hijt27GBtbR0BsLKyii6Xm5MmTWFjY+MQJthdYj6f5+LFlxIQfOmFF0iSHS/9iin4mA4lmPaXUwmUMxNKMO0Ksfe/tx0fBKvp6t//DlPBcir+MjP68LPnlf+yqsTI6D+EAX5fkOFwlIlEFdetW8fm5mbHcVVVmcvlTOM7OnjNNdcSAMvKKujx+Hj22ROHBcFOhf379zMcjnBm3TnMdnVR0zQeueRypkWASihhVoZQgoocYaZmNtXu7qH9u67TUDXq+Twz8y4xPxupZloKU6mdR21gYIhQnuhw1oWSJEGSXUW/4fn9/sJFE2RZtvYCS/D8889h/fqH0dXVBZ/PB0XJ4LLLlqCpqQmyLEO3tq1cLhc0TUNtbS1WrlyJffvfxlNPPgWXy4XA6m8DsgxBDG6WBkPQ3tuP7gfWQ7hc5t6dbYOuQ8gudD30fWhvvQkRLjFXlIYOVzwG4fWOfuucJPv6ennGGeMJgCUlpYxGYxbFK7ly5XcdobOja68NSPK3v/1PJhLV9PkC9PkCnD59BltaWopKpF0Vcrkc582bz8rKKrZmWqiTbL36Bme1WLgASntL2L9rt5MKdsnr+9MupjwlzAQLFkueOFsqJjBvdaujYQBs47Zu3cr6+qUMBsMEBKPRmAPEmWeexQ0bNrCnp8dxTNM0R+QaGho5Z85cyrKHHo+P8+cvZFdXFw3DcECwU2Hnzl0EBP/xbrOX6N13gKlA2eCix29qgSICbL14CXXd6uY0jVp/PzOzL2BaCpnX2EvlUIIpEWTfa6+TJPURlkDDMAj7xB579uzhihUrWFlZ5ZQ9e00wY0YtX3zxxSI22CBks1lee+11dLnclCSZ9fVLqar5oo7RBuHuu++lx+PlnxsaaJBsvXm5UxaVY9f9m1925mu/70Em4TLZUrBPoISr2AyZHQ+aC7chewQjAcCmtj2SySTXrl3HSZOmEBAMBEIMBEKUJJlXXFHPffv2Odfa4kiS9923il6vnwC4fPnyIlHUdZ26rjObzfKMM8bx+uuuM2nd0Mh0NEHFV8a0Te1gJRU5QmXabBok+3bvZcpXMnSJ7IAVYGv935opMJoyeOwLmqY5kSLJzs5O/uQnP+WsWbMpSS76fAHKsoexWClXr17N/v5+BwT7c5s2/cxhzcaNG0kO9gj2Nb/61a8JSNz/9tsmC5bfxRTc5n6AtROUttrk9h/+K498/gtMw1dMfXvHKGh1k7FxzLe1j0oHhgBQWMNtepPkwMAAn3vuOS5YsJCy7KEsewgIzpkzl2++aa4SdV3nwIC5qnv11VdZVVVNQHDLlt8MC8KSJUu49IorTBZ88Gemo9XM+Ab1wBS5cia9caa9cWYCxTtGGYspmVCCSchsvfF2ar29pvOfFoDjAaHrOjdv3syFCy+g2+0hAIZCEa5Z84CTQn19fSTJffv28eyzJzEUCvPgwYNWOgzuLh0+fJilpXHu+tMOkmTrrXeYLAgPvxlS5Lgllhl/OVPw8eiadWYXOMo+4IQAFAJh57INxPPPP8+5c+cRZiXnpZf+DRsbPygCoampiTU10zl58mT29PQ4OmCz4P77V7PeZsH7h5gKVToCp1iODsl5q2lKu2NMeaLs2vSMSXt1+I3PkwJA4VEIRC6X45NPPsUZM2qdznDLli1F4vjRRx9z+vQZrK9f6qSCXX77+/t57rnncfcOkwVHvnob0/BYe4N2KpQ7QKT95eYiSgozHR/HnlfNHSdjFMr/qQEYDojOzk4+9NB6xuPmcnn16jVOBSDJo0c7WFMznQ888KADgv3e5s2beeNXvkqS7N3zNtO+UrYEimmvWOsFJVJt7hRNrGX/vgOfyvlPDYB9FGpEU1MTv/71W529RJsFhmGwvb2dc+bM5e9//3sHHFs3brzxZh56z9wPaFl6tan4BVqgBCqoRKqYgpst8y9mXsl8audPGgC2RhQCsX37dtbVzWRd3UxmMi0OCJlMhsuWXcVUKuUwgSTfeustblj/CEmya+vvmBYB5weTTLDSqvUyjyz7O6pZsyMdTcd3ygGwj8I+wjAMrl69muecM4t79ux1wHn//UNcter+Ii0gyccfe5ytra008nm21M2j4oo4O8MpuNn2jbuclPukre4xBcA+CvVh585dvP76G/jGG284IOzZs4evvPJKkU4kk0ln1+joQxuZgoctkWqm4ePRdY8MbnZoo1P6MQHAPmyK53I5Pvvsszx48KAT8YaGBnZ0dDhsIcmO9nYaJHNNSab8ZUy5wgVlTh11mTvRIQzj1P9nyDAM54bK9vZ2xGIxAHButLDfIwlJkkDrRqi2m5fDe+nFCF9zNaiq5i+7J/mO0c8EAKDgthhJKrr9pfC5/Zp9g5QBAZckQF0v+k3/ZI7PDAB7OA6OdFh3c52q8Zn/Y2TUf4Q4hc4Dp/8ycxqA0wCMtQFjPU4DMNYGjPU4DcBYGzDW4/8APdNRfvUtCW8AAAAASUVORK5CYII="

def get_logo_photo(variant="header"):
    """Base64 se seedha GIF load karo — koi file ya tempfile nahi chahiye."""
    if variant in _logo_cache:
        return _logo_cache[variant]
    b64_map = {
        "header":  _LOGO_HEADER_B64,
        "sidebar": _LOGO_SIDEBAR_B64,
        "splash":  _LOGO_SPLASH_B64,
        "png64":   _LOGO_PNG64_B64,
    }
    b64 = b64_map.get(variant)
    if not b64:
        return None
    try:
        # tk.PhotoImage seedha base64 GIF data accept karta hai — tempfile nahi chahiye
        photo = tk.PhotoImage(data=b64)
        _logo_cache[variant] = photo
        return photo
    except Exception:
        return None

def logo_label(parent, bg_color, variant="header"):
    """
    variant:
      'header'  → dark red bg, 200px wide  (License/Setup screens)
      'sidebar' → white bg,   170px wide  (App sidebar)
      'splash'  → dark red bg, 220px wide (Splash screen)
    """
    photo = get_logo_photo(variant)
    if photo:
        lbl = tk.Label(parent, image=photo, bg=bg_color, bd=0)
        lbl.image = photo  # GC se bachao
        return lbl
    # Fallback — agar kuch bhi kaam na kare
    return tk.Label(parent, text="BhugtanEase",
                    font=("Georgia", 15, "bold"), bg=bg_color, fg=WHITE)

def btn(par, text, cmd, bg=RED, fg=WHITE, font=FBD, px=14, py=7, w=None):
    kw = dict(text=text,command=cmd,bg=bg,fg=fg,font=font,
              relief="flat",cursor="hand2",padx=px,pady=py,bd=0,
              activebackground=bg,activeforeground=fg)
    if w: kw["width"]=w
    return tk.Button(par,**kw)

def entry(par, var, width=30, show=None):
    kw = dict(textvariable=var,font=FB,bg=IBGC,fg=DARK,relief="solid",bd=1,width=width)
    if show: kw["show"]=show
    return tk.Entry(par,**kw)

# ════════════════════════════════════════════════════════════
#  USER AUTH HELPERS
# ════════════════════════════════════════════════════════════
def _hash_password(password):
    """Password ko SHA-256 se hash karo."""
    return hashlib.sha256(password.encode()).hexdigest()

def _ensure_default_admin():
    """Agar koi user nahi hai toh default admin banao."""
    c = db()
    count = c.execute("SELECT COUNT(*) as n FROM users").fetchone()["n"]
    if count == 0:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
                  ("admin", _hash_password("empo123"), "admin"))
        c.commit()

def verify_login(username, password):
    """Username aur password verify karo."""
    c = db()
    row = c.execute("SELECT * FROM users WHERE username=?", (username.strip(),)).fetchone()
    if not row:
        return False
    return row["password_hash"] == _hash_password(password)

# ════════════════════════════════════════════════════════════
#  LOGIN SCREEN
# ════════════════════════════════════════════════════════════
class LoginScreen(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.logged_in = False
        self.title("BhugtanEase — Login")
        self.geometry("480x500"); self.resizable(False, False)
        self.configure(bg=BG)
        self.grab_set()
        self.update_idletasks()
        x = (self.winfo_screenwidth() // 2) - 240
        y = (self.winfo_screenheight() // 2) - 250
        self.geometry(f"480x500+{x}+{y}")
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._build()

    def _on_close(self):
        self.logged_in = False
        self.destroy()

    def _build(self):
        # Pure white background — photo jaisa
        self.configure(bg=WHITE)

        # Logo section (top center)
        logo_frame = tk.Frame(self, bg=WHITE, pady=28)
        logo_frame.pack(fill="x")
        lbl = logo_label(logo_frame, WHITE, variant="header")
        lbl.pack()

        # Subtitle
        tk.Label(logo_frame, text="Billing & Inventory Management",
                 font=("Segoe UI", 11), bg=WHITE, fg="#555555").pack(pady=(6, 0))

        # Separator line
        tk.Frame(self, bg="#d0d0d0", height=1).pack(fill="x")

        # Form area
        card = tk.Frame(self, bg=WHITE, padx=40, pady=24)
        card.pack(fill="both", expand=True)

        # Username
        tk.Label(card, text="Username", font=("Segoe UI", 10),
                 bg=WHITE, fg=DARK, anchor="w").pack(fill="x", pady=(8, 3))
        self.uv = tk.StringVar()
        tk.Entry(card, textvariable=self.uv, font=("Segoe UI", 11),
                 bg=WHITE, fg=DARK, relief="solid", bd=1).pack(fill="x", ipady=9, pady=(0, 14))

        # Password
        tk.Label(card, text="Password", font=("Segoe UI", 10),
                 bg=WHITE, fg=DARK, anchor="w").pack(fill="x", pady=(0, 3))
        self.pv = tk.StringVar()
        tk.Entry(card, textvariable=self.pv, font=("Segoe UI", 11),
                 bg=WHITE, fg=DARK, relief="solid", bd=1,
                 show="*").pack(fill="x", ipady=9, pady=(0, 6))

        # Error message label
        self.msg = tk.StringVar()
        tk.Label(card, textvariable=self.msg, font=("Segoe UI", 9),
                 bg=WHITE, fg="#c53030", wraplength=380).pack(pady=(4, 0))

        # LOGIN button — blue photo jaisa
        tk.Button(card, text="LOGIN", command=self._login,
                  bg="#1d4ed8", fg=WHITE, font=("Segoe UI", 12, "bold"),
                  relief="flat", cursor="hand2", pady=12, bd=0,
                  activebackground="#1e40af").pack(fill="x", pady=(18, 0))

        self.bind("<Return>", lambda e: self._login())
        self.after(100, lambda: self.uv.set(self.uv.get()))

        # Footer — window ke bottom mein, hamesha visible
        tk.Label(self, text="Ashrisha Ecommerce Solution Pvt Ltd",
                 font=("Segoe UI", 9), bg=WHITE, fg="#2563eb").pack(side="bottom", pady=16)

    def _login(self):
        u = self.uv.get().strip()
        p = self.pv.get()
        if not u:
            self.msg.set("❌ Username likhein!"); return
        if not p:
            self.msg.set("❌ Password likhein!"); return
        self.msg.set("⏳ Verify ho raha hai..."); self.update()
        if verify_login(u, p):
            self.msg.set("✅ Login ho gaya!")
            self.logged_in = True
            self.after(600, self.destroy)
        else:
            self.msg.set("❌ Username ya password galat hai!")
            self.pv.set("")

# ════════════════════════════════════════════════════════════
#  LICENSE SCREEN
# ════════════════════════════════════════════════════════════
class LicenseScreen(tk.Toplevel):
    def __init__(self, parent, lic=None):
        super().__init__(parent)
        self.activated = False
        self.title("BhugtanEase — Activate")
        self.geometry("520x480"); self.resizable(False,False)
        self.configure(bg=BG)
        self.grab_set()
        self.update_idletasks()
        x=(self.winfo_screenwidth()//2)-260; y=(self.winfo_screenheight()//2)-240
        self.geometry(f"520x480+{x}+{y}")
        self._build(lic)

    def _build(self, lic):
        # Header
        hdr = tk.Frame(self, bg=DKRED, pady=10)
        hdr.pack(fill="x")
        logo_frame = tk.Frame(hdr, bg=WHITE, padx=10, pady=6)
        logo_frame.pack()
        lbl = logo_label(logo_frame, WHITE, variant="header")
        lbl.pack()
        tk.Label(hdr, text="Restaurant Billing & Inventory", font=FS,
                 bg=DKRED, fg="#ffcccc").pack(pady=(4,0))
        tk.Frame(self, bg=RED, height=3).pack(fill="x")

        if lic and lic.get("status") == "EXPIRED":
            sb = tk.Frame(self, bg="#fff0f0", pady=10)
            sb.pack(fill="x")
            tk.Label(sb, text="⛔  License Expire Ho Gayi!",
                     font=FBD, bg="#fff0f0", fg="#c53030").pack()

        card = tk.Frame(self, bg=WHITE, padx=36, pady=20)
        card.pack(fill="both", expand=True, padx=24, pady=16)

        tk.Label(card, text="Activation", font=FH,
                 bg=WHITE, fg=DARK).pack(anchor="w", pady=(0,4))

        # Shop naam — MANDATORY for activation
        tk.Label(card, text="Aapki Shop ka Naam *  (Serial isi naam ke liye generate hua hoga)",
                 font=FBD, bg=WHITE, fg=DARK).pack(anchor="w", pady=(4,0))
        self.shop_v = tk.StringVar()
        # Pre-fill if already set
        existing_shop = gset("shop_name", "").strip()
        if existing_shop and existing_shop not in ("Mera Restaurant", ""):
            self.shop_v.set(existing_shop)
        tk.Entry(card, textvariable=self.shop_v, font=("Segoe UI",11),
                 bg=IBGC, fg=DARK, relief="solid", bd=1).pack(fill="x", ipady=8, pady=(3,14))

        tk.Label(card, text="Serial Number *  (Format: BE-XXXXXXXX-XXXXXX)",
                 font=FBD, bg=WHITE, fg=DARK).pack(anchor="w")
        self.sv = tk.StringVar()
        tk.Entry(card, textvariable=self.sv, font=("Courier New",11,"bold"),
                 bg=IBGC, fg=DARK, relief="solid", bd=1,
                 width=42).pack(fill="x", ipady=9, pady=(3,6))
        tk.Label(card, text="(Seller se mila hua poora serial yahan paste karein)",
                 font=FS, bg=WHITE, fg=MUTED).pack(anchor="w")

        self.msg = tk.StringVar()
        self.mlbl = tk.Label(card, textvariable=self.msg, font=FS,
                             bg=WHITE, fg="#c53030", wraplength=420)
        self.mlbl.pack(pady=(10,0))

        tk.Button(card, text="✅  ACTIVATE KARO", command=self._activate,
                  bg=RED, fg=WHITE, font=("Segoe UI",12,"bold"),
                  relief="flat", cursor="hand2", pady=12, bd=0,
                  activebackground=DKRED).pack(fill="x", pady=(16,0))
        self.bind("<Return>", lambda e: self._activate())

        tk.Label(self, text="Serial ke liye BhugtanEase se contact karein.",
                 font=FS, bg=BG, fg=MUTED).pack(pady=6)

    def _activate(self):
        shop = self.shop_v.get().strip()
        if not shop:
            self.msg.set("❌ Shop ka naam likhna zaroori hai!"); return
        s = self.sv.get().strip()
        if not s:
            self.msg.set("❌ Serial number likhein!"); return
        if not s.upper().startswith("BE-"):
            s = "BE-" + s
        self.msg.set("⏳ Verify ho raha hai..."); self.update()

        # Pehle shop naam temporarily save karo taaki check ho sake
        old_shop = gset("shop_name", "")
        old_setup = gset("setup_done", "0")
        sset("shop_name", shop)
        sset("setup_done", "1")   # shop check enable karo

        r = save_license(s)

        if not r["valid"]:
            # Restore original values agar activate fail hua
            sset("shop_name", old_shop)
            sset("setup_done", old_setup)
            self.msg.set(f"❌ {r['message']}")
        else:
            self.msg.set(f"✅ {r['message']}"); self.mlbl.config(fg=GREEN)
            self.update()
            self.activated = True
            self.after(1000, self.destroy)

# ════════════════════════════════════════════════════════════
#  SETUP WIZARD
# ════════════════════════════════════════════════════════════
class SetupWizard(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.done = False
        self.title("BhugtanEase — Shop Setup")
        self.resizable(True, True)
        self.configure(bg=BG)
        self.grab_set()
        sw = self.winfo_screenwidth(); sh = self.winfo_screenheight()
        w, h = 520, min(680, sh - 40)
        x = (sw - w) // 2; y = max(0, (sh - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")
        self.minsize(480, 520)
        self._build()

    def _build(self):
        # FIX: window fully initialize hone ke baad hi build karo — blank screen fix
        self.update_idletasks()
        hdr = tk.Frame(self, bg=DKRED, pady=10)
        hdr.pack(fill="x")
        logo_frame = tk.Frame(hdr, bg=WHITE, padx=10, pady=6)
        logo_frame.pack()
        lbl = logo_label(logo_frame, WHITE, variant="header")
        lbl.pack()
        tk.Label(hdr, text="Pehli baar — Shop Setup", font=FS,
                 bg=DKRED, fg="#ffcccc").pack(pady=(4,0))
        tk.Frame(self, bg=RED, height=3).pack(fill="x")

        # FIX: Buttons PEHLE bottom mein pack karo — hamesha visible rahenge
        btn_frame = tk.Frame(self, bg=WHITE, padx=36, pady=14)
        btn_frame.pack(side="bottom", fill="x")
        tk.Frame(btn_frame, bg=BORD, height=1).pack(fill="x", pady=(0,10))

        self.saved = False

        def save():
            if not self.flds["shop_name"].get().strip():
                messagebox.showerror("Error","Shop naam zaroori hai!",parent=self); return
            if not self.flds["phone"].get().strip():
                messagebox.showerror("Error","Phone zaroori hai!",parent=self); return
            for k,v in self.flds.items(): sset(k, v.get().strip())
            self.saved = True
            save_btn.config(text="✅  Saved!", bg=GREEN)
            start_btn.config(state="normal", bg=RED)

        def start():
            if not self.saved:
                messagebox.showwarning("Pehle Save Karo","Pehle Save karo, phir shuru karein!",parent=self); return
            sset("setup_done","1"); _load_settings_cache(); self.done = True; self.destroy()

        save_btn = tk.Button(btn_frame, text="💾  Save Karo", command=save,
                  bg="#2B6CB0", fg=WHITE, font=("Segoe UI",11,"bold"),
                  relief="flat", cursor="hand2", pady=11, bd=0)
        save_btn.pack(fill="x", pady=(0,6))

        start_btn = tk.Button(btn_frame, text="🚀  Save Karke Shuru Karein!", command=start,
                  bg=MUTED, fg=WHITE, font=("Segoe UI",11,"bold"),
                  relief="flat", cursor="hand2", pady=11, bd=0, state="disabled")
        start_btn.pack(fill="x")

        # Form area — upar scroll hogi screen chhoti ho toh
        card = tk.Frame(self, bg=WHITE, padx=36, pady=20)
        card.pack(fill="both", expand=True)

        tk.Label(card, text="🏪  Apni Shop ki Details Bharein",
                 font=FH, bg=WHITE, fg=DARK).pack(anchor="w", pady=(0,10))

        self.flds = {}
        for key, label, hint in [
            ("shop_name", "Shop ka Naam *",    "Jaise: Sharma Dhaba"),
            ("address",   "Address",            "Gali, Sheher, PIN"),
            ("phone",     "Phone Number *",     "9876543210"),
            ("gst",       "GST Number",         "Optional"),
            ("tax",       "Default Tax %",      "Jaise: 5"),
        ]:
            tk.Label(card, text=label, font=FBD, bg=WHITE, fg=DARK, anchor="w").pack(fill="x", pady=(4,1))
            var = tk.StringVar(value=gset(key) or ("5" if key=="tax" else ""))
            tk.Entry(card, textvariable=var, font=FB, bg=IBGC, fg=DARK,
                     relief="solid", bd=1).pack(fill="x", ipady=6)
            tk.Label(card, text=hint, font=FS, bg=WHITE, fg=MUTED, anchor="w").pack(fill="x")
            self.flds[key] = var

# ════════════════════════════════════════════════════════════
#  MAIN APP
# ════════════════════════════════════════════════════════════
class App(tk.Toplevel):
    def __init__(self, parent, lic=None):
        super().__init__(parent)
        self._lic = lic
        shop = gset("shop_name","Restaurant")
        self.title(f"{shop} — BhugtanEase")
        self.geometry("1200x740"); self.minsize(1000,640)
        self.configure(bg=BG)
        self.resizable(True,True)
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        _icon = get_logo_photo("png64")
        if _icon:
            self.iconphoto(True, _icon)
        self._build()
        self.show("billing")
        self.update_idletasks()
        self.after(100, self._post_init)
        self.after(1200, self._check_low_stock_alert)

    def _on_close(self):
        self.master.destroy()

    def _get_low_stock_items(self):
        """Jo materials min stock se neeche hain unki list do."""
        try:
            c = db()
            rows = c.execute(
                "SELECT name, unit, current_stock, min_stock FROM raw_materials "
                "WHERE current_stock <= min_stock ORDER BY (min_stock - current_stock) DESC"
            ).fetchall()
            pass  # conn reused
            return [dict(r) for r in rows]
        except Exception:
            return []

    def _check_low_stock_alert(self):
        """Startup alert — low stock items dikhao."""
        items = self._get_low_stock_items()
        if not items:
            return
        # Sidebar badge update karo
        self._update_low_stock_badge(len(items))
        # Popup dikhao
        self._show_low_stock_popup(items)

    def _update_low_stock_badge(self, count):
        """Sidebar mein ⚠️ badge dikhao ya chhupaao."""
        if not self._low_badge:
            return
        if count > 0:
            self._low_badge.config(text=f"⚠️  {count} item low stock mein")
            self._low_badge.pack(fill="x")
        else:
            self._low_badge.pack_forget()

    def _show_low_stock_popup(self, items):
        """Low stock warning popup."""
        dlg = tk.Toplevel(self)
        dlg.title("⚠️ Low Stock Alert")
        dlg.configure(bg=BG); dlg.grab_set(); dlg.resizable(False, True)
        dlg.update_idletasks()
        sw = dlg.winfo_screenwidth(); sh = dlg.winfo_screenheight()
        h = min(520, 180 + len(items) * 52)
        dlg.geometry(f"460x{h}+{(sw-460)//2}+{(sh-h)//2}")

        # Header
        hdr = tk.Frame(dlg, bg="#c53030", pady=12); hdr.pack(fill="x")
        tk.Label(hdr, text="⚠️  Low Stock Alert!", font=("Segoe UI",14,"bold"),
                 bg="#c53030", fg=WHITE).pack()
        tk.Label(hdr, text="Yeh materials khatam hone wale hain — jaldi order karo!",
                 font=FS, bg="#c53030", fg="#ffcccc").pack(pady=(2,0))
        tk.Frame(dlg, bg=RED, height=2).pack(fill="x")

        # Scrollable list
        wrap = tk.Frame(dlg, bg=BG); wrap.pack(fill="both", expand=True, padx=16, pady=12)
        canvas = tk.Canvas(wrap, bg=BG, highlightthickness=0)
        vsb = tk.Scrollbar(wrap, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canvas.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canvas, bg=BG)
        canvas.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        for mat in items:
            stock = mat["current_stock"] or 0
            min_s = mat["min_stock"] or 0
            unit  = mat["unit"] or ""
            short = min_s - stock  # kitna aur chahiye

            card = tk.Frame(inner, bg="#fff5f5", relief="solid", bd=1, padx=12, pady=8)
            card.pack(fill="x", pady=4)

            left = tk.Frame(card, bg="#fff5f5"); left.pack(side="left", fill="both", expand=True)
            tk.Label(left, text=f"🔴  {mat['name']}", font=FBD, bg="#fff5f5", fg="#c53030").pack(anchor="w")
            tk.Label(left, text=f"Current: {stock:.2f} {unit}   |   Minimum: {min_s:.2f} {unit}",
                     font=FS, bg="#fff5f5", fg=MED).pack(anchor="w")

            right = tk.Frame(card, bg="#fff5f5"); right.pack(side="right")
            tk.Label(right, text=f"Aur {short:.2f} {unit}", font=FBD,
                     bg="#fff5f5", fg="#c53030").pack(anchor="e")
            tk.Label(right, text="mangwao", font=FS, bg="#fff5f5", fg=MED).pack(anchor="e")

        # Buttons
        tk.Frame(dlg, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(dlg, bg=BG, pady=10); bf.pack(fill="x", padx=16)
        btn(bf, "📦 Inventory Dekho", lambda: [dlg.destroy(), self.show("inventory")],
            bg=DARK, px=12).pack(side="left", padx=4)
        btn(bf, "✖  Baad Mein", dlg.destroy, bg=MUTED, px=12).pack(side="right", padx=4)

    def _build(self):
        # Sidebar
        sb = tk.Frame(self, bg=WHITE, width=200)
        sb.pack(side="left", fill="y"); sb.pack_propagate(False)
        tk.Frame(sb, bg=BORD, width=1).pack(side="right", fill="y")

        # Logo area
        lp = tk.Frame(sb, bg=WHITE, pady=10)
        lp.pack(fill="x")
        sidebar_logo = logo_label(lp, WHITE, variant="sidebar")
        sidebar_logo.pack()
        tk.Frame(sb, bg=RED, height=3).pack(fill="x")
        shop = gset("shop_name","My Shop")
        shop_lbl = tk.Label(sb, text=shop, font=("Segoe UI",9,"bold"),
                 bg=WHITE, fg=MED, wraplength=190)
        shop_lbl.pack(pady=(8,14))
        shop_lbl._is_shop_label = True  # FIX: tag se settings update kar sakti hai

        # Nav buttons
        self.nav = {}
        self._low_badge = None
        self._low_badge_frame = None
        for label, key in [
            ("🧾  Billing",         "billing"),
            ("📦  Inventory",       "inventory"),
            ("🛒  Purchase",        "purchase"),
            ("📋  Order History",   "order_history"),
            ("🥬  Material History","material_history"),
            ("📊  Reports",         "reports"),
            ("⚙️  Settings",        "settings"),
            ("🗄️  Backup",          "backup"),
        ]:
            b = tk.Button(sb, text=label, bg=WHITE, fg=MED,
                          font=("Segoe UI",11), relief="flat", anchor="w",
                          padx=18, pady=12, cursor="hand2", bd=0,
                          activebackground=RED, activeforeground=WHITE,
                          command=lambda p=key: self.show(p))
            b.pack(fill="x")
            self.nav[key] = b
            # Inventory ke baad low stock badge frame
            if key == "inventory":
                self._low_badge_frame = tk.Frame(sb, bg=WHITE)
                self._low_badge_frame.pack(fill="x", padx=14, pady=(0,4))
                self._low_badge = tk.Label(
                    self._low_badge_frame, text="", font=("Segoe UI",8,"bold"),
                    bg="#fff5f5", fg="#c53030", relief="solid", bd=1,
                    padx=8, pady=3, cursor="hand2")
                self._low_badge.bind("<Button-1>", lambda e: self._check_low_stock_alert())
                # Badge abhi khali hai — startup check baad mein update karega

        # License info at bottom — SPEED: already loaded lic use karo
        tk.Frame(sb, bg=BORD, height=1).pack(side="bottom", fill="x")
        lic = self._lic or load_license()
        if lic["valid"]:
            days = lic["days"]
            col  = "#c53030" if days<=7 else WARN if days<=30 else GREEN
            tk.Label(sb, text=f"License: {days} din bache",
                     font=FS, bg=WHITE, fg=col).pack(side="bottom", pady=4)
        else:
            tk.Label(sb, text="⛔ License inactive",
                     font=FS, bg=WHITE, fg="#c53030").pack(side="bottom", pady=4)

        self.content = tk.Frame(self, bg=BG)
        self.content.pack(side="left", fill="both", expand=True)

    def _post_init(self):
        try:
            self.state("zoomed")
        except Exception:
            try: self.attributes("-zoomed", True)
            except: pass
        self.update_idletasks()

    def show(self, key):
        for k, b in self.nav.items():
            b.config(bg=RED if k==key else WHITE,
                     fg=WHITE if k==key else MED)
        try:
            self.content.unbind_all("<MouseWheel>")
        except Exception:
            pass
        for w in self.content.winfo_children():
            try: w.destroy()
            except Exception: pass
        _pages = {"billing": BillingPage, "inventory": InventoryPage,
                  "purchase": PurchasePage,
                  "order_history": OrderHistoryPage,
                  "material_history": MaterialHistoryPage,
                  "reports": ReportsPage, "settings": SettingsPage,
                  "backup": BackupPage}
        try:
            _pages[key](self.content, self)
            self.content.update_idletasks()
        except Exception as exc:
            import traceback
            for w in self.content.winfo_children():
                try: w.destroy()
                except: pass
            ef = tk.Frame(self.content, bg="#fff5f5")
            ef.pack(fill="both", expand=True)
            tk.Frame(ef, bg=RED, height=4).pack(fill="x")
            tk.Label(ef, text="Page Error - Yeh Message Copy Karo",
                     font=("Segoe UI",14,"bold"), bg="#fff5f5", fg="#c53030").pack(pady=(30,8))
            tk.Label(ef, text=str(exc), font=("Segoe UI",10),
                     bg="#fff5f5", fg=DARK, wraplength=700).pack(pady=(0,10))
            box = tk.Text(ef, height=10, font=("Courier New",8), wrap="word")
            box.insert("1.0", traceback.format_exc())
            box.config(state="disabled")
            box.pack(fill="x", padx=40, pady=(0,16))
            tk.Button(ef, text="Dobara Try Karo", font=("Segoe UI",11,"bold"),
                      bg=RED, fg=WHITE, relief="flat", padx=24, pady=10,
                      cursor="hand2", command=lambda k=key: self.show(k)).pack()

# ════════════════════════════════════════════════════════════
#  BILLING PAGE
# ════════════════════════════════════════════════════════════
class BillingPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self.cart = []
        self.qty_lbl = {}
        self.sel_cat = tk.StringVar(value="All")
        self._search_after = None  # SPEED: debounce ke liye
        self._build()

    def _build(self):
        # Page header
        hdr = tk.Frame(self, bg=WHITE)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10)
        th.pack(fill="x")
        tk.Label(th, text="🧾  New Bill", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        tk.Label(th, text=gset("shop_name","Restaurant"), font=("Segoe UI",11,"bold"),
                 bg=WHITE, fg=MED).pack(side="right")

        # Main 2-column layout
        main = tk.Frame(self, bg=BG)
        main.pack(fill="both", expand=True, padx=8, pady=6)
        main.columnconfigure(0, weight=3); main.columnconfigure(1, weight=2)
        main.rowconfigure(0, weight=1)

        left = tk.Frame(main, bg=BG)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,6))
        self._menu_panel(left)

        right = tk.Frame(main, bg=BG)
        right.grid(row=0, column=1, sticky="nsew")
        self._bill_panel(right)

    def _menu_panel(self, par):
        # Customer info bar
        info = tk.Frame(par, bg=WHITE, relief="solid", bd=1)
        info.pack(fill="x", pady=(0,6))
        tk.Frame(info, bg=RED, height=3).pack(fill="x")
        row = tk.Frame(info, bg=WHITE, padx=8, pady=8)
        row.pack(fill="x")
        for i, (lbl, attr) in enumerate([("Table:","tv"),("Customer:","cv"),("Phone:","pv")]):
            tk.Label(row, text=lbl, font=FS, bg=WHITE, fg=MED).grid(row=0,column=i*2,padx=(0,2))
            var = tk.StringVar(); setattr(self, attr, var)
            tk.Entry(row, textvariable=var, font=FS, width=13,
                     bg=IBGC, relief="solid", bd=1).grid(row=0,column=i*2+1,padx=(0,10))

        # Category tabs
        cf = tk.Frame(par, bg=BG)
        cf.pack(fill="x", pady=(0,4))
        c = db()
        self.cats = ["All"] + [r["name"] for r in c.execute("SELECT name FROM categories ORDER BY name").fetchall()]
        pass  # conn reused
        self.cat_btns = {}
        for cat in self.cats:
            active = cat == "All"
            b = tk.Button(cf, text=cat, font=FS, relief="flat",
                          bg=RED if active else "#edf2f7",
                          fg=WHITE if active else DARK,
                          padx=8, pady=4, cursor="hand2",
                          command=lambda c=cat: self._filter(c))
            b.pack(side="left", padx=2)
            self.cat_btns[cat] = b

        # Search
        sf = tk.Frame(par, bg=BG)
        sf.pack(fill="x", pady=(0,4))
        tk.Label(sf, text="🔍", bg=BG).pack(side="left")
        self.srch = tk.StringVar()
        self.srch.trace("w", lambda *a: self._search_debounce())
        tk.Entry(sf, textvariable=self.srch, font=FB, bg=IBGC,
                 relief="solid", bd=1, width=30).pack(side="left", padx=4)

        # Menu grid
        outer = tk.Frame(par, bg=WHITE, relief="solid", bd=1)
        outer.pack(fill="both", expand=True)
        canvas = tk.Canvas(outer, bg=WHITE, highlightthickness=0)
        vsb = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self.mgrid = tk.Frame(canvas, bg=WHITE)
        fid = canvas.create_window((0,0), window=self.mgrid, anchor="nw")
        self.mgrid.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(fid, width=e.width))
        self._load_items()

    def _search_debounce(self):
        # SPEED: Har keypress pe re-render nahi — 280ms wait karo phir load karo
        if self._search_after:
            self.after_cancel(self._search_after)
        self._search_after = self.after(280, self._load_items)

    def _filter(self, cat):
        self.sel_cat.set(cat)
        for c, b in self.cat_btns.items():
            b.config(bg=RED if c==cat else "#edf2f7",
                     fg=WHITE if c==cat else DARK)
        self._load_items()

    def _load_items(self):
        for w in self.mgrid.winfo_children(): w.destroy()
        c = db()
        q = ("SELECT m.id,m.name,m.price,c.name as cat "
             "FROM menu_items m LEFT JOIN categories c ON m.category_id=c.id "
             "WHERE m.is_available=1")
        params = []
        cat = self.sel_cat.get()
        if cat != "All": q += " AND c.name=?"; params.append(cat)
        s = self.srch.get().strip()
        if s: q += " AND m.name LIKE ?"; params.append(f"%{s}%")
        items = c.execute(q+" ORDER BY c.name,m.name", params).fetchall()
        pass  # conn reused
        cur = gset("currency","₹"); COLS = 3
        self.qty_lbl = {}
        emap = {"Starters":"🥗","Main Course":"🍛","Breads":"🫓",
                "Drinks":"🥤","Desserts":"🍮","Rice & Biryani":"🍚"}
        for idx, item in enumerate(items):
            r, col = divmod(idx, COLS)
            card = tk.Frame(self.mgrid, bg=CARD, relief="solid", bd=1)
            card.grid(row=r, column=col, padx=4, pady=4, sticky="ew")
            self.mgrid.columnconfigure(col, weight=1)
            tk.Label(card, text=emap.get(item["cat"],"🍽️"),
                     font=("Segoe UI",18), bg=CARD).pack(pady=(8,0))
            tk.Label(card, text=item["name"], font=("Segoe UI",9,"bold"),
                     bg=CARD, fg=DARK, wraplength=100).pack()
            tk.Label(card, text=f"{cur}{item['price']:.0f}",
                     font=("Segoe UI",10,"bold"), bg=CARD, fg=RED).pack()
            bf = tk.Frame(card, bg=CARD); bf.pack(pady=(4,8))
            tk.Button(bf, text="−", font=FBD, bg="#edf2f7", fg=DARK,
                      relief="flat", padx=8, pady=2, cursor="hand2",
                      command=lambda i=dict(item): self._chg(i,-1)).pack(side="left", padx=2)
            ql = tk.Label(bf, text="0", font=FBD, bg=CARD, fg=RED, width=3)
            ql.pack(side="left")
            self.qty_lbl[item["id"]] = ql
            tk.Button(bf, text="+", font=FBD, bg=RED, fg=WHITE,
                      relief="flat", padx=8, pady=2, cursor="hand2",
                      command=lambda i=dict(item): self._chg(i,1)).pack(side="left", padx=2)
        for ci in self.cart:
            if ci["id"] in self.qty_lbl:
                self.qty_lbl[ci["id"]].config(text=str(ci["qty"]))

    def _chg(self, item, delta):
        for ci in self.cart:
            if ci["id"] == item["id"]:
                ci["qty"] += delta
                if ci["qty"] <= 0: self.cart.remove(ci)
                break
        else:
            if delta > 0: self.cart.append({"id":item["id"],"name":item["name"],"price":item["price"],"qty":1})
        found = next((o for o in self.cart if o["id"]==item["id"]), None)
        if item["id"] in self.qty_lbl:
            self.qty_lbl[item["id"]].config(text=str(found["qty"]) if found else "0")
        self._refresh_bill()

    def _bill_panel(self, par):
        tk.Label(par, text="🧾  Current Bill", font=FH,
                 bg=BG, fg=DARK).pack(anchor="w", pady=(0,4))
        outer = tk.Frame(par, bg=WHITE, relief="solid", bd=1)
        outer.pack(fill="both", expand=True)

        cols = ("Item","Qty","Rate","Total")
        self.tree = ttk.Treeview(outer, columns=cols, show="headings", height=12)
        for col, w in zip(cols,[150,50,70,80]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center" if col!="Item" else "w")
        vsb = ttk.Scrollbar(outer, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self._remove)

        # Summary
        sf = tk.Frame(par, bg="#fff5f5", relief="solid", bd=1)
        sf.pack(fill="x", pady=4)
        self.slbl = {}
        for key, label in [("sub","Subtotal:"),("tax","GST:"),
                            ("disc","Discount:"),("total","TOTAL:")]:
            row = tk.Frame(sf, bg="#fff5f5"); row.pack(fill="x", padx=10, pady=2)
            font = ("Segoe UI",12,"bold") if key=="total" else FB
            fg   = RED if key=="total" else DARK
            tk.Label(row, text=label, font=font, bg="#fff5f5", fg=fg).pack(side="left")
            lbl = tk.Label(row, text="₹0.00", font=font, bg="#fff5f5", fg=fg)
            lbl.pack(side="right")
            self.slbl[key] = lbl

        # Options
        opt = tk.Frame(par, bg=BG); opt.pack(fill="x", pady=3)
        for txt, attr, dflt in [("Tax%:","taxv",gset("tax","5")),("Disc%:","discv","0")]:
            tk.Label(opt, text=txt, font=FS, bg=BG).pack(side="left")
            var = tk.StringVar(value=dflt); setattr(self, attr, var)
            tk.Entry(opt, textvariable=var, width=5, font=FS,
                     bg=IBGC, relief="solid", bd=1).pack(side="left", padx=(2,10))
            var.trace("w", lambda *a: self._refresh_bill())
        tk.Label(opt, text="Pay:", font=FS, bg=BG).pack(side="left")
        self.payv = tk.StringVar(value="Cash")
        ttk.Combobox(opt, textvariable=self.payv, values=["Cash","UPI","Card","Credit"],
                     width=7, state="readonly").pack(side="left", padx=4)

        # Action buttons
        bf = tk.Frame(par, bg=BG); bf.pack(fill="x", pady=4)
        btn(bf,"🗑️ Clear",self._clear,bg="#a0aec0",px=10).pack(side="left",padx=4)
        btn(bf,"💾 Save",lambda: self._save(False),bg=DARK,px=10).pack(side="left",padx=4)
        btn(bf,"🖨️ Print",self._print_dialog,bg=RED,px=10).pack(side="right",padx=4)

    def _refresh_bill(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        cur = gset("currency","₹"); sub = 0
        for item in self.cart:
            tot = item["price"]*item["qty"]; sub += tot
            self.tree.insert("","end",values=(item["name"],item["qty"],
                f"{cur}{item['price']:.2f}",f"{cur}{tot:.2f}"))
        try: tp = float(self.taxv.get() or 0)
        except: tp = 0
        try: dp = float(self.discv.get() or 0)
        except: dp = 0
        disc = sub*dp/100; tax = (sub-disc)*tp/100; total = sub-disc+tax
        self.slbl["sub"].config(text=f"{cur}{sub:.2f}")
        self.slbl["tax"].config(text=f"{cur}{tax:.2f}")
        self.slbl["disc"].config(text=f"-{cur}{disc:.2f}")
        self.slbl["total"].config(text=f"{cur}{total:.2f}")

    def _remove(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        idx = self.tree.index(sel[0])
        if 0 <= idx < len(self.cart):
            item = self.cart[idx]
            if item["id"] in self.qty_lbl: self.qty_lbl[item["id"]].config(text="0")
            self.cart.pop(idx); self._refresh_bill()

    def _clear(self):
        self.cart.clear()
        for l in self.qty_lbl.values(): l.config(text="0")
        self._refresh_bill()

    def _totals(self):
        sub = sum(i["price"]*i["qty"] for i in self.cart)
        try: tp = float(self.taxv.get() or 0)
        except: tp = 0
        try: dp = float(self.discv.get() or 0)
        except: dp = 0
        disc = sub*dp/100; tax = (sub-disc)*tp/100
        return sub, tp, tax, dp, disc, sub-disc+tax

    def _print_dialog(self):
        """Print type choose karo — PDF ya Thermal"""
        if not self.cart:
            messagebox.showwarning("Empty","Pehle items add karo!"); return

        dlg = tk.Toplevel(self)
        dlg.title("🖨️ Print Options")
        dlg.configure(bg=BG); dlg.grab_set(); dlg.resizable(False,True)
        dlg.update_idletasks()
        sw=dlg.winfo_screenwidth(); sh=dlg.winfo_screenheight()
        dlg.geometry(f"380x480+{(sw-380)//2}+{(sh-480)//2}")

        hdr = tk.Frame(dlg, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="🖨️  Print Type Chuniye", font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(dlg, bg=RED, height=2).pack(fill="x")

        body = tk.Frame(dlg, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)

        ptype = tk.StringVar(value=gset("default_print_type","thermal"))

        # Thermal option
        tf = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=12, pady=10)
        tf.pack(fill="x", pady=(0,8))
        tk.Radiobutton(tf, text="🖨️  Thermal Printer", variable=ptype, value="thermal",
                       font=FBD, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Label(tf, text="58mm / 80mm thermal roll printer ke liye\n(ESC/POS compatible)",
                 font=FS, bg=WHITE, fg=MUTED, justify="left").pack(anchor="w", padx=20)

        # Width option (only for thermal)
        wf = tk.Frame(body, bg=BG); wf.pack(fill="x", pady=(0,8))
        tk.Label(wf, text="  Paper Width:", font=FS, bg=BG).pack(side="left")
        wvar = tk.StringVar(value=gset("thermal_width","80"))
        ttk.Combobox(wf, textvariable=wvar, values=["58","80"],
                     width=5, state="readonly", font=FS).pack(side="left", padx=6)
        tk.Label(wf, text="mm", font=FS, bg=BG).pack(side="left")

        # PDF option
        pf = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=12, pady=10)
        pf.pack(fill="x", pady=(0,8))
        tk.Radiobutton(pf, text="📄  PDF Bill (A5)", variable=ptype, value="pdf",
                       font=FBD, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Label(pf, text="A5 size PDF generate karke open karega\n(Normal printer ke liye)",
                 font=FS, bg=WHITE, fg=MUTED, justify="left").pack(anchor="w", padx=20)

        # Printer name (optional)
        pnf = tk.Frame(body, bg=BG); pnf.pack(fill="x", pady=(0,4))
        tk.Label(pnf, text="  Printer Name (optional):", font=FS, bg=BG).pack(anchor="w")
        pname_var = tk.StringVar(value=gset("thermal_printer",""))
        # Try to get default printer
        try:
            import win32print
            pname_var.set(win32print.GetDefaultPrinter())
        except: pass
        tk.Entry(pnf, textvariable=pname_var, font=FS, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=4, pady=(2,0))
        tk.Label(pnf, text="  (Khali chhodein = default printer)",
                 font=("Segoe UI",8), bg=BG, fg=MUTED).pack(anchor="w")

        def do_print():
            dlg.destroy()
            pt = ptype.get()
            pn = pname_var.get().strip() or None
            if pt == "thermal":
                self._save(print_bill=True, thermal=True,
                           thermal_width=int(wvar.get()), printer_name=pn)
            else:
                self._save(print_bill=True, thermal=False, printer_name=pn)

        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=(8,0))
        btn(body, "🖨️  PRINT KARO", do_print, bg=RED, py=12).pack(fill="x", pady=(8,4))
        btn(body, "✖  Cancel", dlg.destroy, bg=MUTED, py=6).pack(fill="x")

    def _save(self, print_bill=False, thermal=False, thermal_width=80, printer_name=None):
        if not self.cart:
            messagebox.showwarning("Empty","Pehle items add karo!"); return
        sub, tp, tax, dp, disc, total = self._totals()
        c = db()
        try:
            cur2 = c.cursor()
            cur2.execute("""INSERT INTO orders
                (table_number,customer_name,subtotal,tax_percent,tax_amount,
                 discount_percent,discount_amount,total_amount,payment_method,created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (self.tv.get(), self.cv.get(), sub, tp, tax, dp, disc, total, self.payv.get(),
                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            oid = cur2.lastrowid
            for item in self.cart:
                tot = item["price"]*item["qty"]
                cur2.execute("INSERT INTO order_items (order_id,item_name,quantity,price_at_order,item_total) VALUES (?,?,?,?,?)",
                             (oid, item["name"], item["qty"], item["price"], tot))
            c.commit()
            cur  = gset("currency","₹")
            messagebox.showinfo("✅ Saved!",f"Bill #{oid:04d} save ho gaya!\nTotal: {cur}{total:.2f}")
            if print_bill:
                order = dict(c.execute("SELECT * FROM orders WHERE id=?",(oid,)).fetchone())
                bill_items = [dict(r) for r in c.execute("SELECT * FROM order_items WHERE order_id=?",(oid,)).fetchall()]
                settings_d = {r["key"]:r["value"] for r in c.execute("SELECT * FROM settings").fetchall()}
                bd = os.path.join(BASE_DIR,"bills"); os.makedirs(bd, exist_ok=True)

                if thermal:
                    # ── THERMAL PREVIEW + PRINT ─────────────────────
                    bill_text = generate_thermal_bill(oid, order, bill_items, settings_d, thermal_width)
                    _show_thermal_preview(self, oid, bill_text, printer_name, bd)
                else:
                    # ── PDF PREVIEW (auto open) ─────────────────────
                    import subprocess
                    pdf_path = os.path.join(bd, f"BILL_{oid:04d}.pdf")
                    path = generate_pdf(oid, order, bill_items, settings_d, pdf_path)
                    if path:
                        try:
                            if sys.platform=="win32": os.startfile(path)
                            elif sys.platform=="darwin": subprocess.run(["open",path])
                            else: subprocess.run(["xdg-open",path])
                        except: pass
                        messagebox.showinfo("📄 PDF Preview!", f"PDF khul gaya — Ctrl+P se print karo\n\nFile: {path}")
            self._clear()
            self.tv.set(""); self.cv.set(""); self.pv.set("")
        except Exception as e:
            c.rollback()
            messagebox.showerror("Error",str(e))

# ════════════════════════════════════════════════════════════
#  INVENTORY PAGE
# ════════════════════════════════════════════════════════════
class InventoryPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=WHITE)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10)
        th.pack(fill="x")
        tk.Label(th, text="📦  Inventory & Menu", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        btn(th, "📥 Excel Download", self._export_excel, bg=GREEN, px=10).pack(side="right", padx=4)
        self.add_btn = btn(th, "+ Add", self._header_add, bg=GREEN)
        self.add_btn.pack(side="right")
        self.extra_btn = btn(th, "📥 Stock IN", self._restock, bg=DARK)
        self.extra_btn.pack(side="right", padx=4)
        self.out_btn = btn(th, "📤 Stock OUT", self._stock_out, bg="#c05621")
        self.out_btn.pack(side="right", padx=4)

        # Tabs
        tabbar = tk.Frame(self, bg=WHITE)
        tabbar.pack(fill="x")
        tk.Frame(self, bg=BORD, height=1).pack(fill="x")
        self.tab = tk.StringVar(value="menu")
        self.tbtns = {}
        for label, key in [("🍽️ Menu Items","menu"),("📦 Raw Material Stock","stock"),("📋 Transaction Log","log")]:
            b = tk.Button(tabbar, text=label, font=FBD,
                          bg=WHITE, fg=MED, relief="flat",
                          padx=16, pady=9, cursor="hand2", bd=0,
                          command=lambda k=key: self._switch(k))
            b.pack(side="left")
            self.tbtns[key] = b
        tk.Frame(self, bg=BORD, height=1).pack(fill="x")

        self.body = tk.Frame(self, bg=BG)
        self.body.pack(fill="both", expand=True)
        self._switch("menu")

    def _header_add(self):
        if self.tab.get() == "menu":
            self._add_menu_item()
        elif self.tab.get() == "stock":
            self._add()
        # log tab pe kuch nahi karna

    def _switch(self, key):
        self.tab.set(key)
        for k, b in self.tbtns.items():
            b.config(bg=RED if k==key else WHITE, fg=WHITE if k==key else MED)
        # Header buttons update
        if key == "menu":
            self.add_btn.config(text="+ Add Food Item", state="normal", bg=GREEN)
            self.extra_btn.pack_forget()
            self.out_btn.pack_forget()
        elif key == "stock":
            self.add_btn.config(text="+ Add Material", state="normal", bg=GREEN)
            self.out_btn.pack(side="right", padx=4)
            self.extra_btn.pack(side="right", padx=4)
        else:
            self.add_btn.config(text="", state="disabled", bg=BG)
            self.extra_btn.pack_forget()
            self.out_btn.pack_forget()
        for w in self.body.winfo_children(): w.destroy()
        if key == "menu":   self._menu_tab()
        elif key == "stock": self._stock_tab()
        else: self._log_tab()

    # ── MENU ITEMS TAB ──────────────────────────────────────
    def _menu_tab(self):
        p = self.body

        # Category filter bar
        cf = tk.Frame(p, bg=BG, pady=6); cf.pack(fill="x", padx=12)
        tk.Label(cf, text="Category:", font=FBD, bg=BG, fg=DARK).pack(side="left", padx=(0,6))
        self._mcat_var = tk.StringVar(value="All")
        c = db()
        cats = ["All"] + [r["name"] for r in c.execute("SELECT name FROM categories ORDER BY name").fetchall()]
        pass  # conn reused
        self._cat_btns = {}
        for cat in cats:
            active = cat == "All"
            b = tk.Button(cf, text=cat, font=FS, relief="flat",
                          bg=RED if active else "#edf2f7",
                          fg=WHITE if active else DARK,
                          padx=8, pady=4, cursor="hand2",
                          command=lambda c=cat: self._mcat_filter(c))
            b.pack(side="left", padx=2)
            self._cat_btns[cat] = b

        # Search
        sf = tk.Frame(p, bg=BG); sf.pack(fill="x", padx=12, pady=(0,6))
        tk.Label(sf, text="🔍", bg=BG).pack(side="left")
        self._msrch = tk.StringVar()
        self._msrch.trace("w", lambda *a: self._load_menu())
        tk.Entry(sf, textvariable=self._msrch, font=FB, bg=IBGC,
                 relief="solid", bd=1, width=30).pack(side="left", padx=4)

        # Treeview
        frame = tk.Frame(p, bg=BG); frame.pack(fill="both", expand=True, padx=12, pady=(0,4))
        cols = ("ID","Naam","Category","Price","Available")
        self.mtree = ttk.Treeview(frame, columns=cols, show="headings")
        for col, w in zip(cols, [40, 260, 140, 90, 90]):
            self.mtree.heading(col, text=col)
            self.mtree.column(col, width=w, anchor="center")
        self.mtree.column("Naam", anchor="w")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.mtree.yview)
        self.mtree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.mtree.pack(fill="both", expand=True)
        self.mtree.tag_configure("na", foreground=MUTED)
        self.mtree.bind("<Double-1>", lambda e: self._edit_menu_item())

        bf = tk.Frame(p, bg=BG, pady=6); bf.pack(fill="x", padx=12)
        btn(bf, "✏️ Edit Item",   self._edit_menu_item,   bg=DARK).pack(side="left", padx=4)
        btn(bf, "🗑️ Delete Item", self._del_menu_item,    bg="#c53030").pack(side="left", padx=4)
        btn(bf, "🔄 Available Toggle", self._toggle_avail, bg="#c05621").pack(side="left", padx=4)
        btn(bf, "📂 Category Manage",  self._manage_cats,  bg="#553C9A").pack(side="right", padx=4)

        self._load_menu()

    def _mcat_filter(self, cat):
        self._mcat_var.set(cat)
        for c, b in self._cat_btns.items():
            b.config(bg=RED if c==cat else "#edf2f7", fg=WHITE if c==cat else DARK)
        self._load_menu()

    def _load_menu(self):
        if not hasattr(self, "mtree"): return
        for r in self.mtree.get_children(): self.mtree.delete(r)
        c = db(); cur = gset("currency","₹")
        q = ("SELECT m.id,m.name,m.price,m.is_available,cat.name as cat "
             "FROM menu_items m LEFT JOIN categories cat ON m.category_id=cat.id WHERE 1=1")
        params = []
        cat = self._mcat_var.get() if hasattr(self,"_mcat_var") else "All"
        if cat != "All": q += " AND cat.name=?"; params.append(cat)
        s = self._msrch.get().strip() if hasattr(self,"_msrch") else ""
        if s: q += " AND m.name LIKE ?"; params.append(f"%{s}%")
        for m in c.execute(q+" ORDER BY cat.name,m.name", params).fetchall():
            avail = "✅ Available" if m["is_available"] else "❌ Hidden"
            tag = "ok" if m["is_available"] else "na"
            self.mtree.insert("","end", iid=m["id"], tags=(tag,),
                values=(m["id"], m["name"], m["cat"] or "—",
                        f"{cur}{m['price']:.2f}", avail))
        pass  # conn reused

    def _menu_item_form(self, title, defaults=None):
        """Add/Edit form for a food item. Returns dict or None."""
        win = tk.Toplevel(self)
        win.title(title); win.geometry("420x480")
        win.configure(bg=BG); win.grab_set(); win.resizable(False,True)
        # Centre
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"420x480+{(sx-420)//2}+{(sy-480)//2}")

        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=title, font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=3).pack(fill="x")
        body = tk.Frame(win, bg=BG, padx=28, pady=16); body.pack(fill="both", expand=True)

        c = db()
        cats = [(r["id"], r["name"]) for r in c.execute("SELECT id,name FROM categories ORDER BY name").fetchall()]
        pass  # conn reused
        cat_names = [n for _, n in cats]
        cat_id_map = {n: i for i, n in cats}

        flds = {}
        def row(lbl, key, widget_fn):
            tk.Label(body, text=lbl, font=FBD, bg=BG, anchor="w").pack(fill="x", pady=(8,1))
            var = tk.StringVar(value=str(defaults.get(key,"")) if defaults else "")
            w = widget_fn(var)
            w.pack(fill="x", ipady=6)
            flds[key] = var

        row("Food Item ka Naam *", "name",
            lambda v: tk.Entry(body, textvariable=v, font=FB, bg=IBGC, relief="solid", bd=1))
        row("Category *", "cat",
            lambda v: ttk.Combobox(body, textvariable=v, values=cat_names, state="readonly", font=FB))
        row("Price (₹) *", "price",
            lambda v: tk.Entry(body, textvariable=v, font=FB, bg=IBGC, relief="solid", bd=1))

        tk.Label(body, text="Available hai?", font=FBD, bg=BG, anchor="w").pack(fill="x", pady=(8,1))
        avail_var = tk.BooleanVar(value=bool(defaults.get("is_available",1)) if defaults else True)
        af = tk.Frame(body, bg=BG); af.pack(fill="x")
        tk.Radiobutton(af, text="✅ Haan", variable=avail_var, value=True,
                       font=FB, bg=BG).pack(side="left", padx=(0,16))
        tk.Radiobutton(af, text="❌ Nahi (Hide karo)", variable=avail_var, value=False,
                       font=FB, bg=BG).pack(side="left")

        result = {}
        def on_save():
            n = flds["name"].get().strip()
            cat_n = flds["cat"].get().strip()
            pr = flds["price"].get().strip()
            if not n:
                messagebox.showerror("Error","Naam zaroori hai!",parent=win); return
            if not cat_n or cat_n not in cat_id_map:
                messagebox.showerror("Error","Category select karo!",parent=win); return
            try: price = float(pr)
            except: messagebox.showerror("Error","Price sahi likhein!",parent=win); return
            result.update({"name":n,"cat_id":cat_id_map[cat_n],
                           "price":price,"is_available":int(avail_var.get())})
            win.destroy()

        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=(16,0))
        btn(body, "💾  SAVE KARO", on_save, bg=GREEN, py=14).pack(fill="x", pady=(10,4))
        btn(body, "✖  Cancel", win.destroy, bg=MUTED, py=8).pack(fill="x")
        win.wait_window()
        return result if result else None

    def _add_menu_item(self):
        data = self._menu_item_form("🍽️ Naya Food Item Add Karo")
        if data:
            c = db()
            try:
                c.execute("INSERT INTO menu_items (name,category_id,price,is_available) VALUES (?,?,?,?)",
                          (data["name"], data["cat_id"], data["price"], data["is_available"]))
                c.commit()
                messagebox.showinfo("✅", f"'{data['name']}' menu mein add ho gaya!")
            except Exception as e:
                messagebox.showerror("Error", str(e))
            self._load_menu()

    def _edit_menu_item(self, event=None):
        sel = self.mtree.selection()
        if not sel:
            messagebox.showinfo("","Pehle item select karo.",parent=self); return
        mid = int(sel[0]); c = db()
        m = dict(c.execute("SELECT m.*,cat.name as cat_name FROM menu_items m "
                           "LEFT JOIN categories cat ON m.category_id=cat.id WHERE m.id=?",
                           (mid,)).fetchone())
        pass  # conn reused
        data = self._menu_item_form("✏️ Food Item Edit Karo",
                                    {"name":m["name"],"cat":m["cat_name"],"price":m["price"],
                                     "is_available":m["is_available"]})
        if data:
            c = db()
            c.execute("UPDATE menu_items SET name=?,category_id=?,price=?,is_available=? WHERE id=?",
                      (data["name"], data["cat_id"], data["price"], data["is_available"], mid))
            c.commit()
            self._load_menu()

    def _del_menu_item(self):
        sel = self.mtree.selection()
        if not sel:
            messagebox.showinfo("","Pehle item select karo.",parent=self); return
        name = self.mtree.item(sel[0])["values"][1]
        if messagebox.askyesno("Confirm", f"'{name}' delete karna chahte hain?\n(Bills ka data safe rahega)"):
            c = db()
            c.execute("DELETE FROM menu_items WHERE id=?", (int(sel[0]),))
            c.commit(); self._load_menu()

    def _toggle_avail(self):
        sel = self.mtree.selection()
        if not sel:
            messagebox.showinfo("","Pehle item select karo.",parent=self); return
        mid = int(sel[0]); c = db()
        cur_val = c.execute("SELECT is_available FROM menu_items WHERE id=?",(mid,)).fetchone()[0]
        c.execute("UPDATE menu_items SET is_available=? WHERE id=?", (0 if cur_val else 1, mid))
        c.commit(); self._load_menu()

    def _manage_cats(self):
        """Category add/delete window"""
        win = tk.Toplevel(self)
        win.title("📂 Categories Manage Karo")
        win.geometry("340x380"); win.configure(bg=BG); win.grab_set(); win.resizable(False,False)
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"340x380+{(sx-340)//2}+{(sy-380)//2}")

        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="📂 Categories", font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=3).pack(fill="x")
        body = tk.Frame(win, bg=BG, padx=20, pady=14); body.pack(fill="both", expand=True)

        lb = tk.Listbox(body, font=FB, bg=IBGC, relief="solid", bd=1, selectbackground=RED)
        lb.pack(fill="both", expand=True)

        def refresh_lb():
            lb.delete(0,"end")
            c = db()
            for r in c.execute("SELECT name FROM categories ORDER BY name").fetchall():
                lb.insert("end", r["name"])
            pass  # conn reused
        refresh_lb()

        add_row = tk.Frame(body, bg=BG); add_row.pack(fill="x", pady=(10,0))
        nv = tk.StringVar()
        tk.Entry(add_row, textvariable=nv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(side="left", fill="x", expand=True, ipady=6)

        def add_cat():
            n = nv.get().strip()
            if not n: return
            c = db()
            try:
                c.execute("INSERT INTO categories (name) VALUES (?)", (n,))
                c.commit()
                nv.set(""); refresh_lb()
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=win)
            pass  # conn reused

        def del_cat():
            sel = lb.curselection()
            if not sel: return
            name = lb.get(sel[0])
            if messagebox.askyesno("Confirm", f"'{name}' category delete karein?\n(Iske items uncategorized ho jayenge)", parent=win):
                c = db()
                c.execute("DELETE FROM categories WHERE name=?", (name,))
                c.commit(); refresh_lb()

        btn(add_row, "+ Add", add_cat, bg=GREEN, px=10).pack(side="left", padx=(6,0))
        btn(body, "🗑️ Selected Delete Karo", del_cat, bg="#c53030").pack(fill="x", pady=(6,0))

        win.wait_window()
        # Refresh tab after closing
        self._switch("menu")

    def _stock_tab(self):
        p = self.body
        frame = tk.Frame(p, bg=BG); frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("ID","Material","Unit","Current Stock","Min Stock","Cost/Unit","Status")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings")
        for col, w in zip(cols,[40,220,60,110,90,90,110]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure("low", foreground="#c53030", font=("Segoe UI",9,"bold"))
        self.tree.tag_configure("ok",  foreground="#276749")
        self.tree.bind("<Double-1>", self._edit)
        bf = tk.Frame(p, bg=BG); bf.pack(pady=6)
        btn(bf,"✏️ Edit",self._edit,bg=DARK).pack(side="left",padx=4)
        btn(bf,"🗑️ Delete",self._delete,bg="#c53030").pack(side="left",padx=4)
        self._load_stock()

    def _load_stock(self):
        if not hasattr(self,"tree"): return
        for r in self.tree.get_children(): self.tree.delete(r)
        c = db()
        cur = gset("currency","₹")
        for m in c.execute("SELECT * FROM raw_materials ORDER BY name").fetchall():
            low = m["current_stock"] <= m["min_stock"]
            self.tree.insert("","end",iid=m["id"],tags=("low" if low else "ok",),
                values=(m["id"],m["name"],m["unit"],f"{m['current_stock']:.2f}",
                        f"{m['min_stock']:.2f}",f"{cur}{m['cost_per_unit']:.2f}",
                        "⚠️ Low Stock" if low else "✅ OK"))
        pass  # conn reused

    def _log_tab(self):
        p = self.body
        frame = tk.Frame(p, bg=BG); frame.pack(fill="both", expand=True, padx=12, pady=8)
        cols = ("Date","Material","Type","Qty","Note")
        self.log_tree = ttk.Treeview(frame, columns=cols, show="headings")
        for col, w in zip(cols,[140,200,80,80,200]):
            self.log_tree.heading(col, text=col)
            self.log_tree.column(col, width=w, anchor="center")
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.log_tree.yview)
        self.log_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); self.log_tree.pack(fill="both", expand=True)
        self.log_tree.tag_configure("Used", foreground="#c53030")
        self.log_tree.tag_configure("Added", foreground="#276749")
        c = db()
        for log in c.execute("""SELECT l.*,r.name as mname,r.unit FROM inventory_log l
                JOIN raw_materials r ON l.material_id=r.id
                ORDER BY l.created_at DESC LIMIT 300""").fetchall():
            self.log_tree.insert("","end",tags=(log["type"],),
                values=(log["created_at"][:16] if log["created_at"] else "",
                        f"{log['mname']} ({log['unit']})",
                        log["type"],f"{log['qty']:.2f}",log["note"] or ""))
        pass  # conn reused

    def _mat_form(self, title, defaults=None):
        win = tk.Toplevel(self)
        win.title(title); win.geometry("380x340")
        win.configure(bg=BG); win.grab_set()
        win.resizable(False,False)
        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=title, font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=3).pack(fill="x")
        body = tk.Frame(win, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)
        flds = {}
        for i, (lbl, key) in enumerate(zip(
            ["Naam:","Unit:","Current Stock:","Min Stock:","Cost/Unit (₹):"],
            ["name","unit","stock","min","cost"])):
            tk.Label(body, text=lbl, font=FBD, bg=BG, anchor="w").grid(row=i,column=0,padx=8,pady=5,sticky="w")
            var = tk.StringVar(value=str(defaults.get(key,"")) if defaults else "")
            if key=="unit":
                w = ttk.Combobox(body,textvariable=var,
                                 values=["kg","g","litre","ml","piece","dozen","pack"],width=20)
            else:
                w = tk.Entry(body,textvariable=var,font=FB,bg=IBGC,relief="solid",bd=1,width=22)
            w.grid(row=i,column=1,padx=8,pady=5); flds[key]=var
        result = {}
        def on_save():
            n=flds["name"].get().strip(); u=flds["unit"].get().strip()
            if not n or not u:
                messagebox.showerror("Error","Naam aur unit zaroori hai!",parent=win); return
            try: s=float(flds["stock"].get() or 0); ms=float(flds["min"].get() or 0); cost=float(flds["cost"].get() or 0)
            except: messagebox.showerror("Error","Numbers sahi likhein!",parent=win); return
            result.update({"name":n,"unit":u,"stock":s,"min":ms,"cost":cost}); win.destroy()
        btn(body,"💾  Save",on_save,bg=RED).grid(row=6,column=0,columnspan=2,pady=16)
        win.wait_window(); return result if result else None

    def _add(self):
        data = self._mat_form("+ Add Material")
        if data:
            c = db()
            try:
                c.execute("INSERT INTO raw_materials (name,unit,current_stock,min_stock,cost_per_unit) VALUES (?,?,?,?,?)",
                          (data["name"],data["unit"],data["stock"],data["min"],data["cost"]))
                c.execute("INSERT INTO inventory_log (material_id,type,qty,note,created_at) VALUES (?,?,?,?,?)",
                          (c.execute("SELECT last_insert_rowid()").fetchone()[0],"Added",data["stock"],"Initial stock",
                           datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                c.commit(); messagebox.showinfo("✅",f"'{data['name']}' add ho gaya!")
            except Exception as e: messagebox.showerror("Error",str(e))
            self._switch("stock")

    def _edit(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        mid = int(sel[0]); c = db()
        m = c.execute("SELECT * FROM raw_materials WHERE id=?",(mid,)).fetchone()
        data = self._mat_form("✏️ Edit Material",
                              {"name":m["name"],"unit":m["unit"],"stock":m["current_stock"],
                               "min":m["min_stock"],"cost":m["cost_per_unit"]})
        if data:
            c = db()
            c.execute("UPDATE raw_materials SET name=?,unit=?,current_stock=?,min_stock=?,cost_per_unit=? WHERE id=?",
                      (data["name"],data["unit"],data["stock"],data["min"],data["cost"],mid))
            c.commit(); self._load_stock()

    def _delete(self):
        sel = self.tree.selection()
        if not sel: return
        if messagebox.askyesno("Confirm","Delete karna chahte hain?"):
            c = db()
            c.execute("DELETE FROM raw_materials WHERE id=?",(int(sel[0]),))
            c.commit(); self._load_stock()

    def _restock(self):
        win = tk.Toplevel(self)
        win.title("📥 Stock Add Karo")
        win.configure(bg=BG); win.grab_set(); win.resizable(False, False)
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"400x520+{(sx-400)//2}+{(sy-520)//2}")

        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="📥 Stock Add Karo", font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=3).pack(fill="x")
        body = tk.Frame(win, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)

        c = db()
        mats = c.execute("SELECT id,name,unit,cost_per_unit FROM raw_materials ORDER BY name").fetchall()
        pass  # conn reused
        mat_names      = [f"{m['name']} ({m['unit']})" for m in mats]
        mat_ids        = {f"{m['name']} ({m['unit']})": m["id"] for m in mats}
        mat_costs      = {f"{m['name']} ({m['unit']})": m["cost_per_unit"] or 0 for m in mats}

        tk.Label(body, text="Material:", font=FBD, bg=BG, anchor="w").pack(fill="x")
        mv = tk.StringVar(value=mat_names[0] if mat_names else "")
        cb = ttk.Combobox(body, textvariable=mv, values=mat_names, state="readonly", width=34)
        cb.pack(fill="x", pady=(3,12))

        tk.Label(body, text="Quantity:", font=FBD, bg=BG, anchor="w").pack(fill="x")
        qv = tk.StringVar()
        tk.Entry(body, textvariable=qv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=8, pady=(3,12))

        # Unit Price
        pf = tk.Frame(body, bg=BG); pf.pack(fill="x", pady=(0,4))
        tk.Label(pf, text="Unit Price (₹):", font=FBD, bg=BG, anchor="w").pack(fill="x")
        pv = tk.StringVar()
        tk.Entry(pf, textvariable=pv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=8, pady=(3,0))
        tk.Label(pf, text="  (1 unit ka price — jaise 1kg Atta = ₹35)",
                 font=("Segoe UI",8), bg=BG, fg=MUTED).pack(anchor="w")

        # Total cost display
        total_lbl = tk.Label(body, text="Total Cost: ₹0.00",
                             font=("Segoe UI",11,"bold"), bg="#f0fff4", fg=GREEN,
                             relief="solid", bd=1, pady=6)
        total_lbl.pack(fill="x", pady=(8,0))

        def update_total(*args):
            try:
                q = float(qv.get()) if qv.get() else 0
                p = float(pv.get()) if pv.get() else 0
                total_lbl.config(text=f"Total Cost: ₹{q*p:.2f}")
            except: total_lbl.config(text="Total Cost: ₹0.00")

        qv.trace("w", update_total); pv.trace("w", update_total)

        # Auto-fill price when material changes
        def on_mat_change(*args):
            cost = mat_costs.get(mv.get(), 0)
            if cost: pv.set(str(cost))
        mv.trace("w", on_mat_change)
        on_mat_change()  # initial

        tk.Label(body, text="Note (optional):", font=FBD, bg=BG, anchor="w").pack(fill="x", pady=(10,0))
        nv = tk.StringVar()
        tk.Entry(body, textvariable=nv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=6, pady=(3,0))

        def do_add():
            mid = mat_ids.get(mv.get())
            if not mid: messagebox.showerror("Error","Material select karo!",parent=win); return
            try:
                qty = float(qv.get())
                if qty <= 0: raise ValueError
            except (ValueError, TypeError):
                messagebox.showerror("Error","Quantity sahi likhein!",parent=win); return
            try:
                unit_price = float(pv.get()) if pv.get().strip() else 0
            except: unit_price = 0
            total_cost = qty * unit_price

            c = db()
            c.execute("UPDATE raw_materials SET current_stock=current_stock+?, cost_per_unit=? WHERE id=?",
                      (qty, unit_price if unit_price > 0 else mat_costs.get(mv.get(),0), mid))
            c.execute("INSERT INTO inventory_log (material_id,type,qty,unit_price,total_cost,note,created_at) VALUES (?,?,?,?,?,?,?)",
                      (mid, "Added", qty, unit_price, total_cost, nv.get(),
                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            c.commit()
            messagebox.showinfo("✅", f"{qty} add ho gaya!\nTotal Cost: ₹{total_cost:.2f}", parent=win)
            win.destroy(); self._switch("stock")

        # Fixed bottom buttons
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(win, bg=BG, padx=24, pady=12); bf.pack(fill="x")
        btn(bf, "✅  Stock Add Karo", do_add, bg=GREEN, py=12).pack(fill="x", pady=(0,6))
        btn(bf, "✖  Cancel", win.destroy, bg=MUTED, py=6).pack(fill="x")

    def _stock_out(self):
        """Raw material use/nikalna — stock minus karo."""
        win = tk.Toplevel(self)
        win.title("📤 Stock OUT")
        win.configure(bg=BG); win.grab_set(); win.resizable(False, False)
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"420x560+{(sx-420)//2}+{(sy-560)//2}")

        hdr = tk.Frame(win, bg="#c05621", pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="📤  Stock OUT — Nikala / Use Kiya",
                 font=FH, bg="#c05621", fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=3).pack(fill="x")
        body = tk.Frame(win, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)

        c = db()
        mats = c.execute("SELECT id,name,unit,current_stock FROM raw_materials ORDER BY name").fetchall()
        pass  # conn reused
        mat_names  = [f"{m['name']} ({m['unit']}) — Stock: {m['current_stock']:.2f}" for m in mats]
        mat_ids    = {f"{m['name']} ({m['unit']}) — Stock: {m['current_stock']:.2f}": m["id"]    for m in mats}
        mat_stocks = {f"{m['name']} ({m['unit']}) — Stock: {m['current_stock']:.2f}": m["current_stock"] for m in mats}

        # Material select
        tk.Label(body, text="Material:", font=FBD, bg=BG, anchor="w").pack(fill="x")
        mv = tk.StringVar(value=mat_names[0] if mat_names else "")
        ttk.Combobox(body, textvariable=mv, values=mat_names,
                     state="readonly", width=40).pack(fill="x", pady=(3,10))

        # Reason / Type
        tk.Label(body, text="Kyu Nikala (Type):", font=FBD, bg=BG, anchor="w").pack(fill="x")
        rv = tk.StringVar(value="Rasoi mein use kiya")
        reasons = ["Rasoi mein use kiya", "Kharab ho gaya / Waste",
                   "Kisi ko diya", "Stock correction", "Aur kuch"]
        ttk.Combobox(body, textvariable=rv, values=reasons,
                     state="normal", width=40).pack(fill="x", pady=(3,10))

        # Quantity
        tk.Label(body, text="Quantity (kitna nikala):", font=FBD, bg=BG, anchor="w").pack(fill="x")
        qv = tk.StringVar()
        tk.Entry(body, textvariable=qv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=8, pady=(3,10))

        # Note
        tk.Label(body, text="Note (optional):", font=FBD, bg=BG, anchor="w").pack(fill="x")
        nv = tk.StringVar()
        tk.Entry(body, textvariable=nv, font=FB, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=8, pady=(3,0))

        def do_out():
            sel_mat = mv.get()
            mid = mat_ids.get(sel_mat)
            if not mid:
                messagebox.showerror("Error","Material select karo!",parent=win); return
            try:
                qty = float(qv.get())
                if qty <= 0: raise ValueError
            except (ValueError, TypeError):
                messagebox.showerror("Error","Quantity sahi likhein!",parent=win); return

            cur_stock = mat_stocks.get(sel_mat, 0)
            if qty > cur_stock:
                messagebox.showerror("❌ Error",
                    f"Stock mein sirf {cur_stock:.2f} hai!\nItna nahi nikaal sakte.",
                    parent=win); return

            reason = rv.get().strip() or "Used"
            note   = nv.get().strip()
            c = db()
            c.execute("UPDATE raw_materials SET current_stock=current_stock-? WHERE id=?", (qty, mid))
            c.execute("INSERT INTO inventory_log (material_id,type,qty,note,created_at) VALUES (?,?,?,?,?)",
                      (mid, f"Used — {reason}", qty, note,
                       datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            c.commit()

            messagebox.showinfo("✅ Done!", f"{qty} stock OUT ho gaya!\nReason: {reason}", parent=win)
            win.destroy()
            self._switch("stock")
            if hasattr(self.app, "_check_low_stock_alert"):
                low_items = self.app._get_low_stock_items()
                self.app._update_low_stock_badge(len(low_items))

        # Buttons — hamesha neeche fixed rahenge
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(win, bg=BG, padx=24, pady=12); bf.pack(fill="x")
        btn(bf, "📤  Stock OUT Karo", do_out, bg="#c05621", py=12).pack(fill="x", pady=(0,6))
        btn(bf, "✖  Cancel", win.destroy, bg=MUTED, py=8).pack(fill="x")

# ════════════════════════════════════════════════════════════
#  PURCHASE REGISTER PAGE
# ════════════════════════════════════════════════════════════

    def _export_excel(self):
        """Inventory data Excel mein export karo."""
        c = db()
        # Menu items
        menu = c.execute(
            "SELECT m.name, c.name as cat, m.price, "
            "CASE WHEN m.is_available=1 THEN 'Available' ELSE 'Unavailable' END as status "
            "FROM menu_items m LEFT JOIN categories c ON m.category_id=c.id ORDER BY c.name, m.name"
        ).fetchall()
        # Raw materials
        mats = c.execute(
            "SELECT name, unit, current_stock, min_stock, cost_per_unit FROM raw_materials ORDER BY name"
        ).fetchall()
        pass  # conn reused
        cur = gset("currency","₹")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            try:
                import subprocess, sys
                subprocess.check_call([sys.executable,"-m","pip","install","openpyxl","-q"])
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            except Exception:
                messagebox.showerror("Error","openpyxl install karo: pip install openpyxl"); return

        from tkinter import filedialog
        import datetime as _dt
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"Inventory_{_dt.date.today()}.xlsx", title="Save Karo"
        )
        if not path: return

        wb  = openpyxl.Workbook()
        hdr_fill = PatternFill("solid", fgColor="8B0000")
        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(style="thin", color="CCCCCC")
        brd  = Border(left=thin,right=thin,top=thin,bottom=thin)

        # Sheet 1 — Menu Items
        ws1 = wb.active; ws1.title = "Menu Items"
        for ci, h in enumerate(["Item Name","Category","Price","Status"], 1):
            cell = ws1.cell(row=1,column=ci,value=h)
            cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal="center")
        for ri, r in enumerate(menu, 2):
            fill = PatternFill("solid", fgColor="FFF5F5" if ri%2==0 else "FFFFFF")
            for ci, v in enumerate([r["name"],r["cat"],f"{cur}{r['price']:.2f}",r["status"]], 1):
                cell = ws1.cell(row=ri,column=ci,value=v)
                cell.border=brd; cell.fill=fill
        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 40)
        ws1.freeze_panes="A2"

        # Sheet 2 — Raw Materials
        ws2 = wb.create_sheet("Raw Materials")
        for ci, h in enumerate(["Material","Unit","Current Stock","Min Stock","Cost/Unit","Status"], 1):
            cell = ws2.cell(row=1,column=ci,value=h)
            cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal="center")
        for ri, m in enumerate(mats, 2):
            stock = m["current_stock"] or 0; min_s = m["min_stock"] or 0
            status = "⚠️ Low" if stock<=min_s else "✅ OK"
            fill = PatternFill("solid", fgColor="FFF0F0" if stock<=min_s else ("FFF5F5" if ri%2==0 else "FFFFFF"))
            for ci, v in enumerate([m["name"],m["unit"],stock,min_s,f"{cur}{m['cost_per_unit'] or 0:.2f}",status], 1):
                cell = ws2.cell(row=ri,column=ci,value=v)
                cell.border=brd; cell.fill=fill
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = min(max(len(str(c.value or "")) for c in col)+4, 30)
        ws2.freeze_panes="A2"

        try:
            wb.save(path)
            messagebox.showinfo("✅ Saved!", f"File save ho gayi:\n{path}", parent=self)
            try: import os; os.startfile(path)
            except: pass
        except Exception as e:
            messagebox.showerror("Error", f"Save nahi hua:\n{e}", parent=self)


class PurchasePage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=WHITE); hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10); th.pack(fill="x")
        tk.Label(th, text="🛒  Purchase Register", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        btn(th, "➕ Nai Purchase", self._new_purchase, bg=GREEN, px=12).pack(side="right", padx=4)
        btn(th, "👤 Suppliers",    self._manage_suppliers, bg=DARK, px=10).pack(side="right", padx=4)
        btn(th, "🔄 Refresh",      self._load, bg=MUTED, px=8).pack(side="right", padx=4)

        # Date filter
        fb = tk.Frame(self, bg=BG, padx=12, pady=8); fb.pack(fill="x")
        tk.Label(fb, text="Period:", font=FBD, bg=BG).pack(side="left", padx=(0,6))
        self.date_var = tk.StringVar(value="aaj")
        for lbl, val in [("Aaj","aaj"),("Kal","kal"),("7 Din","7din"),("30 Din","30din"),("Sab","sab")]:
            b = tk.Button(fb, text=lbl, font=FS, relief="flat",
                          bg=RED if val=="aaj" else "#edf2f7",
                          fg=WHITE if val=="aaj" else DARK,
                          padx=10, pady=4, cursor="hand2",
                          command=lambda v=val: self._filter(v))
            b.pack(side="left", padx=2)
            setattr(self, f"_dbtn_{val}", b)

        # Summary cards
        sf = tk.Frame(self, bg="#fffaf0", relief="solid", bd=1)
        sf.pack(fill="x", padx=12, pady=(0,6))
        self.lbl_count  = tk.Label(sf, text="Purchases: 0",       font=FBD, bg="#fffaf0", fg=DARK,    padx=16, pady=6)
        self.lbl_count.pack(side="left")
        self.lbl_total  = tk.Label(sf, text="Total: ₹0",          font=FBD, bg="#fffaf0", fg="#c05621", padx=16)
        self.lbl_total.pack(side="left")
        self.lbl_unpaid = tk.Label(sf, text="Unpaid: ₹0",         font=FBD, bg="#fffaf0", fg=RED,      padx=16)
        self.lbl_unpaid.pack(side="left")
        self.lbl_paid   = tk.Label(sf, text="Paid: ₹0",           font=FBD, bg="#fffaf0", fg=GREEN,    padx=16)
        self.lbl_paid.pack(side="left")

        # Purchase list
        frm = tk.Frame(self, bg=BG); frm.pack(fill="both", expand=True, padx=12, pady=(0,4))
        cols = ("ID","Date","Supplier","Invoice","Items","Total","Payment","Status")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings")
        widths =    [50,  130,     160,       120,     60,     100,      90,       90]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center")
        self.tree.column("Supplier", anchor="w")
        self.tree.column("Invoice",  anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical",   command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure("even",   background="#f9f9f9")
        self.tree.tag_configure("unpaid", foreground=RED)
        self.tree.bind("<Double-1>", self._view_detail)

        # Bottom buttons
        bf = tk.Frame(self, bg=BG, pady=6); bf.pack(fill="x", padx=12)
        btn(bf, "🔍 Detail Dekho",   self._view_detail,   bg=DARK,    px=10).pack(side="left", padx=4)
        btn(bf, "✏️ Edit",            self._edit_purchase, bg="#2b6cb0", px=10).pack(side="left", padx=4)
        btn(bf, "🗑️ Delete",          self._delete,        bg="#c53030", px=10).pack(side="left", padx=4)

        self._load()

    def _filter(self, val):
        self.date_var.set(val)
        for v in ["aaj","kal","7din","30din","sab"]:
            b = getattr(self, f"_dbtn_{v}", None)
            if b: b.config(bg=RED if v==val else "#edf2f7", fg=WHITE if v==val else DARK)
        self._load()

    def _date_range(self):
        today = datetime.date.today()
        v = self.date_var.get()
        if v == "aaj":   return str(today), str(today)
        elif v == "kal":
            d = today - datetime.timedelta(days=1); return str(d), str(d)
        elif v == "7din":  return str(today - datetime.timedelta(days=6)), str(today)
        elif v == "30din": return str(today - datetime.timedelta(days=29)), str(today)
        else: return None, None

    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        c = db(); cur = gset("currency","₹")
        q = ("SELECT p.*, (SELECT COUNT(*) FROM purchase_items WHERE purchase_id=p.id) as item_count "
             "FROM purchases p WHERE 1=1")
        params = []
        d1, d2 = self._date_range()
        if d1: q += " AND DATE(p.created_at)>=?"; params.append(d1)
        if d2: q += " AND DATE(p.created_at)<=?"; params.append(d2)
        q += " ORDER BY p.id DESC"
        rows = c.execute(q, params).fetchall()
        total = paid = unpaid = 0
        for i, r in enumerate(rows):
            amt = r["total_amount"] or 0
            total += amt
            if r["payment_status"] == "Paid": paid += amt
            else: unpaid += amt
            tags = []
            if r["payment_status"] != "Paid": tags.append("unpaid")
            elif i % 2 == 0: tags.append("even")
            self.tree.insert("","end", iid=r["id"], tags=tuple(tags), values=(
                f"#{r['id']:04d}",
                r["created_at"][:16] if r["created_at"] else "",
                r["supplier_name"] or "—",
                r["invoice_no"] or "—",
                r["item_count"],
                f"{cur}{amt:.2f}",
                r["payment_method"] or "Cash",
                r["payment_status"] or "Paid",
            ))
        n = len(rows)
        self.lbl_count.config(text=f"Purchases: {n}")
        self.lbl_total.config(text=f"Total: {cur}{total:.2f}")
        self.lbl_unpaid.config(text=f"Unpaid: {cur}{unpaid:.2f}")
        self.lbl_paid.config(text=f"Paid: {cur}{paid:.2f}")

    def _new_purchase(self):
        self._purchase_form(None)

    def _edit_purchase(self):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("","Pehle ek purchase select karo."); return
        self._purchase_form(int(sel[0]))

    def _purchase_form(self, purchase_id):
        """Nai ya edit purchase form."""
        c = db()
        suppliers = [dict(r) for r in c.execute("SELECT * FROM suppliers ORDER BY name").fetchall()]
        mats      = [dict(r) for r in c.execute("SELECT * FROM raw_materials ORDER BY name").fetchall()]
        existing  = None
        ex_items  = []
        if purchase_id:
            existing = dict(c.execute("SELECT * FROM purchases WHERE id=?", (purchase_id,)).fetchone())
            ex_items = [dict(r) for r in c.execute("SELECT * FROM purchase_items WHERE purchase_id=?", (purchase_id,)).fetchall()]
        pass  # conn reused

        cur = gset("currency","₹")
        win = tk.Toplevel(self)
        win.title("✏️ Purchase Edit" if purchase_id else "➕ Nai Purchase")
        win.configure(bg=BG); win.grab_set(); win.resizable(True, True)
        win.update_idletasks()
        sw=win.winfo_screenwidth(); sh=win.winfo_screenheight()
        win.geometry(f"700x680+{(sw-700)//2}+{(sh-680)//2}")

        # Header
        hdr = tk.Frame(win, bg=GREEN if not purchase_id else "#2b6cb0", pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text="✏️ Purchase Edit" if purchase_id else "➕ Nai Purchase Entry",
                 font=FH, bg=hdr["bg"], fg=WHITE).pack()
        tk.Frame(win, bg=RED, height=2).pack(fill="x")

        # Scrollable body
        canv = tk.Canvas(win, bg=BG, highlightthickness=0); canv.pack(fill="both", expand=True)
        vsb2 = tk.Scrollbar(win, orient="vertical", command=canv.yview); vsb2.pack(side="right", fill="y")
        canv.configure(yscrollcommand=vsb2.set)
        body = tk.Frame(canv, bg=BG, padx=24, pady=16)
        fid  = canv.create_window((0,0), window=body, anchor="nw")
        body.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))
        canv.bind("<Configure>", lambda e: canv.itemconfig(fid, width=e.width))
        def _mw1(e, _c=canv):
            try: _c.yview_scroll(int(-1*(e.delta/120)), "units")
            except Exception: pass
        canv.bind("<MouseWheel>", _mw1)
        win.bind("<MouseWheel>", _mw1)

        # ── Supplier ──
        r1 = tk.Frame(body, bg=BG); r1.pack(fill="x", pady=(0,10))
        tk.Label(r1, text="Supplier:", font=FBD, bg=BG, width=14, anchor="w").pack(side="left")
        sup_names = ["— Koi Nahi —"] + [s["name"] for s in suppliers]
        sv = tk.StringVar(value=existing["supplier_name"] if existing else "— Koi Nahi —")
        ttk.Combobox(r1, textvariable=sv, values=sup_names, state="normal", width=28).pack(side="left")
        btn(r1, "➕ Add", lambda: self._quick_add_supplier(sv, win), bg=DARK, px=6, py=4).pack(side="left", padx=6)

        # ── Invoice No ──
        r2 = tk.Frame(body, bg=BG); r2.pack(fill="x", pady=(0,10))
        tk.Label(r2, text="Invoice No:", font=FBD, bg=BG, width=14, anchor="w").pack(side="left")
        iv = tk.StringVar(value=existing["invoice_no"] if existing else "")
        tk.Entry(r2, textvariable=iv, font=FB, bg=IBGC, relief="solid", bd=1, width=20).pack(side="left", ipady=5)

        # ── Payment ──
        r3 = tk.Frame(body, bg=BG); r3.pack(fill="x", pady=(0,10))
        tk.Label(r3, text="Payment:", font=FBD, bg=BG, width=14, anchor="w").pack(side="left")
        pmv = tk.StringVar(value=existing["payment_method"] if existing else "Cash")
        ttk.Combobox(r3, textvariable=pmv, values=["Cash","UPI","Card","Credit","Cheque"],
                     state="readonly", width=14).pack(side="left", padx=(0,16))
        tk.Label(r3, text="Status:", font=FBD, bg=BG).pack(side="left")
        stv = tk.StringVar(value=existing["payment_status"] if existing else "Paid")
        ttk.Combobox(r3, textvariable=stv, values=["Paid","Unpaid","Partial"],
                     state="readonly", width=10).pack(side="left", padx=6)

        # ── Note ──
        r4 = tk.Frame(body, bg=BG); r4.pack(fill="x", pady=(0,14))
        tk.Label(r4, text="Note:", font=FBD, bg=BG, width=14, anchor="w").pack(side="left")
        nv2 = tk.StringVar(value=existing["note"] if existing else "")
        tk.Entry(r4, textvariable=nv2, font=FB, bg=IBGC, relief="solid", bd=1, width=38).pack(side="left", ipady=4)

        # ── Items Section ──
        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=(0,8))
        th2 = tk.Frame(body, bg=BG); th2.pack(fill="x")
        tk.Label(th2, text="📦 Items (Raw Materials)", font=FBD, bg=BG, fg=DARK).pack(side="left")

        items_frame = tk.Frame(body, bg=BG); items_frame.pack(fill="x", pady=(6,0))
        item_rows = []  # list of dicts: {mat_var, qty_var, price_var, total_lbl, frame}

        mat_names_list = [f"{m['name']} ({m['unit']})" for m in mats]
        mat_by_name    = {f"{m['name']} ({m['unit']})": m for m in mats}

        def add_item_row(mat_val="", qty_val="", price_val=""):
            rf = tk.Frame(items_frame, bg="#f9f9f9", relief="solid", bd=1, padx=8, pady=6)
            rf.pack(fill="x", pady=3)

            mat_v   = tk.StringVar(value=mat_val)
            qty_v   = tk.StringVar(value=str(qty_val) if qty_val else "")
            price_v = tk.StringVar(value=str(price_val) if price_val else "")
            tot_lbl = tk.Label(rf, text="₹0.00", font=FBD, bg="#f9f9f9", fg="#c05621", width=10)

            def upd_total(*a):
                try: t = float(qty_v.get()) * float(price_v.get()); tot_lbl.config(text=f"{cur}{t:.2f}")
                except: tot_lbl.config(text=f"{cur}0.00")
            qty_v.trace("w", upd_total); price_v.trace("w", upd_total)

            def on_mat(*a):
                m = mat_by_name.get(mat_v.get())
                if m and m["cost_per_unit"]: price_v.set(str(m["cost_per_unit"]))
            mat_v.trace("w", on_mat)

            row1 = tk.Frame(rf, bg="#f9f9f9"); row1.pack(fill="x")
            tk.Label(row1, text="Material:", font=FS, bg="#f9f9f9", fg=MED).pack(side="left")
            ttk.Combobox(row1, textvariable=mat_v, values=mat_names_list,
                         state="normal", width=30).pack(side="left", padx=6)

            def del_row():
                item_rows[:] = [x for x in item_rows if x["frame"] is not rf]
                rf.destroy(); upd_grand()
            btn(row1, "✖", del_row, bg="#c53030", px=6, py=2).pack(side="right")

            row2 = tk.Frame(rf, bg="#f9f9f9"); row2.pack(fill="x", pady=(4,0))
            tk.Label(row2, text="Qty:", font=FS, bg="#f9f9f9", fg=MED).pack(side="left")
            tk.Entry(row2, textvariable=qty_v, font=FB, bg=IBGC, relief="solid", bd=1,
                     width=8).pack(side="left", padx=(4,12), ipady=4)
            tk.Label(row2, text="Unit Price:", font=FS, bg="#f9f9f9", fg=MED).pack(side="left")
            tk.Entry(row2, textvariable=price_v, font=FB, bg=IBGC, relief="solid", bd=1,
                     width=10).pack(side="left", padx=(4,12), ipady=4)
            tk.Label(row2, text="Total:", font=FS, bg="#f9f9f9", fg=MED).pack(side="left")
            tot_lbl.pack(side="left")

            row_data = {"mat_var":mat_v,"qty_var":qty_v,"price_var":price_v,"total_lbl":tot_lbl,"frame":rf}
            item_rows.append(row_data)
            qty_v.trace("w", lambda *a: upd_grand())
            price_v.trace("w", lambda *a: upd_grand())
            if qty_val and price_val: upd_total()
            return row_data

        # Grand total
        gt_frame = tk.Frame(body, bg="#fff8f0", relief="solid", bd=1, padx=16, pady=8)
        self._gt_lbl = tk.Label(gt_frame, text="Grand Total: ₹0.00",
                                font=("Segoe UI",14,"bold"), bg="#fff8f0", fg="#c05621")
        self._gt_lbl.pack(side="right")

        def upd_grand(*a):
            total = 0
            for rd in item_rows:
                try: total += float(rd["qty_var"].get()) * float(rd["price_var"].get())
                except: pass
            self._gt_lbl.config(text=f"Grand Total: {cur}{total:.2f}")

        btn(body, "➕ Item Add Karo", lambda: add_item_row(), bg=DARK, px=10, py=6).pack(anchor="w", pady=(6,4))

        # Load existing items if editing
        for ei in ex_items:
            add_item_row(f"{ei['material_name']} ({ei['unit']})", ei["quantity"], ei["unit_price"])

        gt_frame.pack(fill="x", pady=(8,0))
        upd_grand()

        def do_save():
            # Validate items
            valid_items = []
            for rd in item_rows:
                mat_key = rd["mat_var"].get().strip()
                if not mat_key: continue
                mat = mat_by_name.get(mat_key)
                try:
                    qty   = float(rd["qty_var"].get())
                    price = float(rd["price_var"].get())
                    if qty <= 0: raise ValueError
                except:
                    messagebox.showerror("Error", f"'{mat_key}' ka quantity/price sahi likhein!", parent=win); return
                valid_items.append({"mat": mat, "mat_key": mat_key, "qty": qty, "price": price, "total": qty*price})

            if not valid_items:
                messagebox.showerror("Error","Kam se kam ek item daalo!", parent=win); return

            grand = sum(x["total"] for x in valid_items)
            c = db()
            try:
                if purchase_id:
                    # Edit: purane items delete karo
                    c.execute("DELETE FROM purchase_items WHERE purchase_id=?", (purchase_id,))
                    c.execute("""UPDATE purchases SET supplier_name=?,invoice_no=?,payment_method=?,
                                 payment_status=?,total_amount=?,note=? WHERE id=?""",
                              (sv.get() if sv.get()!="— Koi Nahi —" else None,
                               iv.get().strip() or None, pmv.get(), stv.get(),
                               grand, nv2.get().strip() or None, purchase_id))
                    pid = purchase_id
                else:
                    cur2 = c.execute("""INSERT INTO purchases
                                        (supplier_name,invoice_no,payment_method,payment_status,total_amount,note,created_at)
                                        VALUES (?,?,?,?,?,?,?)""",
                                     (sv.get() if sv.get()!="— Koi Nahi —" else None,
                                      iv.get().strip() or None, pmv.get(), stv.get(),
                                      grand, nv2.get().strip() or None,
                                      datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    pid = cur2.lastrowid

                for x in valid_items:
                    mat = x["mat"]
                    mid = mat["id"] if mat else None
                    mname = mat["name"] if mat else x["mat_key"].split(" (")[0]
                    unit  = mat["unit"] if mat else ""
                    c.execute("""INSERT INTO purchase_items
                                 (purchase_id,material_id,material_name,unit,quantity,unit_price,total_cost)
                                 VALUES (?,?,?,?,?,?,?)""",
                              (pid, mid, mname, unit, x["qty"], x["price"], x["total"]))
                    # Stock update + inventory log
                    if mid:
                        c.execute("UPDATE raw_materials SET current_stock=current_stock+?, cost_per_unit=? WHERE id=?",
                                  (x["qty"], x["price"], mid))
                        c.execute("""INSERT INTO inventory_log
                                     (material_id,type,qty,unit_price,total_cost,note,created_at)
                                     VALUES (?,?,?,?,?,?,?)""",
                                  (mid, "Added", x["qty"], x["price"], x["total"],
                                   f"Purchase #{pid:04d}",
                                   datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                c.commit()
                messagebox.showinfo("✅ Saved!", f"Purchase #{pid:04d} save ho gaya!\nTotal: {cur}{grand:.2f}", parent=win)
                win.destroy(); self._load()
                if hasattr(self.app,"_get_low_stock_items"):
                    self.app._update_low_stock_badge(len(self.app._get_low_stock_items()))
            except Exception as e:
                c.rollback(); messagebox.showerror("Error", str(e), parent=win)
            finally:
                pass  # conn reused

        # Fixed bottom
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        bf2 = tk.Frame(win, bg=BG, padx=24, pady=12); bf2.pack(fill="x")
        btn(bf2, "💾  Save Purchase", do_save, bg=GREEN, py=12).pack(side="left", fill="x", expand=True, padx=(0,6))
        btn(bf2, "✖  Cancel",        win.destroy, bg=MUTED, py=12).pack(side="right", fill="x", expand=True)

    def _quick_add_supplier(self, sv, parent_win):
        """Jaldi supplier add karo."""
        win = tk.Toplevel(parent_win)
        win.title("➕ Supplier Add"); win.configure(bg=BG); win.grab_set()
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"360x420+{(sx-360)//2}+{(sy-420)//2}")
        hdr = tk.Frame(win, bg=DARK, pady=8); hdr.pack(fill="x")
        tk.Label(hdr, text="➕ Naya Supplier", font=FH, bg=DARK, fg=WHITE).pack()
        body = tk.Frame(win, bg=BG, padx=20, pady=16); body.pack(fill="both", expand=True)
        fields = {}
        for lbl, key in [("Naam *","name"),("Phone","phone"),("Address","address"),("GST No","gst")]:
            tk.Label(body, text=lbl+":", font=FBD, bg=BG, anchor="w").pack(fill="x")
            v = tk.StringVar(); fields[key] = v
            tk.Entry(body, textvariable=v, font=FB, bg=IBGC, relief="solid", bd=1).pack(fill="x", ipady=6, pady=(2,10))
        def save_sup():
            name = fields["name"].get().strip()
            if not name: messagebox.showerror("Error","Naam likhna zaroori hai!",parent=win); return
            c = db()
            try:
                c.execute("INSERT OR IGNORE INTO suppliers (name,phone,address,gst) VALUES (?,?,?,?)",
                          (name, fields["phone"].get().strip(), fields["address"].get().strip(), fields["gst"].get().strip()))
                c.commit()
                sv.set(name)
                messagebox.showinfo("✅",f"'{name}' add ho gaya!",parent=win)
                win.destroy()
            except Exception as e:
                messagebox.showerror("Error",str(e),parent=win)
            finally: pass  # conn reused
        # Fixed bottom buttons
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(win, bg=BG, padx=16, pady=12); bf.pack(fill="x")
        btn(bf, "💾  Save Supplier", save_sup,    bg=GREEN, py=10).pack(fill="x", pady=(0,6))
        btn(bf, "✖  Cancel",        win.destroy, bg=MUTED, py=8).pack(fill="x")

    def _manage_suppliers(self):
        """Suppliers list dekho aur manage karo."""
        c = db()
        sups = [dict(r) for r in c.execute("SELECT * FROM suppliers ORDER BY name").fetchall()]
        pass  # conn reused
        win = tk.Toplevel(self)
        win.title("👤 Suppliers"); win.configure(bg=BG); win.grab_set()
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"560x420+{(sx-560)//2}+{(sy-420)//2}")
        hdr = tk.Frame(win, bg=DARK, pady=8); hdr.pack(fill="x")
        tk.Label(hdr, text="👤  Suppliers List", font=FH, bg=DARK, fg=WHITE).pack()
        frm = tk.Frame(win, bg=BG, padx=12, pady=12); frm.pack(fill="both", expand=True)
        cols = ("ID","Naam","Phone","Address","GST")
        tree = ttk.Treeview(frm, columns=cols, show="headings", height=12)
        for col, w in zip(cols, [40,160,100,160,100]):
            tree.heading(col, text=col); tree.column(col, width=w, anchor="center")
        tree.column("Naam", anchor="w"); tree.column("Address", anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); tree.pack(fill="both", expand=True)
        for s in sups:
            tree.insert("","end", iid=s["id"], values=(s["id"], s["name"], s["phone"] or "—",
                                                        s["address"] or "—", s["gst"] or "—"))
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        btn(win, "✖ Band Karo", win.destroy, bg=MUTED, py=8).pack(fill="x", padx=16, pady=10)

    def _view_detail(self, event=None):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("","Pehle ek purchase select karo."); return
        pid = int(sel[0]); c = db(); cur = gset("currency","₹")
        purchase = dict(c.execute("SELECT * FROM purchases WHERE id=?", (pid,)).fetchone())
        items    = [dict(r) for r in c.execute("SELECT * FROM purchase_items WHERE purchase_id=?", (pid,)).fetchall()]
        pass  # conn reused
        win = tk.Toplevel(self)
        win.title(f"Purchase #{pid:04d} Detail")
        win.configure(bg=BG); win.grab_set()
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        h = min(620, 280 + len(items)*52)
        win.geometry(f"500x{h}+{(sx-500)//2}+{(sy-h)//2}")
        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"🛒 Purchase #{pid:04d} Detail", font=FH, bg=DKRED, fg=WHITE).pack()
        body = tk.Frame(win, bg=BG, padx=20, pady=14); body.pack(fill="both", expand=True)
        for lbl, val in [
            ("Purchase #", f"#{pid:04d}"),
            ("Date", purchase["created_at"][:16]),
            ("Supplier", purchase["supplier_name"] or "—"),
            ("Invoice", purchase["invoice_no"] or "—"),
            ("Payment", purchase["payment_method"]),
            ("Status", purchase["payment_status"]),
        ]:
            r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=1)
            tk.Label(r, text=f"{lbl}:", font=FBD, bg=BG, fg=MED, width=12, anchor="w").pack(side="left")
            color = RED if val=="Unpaid" else GREEN if val=="Paid" else DARK
            tk.Label(r, text=val, font=FB, bg=BG, fg=color).pack(side="left")
        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=8)
        tk.Label(body, text="Items:", font=FBD, bg=BG, fg=DARK).pack(anchor="w")
        for it in items:
            rf = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=8, pady=4); rf.pack(fill="x", pady=2)
            tk.Label(rf, text=f"{it['material_name']} ({it['unit']})", font=FB, bg=WHITE, fg=DARK).pack(side="left")
            tk.Label(rf, text=f"{it['quantity']} × {cur}{it['unit_price']:.2f} = {cur}{it['total_cost']:.2f}",
                     font=FB, bg=WHITE, fg="#c05621").pack(side="right")
        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=6)
        r = tk.Frame(body, bg=BG); r.pack(fill="x")
        tk.Label(r, text="TOTAL:", font=FBD, bg=BG, fg=MED).pack(side="left")
        tk.Label(r, text=f"{cur}{purchase['total_amount']:.2f}", font=("Segoe UI",14,"bold"),
                 bg=BG, fg="#c05621").pack(side="right")
        if purchase.get("note"):
            tk.Label(body, text=f"Note: {purchase['note']}", font=FS, bg=BG, fg=MED).pack(anchor="w", pady=(4,0))
        tk.Frame(win, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(win, bg=BG, padx=16, pady=10); bf.pack(fill="x")
        btn(bf, "✏️ Edit", lambda: [win.destroy(), self._purchase_form(pid)], bg="#2b6cb0", py=8).pack(side="left", fill="x", expand=True, padx=(0,6))
        btn(bf, "✖ Band Karo", win.destroy, bg=MUTED, py=8).pack(side="right", fill="x", expand=True)

    def _delete(self):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("","Pehle ek purchase select karo."); return
        pid = int(sel[0])
        if messagebox.askyesno("⚠️ Confirm", f"Purchase #{pid:04d} delete karna chahte hain?\nStock wapas nahi hoga!"):
            c = db()
            c.execute("DELETE FROM purchase_items WHERE purchase_id=?", (pid,))
            c.execute("DELETE FROM purchases WHERE id=?", (pid,))
            c.commit()
            self._load()
            messagebox.showinfo("✅", f"Purchase #{pid:04d} delete ho gaya.")


# ════════════════════════════════════════════════════════════
#  ORDER HISTORY PAGE
# ════════════════════════════════════════════════════════════
class OrderHistoryPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=WHITE); hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10); th.pack(fill="x")
        tk.Label(th, text="📋  Order History", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        btn(th, "🔄 Refresh", self._load, bg=DARK, px=10).pack(side="right")

        # Filter bar
        fb = tk.Frame(self, bg=BG, padx=12, pady=8); fb.pack(fill="x")
        tk.Label(fb, text="Date Filter:", font=FBD, bg=BG, fg=DARK).pack(side="left", padx=(0,6))
        self.date_var = tk.StringVar(value="aaj")
        for lbl, val in [("Aaj","aaj"),("Kal","kal"),("7 Din","7din"),("30 Din","30din"),("Sab","sab")]:
            b = tk.Button(fb, text=lbl, font=FS, relief="flat",
                          bg=RED if val=="aaj" else "#edf2f7",
                          fg=WHITE if val=="aaj" else DARK,
                          padx=10, pady=4, cursor="hand2",
                          command=lambda v=val: self._filter(v))
            b.pack(side="left", padx=2)
            setattr(self, f"_dbtn_{val}", b)

        # Search
        sf = tk.Frame(self, bg=BG, padx=12); sf.pack(fill="x", pady=(0,6))
        tk.Label(sf, text="🔍 Customer/Table:", font=FS, bg=BG).pack(side="left")
        self.srch_var = tk.StringVar()
        self.srch_var.trace("w", lambda *a: self._load())
        tk.Entry(sf, textvariable=self.srch_var, font=FB, bg=IBGC,
                 relief="solid", bd=1, width=24).pack(side="left", padx=6)

        # Summary bar
        self.sum_frame = tk.Frame(self, bg="#fff5f5", relief="solid", bd=1)
        self.sum_frame.pack(fill="x", padx=12, pady=(0,6))
        self.lbl_orders  = tk.Label(self.sum_frame, text="Orders: 0",   font=FBD, bg="#fff5f5", fg=DARK, padx=16, pady=6)
        self.lbl_orders.pack(side="left")
        self.lbl_sales   = tk.Label(self.sum_frame, text="Total Sale: ₹0", font=FBD, bg="#fff5f5", fg=GREEN, padx=16)
        self.lbl_sales.pack(side="left")
        self.lbl_avg     = tk.Label(self.sum_frame, text="Average Bill: ₹0", font=FS, bg="#fff5f5", fg=MED, padx=16)
        self.lbl_avg.pack(side="left")

        # Treeview
        frm = tk.Frame(self, bg=BG); frm.pack(fill="both", expand=True, padx=12, pady=(0,4))
        cols = ("Bill#","Date & Time","Table","Customer","Items","Subtotal","Tax","Discount","TOTAL","Payment")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings")
        widths =    [55,    140,         80,     120,      50,       80,       60,    80,        90,     80]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col, command=lambda c=col: self._sort(c))
            self.tree.column(col, width=w, anchor="center")
        self.tree.column("Date & Time", anchor="w")
        self.tree.column("Customer", anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure("even", background="#f9f9f9")
        self.tree.bind("<Double-1>", self._view_detail)

        # Bottom buttons
        bf = tk.Frame(self, bg=BG, pady=6); bf.pack(fill="x", padx=12)
        btn(bf, "🔍 Bill Detail Dekho", self._view_detail, bg=DARK, px=10).pack(side="left", padx=4)
        btn(bf, "🖨️ Bill Print Karo",   self._print_bill,  bg=RED,  px=10).pack(side="left", padx=4)
        btn(bf, "🗑️ Order Delete",      self._delete_order, bg="#c53030", px=10).pack(side="left", padx=4)
        btn(bf, "📥 Excel Download",    self._export_excel, bg=GREEN, px=10).pack(side="right", padx=4)

        self._sort_col = "Bill#"; self._sort_asc = False
        self._load()

    def _filter(self, val):
        self.date_var.set(val)
        for v in ["aaj","kal","7din","30din","sab"]:
            b = getattr(self, f"_dbtn_{v}", None)
            if b: b.config(bg=RED if v==val else "#edf2f7", fg=WHITE if v==val else DARK)
        self._load()

    def _date_range(self):
        today = datetime.date.today()
        v = self.date_var.get()
        if v == "aaj":
            return str(today), str(today)
        elif v == "kal":
            d = today - datetime.timedelta(days=1)
            return str(d), str(d)
        elif v == "7din":
            return str(today - datetime.timedelta(days=6)), str(today)
        elif v == "30din":
            return str(today - datetime.timedelta(days=29)), str(today)
        else:
            return None, None

    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        c = db(); cur = gset("currency","₹")
        q = "SELECT o.*, (SELECT COUNT(*) FROM order_items WHERE order_id=o.id) as item_count FROM orders o WHERE 1=1"
        params = []
        d1, d2 = self._date_range()
        if d1: q += " AND DATE(o.created_at) >= ?"; params.append(d1)
        if d2: q += " AND DATE(o.created_at) <= ?"; params.append(d2)
        srch = self.srch_var.get().strip()
        if srch:
            q += " AND (o.customer_name LIKE ? OR o.table_number LIKE ?)"
            params += [f"%{srch}%", f"%{srch}%"]
        q += " ORDER BY o.id DESC"
        rows = c.execute(q, params).fetchall()
        pass  # conn reused
        total_sale = 0
        for i, r in enumerate(rows):
            total_sale += (r["total_amount"] or 0)
            tag = "even" if i % 2 == 0 else ""
            self.tree.insert("", "end", iid=r["id"], tags=(tag,), values=(
                f"#{r['id']:04d}",
                r["created_at"][:16] if r["created_at"] else "",
                r["table_number"] or "Takeaway",
                r["customer_name"] or "—",
                r["item_count"],
                f"{cur}{(r['subtotal'] or 0):.2f}",
                f"{cur}{(r['tax_amount'] or 0):.2f}",
                f"-{cur}{(r['discount_amount'] or 0):.2f}",
                f"{cur}{(r['total_amount'] or 0):.2f}",
                r["payment_method"] or "Cash",
            ))
        n = len(rows)
        avg = total_sale / n if n > 0 else 0
        self.lbl_orders.config(text=f"Orders: {n}")
        self.lbl_sales.config(text=f"Total Sale: {cur}{total_sale:.2f}")
        self.lbl_avg.config(text=f"Average Bill: {cur}{avg:.2f}")

    def _sort(self, col):
        if self._sort_col == col:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col; self._sort_asc = True
        data = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        data.sort(reverse=not self._sort_asc)
        for idx, (_, k) in enumerate(data):
            self.tree.move(k, "", idx)

    def _view_detail(self, event=None):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("","Pehle ek order select karo."); return
        oid = int(sel[0]); c = db()
        order = dict(c.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone())
        items = [dict(r) for r in c.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()]
        pass  # conn reused
        cur = gset("currency","₹")
        win = tk.Toplevel(self)
        win.title(f"Bill #{oid:04d} Detail")
        win.geometry("480x520"); win.configure(bg=BG); win.grab_set()
        win.update_idletasks()
        sx=win.winfo_screenwidth(); sy=win.winfo_screenheight()
        win.geometry(f"480x520+{(sx-480)//2}+{(sy-520)//2}")
        hdr = tk.Frame(win, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"📋 Bill #{oid:04d} Detail", font=FH, bg=DKRED, fg=WHITE).pack()
        body = tk.Frame(win, bg=BG, padx=20, pady=16); body.pack(fill="both", expand=True)
        info = [
            ("Bill No", f"#{oid:04d}"), ("Date", order["created_at"][:16]),
            ("Table", order["table_number"] or "Takeaway"),
            ("Customer", order["customer_name"] or "—"),
            ("Payment", order["payment_method"] or "Cash"),
        ]
        for lbl, val in info:
            r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=1)
            tk.Label(r, text=f"{lbl}:", font=FBD, bg=BG, fg=MED, width=12, anchor="w").pack(side="left")
            tk.Label(r, text=val, font=FB, bg=BG, fg=DARK).pack(side="left")
        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=8)
        tk.Label(body, text="Items:", font=FBD, bg=BG, fg=DARK).pack(anchor="w")
        for it in items:
            r = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=8, pady=4)
            r.pack(fill="x", pady=2)
            tk.Label(r, text=it["item_name"], font=FB, bg=WHITE, fg=DARK).pack(side="left")
            tk.Label(r, text=f"x{it['quantity']}  {cur}{it['item_total']:.2f}",
                     font=FB, bg=WHITE, fg=RED).pack(side="right")
        tk.Frame(body, bg=BORD, height=1).pack(fill="x", pady=8)
        for lbl, val, color in [
            ("Subtotal", f"{cur}{order['subtotal']:.2f}", DARK),
            ("Tax", f"{cur}{order['tax_amount']:.2f}", DARK),
            ("Discount", f"-{cur}{order['discount_amount']:.2f}", DARK),
            ("TOTAL", f"{cur}{order['total_amount']:.2f}", RED),
        ]:
            r = tk.Frame(body, bg=BG); r.pack(fill="x", pady=1)
            tk.Label(r, text=lbl+":", font=FBD, bg=BG, fg=MED, width=12, anchor="w").pack(side="left")
            tk.Label(r, text=val, font=FBD, bg=BG, fg=color).pack(side="right")
        btn(body, "✖ Close", win.destroy, bg=MUTED, py=8).pack(fill="x", pady=(12,0))
        btn(body, "🖨️ Print Karo", lambda: self._print_bill(oid=oid), bg=RED, py=8).pack(fill="x", pady=(6,0))

    def _delete_order(self):
        sel = self.tree.selection()
        if not sel: messagebox.showinfo("","Pehle ek order select karo."); return
        oid = int(sel[0])
        if messagebox.askyesno("⚠️ Confirm", f"Bill #{oid:04d} permanently delete karna chahte hain?\nYeh action undo nahi hoga!"):
            c = db()
            c.execute("DELETE FROM order_items WHERE order_id=?", (oid,))
            c.execute("DELETE FROM orders WHERE id=?", (oid,))
            c.commit()
            self._load()
            messagebox.showinfo("✅", f"Bill #{oid:04d} delete ho gaya.")

    def _print_bill(self, event=None, oid=None):
        """Order History se bill print karo — print type dialog dikhao."""
        if oid is None:
            sel = self.tree.selection()
            if not sel: messagebox.showinfo("", "Pehle ek bill select karo."); return
            oid = int(sel[0])

        # DB se order aur items load karo
        c = db()
        order_row = c.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone()
        if not order_row:
            messagebox.showerror("Error", f"Bill #{oid:04d} nahi mila!"); return
        order      = dict(order_row)
        bill_items = [dict(r) for r in c.execute("SELECT * FROM order_items WHERE order_id=?", (oid,)).fetchall()]
        settings_d = {r["key"]: r["value"] for r in c.execute("SELECT * FROM settings").fetchall()}
        pass  # conn reused

        # Print type dialog
        dlg = tk.Toplevel(self)
        dlg.title("🖨️ Print Options")
        dlg.configure(bg=BG); dlg.grab_set(); dlg.resizable(False, False)
        dlg.update_idletasks()
        sw = dlg.winfo_screenwidth(); sh = dlg.winfo_screenheight()
        dlg.geometry(f"400x500+{(sw-400)//2}+{(sh-500)//2}")

        hdr = tk.Frame(dlg, bg=DKRED, pady=10); hdr.pack(fill="x")
        tk.Label(hdr, text=f"🖨️  Bill #{oid:04d} Print Karo", font=FH, bg=DKRED, fg=WHITE).pack()
        tk.Frame(dlg, bg=RED, height=2).pack(fill="x")
        body = tk.Frame(dlg, bg=BG, padx=24, pady=16); body.pack(fill="both", expand=True)

        ptype = tk.StringVar(value=gset("default_print_type", "thermal"))

        # Thermal option
        tf = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=12, pady=10)
        tf.pack(fill="x", pady=(0,8))
        tk.Radiobutton(tf, text="🖨️  Thermal Printer", variable=ptype, value="thermal",
                       font=FBD, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Label(tf, text="58mm / 80mm thermal roll printer ke liye\n(ESC/POS compatible)",
                 font=FS, bg=WHITE, fg=MUTED, justify="left").pack(anchor="w", padx=20)

        wf = tk.Frame(body, bg=BG); wf.pack(fill="x", pady=(0,8))
        tk.Label(wf, text="  Paper Width:", font=FS, bg=BG).pack(side="left")
        wvar = tk.StringVar(value=gset("thermal_width", "80"))
        ttk.Combobox(wf, textvariable=wvar, values=["58","80"],
                     width=5, state="readonly").pack(side="left", padx=6)
        tk.Label(wf, text="mm", font=FS, bg=BG).pack(side="left")

        # PDF option
        pf = tk.Frame(body, bg=WHITE, relief="solid", bd=1, padx=12, pady=10)
        pf.pack(fill="x", pady=(0,8))
        tk.Radiobutton(pf, text="📄  PDF Bill (A5)", variable=ptype, value="pdf",
                       font=FBD, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Label(pf, text="A5 size PDF generate karke open karega\n(Normal printer ke liye)",
                 font=FS, bg=WHITE, fg=MUTED, justify="left").pack(anchor="w", padx=20)

        # Printer name
        pnf = tk.Frame(body, bg=BG); pnf.pack(fill="x", pady=(0,4))
        tk.Label(pnf, text="  Printer Name (optional):", font=FS, bg=BG).pack(anchor="w")
        pname_var = tk.StringVar(value=gset("thermal_printer", ""))
        try:
            import win32print
            pname_var.set(win32print.GetDefaultPrinter())
        except: pass
        tk.Entry(pnf, textvariable=pname_var, font=FS, bg=IBGC,
                 relief="solid", bd=1).pack(fill="x", ipady=4, pady=(2,0))
        tk.Label(pnf, text="  (Khali chhodein = default printer)",
                 font=("Segoe UI",8), bg=BG, fg=MUTED).pack(anchor="w")

        def do_print():
            dlg.destroy()
            pt    = ptype.get()
            pn    = pname_var.get().strip() or None
            bd    = os.path.join(BASE_DIR, "bills"); os.makedirs(bd, exist_ok=True)

            if pt == "thermal":
                # Thermal preview dikhao
                bill_text = generate_thermal_bill(oid, order, bill_items, settings_d, int(wvar.get()))
                _show_thermal_preview(self, oid, bill_text, pn, bd)
            else:
                # PDF generate karo aur open karo (preview)
                pdf_path = os.path.join(bd, f"BILL_{oid:04d}.pdf")
                path = generate_pdf(oid, order, bill_items, settings_d, pdf_path)
                if path:
                    try:
                        if sys.platform == "win32": os.startfile(path)
                    except: pass
                    messagebox.showinfo("📄 PDF Preview Khul Gaya!",
                        f"PDF preview khul gaya hai.\nWahan se Print karo (Ctrl+P)\n\nFile: {path}")

        # Buttons — hamesha neeche fixed
        tk.Frame(dlg, bg=BORD, height=1).pack(fill="x")
        bf = tk.Frame(dlg, bg=BG, padx=24, pady=12); bf.pack(fill="x")
        btn(bf, "🖨️  PRINT KARO", do_print, bg=RED, py=12).pack(fill="x", pady=(0,6))
        btn(bf, "✖  Cancel", dlg.destroy, bg=MUTED, py=8).pack(fill="x")


# ════════════════════════════════════════════════════════════
#  MATERIAL HISTORY PAGE
# ════════════════════════════════════════════════════════════

    def _export_excel(self):
        """Order History Excel mein export karo."""
        c = db()
        rows = c.execute(
            "SELECT o.id, o.created_at, o.table_number, o.customer_name, "
            "o.payment_method, o.subtotal, o.tax_amount, o.discount_amount, o.total_amount "
            "FROM orders o ORDER BY o.id DESC"
        ).fetchall()
        pass  # conn reused
        headers = ["Bill No","Date & Time","Table","Customer","Payment","Subtotal","Tax","Discount","Total"]
        data = []
        cur = gset("currency","₹")
        for r in rows:
            data.append((
                f"#{r['id']:04d}", r['created_at'], r['table_number'] or "Takeaway",
                r['customer_name'] or "", r['payment_method'] or "Cash",
                r['subtotal'] or 0, r['tax_amount'] or 0,
                r['discount_amount'] or 0, r['total_amount'] or 0
            ))
        export_to_excel("Order_History", headers, data, parent=self)


class MaterialHistoryPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=WHITE); hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10); th.pack(fill="x")
        tk.Label(th, text="🥬  Raw Material History", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        btn(th, "📥 Excel Download", self._export_excel, bg=GREEN, px=10).pack(side="right", padx=4)
        btn(th, "🔄 Refresh", self._load, bg=DARK, px=10).pack(side="right")

        # Filter bar
        fb = tk.Frame(self, bg=BG, padx=12, pady=8); fb.pack(fill="x")
        tk.Label(fb, text="Date Filter:", font=FBD, bg=BG).pack(side="left", padx=(0,6))
        self.date_var = tk.StringVar(value="aaj")
        for lbl, val in [("Aaj","aaj"),("Kal","kal"),("7 Din","7din"),("30 Din","30din"),("Sab","sab")]:
            b = tk.Button(fb, text=lbl, font=FS, relief="flat",
                          bg=RED if val=="aaj" else "#edf2f7",
                          fg=WHITE if val=="aaj" else DARK,
                          padx=10, pady=4, cursor="hand2",
                          command=lambda v=val: self._filter(v))
            b.pack(side="left", padx=2)
            setattr(self, f"_dbtn_{val}", b)

        # Material filter
        mf = tk.Frame(self, bg=BG, padx=12); mf.pack(fill="x", pady=(0,6))
        tk.Label(mf, text="Material:", font=FBD, bg=BG).pack(side="left", padx=(0,6))
        c = db()
        mats = ["Sab"] + [r["name"] for r in c.execute("SELECT name FROM raw_materials ORDER BY name").fetchall()]
        pass  # conn reused
        self.mat_var = tk.StringVar(value="Sab")
        ttk.Combobox(mf, textvariable=self.mat_var, values=mats, state="readonly",
                     width=20, font=FS).pack(side="left", padx=4)
        btn(mf, "Filter", self._load, bg=DARK, px=8, py=4).pack(side="left", padx=4)

        # Summary bar
        self.sum_frame = tk.Frame(self, bg="#f0fff4", relief="solid", bd=1)
        self.sum_frame.pack(fill="x", padx=12, pady=(0,6))
        self.lbl_entries  = tk.Label(self.sum_frame, text="Entries: 0",         font=FBD, bg="#f0fff4", fg=DARK, padx=16, pady=6)
        self.lbl_entries.pack(side="left")
        self.lbl_added    = tk.Label(self.sum_frame, text="Total Added: 0",      font=FBD, bg="#f0fff4", fg=GREEN, padx=16)
        self.lbl_added.pack(side="left")
        self.lbl_used     = tk.Label(self.sum_frame, text="Total Used/Removed: 0", font=FBD, bg="#f0fff4", fg=RED, padx=16)
        self.lbl_used.pack(side="left")

        # Treeview
        frm = tk.Frame(self, bg=BG); frm.pack(fill="both", expand=True, padx=12, pady=(0,4))
        cols = ("ID","Date & Time","Material","Type","Quantity","Note")
        self.tree = ttk.Treeview(frm, columns=cols, show="headings")
        widths =    [50,   140,         180,       100,   100,      260]
        for col, w in zip(cols, widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center")
        self.tree.column("Date & Time", anchor="w")
        self.tree.column("Material", anchor="w")
        self.tree.column("Note", anchor="w")
        vsb = ttk.Scrollbar(frm, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(frm, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        self.tree.tag_configure("Added", foreground=GREEN)
        self.tree.tag_configure("Used",  foreground=RED)
        self.tree.tag_configure("even",  background="#f9f9f9")
        self._load()

    def _filter(self, val):
        self.date_var.set(val)
        for v in ["aaj","kal","7din","30din","sab"]:
            b = getattr(self, f"_dbtn_{v}", None)
            if b: b.config(bg=RED if v==val else "#edf2f7", fg=WHITE if v==val else DARK)
        self._load()

    def _date_range(self):
        today = datetime.date.today()
        v = self.date_var.get()
        if v == "aaj":   return str(today), str(today)
        elif v == "kal":
            d = today - datetime.timedelta(days=1); return str(d), str(d)
        elif v == "7din":  return str(today - datetime.timedelta(days=6)), str(today)
        elif v == "30din": return str(today - datetime.timedelta(days=29)), str(today)
        else: return None, None

    def _load(self):
        for r in self.tree.get_children(): self.tree.delete(r)
        c = db()
        q = ("SELECT il.*, rm.name as mat_name, rm.unit "
             "FROM inventory_log il "
             "LEFT JOIN raw_materials rm ON il.material_id=rm.id WHERE 1=1")
        params = []
        d1, d2 = self._date_range()
        if d1: q += " AND DATE(il.created_at) >= ?"; params.append(d1)
        if d2: q += " AND DATE(il.created_at) <= ?"; params.append(d2)
        mat = self.mat_var.get()
        if mat != "Sab": q += " AND rm.name=?"; params.append(mat)
        q += " ORDER BY il.id DESC"
        rows = c.execute(q, params).fetchall()
        total_added = 0; total_used = 0
        for i, r in enumerate(rows):
            typ = r["type"] or "Added"
            qty = r["qty"] or 0
            unit = r["unit"] or ""
            if "added" in typ.lower() or "add" in typ.lower():
                total_added += qty
            else:
                total_used += qty
            tag_color = "Added" if "add" in typ.lower() else "Used"
            tag_row = "even" if i % 2 == 0 else ""
            self.tree.insert("", "end", tags=(tag_color, tag_row), values=(
                r["id"],
                r["created_at"][:16] if r["created_at"] else "",
                r["mat_name"] or "—",
                typ,
                f"{qty} {unit}",
                r["note"] or "—",
            ))
        self.lbl_entries.config(text=f"Entries: {len(rows)}")
        self.lbl_added.config(text=f"Total Added: {total_added:.2f}")
        self.lbl_used.config(text=f"Total Used/Removed: {total_used:.2f}")


# ════════════════════════════════════════════════════════════
#  REPORTS PAGE
# ════════════════════════════════════════════════════════════

    def _export_excel(self):
        """Material History Excel mein export karo."""
        c = db()
        rows = c.execute(
            "SELECT il.created_at, rm.name as material, il.type, il.qty, rm.unit, "
            "il.unit_price, il.total_cost, il.note "
            "FROM inventory_log il LEFT JOIN raw_materials rm ON il.material_id=rm.id "
            "ORDER BY il.id DESC"
        ).fetchall()
        pass  # conn reused
        cur = gset("currency","₹")
        headers = ["Date & Time","Material","Type","Quantity","Unit","Unit Price","Total Cost","Note"]
        data = [(
            r["created_at"], r["material"] or "", r["type"] or "",
            r["qty"] or 0, r["unit"] or "",
            f"{cur}{r['unit_price'] or 0:.2f}", f"{cur}{r['total_cost'] or 0:.2f}",
            r["note"] or ""
        ) for r in rows]
        export_to_excel("Material_History", headers, data, parent=self)


class ReportsPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        hdr = tk.Frame(self, bg=WHITE); hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10); th.pack(fill="x")
        tk.Label(th, text="📊  Sales & Purchase Reports", font=FH, bg=WHITE, fg=DARK).pack(side="left")
        btn(th, "📥 Excel Download", self._export_excel, bg=GREEN, px=10).pack(side="right", padx=4)
        btn(th, "🔄 Refresh", self._load_all, bg=DARK, px=10).pack(side="right")

        # Date filter
        fb = tk.Frame(self, bg=BG, padx=12, pady=8); fb.pack(fill="x")
        tk.Label(fb, text="Period:", font=FBD, bg=BG).pack(side="left", padx=(0,6))
        self.date_var = tk.StringVar(value="aaj")
        for lbl, val in [("Aaj","aaj"),("Kal","kal"),("7 Din","7din"),("30 Din","30din"),("Sab","sab")]:
            b = tk.Button(fb, text=lbl, font=FS, relief="flat",
                          bg=RED if val=="aaj" else "#edf2f7",
                          fg=WHITE if val=="aaj" else DARK,
                          padx=10, pady=4, cursor="hand2",
                          command=lambda v=val: self._filter(v))
            b.pack(side="left", padx=2)
            setattr(self, f"_dbtn_{val}", b)

        # Scrollable content
        canv_wrap = tk.Frame(self, bg=BG); canv_wrap.pack(fill="both", expand=True)
        canv = tk.Canvas(canv_wrap, bg=BG, highlightthickness=0)
        vsb  = tk.Scrollbar(canv_wrap, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y"); canv.pack(side="left", fill="both", expand=True)
        self.outer = tk.Frame(canv, bg=BG)
        fid = canv.create_window((0,0), window=self.outer, anchor="nw")
        self.outer.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))
        canv.bind("<Configure>", lambda e: canv.itemconfig(fid, width=e.width))
        def _mw2(e, _c=canv):
            try: _c.yview_scroll(int(-1*(e.delta/120)), "units")
            except Exception: pass
        canv.bind("<MouseWheel>", _mw2)
        self.bind("<MouseWheel>", _mw2)
        self._load_all()

    def _filter(self, val):
        self.date_var.set(val)
        for v in ["aaj","kal","7din","30din","sab"]:
            b = getattr(self, f"_dbtn_{v}", None)
            if b: b.config(bg=RED if v==val else "#edf2f7", fg=WHITE if v==val else DARK)
        self._load_all()

    def _date_range(self):
        today = datetime.date.today()
        v = self.date_var.get()
        if v == "aaj":   return str(today), str(today)
        elif v == "kal":
            d = today - datetime.timedelta(days=1); return str(d), str(d)
        elif v == "7din":  return str(today - datetime.timedelta(days=6)), str(today)
        elif v == "30din": return str(today - datetime.timedelta(days=29)), str(today)
        else: return None, None

    def _card(self, label, value, color=DARK, bg="#ffffff"):
        """Stat card banana"""
        f = tk.Frame(self.row_frame, bg=bg, relief="solid", bd=1, padx=20, pady=16)
        f.pack(side="left", expand=True, fill="both", padx=6, pady=4)
        tk.Label(f, text=label, font=FS, bg=bg, fg=MED).pack(anchor="w")
        tk.Label(f, text=value, font=("Segoe UI",18,"bold"), bg=bg, fg=color).pack(anchor="w", pady=(4,0))

    def _load_all(self):
        for w in self.outer.winfo_children(): w.destroy()
        c = db(); cur = gset("currency","₹")
        d1, d2 = self._date_range()

        # ── SALES SUMMARY ──────────────────────────────────
        q = "SELECT COUNT(*) as cnt, SUM(total_amount) as total, SUM(tax_amount) as tax, SUM(discount_amount) as disc, SUM(subtotal) as sub FROM orders WHERE 1=1"
        params = []
        if d1: q += " AND DATE(created_at)>=?"; params.append(d1)
        if d2: q += " AND DATE(created_at)<=?"; params.append(d2)
        row = dict(c.execute(q, params).fetchone())
        total_sale  = row["total"] or 0
        total_tax   = row["tax"] or 0
        total_disc  = row["disc"] or 0
        total_sub   = row["sub"] or 0
        order_count = row["cnt"] or 0
        avg_bill    = total_sale / order_count if order_count > 0 else 0

        # Section header
        tk.Label(self.outer, text="💰  Sales Summary", font=FH, bg=BG, fg=DARK).pack(anchor="w", padx=18, pady=(14,4))
        tk.Frame(self.outer, bg=RED, height=2).pack(fill="x", padx=12, pady=(0,6))

        self.row_frame = tk.Frame(self.outer, bg=BG); self.row_frame.pack(fill="x", padx=6)
        self._card("Total Orders",     str(order_count),              color=DARK,  bg="#edf2f7")
        self._card("Total Sale",       f"{cur}{total_sale:.2f}",      color=GREEN, bg="#f0fff4")
        self._card("Average Bill",     f"{cur}{avg_bill:.2f}",        color=DARK,  bg="#fffbf0")
        self._card("Total Tax (GST)",  f"{cur}{total_tax:.2f}",       color=DARK,  bg="#fff5f5")
        self._card("Total Discount",   f"{cur}{total_disc:.2f}",      color=RED,   bg="#fff5f5")

        # ── PAYMENT METHOD BREAKUP ──────────────────────────
        tk.Label(self.outer, text="💳  Payment Method Breakup", font=FH, bg=BG, fg=DARK).pack(anchor="w", padx=18, pady=(18,4))
        tk.Frame(self.outer, bg=RED, height=2).pack(fill="x", padx=12, pady=(0,6))
        q2 = "SELECT payment_method, COUNT(*) as cnt, SUM(total_amount) as total FROM orders WHERE 1=1"
        params2 = []
        if d1: q2 += " AND DATE(created_at)>=?"; params2.append(d1)
        if d2: q2 += " AND DATE(created_at)<=?"; params2.append(d2)
        q2 += " GROUP BY payment_method"
        pay_rows = c.execute(q2, params2).fetchall()
        pf = tk.Frame(self.outer, bg=BG); pf.pack(fill="x", padx=12, pady=(0,8))
        pay_colors = {"Cash":"#f0fff4","UPI":"#ebf8ff","Card":"#faf5ff","Credit":"#fffbf0"}
        for pr in pay_rows:
            meth = pr["payment_method"] or "Cash"
            bg_c = pay_colors.get(meth, "#f9f9f9")
            pcard = tk.Frame(pf, bg=bg_c, relief="solid", bd=1, padx=16, pady=10)
            pcard.pack(side="left", expand=True, fill="both", padx=6)
            tk.Label(pcard, text=meth, font=FBD, bg=bg_c, fg=DARK).pack(anchor="w")
            tk.Label(pcard, text=f"{pr['cnt']} orders", font=FS, bg=bg_c, fg=MED).pack(anchor="w")
            tk.Label(pcard, text=f"{cur}{(pr['total'] or 0):.2f}", font=("Segoe UI",14,"bold"), bg=bg_c, fg=GREEN).pack(anchor="w")

        # ── TOP SELLING ITEMS ───────────────────────────────
        tk.Label(self.outer, text="🏆  Top Selling Items", font=FH, bg=BG, fg=DARK).pack(anchor="w", padx=18, pady=(18,4))
        tk.Frame(self.outer, bg=RED, height=2).pack(fill="x", padx=12, pady=(0,6))
        q3 = ("SELECT oi.item_name, SUM(oi.quantity) as total_qty, SUM(oi.item_total) as total_rev "
              "FROM order_items oi JOIN orders o ON oi.order_id=o.id WHERE 1=1")
        params3 = []
        if d1: q3 += " AND DATE(o.created_at)>=?"; params3.append(d1)
        if d2: q3 += " AND DATE(o.created_at)<=?"; params3.append(d2)
        q3 += " GROUP BY oi.item_name ORDER BY total_qty DESC LIMIT 10"
        top_items = c.execute(q3, params3).fetchall()
        tf = tk.Frame(self.outer, bg=BG); tf.pack(fill="x", padx=12, pady=(0,8))
        cols = ("Rank","Item","Total Qty Sold","Total Revenue")
        tree = ttk.Treeview(tf, columns=cols, show="headings", height=min(10, max(3, len(top_items))))
        for col, w in zip(cols, [60, 280, 120, 130]):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor="center")
        tree.column("Item", anchor="w")
        for i, r in enumerate(top_items):
            tree.insert("","end", values=(f"#{i+1}", r["item_name"],
                                          r["total_qty"], f"{cur}{(r['total_rev'] or 0):.2f}"))
        tree.pack(fill="x")

        # ── RAW MATERIAL PURCHASE SUMMARY ──────────────────
        tk.Label(self.outer, text="📦  Raw Material Purchase Summary", font=FH, bg=BG, fg=DARK).pack(anchor="w", padx=18, pady=(18,4))
        tk.Frame(self.outer, bg=RED, height=2).pack(fill="x", padx=12, pady=(0,6))
        mats = c.execute("SELECT * FROM raw_materials ORDER BY name").fetchall()

        # Purchase totals from inventory_log
        q4 = ("SELECT material_id, SUM(qty) as total_qty, SUM(total_cost) as total_cost "
              "FROM inventory_log WHERE type LIKE '%Add%' OR type LIKE '%add%' GROUP BY material_id")
        add_map = {r["material_id"]: {"qty": r["total_qty"], "cost": r["total_cost"] or 0}
                   for r in c.execute(q4).fetchall()}

        # Grand total purchase
        grand_purchase = sum(v["cost"] for v in add_map.values())
        pass  # conn reused

        # Purchase summary cards
        self.row_frame = tk.Frame(self.outer, bg=BG); self.row_frame.pack(fill="x", padx=6, pady=(0,8))
        self._card("Total Purchase Amount", f"{cur}{grand_purchase:.2f}", color="#c05621", bg="#fffaf0")
        self._card("Total Materials",       str(len(mats)),               color=DARK,      bg="#edf2f7")

        # Material wise table
        mf = tk.Frame(self.outer, bg=BG); mf.pack(fill="x", padx=12, pady=(0,16))
        cols2 = ("Material","Unit","Unit Price","Current Stock","Min Stock","Total Purchased","Total Cost","Status")
        tree2 = ttk.Treeview(mf, columns=cols2, show="headings", height=min(12, max(3, len(mats))))
        widths2 = [150, 60, 90, 110, 90, 120, 110, 100]
        for col, w in zip(cols2, widths2):
            tree2.heading(col, text=col)
            tree2.column(col, width=w, anchor="center")
        tree2.column("Material", anchor="w")
        tree2.tag_configure("low", foreground=RED)
        tree2.tag_configure("ok",  foreground=GREEN)
        hsb2 = ttk.Scrollbar(mf, orient="horizontal", command=tree2.xview)
        tree2.configure(xscrollcommand=hsb2.set)
        hsb2.pack(side="bottom", fill="x")
        for m in mats:
            stock    = m["current_stock"] or 0
            min_s    = m["min_stock"] or 0
            cpu      = m["cost_per_unit"] or 0
            info     = add_map.get(m["id"], {"qty":0,"cost":0})
            status   = "⚠️ Low" if stock <= min_s else "✅ OK"
            tag      = "low" if stock <= min_s else "ok"
            tree2.insert("","end", tags=(tag,), values=(
                m["name"],
                m["unit"],
                f"{cur}{cpu:.2f}",
                f"{stock:.2f}",
                f"{min_s:.2f}",
                f"{info['qty']:.2f}",
                f"{cur}{info['cost']:.2f}",
                status,
            ))
        tree2.pack(fill="x")


# ════════════════════════════════════════════════════════════
#  SETTINGS PAGE
# ════════════════════════════════════════════════════════════

    def _export_excel(self):
        """Reports Excel mein export karo."""
        c = db()
        cur = gset("currency","₹")
        d1, d2 = self._date_range()
        q = "SELECT o.id, o.created_at, o.table_number, o.customer_name, o.payment_method, o.subtotal, o.tax_amount, o.discount_amount, o.total_amount FROM orders o WHERE 1=1"
        params = []
        if d1: q += " AND DATE(o.created_at)>=?"; params.append(d1)
        if d2: q += " AND DATE(o.created_at)<=?"; params.append(d2)
        q += " ORDER BY o.id DESC"
        rows = c.execute(q, params).fetchall()
        pass  # conn reused
        headers = ["Bill No","Date & Time","Table","Customer","Payment","Subtotal","Tax","Discount","Total"]
        data = [(
            f"#{r['id']:04d}", r["created_at"], r["table_number"] or "Takeaway",
            r["customer_name"] or "", r["payment_method"] or "Cash",
            r["subtotal"] or 0, r["tax_amount"] or 0,
            r["discount_amount"] or 0, r["total_amount"] or 0
        ) for r in rows]
        period = self.date_var.get()
        export_to_excel(f"Sales_Report_{period}", headers, data, parent=self)


class SettingsPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        # Header
        hdr = tk.Frame(self, bg=WHITE)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10)
        th.pack(fill="x")
        tk.Label(th, text="⚙️  Shop Settings", font=FH, bg=WHITE, fg=DARK).pack(side="left")

        # Scrollable content — Canvas + Scrollbar
        canv_wrap = tk.Frame(self, bg=BG)
        canv_wrap.pack(fill="both", expand=True)
        canv = tk.Canvas(canv_wrap, bg=BG, highlightthickness=0)
        vsb  = tk.Scrollbar(canv_wrap, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canv.pack(side="left", fill="both", expand=True)
        outer = tk.Frame(canv, bg=BG)
        fid = canv.create_window((0, 0), window=outer, anchor="nw")
        def _on_frame_configure(e):
            canv.configure(scrollregion=canv.bbox("all"))
        def _on_canvas_configure(e):
            canv.itemconfig(fid, width=e.width)
        outer.bind("<Configure>", _on_frame_configure)
        canv.bind("<Configure>", _on_canvas_configure)
        def _mw3(e, _c=canv):
            try: _c.yview_scroll(int(-1*(e.delta/120)), "units")
            except Exception: pass
        canv.bind("<MouseWheel>", _mw3)
        self.bind("<MouseWheel>", _mw3)
        outer.pack_configure = lambda **kw: None  # dummy, outer is window not packed directly
        # Padding frame inside canvas
        outer_pad = tk.Frame(outer, bg=BG)
        outer_pad.pack(fill="both", expand=True, padx=20, pady=16)
        outer = outer_pad

        # Shop Info Card
        card = tk.Frame(outer, bg=WHITE, relief="solid", bd=1, padx=24, pady=18)
        card.pack(fill="x", pady=(0, 14))
        tk.Label(card, text="🏪  Shop ki Jaankari", font=FH, bg=WHITE, fg=DARK).pack(anchor="w", pady=(0, 12))
        tk.Frame(card, bg=RED, height=2).pack(fill="x", pady=(0, 12))

        self.flds = {}
        fields = [
            ("shop_name", "Shop ka Naam *",      "Jaise: Sharma Dhaba"),
            ("address",   "Address",              "Gali, Sheher, PIN"),
            ("phone",     "Phone Number",         "9876543210"),
            ("gst",       "GST Number",           "GSTIN..."),
        ]
        for key, label, hint in fields:
            tk.Label(card, text=label, font=FBD, bg=WHITE, fg=DARK, anchor="w").pack(fill="x", pady=(6,1))
            var = tk.StringVar(value=gset(key, ""))
            e = tk.Entry(card, textvariable=var, font=FB, bg=IBGC, fg=DARK,
                         relief="solid", bd=1)
            e.pack(fill="x", ipady=7)
            tk.Label(card, text=hint, font=FS, bg=WHITE, fg=MUTED, anchor="w").pack(fill="x")
            self.flds[key] = var

        # Billing Settings Card
        card2 = tk.Frame(outer, bg=WHITE, relief="solid", bd=1, padx=24, pady=18)
        card2.pack(fill="x", pady=(0, 14))
        tk.Label(card2, text="💰  Billing Settings", font=FH, bg=WHITE, fg=DARK).pack(anchor="w", pady=(0, 12))
        tk.Frame(card2, bg=RED, height=2).pack(fill="x", pady=(0, 12))

        row1 = tk.Frame(card2, bg=WHITE); row1.pack(fill="x", pady=4)
        # Tax
        tk.Label(row1, text="Default Tax %:", font=FBD, bg=WHITE, fg=DARK, width=18, anchor="w").pack(side="left")
        self.flds["tax"] = tk.StringVar(value=gset("tax", "5"))
        tk.Entry(row1, textvariable=self.flds["tax"], font=FB, bg=IBGC,
                 relief="solid", bd=1, width=10).pack(side="left", padx=(0, 30))
        # Currency
        tk.Label(row1, text="Currency Symbol:", font=FBD, bg=WHITE, fg=DARK, width=18, anchor="w").pack(side="left")
        self.flds["currency"] = tk.StringVar(value=gset("currency", "₹"))
        ttk.Combobox(row1, textvariable=self.flds["currency"],
                     values=["₹", "$", "€", "£", "¥"], width=8,
                     state="readonly").pack(side="left")

        # Thermal Printer Settings
        tk.Frame(card2, bg=BORD, height=1).pack(fill="x", pady=(12,8))
        tk.Label(card2, text="🖨️  Thermal Printer Settings", font=FBD, bg=WHITE, fg=DARK).pack(anchor="w", pady=(0,6))
        trow1 = tk.Frame(card2, bg=WHITE); trow1.pack(fill="x", pady=3)
        tk.Label(trow1, text="Default Print Type:", font=FS, bg=WHITE, fg=DARK, width=18, anchor="w").pack(side="left")
        self.flds["default_print_type"] = tk.StringVar(value=gset("default_print_type","thermal"))
        ttk.Combobox(trow1, textvariable=self.flds["default_print_type"],
                     values=["thermal","pdf"], width=10, state="readonly").pack(side="left", padx=(0,20))
        tk.Label(trow1, text="Paper Width:", font=FS, bg=WHITE, fg=DARK, anchor="w").pack(side="left")
        self.flds["thermal_width"] = tk.StringVar(value=gset("thermal_width","80"))
        ttk.Combobox(trow1, textvariable=self.flds["thermal_width"],
                     values=["58","80"], width=5, state="readonly").pack(side="left")
        tk.Label(trow1, text="mm", font=FS, bg=WHITE, fg=DARK).pack(side="left", padx=4)
        trow2 = tk.Frame(card2, bg=WHITE); trow2.pack(fill="x", pady=3)
        tk.Label(trow2, text="Printer Name:", font=FS, bg=WHITE, fg=DARK, width=18, anchor="w").pack(side="left")
        self.flds["thermal_printer"] = tk.StringVar(value=gset("thermal_printer",""))
        # Try auto-detect default printer
        if not gset("thermal_printer",""):
            try:
                import win32print
                self.flds["thermal_printer"].set(win32print.GetDefaultPrinter())
            except: pass
        tk.Entry(trow2, textvariable=self.flds["thermal_printer"], font=FS, bg=IBGC,
                 relief="solid", bd=1, width=36).pack(side="left", padx=(0,8))
        tk.Label(trow2, text="(Khali = default)", font=("Segoe UI",8), bg=WHITE, fg=MUTED).pack(side="left")

        # License Card
        card3 = tk.Frame(outer, bg=WHITE, relief="solid", bd=1, padx=24, pady=18)
        card3.pack(fill="x", pady=(0, 14))
        tk.Label(card3, text="🔑  License Info", font=FH, bg=WHITE, fg=DARK).pack(anchor="w", pady=(0, 12))
        tk.Frame(card3, bg=RED, height=2).pack(fill="x", pady=(0, 12))

        lic = load_license()
        if lic["valid"]:
            days = lic["days"]
            col = "#c53030" if days <= 7 else WARN if days <= 30 else GREEN
            tk.Label(card3, text=f"✅ License Active — {days} din bache",
                     font=FBD, bg=WHITE, fg=col).pack(anchor="w")
            tk.Label(card3, text=f"Shop: {gset('shop_name','')}  |  Expiry: {lic.get('expiry','')}  |  Type: {lic.get('type','')}",
                     font=FS, bg=WHITE, fg=MED).pack(anchor="w", pady=(4, 0))
        else:
            tk.Label(card3, text=f"⛔ License Inactive: {lic.get('message','')}",
                     font=FBD, bg=WHITE, fg="#c53030").pack(anchor="w")

        lrow = tk.Frame(card3, bg=WHITE); lrow.pack(fill="x", pady=(10, 0))
        tk.Label(lrow, text="Naya Serial:", font=FBD, bg=WHITE, fg=DARK).pack(side="left", padx=(0, 8))
        self.lic_var = tk.StringVar()
        tk.Entry(lrow, textvariable=self.lic_var, font=("Courier New", 10),
                 bg=IBGC, relief="solid", bd=1, width=36).pack(side="left", padx=(0, 8))
        btn(lrow, "🔑 Update License", self._update_license, bg=DARK, px=10).pack(side="left")

        # Save Button
        btn_row = tk.Frame(outer, bg=BG); btn_row.pack(fill="x", pady=4)
        btn(btn_row, "💾  Settings Save Karo", self._save, bg=RED, px=20, py=10).pack(side="left")
        self.status_lbl = tk.Label(btn_row, text="", font=FBD, bg=BG, fg=GREEN)
        self.status_lbl.pack(side="left", padx=16)

    def _save(self):
        if not self.flds["shop_name"].get().strip():
            messagebox.showerror("Error", "Shop naam zaroori hai!"); return
        for k, v in self.flds.items():
            sset(k, v.get().strip())
        # Update window title
        shop = gset("shop_name", "Restaurant")
        self.app.title(f"{shop} — BhugtanEase")
        # Update shop name label in sidebar
        # FIX: font ko string ke saath compare nahi karte - tag ya widget store karke update karo
        for w in self.app.winfo_children():
            if isinstance(w, tk.Frame):
                for child in w.winfo_children():
                    if isinstance(child, tk.Label) and hasattr(child, '_is_shop_label'):
                        child.config(text=shop)
        self.status_lbl.config(text="✅ Settings save ho gayi!")
        self.after(3000, lambda: self.status_lbl.config(text=""))

    def _update_license(self):
        serial = self.lic_var.get().strip()
        if not serial:
            messagebox.showerror("Error", "Serial number likhein!"); return
        r = save_license(serial)
        if r["valid"]:
            messagebox.showinfo("✅ Success", f"License update ho gayi!\n{r['message']}")
            self.app.show("settings")
        else:
            messagebox.showerror("❌ Error", r["message"])


# ════════════════════════════════════════════════════════════
#  BACKUP & RESTORE PAGE
# ════════════════════════════════════════════════════════════
class BackupPage(tk.Frame):
    def __init__(self, parent, app):
        super().__init__(parent, bg=BG)
        self.pack(fill="both", expand=True)
        self.app = app
        self._build()

    def _build(self):
        import datetime as _dt, shutil, os, zipfile, sys, subprocess

        db_path  = DB_PATH
        db_exists = os.path.exists(db_path)
        db_size   = os.path.getsize(db_path) if db_exists else 0

        # ── Header ───────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=WHITE)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=RED, height=3).pack(fill="x")
        th = tk.Frame(hdr, bg=WHITE, padx=16, pady=10)
        th.pack(fill="x")
        tk.Label(th, text="🗄️  Backup & Restore — Data Suraksha",
                 font=FH, bg=WHITE, fg=DARK).pack(side="left")

        # ── Scrollable area ───────────────────────────────────────────────────
        canv_wrap = tk.Frame(self, bg=BG)
        canv_wrap.pack(fill="both", expand=True)
        canv = tk.Canvas(canv_wrap, bg=BG, highlightthickness=0)
        vsb  = tk.Scrollbar(canv_wrap, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canv.pack(side="left", fill="both", expand=True)
        outer_raw = tk.Frame(canv, bg=BG)
        fid = canv.create_window((0,0), window=outer_raw, anchor="nw")
        outer_raw.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))
        canv.bind("<Configure>",      lambda e: canv.itemconfig(fid, width=e.width))
        canv.bind("<MouseWheel>",     lambda e: canv.yview_scroll(int(-1*(e.delta/120)),"units"))
        p = tk.Frame(outer_raw, bg=BG, padx=22, pady=16)
        p.pack(fill="both", expand=True)

        # ── Info banner ───────────────────────────────────────────────────────
        info = tk.Frame(p, bg="#EBF4FF", relief="solid", bd=1)
        info.pack(fill="x", pady=(0,14))
        tk.Label(info,
                 text="⚠️  Laptop kharab ho, data delete ho, ya software delete ho jaye —\n"
                      "Backup se sara data wapas aa jaayega! Roz ya hafta mein ek baar backup lo.",
                 font=FS, bg="#EBF4FF", fg="#1A365D",
                 padx=14, pady=10, justify="left", wraplength=820).pack(anchor="w")

        # ── DB Status card ────────────────────────────────────────────────────
        sc = tk.Frame(p, bg=WHITE, relief="solid", bd=1)
        sc.pack(fill="x", pady=(0,14))
        sc.columnconfigure(0, weight=1); sc.columnconfigure(1, weight=1); sc.columnconfigure(2, weight=1)
        for col, (title, val, color) in enumerate([
            ("Database File", os.path.basename(db_path), RED),
            ("DB Size",       f"{db_size/1024:.1f} KB" if db_exists else "Not Found",
                              GREEN if db_exists else "#c53030"),
            ("Location",      os.path.dirname(db_path)[:40] + ("..." if len(os.path.dirname(db_path))>40 else ""), MED),
        ]):
            f2 = tk.Frame(sc, bg=WHITE); f2.grid(row=0, column=col, sticky="nsew", padx=1)
            tk.Label(f2, text=title, font=FS,  bg=WHITE, fg=MUTED).pack(pady=6)
            tk.Label(f2, text=val,   font=FBD, bg=WHITE, fg=color).pack(pady=(0,10))

        # Read saved settings
        try:
            _last_bk     = gset("last_backup_date", "")
            _saved_folder = gset("auto_backup_folder", "")
        except Exception:
            _last_bk = ""; _saved_folder = ""

        desktop       = os.path.join(os.path.expanduser("~"), "Desktop")
        default_folder = _saved_folder if (_saved_folder and os.path.isdir(_saved_folder)) else desktop

        # ══════════════════════════════════════════
        # SECTION 1 — BACKUP BANAO
        # ══════════════════════════════════════════
        bk_card = tk.Frame(p, bg=WHITE, relief="solid", bd=1, padx=22, pady=18)
        bk_card.pack(fill="x", pady=(0,14))
        tk.Label(bk_card, text="📦  Backup Banao (Local / Pendrive)",
                 font=FH, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Frame(bk_card, bg=RED, height=2).pack(fill="x", pady=(6,12))

        now_str      = _dt.datetime.now().strftime("%Y-%m-%d_%H-%M")
        default_name = f"BhugtanEase_Backup_{now_str}.zip"

        r1 = tk.Frame(bk_card, bg=WHITE); r1.pack(fill="x", pady=4)
        tk.Label(r1, text="Backup folder:", font=FBD, bg=WHITE, fg=DARK, width=16, anchor="w").pack(side="left")
        v_folder = tk.StringVar(value=default_folder)
        tk.Entry(r1, textvariable=v_folder, font=FS, bg=IBGC,
                 relief="solid", bd=1, width=50).pack(side="left", padx=(6,6))
        def _browse():
            from tkinter import filedialog
            f = filedialog.askdirectory(title="Backup kahaan save karein?")
            if f: v_folder.set(f)
        btn(r1, "📁 Browse", _browse, bg=DARK, px=10).pack(side="left")

        r2 = tk.Frame(bk_card, bg=WHITE); r2.pack(fill="x", pady=4)
        tk.Label(r2, text="File naam:", font=FBD, bg=WHITE, fg=DARK, width=16, anchor="w").pack(side="left")
        v_bkname = tk.StringVar(value=default_name)
        tk.Entry(r2, textvariable=v_bkname, font=FS, bg=IBGC,
                 relief="solid", bd=1, width=42).pack(side="left", padx=(6,0))

        bk_msg = tk.StringVar(value="")
        bk_lbl = tk.Label(bk_card, textvariable=bk_msg, font=FBD,
                           bg=WHITE, fg=GREEN, wraplength=820, justify="left")
        bk_lbl.pack(anchor="w", pady=(6,0))

        def do_backup():
            folder = v_folder.get().strip()
            fname  = v_bkname.get().strip() or default_name
            if not os.path.isdir(folder):
                bk_msg.set("❌ Folder nahi mila — Browse se select karo!")
                bk_lbl.config(fg="#c53030"); return
            if not fname.endswith(".zip"): fname += ".zip"
            zip_path = os.path.join(folder, fname)
            try:
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.write(db_path, os.path.basename(db_path))
                size_kb = os.path.getsize(zip_path) / 1024
                sset("last_backup_date",  _dt.date.today().isoformat())
                sset("auto_backup_folder", folder)
                bk_msg.set(f"✅ Backup ready!   📁 {zip_path}   💾 {size_kb:.1f} KB")
                bk_lbl.config(fg=GREEN)
                if sys.platform == "win32":
                    subprocess.Popen(f'explorer /select,"{zip_path}"', shell=True)
                load_history()
            except Exception as e:
                bk_msg.set(f"❌ Backup fail: {e}"); bk_lbl.config(fg="#c53030")

        btn(bk_card, "🗄️  Backup Banao", do_backup, bg=GREEN, px=20, py=10).pack(anchor="w", pady=(10,0))

        # ══════════════════════════════════════════
        # SECTION 2 — GOOGLE DRIVE BACKUP
        # ══════════════════════════════════════════
        gd_card = tk.Frame(p, bg=WHITE, relief="solid", bd=1, padx=22, pady=18)
        gd_card.pack(fill="x", pady=(0,14))
        tk.Label(gd_card, text="☁️  Google Drive Backup",
                 font=FH, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Frame(gd_card, bg="#1a73e8", height=2).pack(fill="x", pady=(6,12))

        def _find_gdrive():
            username = os.environ.get("USERNAME","")
            candidates = [
                "G:\\My Drive", "G:\\",
                f"C:\\Users\\{username}\\Google Drive",
                f"C:\\Users\\{username}\\My Drive",
                "D:\\Google Drive", "D:\\My Drive",
            ]
            try:
                import winreg
                key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                    r"SOFTWARE\Google\DriveFS", 0, winreg.KEY_READ)
                val, _ = winreg.QueryValueEx(key, "DefaultMountPoint")
                winreg.CloseKey(key)
                if val: candidates.insert(0, os.path.join(val, "My Drive")); candidates.insert(0, val)
            except Exception: pass
            for c in candidates:
                if c and os.path.isdir(c): return c
            return ""

        gd_path       = gset("gdrive_folder","") or _find_gdrive()
        gd_found      = bool(gd_path and os.path.isdir(gd_path))
        gd_status_txt = f"✅ Google Drive mila: {gd_path}" if gd_found else "❌ Google Drive nahi mila — manually select karo"
        gd_info_lbl   = tk.Label(gd_card, text=gd_status_txt, font=FBD,
                                  bg=WHITE, fg=GREEN if gd_found else "#c53030")
        gd_info_lbl.pack(anchor="w", pady=(0,8))

        gr1 = tk.Frame(gd_card, bg=WHITE); gr1.pack(fill="x", pady=4)
        tk.Label(gr1, text="Drive folder:", font=FBD, bg=WHITE, fg=DARK, width=16, anchor="w").pack(side="left")
        v_gd = tk.StringVar(value=gd_path if gd_found else "")
        tk.Entry(gr1, textvariable=v_gd, font=FS, bg=IBGC,
                 relief="solid", bd=1, width=50).pack(side="left", padx=(6,6))
        def _browse_gd():
            from tkinter import filedialog
            f = filedialog.askdirectory(title="Google Drive folder select karo",
                                         initialdir=gd_path or "G:\\")
            if f: v_gd.set(f)
        btn(gr1, "📁 Browse", _browse_gd, bg="#1a73e8", px=10).pack(side="left")

        gr2 = tk.Frame(gd_card, bg=WHITE); gr2.pack(fill="x", pady=4)
        tk.Label(gr2, text="Sub-folder:", font=FBD, bg=WHITE, fg=DARK, width=16, anchor="w").pack(side="left")
        v_gd_sub = tk.StringVar(value="BhugtanEase_Backup")
        tk.Entry(gr2, textvariable=v_gd_sub, font=FS, bg=IBGC,
                 relief="solid", bd=1, width=28).pack(side="left", padx=(6,0))
        tk.Label(gr2, text="(Drive mein is naam ka folder banega)",
                 font=("Segoe UI",8), bg=WHITE, fg=MUTED).pack(side="left", padx=8)

        gd_msg = tk.StringVar(value="")
        gd_lbl = tk.Label(gd_card, textvariable=gd_msg, font=FBD,
                           bg=WHITE, fg=GREEN, wraplength=820, justify="left")
        gd_lbl.pack(anchor="w", pady=(6,0))

        def backup_to_gdrive():
            root_f = v_gd.get().strip()
            sub    = v_gd_sub.get().strip() or "BhugtanEase_Backup"
            if not root_f or not os.path.isdir(root_f):
                gd_msg.set("❌ Google Drive folder nahi mila! Browse se select karo.")
                gd_lbl.config(fg="#c53030"); return
            save_dir = os.path.join(root_f, sub)
            try: os.makedirs(save_dir, exist_ok=True)
            except Exception as e:
                gd_msg.set(f"❌ Folder nahi bana: {e}"); gd_lbl.config(fg="#c53030"); return
            ts       = _dt.datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_path = os.path.join(save_dir, f"BhugtanEase_{ts}.zip")
            try:
                with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.write(db_path, os.path.basename(db_path))
                size_kb = os.path.getsize(zip_path) / 1024
                sset("last_backup_date", _dt.date.today().isoformat())
                sset("gdrive_folder",    save_dir)
                gd_msg.set(
                    f"✅ Google Drive backup ho gaya!\n"
                    f"📁 {zip_path}   ☁️  Thodi der mein sync hoga — {size_kb:.1f} KB")
                gd_lbl.config(fg=GREEN)
                if sys.platform == "win32":
                    subprocess.Popen(f'explorer /select,"{zip_path}"', shell=True)
                load_history()
            except Exception as e:
                gd_msg.set(f"❌ Backup fail: {e}"); gd_lbl.config(fg="#c53030")

        btn_row_gd = tk.Frame(gd_card, bg=WHITE); btn_row_gd.pack(anchor="w", pady=(10,0))
        btn(btn_row_gd, "☁️  Google Drive pe Backup Karo",
            backup_to_gdrive, bg="#1a73e8", px=20, py=10).pack(side="left")
        def _redetect():
            found = _find_gdrive()
            v_gd.set(found or "")
            gd_info_lbl.config(
                text=f"✅ Google Drive mila: {found}" if found else "❌ Nahi mila",
                fg=GREEN if found else "#c53030")
        btn(btn_row_gd, "🔍 Drive Dhundho", _redetect, bg=DARK, px=10).pack(side="left", padx=8)

        # ══════════════════════════════════════════
        # SECTION 3 — RESTORE
        # ══════════════════════════════════════════
        rs_card = tk.Frame(p, bg=WHITE, relief="solid", bd=1, padx=22, pady=18)
        rs_card.pack(fill="x", pady=(0,14))
        tk.Label(rs_card, text="♻️  Restore — Backup se Data Wapas Lao",
                 font=FH, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Frame(rs_card, bg="#c53030", height=2).pack(fill="x", pady=(6,12))

        tk.Label(rs_card,
                 text="⚠️  WARNING: Restore karne se current data REPLACE ho jaayega!\n"
                      "Restore se pehle ek fresh backup zaroor le lo.",
                 font=FS, bg="#FFF5F5", fg="#9B2C2C",
                 padx=12, pady=8, wraplength=820, justify="left").pack(fill="x", pady=(0,10))

        rs_msg = tk.StringVar(value="")
        rs_lbl = tk.Label(rs_card, textvariable=rs_msg, font=FBD,
                           bg=WHITE, wraplength=820, justify="left")
        rs_lbl.pack(anchor="w", pady=(0,8))

        def do_restore():
            from tkinter import filedialog
            zip_path = filedialog.askopenfilename(
                title="Backup ZIP file select karo",
                filetypes=[("ZIP Backup","*.zip"),("All Files","*.*")])
            if not zip_path: return
            try:
                with zipfile.ZipFile(zip_path,"r") as zf:
                    db_in_zip = [n for n in zf.namelist() if n.endswith(".db")]
            except Exception as e:
                rs_msg.set(f"❌ Invalid file: {e}"); rs_lbl.config(fg="#c53030"); return
            if not db_in_zip:
                rs_msg.set("❌ Is ZIP mein koi .db file nahi mili!"); rs_lbl.config(fg="#c53030"); return
            if not messagebox.askyesno("⚠️ Confirm Restore",
                f"File: {os.path.basename(zip_path)}\n\n"
                "Restore karne se CURRENT DATA REPLACE hoga!\n\nKya aap sure hain?",
                icon="warning"): return
            # Safety backup pehle
            safety = os.path.join(os.path.dirname(db_path),
                                   f"SAFETY_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.zip")
            try:
                with zipfile.ZipFile(safety,"w",zipfile.ZIP_DEFLATED) as zf:
                    zf.write(db_path, os.path.basename(db_path))
            except Exception: pass
            try:
                tmp = os.path.join(os.path.dirname(db_path),"__tmp_restore__.db")
                with zipfile.ZipFile(zip_path,"r") as zf:
                    with zf.open(db_in_zip[0]) as src, open(tmp,"wb") as dst:
                        dst.write(src.read())
                shutil.move(tmp, db_path)
                rs_msg.set("✅ Restore ho gaya! Ab software band karke dobara kholo.")
                rs_lbl.config(fg=GREEN)
                messagebox.showinfo("✅ Restore Done!",
                    "Data restore ho gaya!\n\nAb software BAND karke DOBARA kholo.")
            except Exception as e:
                rs_msg.set(f"❌ Restore fail: {e}"); rs_lbl.config(fg="#c53030")

        btn(rs_card, "♻️  Backup se Restore Karo", do_restore, bg="#c53030", px=20, py=10).pack(anchor="w")

        # ══════════════════════════════════════════
        # SECTION 4 — RECENT BACKUPS
        # ══════════════════════════════════════════
        hist_card = tk.Frame(p, bg=WHITE, relief="solid", bd=1, padx=22, pady=18)
        hist_card.pack(fill="x", pady=(0,14))
        tk.Label(hist_card, text="📋  Recent Backups",
                 font=FH, bg=WHITE, fg=DARK).pack(anchor="w")
        tk.Frame(hist_card, bg=RED, height=2).pack(fill="x", pady=(6,12))

        # Last backup status
        last_bk_txt   = f"Aakhri backup: {_last_bk}" if _last_bk else "⚠️  Kabhi backup nahi liya!"
        last_bk_color = "#c53030" if not _last_bk else (GREEN if _last_bk >= _dt.date.today().isoformat() else WARN)
        tk.Label(hist_card, text=last_bk_txt, font=FBD,
                 bg=WHITE, fg=last_bk_color).pack(anchor="w", pady=(0,10))

        # Table header
        cols   = [("File Name",50), ("Size",10), ("Date",18), ("Open",6)]
        tbl_hdr = tk.Frame(hist_card, bg=DARK)
        tbl_hdr.pack(fill="x")
        for col_name, w in cols:
            tk.Label(tbl_hdr, text=col_name, font=FBD, bg=DARK, fg=WHITE,
                     width=w, anchor="w", padx=8, pady=6).pack(side="left")

        tbl_body = tk.Frame(hist_card, bg=WHITE)
        tbl_body.pack(fill="x")

        def load_history():
            for w in tbl_body.winfo_children(): w.destroy()
            folder = v_folder.get().strip()
            if not folder or not os.path.isdir(folder):
                tk.Label(tbl_body, text="Folder select karo upar se.",
                         font=FS, bg=WHITE, fg=MUTED, pady=8).pack(anchor="w", padx=10)
                return
            try:
                bk_files = sorted(
                    [f for f in os.listdir(folder)
                     if f.startswith("BhugtanEase") and f.endswith(".zip")],
                    reverse=True)[:10]
                if not bk_files:
                    tk.Label(tbl_body, text="Is folder mein koi backup nahi mila.",
                             font=FS, bg=WHITE, fg=MUTED, pady=10).pack(anchor="w", padx=10)
                    return
                for i, fn in enumerate(bk_files):
                    fp    = os.path.join(folder, fn)
                    sz    = f"{os.path.getsize(fp)/1024:.1f} KB"
                    mtime = _dt.datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%d-%b-%Y %H:%M")
                    bg    = WHITE if i%2==0 else "#F7FAFC"
                    row   = tk.Frame(tbl_body, bg=bg); row.pack(fill="x")
                    tk.Label(row, text=fn,    font=FS, bg=bg, fg=DARK,   width=50, anchor="w", padx=8, pady=5).pack(side="left")
                    tk.Label(row, text=sz,    font=FS, bg=bg, fg=MED,    width=10, anchor="w", padx=8).pack(side="left")
                    tk.Label(row, text=mtime, font=FS, bg=bg, fg=MED,    width=18, anchor="w", padx=8).pack(side="left")
                    tk.Button(row, text="📁", font=FS, bg="#1a73e8", fg=WHITE,
                              relief="flat", cursor="hand2", bd=0, padx=8,
                              command=lambda _fp=fp: subprocess.Popen(
                                  f'explorer /select,"{_fp}"', shell=True)
                              ).pack(side="left", padx=4)
            except Exception as e:
                tk.Label(tbl_body, text=f"Error: {e}", font=FS,
                         bg=WHITE, fg="#c53030", pady=8).pack(anchor="w", padx=10)

        btn(hist_card, "🔄 Refresh", load_history, bg=DARK, px=12, py=6).pack(anchor="w", pady=(0,8))
        load_history()


# ════════════════════════════════════════════════════════════
#  LAUNCHER
# ════════════════════════════════════════════════════════════
def start():
    init_db()
    _load_settings_cache()   # SPEED: ek baar saari settings load karo

    # License check
    lic = load_license()
    if not lic["valid"]:
        act = LicenseScreen(lic)
        act.mainloop()
        if not act.activated: return
        lic = load_license()
        if not lic["valid"]: return

    # Setup wizard
    if gset("setup_done") != "1":
        wiz = SetupWizard()
        wiz.mainloop()
        if not wiz.done: return
        _load_settings_cache()  # FIX: wizard ke baad cache refresh karo

    # Main app — license pass karo taaki sidebar dobara load na kare
    app = App(lic)
    app.mainloop()

if __name__ == "__main__":
    # Root window — screen se bahar rakho, bilkul invisible
    _root = tk.Tk()
    _root.overrideredirect(True)          # koi border/titlebar nahi
    _root.geometry("1x1+-10000+-10000")   # screen se bilkul bahar
    _root.attributes("-alpha", 0)         # transparent
    _root.update_idletasks()              # turant apply karo

    # Safe check helper — destroy ke baad winfo_exists() crash karti hai
    def _alive():
        try:
            return bool(_root.winfo_exists())
        except Exception:
            return False

    def _quit():
        try:
            _root.destroy()
        except Exception:
            pass

    # Seedha launch — no after() delay
    init_db()
    _load_settings_cache()
    _ensure_default_admin()
    lic = load_license()

    # LICENSE CHECK
    if not lic["valid"]:
        act = LicenseScreen(_root, lic)
        _root.wait_window(act)
        if not act.activated:
            _quit()
        else:
            lic = load_license()
            if not lic["valid"]:
                _quit()

    if _alive():
        _load_settings_cache()

        # LOGIN CHECK
        login = LoginScreen(_root)
        _root.wait_window(login)
        if not login.logged_in:
            _quit()

    if _alive():
        # SETUP WIZARD
        if gset("setup_done") != "1":
            wiz = SetupWizard(_root)
            _root.wait_window(wiz)
            if not wiz.done:
                _quit()
            else:
                _load_settings_cache()

    if _alive():
        # MAIN APP
        _root.overrideredirect(False)
        _root.attributes("-alpha", 1)
        _root.geometry("")
        _root.deiconify()
        app = App(_root, lic)
        _root.mainloop()
